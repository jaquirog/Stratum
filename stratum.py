#!/usr/bin/env python3
# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                                                                          ║
# ║   ███████╗████████╗██████╗  █████╗ ████████╗██╗   ██╗███╗   ███╗       ║
# ║   ██╔════╝╚══██╔══╝██╔══██╗██╔══██╗╚══██╔══╝██║   ██║████╗ ████║       ║
# ║   ███████╗   ██║   ██████╔╝███████║   ██║   ██║   ██║██╔████╔██║       ║
# ║   ╚════██║   ██║   ██╔══██╗██╔══██║   ██║   ██║   ██║██║╚██╔╝██║       ║
# ║   ███████║   ██║   ██║  ██║██║  ██║   ██║   ╚██████╔╝██║ ╚═╝ ██║       ║
# ║   ╚══════╝   ╚═╝   ╚═╝  ╚═╝╚═╝  ╚═╝   ╚═╝    ╚═════╝ ╚═╝     ╚═╝       ║
# ║                                                                          ║
# ║   System Intelligence Engine  v1.0.0                                    ║
# ║   Autor : Javier Alfonso Quiroga Sarmiento                              ║
# ║   Marca : Estratega de Negocios: Arquitecto y Ejecutor de               ║
# ║           Ecosistemas Tecnológicos | Data Analytics |                   ║
# ║           Financial BI | ERP Solutions                                  ║
# ║                                                                          ║
# ║   Manifiesto: STRATUM no reemplaza la inteligencia artificial.          ║
# ║   La potencia. Somos aliados de la IA, no su competencia.               ║
# ║   STRATUM es el traductor entre tu proyecto real y Claude/Gemini.       ║
# ║                                                                          ║
# ║   Uso:                                                                   ║
# ║     python3 stratum.py --ruta /tu/proyecto                              ║
# ║     python3 stratum.py --ruta . --puerto 7474                           ║
# ║     python3 stratum.py --ruta . --red   (modo red local)                ║
# ║                                                                          ║
# ║   Seguridad: Este programa no realiza ninguna llamada a internet.       ║
# ║   Todo el procesamiento ocurre localmente en tu máquina.                ║
# ║   Audita este archivo para verificarlo.                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 0 — IMPORTS Y DEPENDENCIAS (solo stdlib de Python)
# Sin dependencias externas requeridas para el núcleo de STRATUM
# ════════════════════════════════════════════════════════════════════════════

import os
import unittest
import tempfile
import sys
import ast
import re
import json
import time
import uuid
import math
import hmac
import shutil
import hashlib
import logging
import argparse
import zipfile
import threading
import webbrowser
import http.server
import urllib.parse
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict, Counter
from typing import Optional

# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 1 — CONFIGURACIÓN GLOBAL Y CONSTANTES
# Todas las constantes del sistema en un solo lugar para fácil mantenimiento
# ════════════════════════════════════════════════════════════════════════════

# ─── Identificación del producto ────────────────────────────────────────────
STRATUM_VERSION        = "1.0.0"
STRATUM_NOMBRE         = "STRATUM System Intelligence Engine"
STRATUM_AUTOR          = "Javier Alfonso Quiroga Sarmiento"
STRATUM_MARCA          = (
    "Estratega de Negocios: Arquitecto y Ejecutor de Ecosistemas "
    "Tecnológicos | Data Analytics | Financial BI | ERP Solutions"
)
STRATUM_MANIFIESTO     = (
    "STRATUM no reemplaza la inteligencia artificial. La potencia. "
    "Somos aliados de la IA, no su competencia."
)

# ─── Configuración de red y servidor ────────────────────────────────────────
HOST_LOCAL             = "127.0.0.1"   # Nunca expuesto a internet
HOST_RED               = "0.0.0.0"    # Solo con flag --red explícito
PUERTO_DEFAULT         = 7474
ABRIR_NAVEGADOR        = True

# ─── Sistema de licencias y lanzamiento ─────────────────────────────────────
# IMPORTANTE: Cambiar LAUNCH_DATE al día real de lanzamiento público
LAUNCH_DATE            = "2025-12-01"  # Fecha de lanzamiento oficial
FREE_PERIOD_DAYS       = 365           # 12 meses completamente gratis
WARNING_DAYS_BEFORE    = 60            # Aviso de fin de período gratuito
LICENSE_FILE           = ".stratum_license.dat"

# ─── Precios por región (USD y COP) ─────────────────────────────────────────
PRECIOS = {
    "usa_canada_europa": {
        "starter":      {"mes": 29,  "anio": 290},
        "professional": {"mes": 79,  "anio": 790},
        "enterprise":   {"mes": 199, "anio": 1990},
        "donacion":     5,
        "moneda":       "USD",
    },
    "latinoamerica": {
        "starter":      {"mes": 12,  "anio": 120},
        "professional": {"mes": 29,  "anio": 290},
        "enterprise":   {"mes": 69,  "anio": 690},
        "donacion":     3,
        "moneda":       "USD",
    },
    "colombia": {
        "starter":      {"mes": 35_000,  "anio": 350_000},
        "professional": {"mes": 89_000,  "anio": 890_000},
        "enterprise":   {"mes": 199_000, "anio": 1_990_000},
        "donacion":     12_000,
        "moneda":       "COP",
    },
}

# ─── Límites por tier ────────────────────────────────────────────────────────
LIMITES_TIER = {
    "trial":        {"max_archivos": 9999, "max_tablas": 9999, "sesiones": 3},
    "free":         {"max_archivos": 9999, "max_tablas": 9999, "sesiones": -1},
    "starter":      {"max_archivos": 100,  "max_tablas": 10,   "sesiones": -1},
    "professional": {"max_archivos": 500,  "max_tablas": 50,   "sesiones": -1},
    "enterprise":   {"max_archivos": -1,   "max_tablas": -1,   "sesiones": -1},
    "early_free":   {"max_archivos": 9999, "max_tablas": 9999, "sesiones": -1},
}

# ─── Carpetas ignoradas en el escaneo ───────────────────────────────────────
CARPETAS_IGNORADAS = {
    ".git", "__pycache__", "venv", ".venv", "env", ".env",
    "node_modules", ".idea", ".vscode", "dist", "build",
    ".mypy_cache", ".pytest_cache", ".tox", ".eggs",
    "htmlcov", "site-packages", ".next", "coverage",
    ".nuxt", "target", "vendor", "tmp", "temp", "logs",
    ".stratum_output",
}

# ─── Extensiones válidas por categoría ──────────────────────────────────────
EXTENSIONES_CODIGO = {
    ".py", ".js", ".jsx", ".ts", ".tsx",
    ".java", ".cs", ".php", ".rb", ".r", ".R",
}
EXTENSIONES_DATOS = {
    ".sql", ".json", ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".env.example",
}
EXTENSIONES_WEB = {
    ".html", ".htm", ".css", ".scss", ".sass",
}
EXTENSIONES_DOCS = {
    ".md", ".rst", ".txt",
}
EXTENSIONES_EXCEL = {
    ".xlsx", ".xlsm", ".xls", ".csv",
}
EXTENSIONES_SCRIPTS = {
    ".sh", ".bash", ".zsh", ".bat", ".ps1",
}
EXTENSIONES_CLOUD = {
    ".tf", ".hcl", ".dockerfile", ".Dockerfile",
    ".graphql", ".prisma", ".proto",
}

EXTENSIONES_VALIDAS = (
    EXTENSIONES_CODIGO | EXTENSIONES_DATOS | EXTENSIONES_WEB |
    EXTENSIONES_DOCS   | EXTENSIONES_EXCEL | EXTENSIONES_SCRIPTS |
    EXTENSIONES_CLOUD
)

# ─── Archivos ignorados explícitamente ──────────────────────────────────────
ARCHIVOS_IGNORADOS = {
    ".env", ".env.local", ".env.production", ".env.development",
    "package-lock.json", "yarn.lock", "poetry.lock", "Pipfile.lock",
    ".gitignore", ".gitattributes", ".editorconfig",
    "contexto_proyecto.txt", ".stratum_license.dat",
}

# ─── Límites de tamaño para archivos ────────────────────────────────────────
MAX_MB_POR_ARCHIVO     = 5      # MB máximo por archivo individual
MAX_MB_TOTAL_CONTEXTO  = 10     # MB máximo del contexto exportado para Claude
MAX_LINEAS_PREVIEW     = 200    # Líneas máximas en preview de archivos grandes

# ─── Configuración de logging ────────────────────────────────────────────────
LOG_FILE               = ".stratum_audit.log"
LOG_LEVEL              = logging.INFO


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 2 — SEGURIDAD Y AUTOVERIFICACIÓN
# SHA-256 del archivo propio + logging de auditoría
# El servidor NUNCA escucha en una IP externa sin permiso explícito
# ════════════════════════════════════════════════════════════════════════════

class SeguridadStratum:
    """
    Motor de seguridad de STRATUM.
    Responsabilidades:
      - Calcular y verificar el hash SHA-256 del propio script
      - Registrar cada operación en el log de auditoría local
      - Detectar manipulación del archivo ejecutable
      - Garantizar que el servidor solo escuche en localhost
    """

    # ─── Configurar el sistema de logging de auditoría ──────────────────────
    @staticmethod
    def configurar_logging(ruta_trabajo: Path) -> logging.Logger:
        """
        Crea el logger de auditoría local.
        El log vive en la carpeta de trabajo, nunca en el proyecto analizado.
        """
        logger = logging.getLogger("STRATUM_AUDIT")
        logger.setLevel(LOG_LEVEL)

        # Evitar duplicar handlers si se llama múltiples veces
        if logger.handlers:
            return logger

        # Handler para archivo de auditoría local
        log_path = ruta_trabajo / LOG_FILE
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setLevel(LOG_LEVEL)

        # Handler para consola
        ch = logging.StreamHandler()
        ch.setLevel(logging.WARNING)

        # Formato de log con timestamp
        fmt = logging.Formatter(
            "%(asctime)s | %(levelname)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S"
        )
        fh.setFormatter(fmt)
        ch.setFormatter(fmt)

        logger.addHandler(fh)
        logger.addHandler(ch)
        return logger

    # ─── Calcular el hash SHA-256 del propio archivo ─────────────────────────
    @staticmethod
    def calcular_hash_propio() -> str:
        """
        Lee el archivo stratum.py actual y calcula su SHA-256.
        Este hash puede ser publicado oficialmente para que los clientes
        verifiquen que el archivo no fue alterado.
        """
        ruta_propia = Path(__file__).resolve()
        sha256 = hashlib.sha256()
        with open(ruta_propia, "rb") as f:
            for bloque in iter(lambda: f.read(65536), b""):
                sha256.update(bloque)
        return sha256.hexdigest()

    # ─── Verificar integridad del archivo contra hash conocido ───────────────
    @staticmethod
    def verificar_integridad(hash_conocido: Optional[str] = None) -> dict:
        """
        Compara el hash actual con un hash de referencia.
        Si no se provee hash_conocido, solo reporta el hash actual.
        Retorna: dict con hash_actual, verificado, mensaje
        """
        hash_actual = SeguridadStratum.calcular_hash_propio()

        if hash_conocido is None:
            return {
                "hash_actual":  hash_actual,
                "verificado":   None,
                "mensaje":      "Hash calculado. No se proporcionó hash de referencia.",
            }

        # Comparación segura que evita timing attacks
        verificado = hmac.compare_digest(hash_actual, hash_conocido.lower())
        return {
            "hash_actual": hash_actual,
            "verificado":  verificado,
            "mensaje":     "✅ Archivo íntegro" if verificado else "⚠️  Hash no coincide — archivo posiblemente modificado",
        }

    # ─── Generar ID único de máquina para licenciamiento ─────────────────────
    @staticmethod
    def obtener_machine_id() -> str:
        """
        Genera un identificador único y estable para esta máquina.
        Basado en características del sistema, no en datos personales.
        """
        componentes = [
            str(os.cpu_count()),
            sys.platform,
            str(Path.home()),
        ]
        raw = "|".join(componentes)
        return hashlib.sha256(raw.encode()).hexdigest()[:32]

    # ─── Validar que el host del servidor sea seguro ─────────────────────────
    @staticmethod
    def validar_host(host: str, modo_red: bool) -> str:
        """
        Garantiza que el servidor nunca escuche en red externa
        sin que el usuario haya solicitado explícitamente --red.
        """
        if not modo_red:
            return HOST_LOCAL  # Siempre localhost por defecto

        # Modo red: solo en red local, nunca en internet
        print("\n⚠️  MODO RED ACTIVADO")
        print("   STRATUM escuchará en la red local de este equipo.")
        print("   Asegúrese de estar en una red de confianza.")
        print("   Sus proyectos NO serán enviados a internet.\n")
        return HOST_RED


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 3 — SISTEMA DE LICENCIAS Y RELOJ DE LANZAMIENTO
# Controla: período gratuito, early adopters, tiers, descuentos
# Todo opera localmente — no hay servidor de validación en v1.0
# ════════════════════════════════════════════════════════════════════════════

class SistemaLicencias:
    """
    Motor de licenciamiento local de STRATUM.
    Responsabilidades:
      - Controlar el período gratuito de 12 meses
      - Identificar early adopters (instalaron durante período free)
      - Gestionar sesiones de trial
      - Aplicar descuento permanente del 50% a early adopters
      - Detectar región del usuario para precios PPP
    """

    def __init__(self, ruta_trabajo: Path):
        # ─── Ruta del archivo de licencia cifrado ───────────────────────────
        self.ruta_licencia = ruta_trabajo / LICENSE_FILE
        self.machine_id    = SeguridadStratum.obtener_machine_id()
        self.datos         = self._cargar_o_crear_licencia()

    # ─── Cargar licencia existente o crear una nueva ─────────────────────────
    def _cargar_o_crear_licencia(self) -> dict:
        """
        Lee el archivo de licencia si existe.
        Si no existe, crea uno nuevo para este usuario.
        El archivo usa JSON con checksum para detectar manipulación.
        """
        if self.ruta_licencia.exists():
            try:
                contenido = self.ruta_licencia.read_text(encoding="utf-8")
                datos_raw = json.loads(contenido)

                # Verificar integridad del archivo de licencia
                checksum_guardado  = datos_raw.pop("_checksum", "")
                checksum_calculado = self._calcular_checksum(datos_raw)

                if not hmac.compare_digest(checksum_guardado, checksum_calculado):
                    print("⚠️  Licencia modificada externamente. Reiniciando...")
                    return self._crear_licencia_nueva()

                return datos_raw

            except (json.JSONDecodeError, KeyError):
                return self._crear_licencia_nueva()

        return self._crear_licencia_nueva()

    # ─── Crear licencia nueva para primera instalación ───────────────────────
    def _crear_licencia_nueva(self) -> dict:
        """
        Inicializa los datos de licencia para un nuevo usuario.
        Registra la fecha de instalación para identificar early adopters.
        """
        datos = {
            "machine_id":       self.machine_id,
            "install_date":     date.today().isoformat(),
            "tier":             "free",
            "trial_sesiones":   0,
            "historial":        [],
            "license_key":      None,
            "expiry":           None,
            "region":           self._detectar_region(),
            "early_adopter":    self._es_early_adopter(date.today().isoformat()),
        }
        self._guardar_licencia(datos)
        return datos

    # ─── Guardar licencia con checksum anti-manipulación ─────────────────────
    def _guardar_licencia(self, datos: dict) -> None:
        """
        Escribe el archivo de licencia con checksum para detectar edición manual.
        """
        datos_con_checksum = dict(datos)
        datos_con_checksum["_checksum"] = self._calcular_checksum(datos)
        self.ruta_licencia.write_text(
            json.dumps(datos_con_checksum, indent=2, ensure_ascii=False),
            encoding="utf-8"
        )

    # ─── Calcular checksum de los datos de licencia ──────────────────────────
    def _calcular_checksum(self, datos: dict) -> str:
        """
        Genera un hash de los datos de licencia para detectar manipulación.
        Usa el machine_id como salt para que el checksum sea único por máquina.
        """
        contenido = json.dumps(datos, sort_keys=True, ensure_ascii=False)
        salted    = f"{self.machine_id}:{contenido}"
        return hashlib.sha256(salted.encode()).hexdigest()

    # ─── Detectar si es early adopter ────────────────────────────────────────
    def _es_early_adopter(self, install_date_str: str) -> bool:
        """
        Un early adopter es quien instaló STRATUM durante el período gratuito.
        Recibirá 50% de descuento permanente cuando comience el cobro.
        """
        try:
            install_date = date.fromisoformat(install_date_str)
            launch_date  = date.fromisoformat(LAUNCH_DATE)
            fin_periodo  = date.fromordinal(
                launch_date.toordinal() + FREE_PERIOD_DAYS
            )
            return install_date <= fin_periodo
        except ValueError:
            return False

    # ─── Detectar región del usuario ─────────────────────────────────────────
    def _detectar_region(self) -> str:
        """
        Detecta la región basándose en el sistema operativo y locale.
        El usuario puede cambiarla manualmente en la interfaz.
        Por defecto Colombia para el mercado principal del autor.
        """
        try:
            import locale
            loc = locale.getlocale()[0] or ""
            if "es_CO" in loc:
                return "colombia"
            if any(x in loc for x in ["es_", "pt_BR", "pt_PT"]):
                return "latinoamerica"
        except Exception:
            pass
        return "colombia"  # Default: mercado principal

    # ─── Calcular días desde el lanzamiento ──────────────────────────────────
    def dias_desde_lanzamiento(self) -> int:
        """
        Calcula cuántos días han pasado desde la fecha oficial de lanzamiento.
        Negativo si STRATUM aún no ha sido lanzado oficialmente.
        """
        try:
            launch = date.fromisoformat(LAUNCH_DATE)
            return (date.today() - launch).days
        except ValueError:
            return 0

    # ─── Obtener estado actual del sistema de licencias ──────────────────────
    def obtener_estado(self) -> dict:
        """
        Retorna el estado completo de la licencia actual.
        Incluye: tier, días restantes de gratuidad, descuento, precios.
        """
        dias = self.dias_desde_lanzamiento()
        region = self.datos.get("region", "colombia")
        es_early = self.datos.get("early_adopter", False)

        # Determinar modo actual del sistema
        if dias < 0:
            # Antes del lanzamiento oficial — modo desarrollo
            modo = "pre_launch"
            mensaje = "Versión de desarrollo — lanzamiento próximo"

        elif dias < (FREE_PERIOD_DAYS - WARNING_DAYS_BEFORE):
            # Período gratuito activo
            dias_restantes = FREE_PERIOD_DAYS - dias
            modo = "gratuito"
            mensaje = f"Período gratuito activo — {dias_restantes} días restantes"

        elif dias < FREE_PERIOD_DAYS:
            # Últimos 60 días del período gratuito
            dias_restantes = FREE_PERIOD_DAYS - dias
            modo = "gratuito_warning"
            mensaje = (
                f"⚠️  El período gratuito termina en {dias_restantes} días. "
                f"Próximamente se requerirá suscripción."
            )
        else:
            # Período de pago activo
            modo = "pago"
            mensaje = "Período gratuito finalizado — se requiere suscripción"

        # Calcular precios con descuento early adopter si aplica
        precios_region = PRECIOS.get(region, PRECIOS["colombia"]).copy()
        descuento = 0.50 if es_early and modo == "pago" else 0.0

        return {
            "tier":           self.datos.get("tier", "free"),
            "modo":           modo,
            "mensaje":        mensaje,
            "dias_activo":    max(0, dias),
            "early_adopter":  es_early,
            "descuento":      descuento,
            "region":         region,
            "precios":        precios_region,
            "machine_id":     self.machine_id[:8] + "...",  # Solo parcial por seguridad
        }

    # ─── Registrar una sesión de análisis ────────────────────────────────────
    def registrar_sesion(self, ruta_proyecto: str, n_archivos: int) -> None:
        """
        Guarda un registro de cada análisis realizado.
        Útil para estadísticas de uso y auditoría.
        """
        self.datos.setdefault("historial", []).append({
            "fecha":      date.today().isoformat(),
            "proyecto":   Path(ruta_proyecto).name,
            "archivos":   n_archivos,
            "timestamp":  int(time.time()),
        })
        # Mantener solo los últimos 100 registros
        self.datos["historial"] = self.datos["historial"][-100:]
        self._guardar_licencia(self.datos)


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 4 — SCANNER UNIVERSAL
# Recorre cualquier proyecto sin importar tamaño ni lenguaje
# Genera estructura jerárquica, lista de archivos y métricas iniciales
# ════════════════════════════════════════════════════════════════════════════

class ScannerUniversal:
    """
    Motor de escaneo universal de proyectos.
    Responsabilidades:
      - Recorrer la estructura de directorios del proyecto
      - Filtrar carpetas y archivos irrelevantes
      - Clasificar archivos por tipo/lenguaje
      - Manejar proyectos de cualquier tamaño con chunking
      - Generar árbol visual de la estructura
      - Leer contenido de archivos de forma segura
    """

    def __init__(self, ruta_proyecto: str):
        self.ruta_base  = Path(ruta_proyecto).resolve()
        self.archivos   = []       # Lista completa de archivos encontrados
        self.estructura = {}       # Árbol jerárquico del proyecto
        self.estadisticas = {
            "total_archivos":  0,
            "total_carpetas":  0,
            "total_bytes":     0,
            "por_extension":   defaultdict(int),
            "por_categoria":   defaultdict(int),
        }

    # ─── Verificar si una carpeta debe ignorarse ──────────────────────────────
    def _ignorar_carpeta(self, nombre: str) -> bool:
        """
        Determina si una carpeta debe excluirse del análisis.
        Ignora: dependencias, caché, control de versiones, builds.
        """
        return nombre in CARPETAS_IGNORADAS or nombre.startswith(".")

    # ─── Verificar si un archivo debe ignorarse ───────────────────────────────
    def _ignorar_archivo(self, nombre: str) -> bool:
        """
        Determina si un archivo debe excluirse del análisis.
        Ignora: lockfiles, archivos de entorno, compilados.
        """
        if nombre in ARCHIVOS_IGNORADOS:
            return True
        # Verificar patrones con wildcard (*.pyc, *.min.js, etc.)
        patrones = {"*.pyc", "*.pyo", "*.pyd", "*.so", "*.dll",
                    "*.min.js", "*.min.css", "*.map", "*.log",
                    "*.class", "*.jar", "*.war"}
        ext = Path(nombre).suffix.lower()
        sufijos_binarios = {".pyc", ".pyo", ".pyd", ".so", ".dll",
                            ".class", ".jar", ".war", ".exe", ".bin"}
        return ext in sufijos_binarios

    # ─── Determinar la categoría de un archivo ───────────────────────────────
    def _categorizar_archivo(self, ruta: Path) -> str:
        """
        Clasifica un archivo en su categoría para estadísticas y análisis.
        Retorna: 'codigo', 'datos', 'web', 'docs', 'excel', 'scripts', 'cloud', 'otro'
        """
        ext = ruta.suffix.lower()
        if ext in EXTENSIONES_CODIGO:  return "codigo"
        if ext in EXTENSIONES_DATOS:   return "datos"
        if ext in EXTENSIONES_WEB:     return "web"
        if ext in EXTENSIONES_DOCS:    return "docs"
        if ext in EXTENSIONES_EXCEL:   return "excel"
        if ext in EXTENSIONES_SCRIPTS: return "scripts"
        if ext in EXTENSIONES_CLOUD:   return "cloud"
        return "otro"

    # ─── Detectar el lenguaje de programación de un archivo ──────────────────
    def _detectar_lenguaje(self, ruta: Path) -> str:
        """
        Mapea la extensión del archivo a su lenguaje de programación.
        """
        mapa = {
            ".py":   "Python",   ".js":  "JavaScript", ".jsx": "JavaScript",
            ".ts":   "TypeScript", ".tsx": "TypeScript",
            ".java": "Java",     ".cs":  "C#",
            ".php":  "PHP",      ".rb":  "Ruby",
            ".r":    "R",        ".R":   "R",
            ".sql":  "SQL",      ".html": "HTML",
            ".css":  "CSS",      ".scss": "SCSS",
            ".json": "JSON",     ".yaml": "YAML",
            ".yml":  "YAML",     ".md":  "Markdown",
            ".sh":   "Shell",    ".bat": "Batch",
            ".tf":   "Terraform", ".xlsx": "Excel",
            ".xlsm": "Excel+VBA", ".csv": "CSV",
        }
        return mapa.get(ruta.suffix.lower(), "Desconocido")

    # ─── Escanear el proyecto completo ───────────────────────────────────────
    def escanear(self) -> dict:
        """
        Recorre todo el proyecto y construye el inventario completo.
        Maneja proyectos grandes sin cargar todo en memoria.
        Retorna: dict con archivos, estructura, estadísticas.
        """
        if not self.ruta_base.exists():
            raise FileNotFoundError(f"Ruta no encontrada: {self.ruta_base}")
        if not self.ruta_base.is_dir():
            raise NotADirectoryError(f"No es una carpeta: {self.ruta_base}")

        print(f"\n  📂 Escaneando: {self.ruta_base}")
        print(f"  {'─' * 50}")

        # Recorrer el árbol de directorios
        for raiz, carpetas, ficheros in os.walk(self.ruta_base):
            raiz_path = Path(raiz)

            # Filtrar carpetas ignoradas in-place (evita descender a ellas)
            carpetas[:] = sorted([
                c for c in carpetas
                if not self._ignorar_carpeta(c)
            ])

            self.estadisticas["total_carpetas"] += len(carpetas)

            # Procesar cada archivo en la carpeta actual
            for fichero in sorted(ficheros):
                if self._ignorar_archivo(fichero):
                    continue

                ruta_archivo = raiz_path / fichero
                ext = ruta_archivo.suffix.lower()

                # Solo procesar extensiones válidas
                if ext not in EXTENSIONES_VALIDAS:
                    continue

                # Obtener tamaño del archivo
                try:
                    tamanio = ruta_archivo.stat().st_size
                except OSError:
                    continue

                # Construir registro del archivo
                ruta_rel = ruta_archivo.relative_to(self.ruta_base)
                categoria = self._categorizar_archivo(ruta_archivo)
                lenguaje  = self._detectar_lenguaje(ruta_archivo)

                registro = {
                    "ruta_abs":   str(ruta_archivo),
                    "ruta_rel":   str(ruta_rel),
                    "nombre":     fichero,
                    "extension":  ext,
                    "categoria":  categoria,
                    "lenguaje":   lenguaje,
                    "tamanio_b":  tamanio,
                    "tamanio_kb": round(tamanio / 1024, 2),
                }

                self.archivos.append(registro)

                # Actualizar estadísticas
                self.estadisticas["total_archivos"] += 1
                self.estadisticas["total_bytes"]    += tamanio
                self.estadisticas["por_extension"][ext] += 1
                self.estadisticas["por_categoria"][categoria] += 1

        # Completar estadísticas globales
        self.estadisticas["total_mb"]      = round(
            self.estadisticas["total_bytes"] / (1024 * 1024), 2
        )
        self.estadisticas["por_extension"] = dict(self.estadisticas["por_extension"])
        self.estadisticas["por_categoria"] = dict(self.estadisticas["por_categoria"])

        print(f"  ✅ {self.estadisticas['total_archivos']} archivos encontrados")
        print(f"  📦 {self.estadisticas['total_mb']} MB analizados")

        return {
            "archivos":      self.archivos,
            "estadisticas":  self.estadisticas,
            "ruta_base":     str(self.ruta_base),
            "nombre":        self.ruta_base.name,
        }

    # ─── Leer el contenido de un archivo de forma segura ─────────────────────
    def leer_archivo(self, ruta: str, max_mb: float = MAX_MB_POR_ARCHIVO) -> dict:
        """
        Lee el contenido de un archivo con manejo de errores y límites de tamaño.
        Archivos grandes se truncan con nota informativa.
        Retorna: dict con contenido, truncado, nota, encoding.
        """
        ruta_path = Path(ruta)
        max_bytes = int(max_mb * 1024 * 1024)
        truncado  = False
        nota      = ""

        try:
            tamanio = ruta_path.stat().st_size

            if tamanio == 0:
                return {"contenido": "", "truncado": False, "nota": "Archivo vacío", "ok": True}

            with open(ruta_path, "r", encoding="utf-8", errors="replace") as f:
                if tamanio > max_bytes:
                    contenido = f.read(max_bytes)
                    truncado  = True
                    nota      = (
                        f"[TRUNCADO: {tamanio/1024/1024:.1f} MB total, "
                        f"se leyeron los primeros {max_mb} MB]"
                    )
                else:
                    contenido = f.read()

            return {
                "contenido": contenido,
                "truncado":  truncado,
                "nota":      nota,
                "ok":        True,
                "lineas":    contenido.count("\n"),
            }

        except PermissionError:
            return {"contenido": "", "truncado": False, "nota": "Sin permiso de lectura", "ok": False}
        except Exception as e:
            return {"contenido": "", "truncado": False, "nota": f"Error: {e}", "ok": False}

    # ─── Generar árbol visual del proyecto ───────────────────────────────────
    def generar_arbol_visual(self, max_nivel: int = 6) -> str:
        """
        Genera una representación visual tipo árbol del proyecto.
        Respeta los mismos filtros que el escaneo principal.
        Útil para el reporte y para el contexto exportado a Claude.
        """
        lineas = [f"{self.ruta_base.name}/"]
        self._construir_arbol_recursivo(self.ruta_base, "", lineas, 0, max_nivel)
        return "\n".join(lineas)

    # ─── Función recursiva para construir el árbol ───────────────────────────
    def _construir_arbol_recursivo(
        self, ruta: Path, prefijo: str, lineas: list, nivel: int, max_nivel: int
    ) -> None:
        """
        Construcción recursiva del árbol visual con conectores ASCII.
        Usa └── para último elemento y ├── para los demás.
        """
        if nivel >= max_nivel:
            return

        try:
            entradas = sorted(ruta.iterdir(), key=lambda x: (x.is_file(), x.name))
        except PermissionError:
            return

        # Filtrar entradas irrelevantes
        entradas = [
            e for e in entradas
            if not (e.is_dir() and self._ignorar_carpeta(e.name))
            and not (e.is_file() and self._ignorar_archivo(e.name))
            and not (e.is_file() and e.suffix.lower() not in EXTENSIONES_VALIDAS)
        ]

        for i, entrada in enumerate(entradas):
            es_ultimo  = (i == len(entradas) - 1)
            conector   = "└── " if es_ultimo else "├── "
            extension  = "    " if es_ultimo else "│   "
            icono      = "📁 " if entrada.is_dir() else "📄 "

            lineas.append(f"{prefijo}{conector}{icono}{entrada.name}")

            if entrada.is_dir():
                self._construir_arbol_recursivo(
                    entrada, prefijo + extension, lineas, nivel + 1, max_nivel
                )

    # ─── Filtrar archivos por categoría ──────────────────────────────────────
    def filtrar_por_categoria(self, categoria: str) -> list:
        """
        Retorna solo los archivos que pertenecen a una categoría específica.
        Categorías: 'codigo', 'datos', 'web', 'docs', 'excel', 'scripts', 'cloud'
        """
        return [a for a in self.archivos if a["categoria"] == categoria]

    # ─── Filtrar archivos por lenguaje ────────────────────────────────────────
    def filtrar_por_lenguaje(self, lenguaje: str) -> list:
        """
        Retorna archivos de un lenguaje específico (Python, SQL, Java, etc.)
        """
        return [a for a in self.archivos if a["lenguaje"] == lenguaje]


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 5 — AST ENGINE PYTHON
# Análisis estático profundo de archivos Python usando el módulo ast stdlib
# Extrae: imports, clases, funciones, complejidad ciclomática, problemas
# Cero dependencias externas — todo con la librería estándar de Python
# ════════════════════════════════════════════════════════════════════════════

class ASTEnginePython:
    """
    Motor de análisis estático para código Python.
    Responsabilidades:
      - Parsear cada archivo .py con el módulo ast estándar
      - Extraer imports, clases, funciones, argumentos, decoradores
      - Calcular complejidad ciclomática por función (estilo radon)
      - Detectar funciones e imports definidos pero nunca utilizados
      - Construir un grafo de llamadas entre funciones del proyecto
      - Reportar archivos con errores de sintaxis sin abortar
    """

    # ─── Inicialización con la lista de archivos del Scanner ────────────────
    # Umbrales por defecto (generic) — se sobreescriben si hay framework detectado
    _UMBRALES_DEFAULT = {
        "complejidad_critica":    10,
        "lineas_funcion_critico": 300,
        "lineas_funcion_alto":    150,
        "clases_cero_es_ok":      None,
    }

    def __init__(self, archivos_python: list, scanner=None, umbrales_fw: dict = None):
        """
        Recibe la lista de archivos Python ya filtrada por ScannerUniversal.
        El scanner opcional se usa para reutilizar leer_archivo() y respetar
        los límites de tamaño definidos globalmente.
        umbrales_fw: dict del FrameworkDetector con lineas_funcion_critico,
                     lineas_funcion_alto y clases_cero_es_ok.
        """
        self.archivos_python = archivos_python
        self.scanner         = scanner
        self.resultado       = {}   # {ruta_rel: {imports, clases, funciones, ...}}
        self.indice_global   = {    # Catálogo cross-archivo para detecciones
            "funciones_definidas": set(),
            "clases_definidas":    set(),
            "nombres_referenciados": set(),
        }
        self.errores_sintaxis = []  # [{ruta, linea, mensaje}]
        self.grafo_llamadas   = defaultdict(set)  # {funcion: {llama_a_estas}}

        # Umbrales framework-aware (Streamlit calibrado con datos reales — QS Ingeniería)
        base = dict(self._UMBRALES_DEFAULT)
        if umbrales_fw:
            base["lineas_funcion_critico"] = umbrales_fw.get(
                "lineas_funcion_critico", base["lineas_funcion_critico"])
            base["lineas_funcion_alto"]    = umbrales_fw.get(
                "lineas_funcion_alto",    base["lineas_funcion_alto"])
            base["clases_cero_es_ok"]      = umbrales_fw.get(
                "clases_cero_es_ok",      base["clases_cero_es_ok"])
            # Dos poblaciones CC (Streamlit — calibrado con ~13K LOC QS Ingeniería):
            #   vista_*/modulo_*/render* → CC ≤ 25 (UI overhead demostrado)
            #   funciones de lógica/utilidad → CC ≤ 10 (McCabe estándar)
            #   CC > 25 → CRÍTICO siempre, sin excepción de tipo
            base["cc_umbral_ui"]     = umbrales_fw.get("cc_umbral_ui",    10)
            base["cc_umbral_logica"] = umbrales_fw.get("cc_umbral_logica",10)
            base["patron_ui"]        = umbrales_fw.get("patrones_ui",     None)
        self._umbrales = base

    # ─── Punto de entrada principal del análisis AST ────────────────────────
    def analizar(self) -> dict:
        """
        Ejecuta el análisis completo en dos pasadas:
          Pasada 1: parsear cada archivo y extraer estructura
          Pasada 2: cruzar referencias para detectar no usados
        Retorna: dict {ruta_rel: análisis} + metadatos agregados
        """
        if not self.archivos_python:
            return {"archivos": {}, "resumen": self._resumen_vacio(), "errores": []}

        # ─── PASADA 1: parseo y extracción por archivo ──────────────────────
        for archivo in self.archivos_python:
            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]
            self.resultado[ruta_rel] = self._analizar_archivo(ruta_abs, ruta_rel)

        # ─── PASADA 2: detección cross-archivo de no usados ─────────────────
        self._detectar_no_usados_global()

        # ─── Construir resumen agregado ─────────────────────────────────────
        return {
            "archivos":     self.resultado,
            "resumen":      self._construir_resumen(),
            "errores":      self.errores_sintaxis,
            "grafo_llamadas": {k: sorted(v) for k, v in self.grafo_llamadas.items()},
        }

    # ─── Análisis de un archivo individual ──────────────────────────────────
    def _analizar_archivo(self, ruta_abs: str, ruta_rel: str) -> dict:
        """
        Lee un archivo .py, lo parsea con ast y extrae toda su estructura.
        Si el parseo falla, registra el error sin abortar el análisis global.
        """
        # ─── Leer contenido respetando límites del scanner ──────────────────
        try:
            with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                contenido = f.read()
        except Exception as e:
            self.errores_sintaxis.append({
                "ruta": ruta_rel, "linea": 0, "mensaje": f"Error al leer: {e}"
            })
            return self._resultado_archivo_vacio(error=str(e))

        # ─── Parsear con ast — captura SyntaxError sin abortar ──────────────
        try:
            tree = ast.parse(contenido, filename=ruta_rel)
        except SyntaxError as e:
            self.errores_sintaxis.append({
                "ruta": ruta_rel, "linea": e.lineno or 0, "mensaje": str(e.msg)
            })
            return self._resultado_archivo_vacio(error=f"SyntaxError línea {e.lineno}")

        # ─── Extraer cada componente del árbol AST ──────────────────────────
        imports     = self._extraer_imports(tree)
        clases      = self._extraer_clases(tree)
        funciones   = self._extraer_funciones_top_level(tree)
        referencias = self._extraer_referencias(tree)

        # ─── Indexar definiciones globales para pasada 2 ────────────────────
        for func in funciones:
            self.indice_global["funciones_definidas"].add(
                f"{ruta_rel}::{func['nombre']}"
            )
        for cls in clases:
            self.indice_global["clases_definidas"].add(
                f"{ruta_rel}::{cls['nombre']}"
            )
        self.indice_global["nombres_referenciados"].update(referencias)

        # ─── Construir grafo de llamadas dentro del archivo ─────────────────
        for func in funciones:
            for llamada in func.get("llama_a", []):
                self.grafo_llamadas[f"{ruta_rel}::{func['nombre']}"].add(llamada)
        for cls in clases:
            for metodo in cls.get("metodos", []):
                clave = f"{ruta_rel}::{cls['nombre']}.{metodo['nombre']}"
                for llamada in metodo.get("llama_a", []):
                    self.grafo_llamadas[clave].add(llamada)

        # ─── Calcular métricas agregadas del archivo ────────────────────────
        complejidad_total = (
            sum(f["complejidad"] for f in funciones) +
            sum(m["complejidad"] for c in clases for m in c["metodos"])
        )
        n_metodos = sum(len(c["metodos"]) for c in clases)
        n_funciones_total = len(funciones) + n_metodos

        return {
            "imports":           imports,
            "clases":            clases,
            "funciones":         funciones,
            "lineas":            contenido.count("\n") + 1,
            "complejidad_total": complejidad_total,
            "complejidad_promedio": (
                round(complejidad_total / n_funciones_total, 2)
                if n_funciones_total else 0
            ),
            "n_clases":          len(clases),
            "n_funciones":       len(funciones),
            "n_metodos":         n_metodos,
            "n_imports":         len(imports),
            "problemas":         [],   # se llena en pasada 2
            "ok":                True,
        }

    # ─── Extraer todos los imports del módulo ───────────────────────────────
    def _extraer_imports(self, tree: ast.AST) -> list:
        """
        Recoge todos los import y from..import del archivo.
        Captura nombre original, alias, módulo de origen, línea.
        """
        imports = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    imports.append({
                        "tipo":    "import",
                        "modulo":  alias.name,
                        "nombre":  alias.name,
                        "alias":   alias.asname,
                        "usado_como": alias.asname or alias.name.split(".")[0],
                        "linea":   node.lineno,
                    })
            elif isinstance(node, ast.ImportFrom):
                modulo = node.module or ""
                for alias in node.names:
                    imports.append({
                        "tipo":    "from_import",
                        "modulo":  modulo,
                        "nombre":  alias.name,
                        "alias":   alias.asname,
                        "usado_como": alias.asname or alias.name,
                        "linea":   node.lineno,
                    })
        return imports

    # ─── Extraer clases con sus métodos y decoradores ───────────────────────
    def _extraer_clases(self, tree: ast.AST) -> list:
        """
        Encuentra todas las clases de nivel superior del módulo.
        No procesa clases anidadas dentro de funciones (caso raro).
        """
        clases = []
        for node in ast.iter_child_nodes(tree):
            if isinstance(node, ast.ClassDef):
                metodos = []
                for item in node.body:
                    if isinstance(item, (ast.FunctionDef, ast.AsyncFunctionDef)):
                        metodos.append(self._inspeccionar_funcion(item, es_metodo=True))
                clases.append({
                    "nombre":      node.name,
                    "linea":       node.lineno,
                    "bases":       [self._nombre_base(b) for b in node.bases],
                    "decoradores": [self._nombre_base(d) for d in node.decorator_list],
                    "metodos":     metodos,
                    "n_metodos":   len(metodos),
                    "docstring":   ast.get_docstring(node) is not None,
                })
        return clases

    # ─── Extraer funciones de nivel superior (no métodos) ───────────────────
    def _extraer_funciones_top_level(self, tree: ast.AST) -> list:
        """
        Solo funciones declaradas a nivel de módulo.
        Las funciones anidadas dentro de otras se ignoran para mantener
        el análisis estable y evitar inflar el reporte.
        """
        funciones = []
        for node in ast.iter_child_nodes(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                funciones.append(self._inspeccionar_funcion(node, es_metodo=False))
        return funciones

    # ─── Inspección detallada de una función o método ───────────────────────
    def _inspeccionar_funcion(self, node, es_metodo: bool = False) -> dict:
        """
        Extrae firma completa, complejidad, docstring y llamadas internas.
        """
        # ─── Argumentos: posicionales, defaults, *args, **kwargs ────────────
        args_pos = [a.arg for a in node.args.args]
        n_defaults = len(node.args.defaults)
        args_obligatorios = (
            args_pos[:-n_defaults] if n_defaults else args_pos
        )
        args_con_default = args_pos[-n_defaults:] if n_defaults else []
        vararg   = node.args.vararg.arg if node.args.vararg else None
        kwarg    = node.args.kwarg.arg if node.args.kwarg else None
        kwonly   = [a.arg for a in node.args.kwonlyargs]

        # ─── Llamadas internas para el grafo ────────────────────────────────
        llama_a = self._extraer_llamadas(node)

        return {
            "nombre":             node.name,
            "linea":              node.lineno,
            "es_async":           isinstance(node, ast.AsyncFunctionDef),
            "es_metodo":          es_metodo,
            "args_obligatorios":  args_obligatorios,
            "args_con_default":   args_con_default,
            "args_vararg":        vararg,
            "args_kwarg":         kwarg,
            "args_kwonly":        kwonly,
            "n_args":             len(args_pos) + len(kwonly) + (1 if vararg else 0) + (1 if kwarg else 0),
            "decoradores":        [self._nombre_base(d) for d in node.decorator_list],
            "complejidad":        self._calcular_complejidad(node),
            "docstring":          ast.get_docstring(node) is not None,
            "llama_a":            llama_a,
            "lineas":             (node.end_lineno or node.lineno) - node.lineno + 1,
        }

    # ─── Calcular complejidad ciclomática (estilo radon) ────────────────────
    def _calcular_complejidad(self, node) -> int:
        """
        Suma 1 por cada nodo de bifurcación. Base = 1.
        Cuenta: If, For, AsyncFor, While, ExceptHandler, IfExp,
                BoolOp (n-1 por operando), comprehensions con if,
                Match.case (Python 3.10+), Assert.
        """
        complejidad = 1
        for child in ast.walk(node):
            if isinstance(child, (ast.If, ast.For, ast.AsyncFor, ast.While,
                                  ast.ExceptHandler, ast.IfExp, ast.Assert)):
                complejidad += 1
            elif isinstance(child, ast.BoolOp):
                # `a and b and c` añade 2 (n-1 puntos de decisión)
                complejidad += len(child.values) - 1
            elif isinstance(child, (ast.ListComp, ast.SetComp,
                                    ast.DictComp, ast.GeneratorExp)):
                # Cada cláusula `if` dentro del comprehension suma 1
                for gen in child.generators:
                    complejidad += len(gen.ifs)
            elif hasattr(ast, "Match") and isinstance(child, ast.Match):
                # match...case añade 1 por cada case (Python 3.10+)
                complejidad += len(child.cases)
        return complejidad

    # ─── Extraer nombres de funciones llamadas dentro de un nodo ────────────
    def _extraer_llamadas(self, node) -> list:
        """
        Recorre el cuerpo del nodo y extrae todos los Call.
        Resuelve nombres simples (foo()) y atributos (obj.metodo()).
        """
        llamadas = set()
        for child in ast.walk(node):
            if isinstance(child, ast.Call):
                nombre = self._nombre_base(child.func)
                if nombre:
                    llamadas.add(nombre)
        return sorted(llamadas)

    # ─── Extraer todas las referencias a nombres del archivo completo ───────
    def _extraer_referencias(self, tree: ast.AST) -> set:
        """
        Recoge todos los Name y Attribute usados.
        Sirve para detectar imports y funciones nunca utilizadas.
        """
        referencias = set()
        for node in ast.walk(tree):
            if isinstance(node, ast.Name):
                referencias.add(node.id)
            elif isinstance(node, ast.Attribute):
                base = self._nombre_base(node)
                if base:
                    referencias.add(base)
                    # También añadir solo la primera parte (módulo raíz)
                    referencias.add(base.split(".")[0])
        return referencias

    # ─── Resolver el nombre de un nodo Name o Attribute ─────────────────────
    def _nombre_base(self, node) -> str:
        """
        Convierte un nodo AST a su nombre como string.
        Maneja Name, Attribute encadenado, Call, Subscript.
        Retorna cadena vacía si no es resoluble.
        """
        if node is None:
            return ""
        if isinstance(node, ast.Name):
            return node.id
        if isinstance(node, ast.Attribute):
            base = self._nombre_base(node.value)
            return f"{base}.{node.attr}" if base else node.attr
        if isinstance(node, ast.Call):
            return self._nombre_base(node.func)
        if isinstance(node, ast.Subscript):
            return self._nombre_base(node.value)
        if isinstance(node, ast.Constant):
            return str(node.value)
        return ""

    # ─── Detectar imports y funciones nunca utilizados (cross-archivo) ──────
    def _detectar_no_usados_global(self) -> None:
        """
        Pasada 2: cruza definiciones contra referencias globales.
        Hipótesis conservadoras para evitar falsos positivos:
          - Funciones __dunder__ siempre se consideran usadas (protocolos)
          - Funciones que empiezan con _ se ignoran (pueden ser API privada)
          - Imports estándar (typing, __future__) se ignoran
        """
        nombres_usados = self.indice_global["nombres_referenciados"]

        # ─── Imports nunca usados por archivo ───────────────────────────────
        for ruta_rel, datos in self.resultado.items():
            if not datos.get("ok"):
                continue
            for imp in datos.get("imports", []):
                nombre_uso = imp.get("usado_como", "")
                if not nombre_uso:
                    continue
                # Ignorar imports de sistema típicos
                if imp.get("modulo", "") in {"__future__", "typing"}:
                    continue
                if nombre_uso not in nombres_usados:
                    datos["problemas"].append({
                        "tipo":    "import_no_usado",
                        "linea":   imp.get("linea", 0),
                        "detalle": f"Import '{nombre_uso}' nunca utilizado",
                    })

            # ─── Funciones de nivel superior nunca llamadas ─────────────────
            for func in datos.get("funciones", []):
                nombre = func["nombre"]
                if nombre.startswith("_"):
                    continue   # Privadas o dunder: no las marcamos
                if nombre in {"main", "setup", "run", "handler"}:
                    continue   # Convenciones de entrypoint
                if nombre not in nombres_usados:
                    datos["problemas"].append({
                        "tipo":    "funcion_no_usada",
                        "linea":   func["linea"],
                        "detalle": f"Función '{nombre}' definida pero nunca llamada",
                    })

            # ─── Complejidad CC y longitud — dos poblaciones (calibrado QS Ingeniería) ──
            umbral_lc     = self._umbrales["lineas_funcion_critico"]
            umbral_la     = self._umbrales["lineas_funcion_alto"]
            cc_ui         = self._umbrales.get("cc_umbral_ui",    10)
            cc_logica     = self._umbrales.get("cc_umbral_logica",10)
            patron_ui_raw = self._umbrales.get("patron_ui", None)
            re_ui = re.compile(patron_ui_raw, re.IGNORECASE) if patron_ui_raw else None

            def _umbral_cc_para(nombre_fn: str) -> tuple:
                """Retorna (umbral, etiqueta_tipo) según nombre de función."""
                if re_ui and re_ui.match(nombre_fn):
                    return cc_ui, "UI/vista (umbral calibrado)"
                return cc_logica, "lógica/utilidad (McCabe)"

            def _check_fn(nombre_f, cc, lines, linea, es_metodo=False):
                umbral_cc, tipo_cc = _umbral_cc_para(nombre_f)
                prefix = f"Método '{nombre_f}'" if es_metodo else f"Función '{nombre_f}'"
                # CC — siempre CRÍTICO si > 25, ALTO si > umbral propio
                if cc > 25:
                    datos["problemas"].append({
                        "tipo": "complejidad_alta", "severidad": "CRITICO",
                        "linea": linea,
                        "detalle": (f"{prefix} CC={cc} (>25 = CRÍTICO siempre, "
                                    f"independiente del framework)"),
                    })
                elif cc > umbral_cc:
                    datos["problemas"].append({
                        "tipo": "complejidad_alta", "severidad": "ALTO",
                        "linea": linea,
                        "detalle": (f"{prefix} CC={cc} (>{umbral_cc} para {tipo_cc})"),
                    })
                # Longitud
                if lines > umbral_lc:
                    datos["problemas"].append({
                        "tipo": "funcion_muy_larga", "severidad": "CRITICO",
                        "linea": linea,
                        "detalle": (f"{prefix} tiene {lines} líneas "
                                    f"(>{umbral_lc} = God-Function, CRÍTICO)"),
                    })
                elif lines > umbral_la:
                    datos["problemas"].append({
                        "tipo": "funcion_larga", "severidad": "ALTO",
                        "linea": linea,
                        "detalle": (f"{prefix} tiene {lines} líneas "
                                    f"(>{umbral_la} = revisar extracción de submódulos)"),
                    })

            for func in datos.get("funciones", []):
                _check_fn(func["nombre"], func["complejidad"],
                          func.get("lineas", 0), func["linea"])

            for cls in datos.get("clases", []):
                for met in cls.get("metodos", []):
                    _check_fn(f"{cls['nombre']}.{met['nombre']}",
                              met["complejidad"], met.get("lineas", 0),
                              met["linea"], es_metodo=True)

    # ─── Histograma de distribución de complejidad ciclomática ─────────────────
    def _calcular_histograma_cc(self) -> dict:
        """
        Distribuye todas las funciones y métodos en rangos de CC.
        Rangos: 1-5 (simple), 6-10 (aceptable), 11-20 (zona gris),
                21-50 (problemática), 50+ (crítica).
        Incluye lista de las 10 funciones más complejas con nombre y archivo.
        """
        rangos = {"1-5": 0, "6-10": 0, "11-20": 0, "21-50": 0, "50+": 0}
        top_funciones = []

        for ruta_rel, datos in self.resultado.items():
            if not datos.get("ok"):
                continue
            todas = list(datos.get("funciones", []))
            for cls in datos.get("clases", []):
                for met in cls.get("metodos", []):
                    todas.append({**met, "nombre": f"{cls['nombre']}.{met['nombre']}"})

            for f in todas:
                cc = f.get("complejidad", 1)
                if   cc <= 5:  rangos["1-5"]   += 1
                elif cc <= 10: rangos["6-10"]  += 1
                elif cc <= 20: rangos["11-20"] += 1
                elif cc <= 50: rangos["21-50"] += 1
                else:          rangos["50+"]   += 1

                if cc > 10:
                    top_funciones.append({
                        "archivo":  ruta_rel,
                        "funcion":  f.get("nombre", "?"),
                        "cc":       cc,
                        "lineas":   f.get("lineas", 0),
                        "linea":    f.get("linea", 0),
                    })

        top_funciones.sort(key=lambda x: x["cc"], reverse=True)
        return {
            "rangos":        rangos,
            "top_complejas": top_funciones[:10],
            "zona_gris_n":   rangos["11-20"],
            "criticas_n":    rangos["21-50"] + rangos["50+"],
        }

    # ─── Construir resumen agregado del análisis Python ─────────────────────
    def _construir_resumen(self) -> dict:
        """
        Resumen de alto nivel para mostrar en el dashboard.
        """
        n_archivos_ok = sum(1 for d in self.resultado.values() if d.get("ok"))
        return {
            "archivos_analizados":   len(self.resultado),
            "archivos_ok":           n_archivos_ok,
            "archivos_con_error":    len(self.resultado) - n_archivos_ok,
            "total_clases":          sum(d.get("n_clases", 0) for d in self.resultado.values()),
            "total_funciones":       sum(d.get("n_funciones", 0) for d in self.resultado.values()),
            "total_metodos":         sum(d.get("n_metodos", 0) for d in self.resultado.values()),
            "total_imports":         sum(d.get("n_imports", 0) for d in self.resultado.values()),
            "complejidad_total":     sum(d.get("complejidad_total", 0) for d in self.resultado.values()),
            "complejidad_promedio_fn": round(
                sum(d.get("complejidad_total", 0) for d in self.resultado.values()) /
                max(1, sum(d.get("n_funciones", 0) + d.get("n_metodos", 0) for d in self.resultado.values())),
                1
            ),
            "umbral_cc_usado":       self._umbrales["complejidad_critica"],
            "umbral_lineas_critico": self._umbrales["lineas_funcion_critico"],
            "umbral_lineas_alto":    self._umbrales["lineas_funcion_alto"],
            "total_problemas":       sum(len(d.get("problemas", [])) for d in self.resultado.values()),
            "errores_sintaxis":      len(self.errores_sintaxis),
        }

    # ─── Plantillas vacías para casos límite ────────────────────────────────
    def _resultado_archivo_vacio(self, error: str = "") -> dict:
        """Retorna estructura completa con valores cero cuando hay error."""
        return {
            "imports": [], "clases": [], "funciones": [],
            "lineas": 0, "complejidad_total": 0, "complejidad_promedio": 0,
            "n_clases": 0, "n_funciones": 0, "n_metodos": 0, "n_imports": 0,
            "problemas": [], "ok": False, "error": error,
        }

    def _resumen_vacio(self) -> dict:
        """Resumen para cuando no hay archivos Python en el proyecto."""
        return {
            "archivos_analizados": 0, "archivos_ok": 0, "archivos_con_error": 0,
            "total_clases": 0, "total_funciones": 0, "total_metodos": 0,
            "total_imports": 0, "complejidad_total": 0, "total_problemas": 0,
            "errores_sintaxis": 0,
        }



# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 6 — HEURISTIC ENGINE MULTI-LENGUAJE
# Análisis basado en regex + patrones para JS/TS, Java, C#, PHP, R
# Limpieza obligatoria de comentarios y enmascarado de strings ANTES del regex
# Solo declaraciones de nivel superior y nivel 1 de anidamiento
# Output con la misma forma que ASTEnginePython para consistencia
# ════════════════════════════════════════════════════════════════════════════

class HeuristicEngine:
    """
    Motor heurístico para lenguajes sin AST stdlib en Python.
    Responsabilidades:
      - Limpiar código (comentarios, strings) antes del análisis
      - Detectar imports/requires/using/library por lenguaje
      - Detectar clases, interfaces, namespaces, packages
      - Detectar funciones y métodos de nivel superior
      - Mantener el formato de salida compatible con ASTEnginePython
      - Reportar de forma conservadora — preferir falsos negativos
    """

    # ─── Inicialización con la lista completa de archivos del Scanner ───────
    def __init__(self, archivos: list, scanner=None):
        """
        Recibe TODOS los archivos del scanner. Internamente filtra los
        lenguajes soportados y procesa cada uno con su parser específico.
        """
        self.archivos       = archivos
        self.scanner        = scanner
        self.resultado      = {}    # {ruta_rel: {imports, clases, funciones, ...}}
        self.lenguajes_soportados = {
            "JavaScript": self._parsear_js_ts,
            "TypeScript": self._parsear_js_ts,
            "Java":       self._parsear_java,
            "C#":         self._parsear_csharp,
            "PHP":        self._parsear_php,
            "R":          self._parsear_r,
        }

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Procesa cada archivo con el parser correspondiente a su lenguaje.
        Retorna estructura idéntica a ASTEnginePython para consistencia.
        """
        for archivo in self.archivos:
            lang = archivo.get("lenguaje", "")
            if lang not in self.lenguajes_soportados:
                continue

            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception as e:
                self.resultado[ruta_rel] = self._resultado_vacio(error=str(e))
                continue

            # Limpieza universal antes de pasar al parser específico
            limpio = self._limpiar_codigo(contenido, lang)
            parser = self.lenguajes_soportados[lang]
            self.resultado[ruta_rel] = parser(limpio, contenido, lang)

        return {
            "archivos": self.resultado,
            "resumen":  self._construir_resumen(),
        }

    # ════════════════════════════════════════════════════════════════════════
    # LIMPIEZA UNIVERSAL DE CÓDIGO — SE EJECUTA SIEMPRE ANTES DE REGEX
    # ════════════════════════════════════════════════════════════════════════

    # ─── Eliminar comentarios y enmascarar strings por lenguaje ─────────────
    def _limpiar_codigo(self, texto: str, lenguaje: str) -> str:
        """
        Devuelve el texto sin comentarios y con strings reemplazados por
        un placeholder neutro. Esto previene que regex capture palabras
        clave dentro de strings o comentarios.

        Estrategia: enmascarar primero strings (para no borrar // dentro
        de un string), luego eliminar comentarios.
        """
        # ─── Paso 1: enmascarar strings con placeholder ─────────────────────
        # Soporta: "...", '...', `...` (backticks), strings multilinea Python no aplican aquí
        def reemplazar_string(match):
            return '"' + ("S" * (len(match.group(0)) - 2)) + '"'

        # Strings con escape (\" dentro)
        patron_string = r'"(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\''
        if lenguaje in ("JavaScript", "TypeScript"):
            patron_string = r'"(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\'|`(?:\\.|[^`\\])*`'
        texto = re.sub(patron_string, reemplazar_string, texto, flags=re.DOTALL)

        # ─── Paso 2: eliminar comentarios según lenguaje ────────────────────
        if lenguaje in ("JavaScript", "TypeScript", "Java", "C#", "PHP"):
            # Comentarios /* ... */ (multilínea)
            texto = re.sub(r"/\*.*?\*/", " ", texto, flags=re.DOTALL)
            # Comentarios // hasta fin de línea
            texto = re.sub(r"//[^\n]*", " ", texto)
            if lenguaje == "PHP":
                # PHP además acepta # como comentario
                texto = re.sub(r"(?<!\$)#[^\n]*", " ", texto)
        elif lenguaje == "R":
            # R solo usa # para comentarios
            texto = re.sub(r"#[^\n]*", " ", texto)

        return texto

    # ════════════════════════════════════════════════════════════════════════
    # PARSERS POR LENGUAJE — CADA UNO USA EL TEXTO YA LIMPIO
    # ════════════════════════════════════════════════════════════════════════

    # ─── JavaScript / TypeScript ────────────────────────────────────────────
    def _parsear_js_ts(self, limpio: str, original: str, lang: str) -> dict:
        """
        Detecta: import/require, export, class, interface (TS),
        function/const/let arrow functions de nivel superior y métodos de clase.
        """
        # Imports: ES6 + CommonJS
        imports = []
        for m in re.finditer(
            r"^\s*import\s+(?:(\*\s+as\s+\w+|\{[^}]+\}|\w+)\s+from\s+)?[\"']([^\"']+)[\"']",
            limpio, re.MULTILINE
        ):
            imports.append({
                "tipo": "import", "modulo": m.group(2),
                "nombre": (m.group(1) or "").strip(),
                "linea": limpio[:m.start()].count("\n") + 1,
            })
        for m in re.finditer(
            r"(?:const|let|var)\s+(\w+|\{[^}]+\})\s*=\s*require\([\"']([^\"']+)[\"']\)",
            limpio
        ):
            imports.append({
                "tipo": "require", "modulo": m.group(2), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })

        # Clases (incluye interfaces de TS)
        clases = []
        patron_class = r"^\s*(?:export\s+(?:default\s+)?)?(?:abstract\s+)?(class|interface)\s+(\w+)(?:\s+extends\s+([\w.<>,\s]+))?(?:\s+implements\s+([\w.<>,\s]+))?\s*\{"
        for m in re.finditer(patron_class, limpio, re.MULTILINE):
            tipo, nombre, extiende, implementa = m.group(1), m.group(2), m.group(3), m.group(4)
            inicio = m.end() - 1   # posición del { de apertura
            cuerpo = self._extraer_bloque_balanceado(limpio, inicio)
            metodos = self._extraer_metodos_js(cuerpo) if cuerpo else []
            clases.append({
                "nombre":      nombre,
                "tipo":        tipo,
                "linea":       limpio[:m.start()].count("\n") + 1,
                "extiende":    [x.strip() for x in (extiende or "").split(",") if x.strip()],
                "implementa":  [x.strip() for x in (implementa or "").split(",") if x.strip()],
                "metodos":     metodos,
                "n_metodos":   len(metodos),
            })

        # Funciones de nivel superior
        funciones = []
        # function nombre(...)
        for m in re.finditer(
            r"^\s*(?:export\s+(?:default\s+)?)?(?:async\s+)?function\s*\*?\s*(\w+)\s*\(([^)]*)\)",
            limpio, re.MULTILINE
        ):
            funciones.append({
                "nombre": m.group(1),
                "linea":  limpio[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": "async" in m.group(0),
                "complejidad": 1,  # heurística básica, no calculamos en regex
            })
        # const nombre = (...) => o const nombre = function(...)
        for m in re.finditer(
            r"^\s*(?:export\s+)?(?:const|let|var)\s+(\w+)\s*=\s*(?:async\s+)?(?:function\s*\*?\s*\(([^)]*)\)|\(([^)]*)\)\s*=>|\w+\s*=>)",
            limpio, re.MULTILINE
        ):
            args_raw = m.group(2) or m.group(3) or ""
            funciones.append({
                "nombre": m.group(1),
                "linea":  limpio[:m.start()].count("\n") + 1,
                "n_args": len([a for a in args_raw.split(",") if a.strip()]),
                "es_async": "async" in m.group(0),
                "complejidad": 1,
            })

        return self._formatear_resultado(
            imports, clases, funciones, original, lang
        )

    # ─── Java ───────────────────────────────────────────────────────────────
    def _parsear_java(self, limpio: str, original: str, lang: str) -> dict:
        """Detecta: package, import, class, interface, enum, métodos, anotaciones."""
        imports = []
        for m in re.finditer(r"^\s*import\s+(?:static\s+)?([\w.*]+)\s*;", limpio, re.MULTILINE):
            imports.append({
                "tipo": "import", "modulo": m.group(1), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })

        clases = []
        patron = r"^\s*(?:@\w+(?:\([^)]*\))?\s+)*(?:public\s+|private\s+|protected\s+)?(?:abstract\s+|final\s+|static\s+)*(class|interface|enum)\s+(\w+)(?:\s*<[^>]+>)?(?:\s+extends\s+([\w.<>,\s]+?))?(?:\s+implements\s+([\w.<>,\s]+?))?\s*\{"
        for m in re.finditer(patron, limpio, re.MULTILINE):
            tipo, nombre = m.group(1), m.group(2)
            inicio = m.end() - 1
            cuerpo = self._extraer_bloque_balanceado(limpio, inicio)
            metodos = self._extraer_metodos_java(cuerpo) if cuerpo else []
            clases.append({
                "nombre":     nombre,
                "tipo":       tipo,
                "linea":      limpio[:m.start()].count("\n") + 1,
                "extiende":   [x.strip() for x in (m.group(3) or "").split(",") if x.strip()],
                "implementa": [x.strip() for x in (m.group(4) or "").split(",") if x.strip()],
                "metodos":    metodos,
                "n_metodos":  len(metodos),
            })

        return self._formatear_resultado(
            imports, clases, [], original, lang
        )

    # ─── C# ─────────────────────────────────────────────────────────────────
    def _parsear_csharp(self, limpio: str, original: str, lang: str) -> dict:
        """Detecta: using, namespace, class, interface, struct, record, métodos."""
        imports = []
        for m in re.finditer(r"^\s*using\s+(?:static\s+)?([\w.]+)\s*;", limpio, re.MULTILINE):
            imports.append({
                "tipo": "using", "modulo": m.group(1), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })

        clases = []
        patron = r"^\s*(?:\[[^\]]+\]\s*)*(?:public\s+|private\s+|protected\s+|internal\s+)?(?:abstract\s+|sealed\s+|static\s+|partial\s+)*(class|interface|struct|record)\s+(\w+)(?:\s*<[^>]+>)?(?:\s*:\s*([\w.<>,\s]+?))?\s*\{"
        for m in re.finditer(patron, limpio, re.MULTILINE):
            tipo, nombre = m.group(1), m.group(2)
            inicio = m.end() - 1
            cuerpo = self._extraer_bloque_balanceado(limpio, inicio)
            metodos = self._extraer_metodos_csharp(cuerpo) if cuerpo else []
            clases.append({
                "nombre":     nombre,
                "tipo":       tipo,
                "linea":      limpio[:m.start()].count("\n") + 1,
                "extiende":   [x.strip() for x in (m.group(3) or "").split(",") if x.strip()],
                "implementa": [],
                "metodos":    metodos,
                "n_metodos":  len(metodos),
            })

        return self._formatear_resultado(
            imports, clases, [], original, lang
        )

    # ─── PHP ────────────────────────────────────────────────────────────────
    def _parsear_php(self, limpio: str, original: str, lang: str) -> dict:
        """Detecta: namespace, use, require/include, class, interface, trait, function."""
        imports = []
        for m in re.finditer(r"^\s*use\s+([\w\\]+)(?:\s+as\s+\w+)?\s*;", limpio, re.MULTILINE):
            imports.append({
                "tipo": "use", "modulo": m.group(1), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })
        for m in re.finditer(
            r"(?:require|include)(?:_once)?\s*\(?\s*[\"']([^\"']+)[\"']\s*\)?\s*;",
            limpio
        ):
            imports.append({
                "tipo": "require", "modulo": m.group(1), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })

        clases = []
        patron = r"^\s*(?:abstract\s+|final\s+)?(class|interface|trait)\s+(\w+)(?:\s+extends\s+([\w\\]+))?(?:\s+implements\s+([\w\\,\s]+))?\s*\{"
        for m in re.finditer(patron, limpio, re.MULTILINE):
            tipo, nombre = m.group(1), m.group(2)
            inicio = m.end() - 1
            cuerpo = self._extraer_bloque_balanceado(limpio, inicio)
            metodos = self._extraer_metodos_php(cuerpo) if cuerpo else []
            clases.append({
                "nombre":     nombre,
                "tipo":       tipo,
                "linea":      limpio[:m.start()].count("\n") + 1,
                "extiende":   [m.group(3)] if m.group(3) else [],
                "implementa": [x.strip() for x in (m.group(4) or "").split(",") if x.strip()],
                "metodos":    metodos,
                "n_metodos":  len(metodos),
            })

        funciones = []
        for m in re.finditer(
            r"^\s*function\s+(\w+)\s*\(([^)]*)\)", limpio, re.MULTILINE
        ):
            funciones.append({
                "nombre": m.group(1),
                "linea":  limpio[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": False,
                "complejidad": 1,
            })

        return self._formatear_resultado(
            imports, clases, funciones, original, lang
        )

    # ─── R ──────────────────────────────────────────────────────────────────
    def _parsear_r(self, limpio: str, original: str, lang: str) -> dict:
        """Detecta: library(), require(), source(), funciones nombre <- function()."""
        imports = []
        for m in re.finditer(r"\b(?:library|require)\(([\"']?)(\w+)\1\)", limpio):
            imports.append({
                "tipo": "library", "modulo": m.group(2), "nombre": m.group(2),
                "linea": limpio[:m.start()].count("\n") + 1,
            })
        for m in re.finditer(r"\bsource\([\"']([^\"']+)[\"']\)", limpio):
            imports.append({
                "tipo": "source", "modulo": m.group(1), "nombre": m.group(1),
                "linea": limpio[:m.start()].count("\n") + 1,
            })

        # R no tiene clases en el sentido tradicional (S3/S4 son distintos)
        clases = []

        # Funciones: nombre <- function(args)
        funciones = []
        for m in re.finditer(
            r"^\s*(\w+)\s*(?:<-|=)\s*function\s*\(([^)]*)\)", limpio, re.MULTILINE
        ):
            funciones.append({
                "nombre": m.group(1),
                "linea":  limpio[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": False,
                "complejidad": 1,
            })

        return self._formatear_resultado(
            imports, clases, funciones, original, lang
        )

    # ════════════════════════════════════════════════════════════════════════
    # EXTRACTORES DE MÉTODOS POR LENGUAJE (sobre cuerpo de clase ya extraído)
    # ════════════════════════════════════════════════════════════════════════

    # ─── Métodos JS/TS dentro de una clase ──────────────────────────────────
    def _extraer_metodos_js(self, cuerpo: str) -> list:
        """Captura métodos: nombre(args) {, nombre = (args) =>, async nombre(...)."""
        metodos = []
        # método clásico: opcional async/static, nombre(args)
        for m in re.finditer(
            r"^\s*(?:public\s+|private\s+|protected\s+|static\s+|async\s+)*(\w+)\s*\(([^)]*)\)\s*[:{]",
            cuerpo, re.MULTILINE
        ):
            nombre = m.group(1)
            if nombre in {"if", "for", "while", "switch", "return", "constructor"}:
                if nombre != "constructor":
                    continue
            metodos.append({
                "nombre": nombre,
                "linea":  cuerpo[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": "async" in m.group(0),
                "complejidad": 1,
            })
        return metodos

    # ─── Métodos Java ───────────────────────────────────────────────────────
    def _extraer_metodos_java(self, cuerpo: str) -> list:
        """Captura métodos Java: modifiers tipoRetorno nombre(args)."""
        metodos = []
        patron = r"^\s*(?:@\w+(?:\([^)]*\))?\s+)*(?:public\s+|private\s+|protected\s+)?(?:static\s+|final\s+|abstract\s+|synchronized\s+|native\s+)*(?:<[^>]+>\s+)?[\w<>\[\].]+\s+(\w+)\s*\(([^)]*)\)\s*(?:throws\s+[\w.,\s]+)?\s*[{;]"
        for m in re.finditer(patron, cuerpo, re.MULTILINE):
            nombre = m.group(1)
            if nombre in {"if", "for", "while", "switch", "return", "new"}:
                continue
            metodos.append({
                "nombre": nombre,
                "linea":  cuerpo[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": False,
                "complejidad": 1,
            })
        return metodos

    # ─── Métodos C# ─────────────────────────────────────────────────────────
    def _extraer_metodos_csharp(self, cuerpo: str) -> list:
        """Captura métodos C#: similar a Java pero con override, virtual, async."""
        metodos = []
        patron = r"^\s*(?:\[[^\]]+\]\s*)*(?:public\s+|private\s+|protected\s+|internal\s+)?(?:static\s+|virtual\s+|override\s+|sealed\s+|abstract\s+|async\s+|partial\s+)*(?:<[^>]+>\s+)?[\w<>\[\].?]+\s+(\w+)\s*\(([^)]*)\)\s*[{=]"
        for m in re.finditer(patron, cuerpo, re.MULTILINE):
            nombre = m.group(1)
            if nombre in {"if", "for", "foreach", "while", "switch", "return", "new"}:
                continue
            metodos.append({
                "nombre": nombre,
                "linea":  cuerpo[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": "async" in m.group(0),
                "complejidad": 1,
            })
        return metodos

    # ─── Métodos PHP ────────────────────────────────────────────────────────
    def _extraer_metodos_php(self, cuerpo: str) -> list:
        """Captura métodos PHP dentro de una clase."""
        metodos = []
        patron = r"^\s*(?:public\s+|private\s+|protected\s+)?(?:static\s+|abstract\s+|final\s+)*function\s+(\w+)\s*\(([^)]*)\)"
        for m in re.finditer(patron, cuerpo, re.MULTILINE):
            metodos.append({
                "nombre": m.group(1),
                "linea":  cuerpo[:m.start()].count("\n") + 1,
                "n_args": len([a for a in m.group(2).split(",") if a.strip()]),
                "es_async": False,
                "complejidad": 1,
            })
        return metodos

    # ════════════════════════════════════════════════════════════════════════
    # UTILIDADES INTERNAS
    # ════════════════════════════════════════════════════════════════════════

    # ─── Extraer el cuerpo de un bloque {...} balanceando llaves ────────────
    def _extraer_bloque_balanceado(self, texto: str, inicio: int) -> str:
        """
        Dado el índice de una { de apertura, retorna el contenido hasta la
        } de cierre correspondiente. Maneja anidación y respeta los strings
        que ya fueron enmascarados por _limpiar_codigo.
        """
        if inicio >= len(texto) or texto[inicio] != "{":
            return ""
        nivel = 0
        for i in range(inicio, len(texto)):
            c = texto[i]
            if c == "{":
                nivel += 1
            elif c == "}":
                nivel -= 1
                if nivel == 0:
                    return texto[inicio + 1 : i]
        return texto[inicio + 1 :]   # bloque sin cerrar — devolvemos lo que hay

    # ─── Formatear el resultado de un archivo en el formato común ───────────
    def _formatear_resultado(
        self, imports: list, clases: list, funciones: list,
        original: str, lang: str
    ) -> dict:
        """
        Estructura idéntica a ASTEnginePython._analizar_archivo() para
        que el resto del sistema pueda tratar ambas salidas de forma uniforme.
        """
        n_metodos = sum(len(c.get("metodos", [])) for c in clases)
        complejidad_total = (
            sum(f.get("complejidad", 1) for f in funciones) +
            sum(m.get("complejidad", 1) for c in clases for m in c.get("metodos", []))
        )
        return {
            "lenguaje":          lang,
            "imports":           imports,
            "clases":            clases,
            "funciones":         funciones,
            "lineas":            original.count("\n") + 1,
            "complejidad_total": complejidad_total,
            "complejidad_promedio": round(
                complejidad_total / max(1, len(funciones) + n_metodos), 2
            ),
            "n_clases":          len(clases),
            "n_funciones":       len(funciones),
            "n_metodos":         n_metodos,
            "n_imports":         len(imports),
            "problemas":         [],
            "ok":                True,
        }

    # ─── Resultado vacío para archivos con error de lectura ─────────────────
    def _resultado_vacio(self, error: str = "") -> dict:
        return {
            "imports": [], "clases": [], "funciones": [],
            "lineas": 0, "complejidad_total": 0, "complejidad_promedio": 0,
            "n_clases": 0, "n_funciones": 0, "n_metodos": 0, "n_imports": 0,
            "problemas": [], "ok": False, "error": error,
        }

    # ─── Resumen agregado por lenguaje ──────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Resumen agregado para mostrar en el dashboard."""
        por_lenguaje = defaultdict(lambda: {
            "archivos": 0, "clases": 0, "funciones": 0,
            "metodos": 0, "imports": 0, "complejidad": 0,
        })
        for datos in self.resultado.values():
            if not datos.get("ok"):
                continue
            lang = datos.get("lenguaje", "Desconocido")
            por_lenguaje[lang]["archivos"]    += 1
            por_lenguaje[lang]["clases"]      += datos.get("n_clases", 0)
            por_lenguaje[lang]["funciones"]   += datos.get("n_funciones", 0)
            por_lenguaje[lang]["metodos"]     += datos.get("n_metodos", 0)
            por_lenguaje[lang]["imports"]     += datos.get("n_imports", 0)
            por_lenguaje[lang]["complejidad"] += datos.get("complejidad_total", 0)
        return {
            "archivos_analizados": len(self.resultado),
            "por_lenguaje":        dict(por_lenguaje),
        }



# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 7 — SQL PARSER MULTI-DIALECTO
# Analiza archivos .sql sin dependencias externas
# Dialectos soportados: PostgreSQL, Supabase, MySQL/MariaDB, SQL Server, SQLite
# Limpieza obligatoria de comentarios antes del parsing
# Identificadores entrecomillados: "x" (PG), `x` (MySQL), [x] (MSSQL)
# ════════════════════════════════════════════════════════════════════════════

class SQLParser:
    """
    Parser de SQL multi-dialecto basado en regex disciplinado.
    Responsabilidades:
      - Limpiar comentarios -- y /* */ antes de cualquier extracción
      - Detectar tablas, vistas, índices, funciones, triggers, secuencias
      - Detectar políticas RLS de Supabase / PostgreSQL
      - Detectar foreign keys (inline, constraint nombrado, ALTER TABLE)
      - Detectar operaciones DML (INSERT/SELECT/UPDATE/DELETE) por tabla
      - Reportar el dialecto inferido por archivo
    """

    # ─── Inicialización con archivos SQL del Scanner ────────────────────────
    def __init__(self, archivos: list, scanner=None):
        """
        Recibe la lista completa del scanner. Filtra internamente .sql.
        """
        self.archivos    = [a for a in archivos if a.get("lenguaje") == "SQL"]
        self.scanner     = scanner
        self.resultado   = {}   # {ruta_rel: análisis}
        # Catálogo global cross-archivo
        self.catalogo = {
            "tablas":    {},   # {nombre_tabla: {ruta, linea, columnas, dialecto}}
            "vistas":    {},
            "funciones": {},
            "fk":        [],   # [{from_tabla, from_col, to_tabla, to_col, ruta}]
            "policies":  [],   # [{nombre, tabla, ruta}]
        }

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Procesa cada archivo SQL y construye el catálogo agregado.
        """
        for archivo in self.archivos:
            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception as e:
                self.resultado[ruta_rel] = self._resultado_vacio(error=str(e))
                continue

            limpio   = self._limpiar_sql(contenido)
            dialecto = self._detectar_dialecto(limpio)
            self.resultado[ruta_rel] = self._analizar_archivo(
                limpio, contenido, dialecto, ruta_rel
            )

        return {
            "archivos":  self.resultado,
            "catalogo":  self.catalogo,
            "resumen":   self._construir_resumen(),
        }

    # ════════════════════════════════════════════════════════════════════════
    # FASE 1 — LIMPIEZA OBLIGATORIA
    # ════════════════════════════════════════════════════════════════════════

    # ─── Eliminar comentarios SQL (-- y /* */) ──────────────────────────────
    def _limpiar_sql(self, texto: str) -> str:
        """
        Quita comentarios -- (línea) y /* */ (multilínea) y normaliza
        saltos de línea. Strings con ' o " se preservan tal cual porque
        SQL Parser sí los necesita para detectar nombres entre comillas.
        """
        # Eliminar /* ... */ multilínea
        texto = re.sub(r"/\*.*?\*/", " ", texto, flags=re.DOTALL)
        # Eliminar -- hasta fin de línea (respetar -- dentro de string es difícil
        # con regex; sacrificamos esos casos extremos por simplicidad)
        texto = re.sub(r"(?<!')--[^\n]*", "", texto)
        return texto

    # ─── Detectar dialecto SQL por características distintivas ──────────────
    def _detectar_dialecto(self, texto: str) -> str:
        """
        Heurística rápida: busca marcadores únicos de cada dialecto.
        Retorna: 'postgresql', 'mysql', 'sqlserver', 'sqlite', 'desconocido'.
        """
        t = texto.lower()
        # Supabase es PostgreSQL, lo reportamos por separado si aparece RLS
        if re.search(r"\bcreate\s+policy\b|\benable\s+row\s+level\s+security\b", t):
            return "supabase"
        if re.search(r"\bserial\b|\bjsonb\b|\breturning\b|::\w+", t):
            return "postgresql"
        if re.search(r"\bauto_increment\b|engine\s*=\s*\w+|`\w+`", t):
            return "mysql"
        if re.search(r"\bidentity\b\s*\(|\bgo\b\s*\n|\[\w+\]", t):
            return "sqlserver"
        if re.search(r"\bautoincrement\b|\bpragma\b|\bwithout\s+rowid\b", t):
            return "sqlite"
        return "desconocido"

    # ════════════════════════════════════════════════════════════════════════
    # FASE 2 — EXTRACCIÓN DE OBJETOS DDL
    # ════════════════════════════════════════════════════════════════════════

    # ─── Análisis completo de un archivo SQL ya limpio ──────────────────────
    def _analizar_archivo(
        self, limpio: str, original: str, dialecto: str, ruta_rel: str
    ) -> dict:
        """
        Extrae todos los objetos DDL y operaciones DML del archivo.
        Alimenta el catálogo global con las definiciones encontradas.
        """
        tablas   = self._extraer_tablas(limpio, dialecto, ruta_rel)
        vistas   = self._extraer_vistas(limpio, ruta_rel)
        indices  = self._extraer_indices(limpio)
        funcs    = self._extraer_funciones(limpio, ruta_rel)
        triggers = self._extraer_triggers(limpio)
        secuencs = self._extraer_secuencias(limpio)
        policies = self._extraer_policies(limpio, ruta_rel)
        fks      = self._extraer_fks_alter(limpio, ruta_rel)
        dml      = self._extraer_dml(limpio)

        # Sumar FKs de las tablas al catálogo global
        for t in tablas:
            for fk in t.get("foreign_keys", []):
                self.catalogo["fk"].append({
                    "from_tabla": t["nombre"],
                    "from_col":   fk["columna"],
                    "to_tabla":   fk["referencia_tabla"],
                    "to_col":     fk["referencia_columna"],
                    "ruta":       ruta_rel,
                })
        for fk in fks:
            self.catalogo["fk"].append(fk)

        return {
            "dialecto":       dialecto,
            "tablas":         tablas,
            "vistas":         vistas,
            "indices":        indices,
            "funciones":      funcs,
            "triggers":       triggers,
            "secuencias":     secuencs,
            "policies":       policies,
            "dml":            dml,
            "n_tablas":       len(tablas),
            "n_vistas":       len(vistas),
            "n_indices":      len(indices),
            "n_funciones":    len(funcs),
            "n_triggers":     len(triggers),
            "n_policies":     len(policies),
            "lineas":         original.count("\n") + 1,
            "ok":             True,
        }

    # ─── Limpiar nombre de identificador (quita comillas/brackets) ──────────
    def _limpiar_identificador(self, nombre: str) -> str:
        """Quita "..." (PG), `...` (MySQL), [...] (MSSQL), espacios."""
        if not nombre:
            return ""
        nombre = nombre.strip()
        if len(nombre) >= 2:
            if (nombre[0] == nombre[-1] and nombre[0] in '"`') or \
               (nombre[0] == "[" and nombre[-1] == "]"):
                return nombre[1:-1]
        return nombre

    # ─── Patrón regex que captura cualquier identificador SQL ───────────────
    # Soporta: nombre, "nombre", `nombre`, [nombre], schema.nombre
    _IDENT = r'(?:"[^"]+"|`[^`]+`|\[[^\]]+\]|\w+)(?:\s*\.\s*(?:"[^"]+"|`[^`]+`|\[[^\]]+\]|\w+))?'

    # ─── Extraer CREATE TABLE ───────────────────────────────────────────────
    def _extraer_tablas(self, texto: str, dialecto: str, ruta_rel: str) -> list:
        """
        Captura cada CREATE TABLE [IF NOT EXISTS] nombre (columnas...).
        Extrae columnas, tipos, nullability y FKs inline.
        """
        tablas = []
        patron = re.compile(
            r"CREATE\s+(?:TEMPORARY\s+|TEMP\s+|UNLOGGED\s+)?TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?"
            r"(" + self._IDENT + r")\s*\(",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            nombre_raw = m.group(1)
            nombre = self._limpiar_identificador(nombre_raw.split(".")[-1])
            schema = (
                self._limpiar_identificador(nombre_raw.split(".")[0])
                if "." in nombre_raw else None
            )

            # Extraer cuerpo entre paréntesis balanceados
            cuerpo = self._extraer_parens_balanceados(texto, m.end() - 1)
            columnas, fks_inline = self._parsear_columnas_tabla(cuerpo)

            datos = {
                "nombre":       nombre,
                "schema":       schema,
                "columnas":     columnas,
                "n_columnas":   len(columnas),
                "foreign_keys": fks_inline,
                "linea":        texto[:m.start()].count("\n") + 1,
                "dialecto":     dialecto,
            }
            tablas.append(datos)
            # Agregar al catálogo global (clave: nombre simple)
            self.catalogo["tablas"][nombre] = {
                "ruta":      ruta_rel,
                "linea":     datos["linea"],
                "columnas":  [c["nombre"] for c in columnas],
                "dialecto":  dialecto,
            }
        return tablas

    # ─── Parsear columnas dentro del cuerpo de CREATE TABLE ─────────────────
    def _parsear_columnas_tabla(self, cuerpo: str) -> tuple:
        """
        Recibe el contenido entre () del CREATE TABLE y separa por comas
        respetando paréntesis anidados. Distingue columnas de constraints.
        """
        partes = self._split_top_level(cuerpo, ",")
        columnas, fks = [], []
        for parte in partes:
            parte = parte.strip()
            if not parte:
                continue
            up = parte.upper()
            # ─── CONSTRAINT FOREIGN KEY ─────────────────────────────────────
            if up.startswith("CONSTRAINT") or up.startswith("FOREIGN KEY"):
                m = re.search(
                    r"FOREIGN\s+KEY\s*\(\s*(" + self._IDENT + r")\s*\)\s*"
                    r"REFERENCES\s+(" + self._IDENT + r")\s*\(\s*(" + self._IDENT + r")\s*\)",
                    parte, re.IGNORECASE
                )
                if m:
                    fks.append({
                        "columna":            self._limpiar_identificador(m.group(1)),
                        "referencia_tabla":   self._limpiar_identificador(m.group(2).split(".")[-1]),
                        "referencia_columna": self._limpiar_identificador(m.group(3)),
                        "tipo":               "constraint",
                    })
                continue
            # Saltar PRIMARY KEY (...), UNIQUE (...), CHECK (...) compuestos
            if up.startswith(("PRIMARY KEY", "UNIQUE", "CHECK", "INDEX", "KEY ")):
                continue

            # ─── Es una columna ─────────────────────────────────────────────
            tokens = parte.split(None, 2)
            if len(tokens) < 2:
                continue
            nombre_col = self._limpiar_identificador(tokens[0])
            tipo_col   = tokens[1]
            resto      = tokens[2] if len(tokens) > 2 else ""

            es_pk      = bool(re.search(r"\bPRIMARY\s+KEY\b", parte, re.IGNORECASE))
            es_null    = not bool(re.search(r"\bNOT\s+NULL\b", parte, re.IGNORECASE))
            es_unique  = bool(re.search(r"\bUNIQUE\b", parte, re.IGNORECASE))

            # ─── FK inline en la columna ────────────────────────────────────
            fk_inline = re.search(
                r"REFERENCES\s+(" + self._IDENT + r")\s*\(\s*(" + self._IDENT + r")\s*\)",
                parte, re.IGNORECASE
            )
            if fk_inline:
                fks.append({
                    "columna":            nombre_col,
                    "referencia_tabla":   self._limpiar_identificador(fk_inline.group(1).split(".")[-1]),
                    "referencia_columna": self._limpiar_identificador(fk_inline.group(2)),
                    "tipo":               "inline",
                })

            columnas.append({
                "nombre":     nombre_col,
                "tipo":       tipo_col,
                "nullable":   es_null,
                "primary":    es_pk,
                "unique":     es_unique,
            })
        return columnas, fks

    # ─── Extraer CREATE VIEW ────────────────────────────────────────────────
    def _extraer_vistas(self, texto: str, ruta_rel: str) -> list:
        """Captura CREATE [OR REPLACE] [MATERIALIZED] VIEW ..."""
        vistas = []
        patron = re.compile(
            r"CREATE\s+(?:OR\s+REPLACE\s+)?(?:MATERIALIZED\s+)?VIEW\s+(?:IF\s+NOT\s+EXISTS\s+)?"
            r"(" + self._IDENT + r")",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            nombre_raw = m.group(1)
            nombre = self._limpiar_identificador(nombre_raw.split(".")[-1])
            vistas.append({
                "nombre": nombre,
                "linea":  texto[:m.start()].count("\n") + 1,
            })
            self.catalogo["vistas"][nombre] = {"ruta": ruta_rel}
        return vistas

    # ─── Extraer CREATE INDEX ───────────────────────────────────────────────
    def _extraer_indices(self, texto: str) -> list:
        """Captura CREATE [UNIQUE] INDEX ... ON tabla(col1, col2)."""
        indices = []
        patron = re.compile(
            r"CREATE\s+(?:UNIQUE\s+)?INDEX\s+(?:IF\s+NOT\s+EXISTS\s+)?"
            r"(" + self._IDENT + r")?\s*ON\s+(" + self._IDENT + r")\s*\(([^)]+)\)",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            indices.append({
                "nombre": self._limpiar_identificador((m.group(1) or "").split(".")[-1]),
                "tabla":  self._limpiar_identificador(m.group(2).split(".")[-1]),
                "columnas": [c.strip() for c in m.group(3).split(",")],
                "linea":   texto[:m.start()].count("\n") + 1,
            })
        return indices

    # ─── Extraer CREATE FUNCTION / PROCEDURE ────────────────────────────────
    def _extraer_funciones(self, texto: str, ruta_rel: str) -> list:
        """Captura CREATE [OR REPLACE] FUNCTION/PROCEDURE."""
        funcs = []
        patron = re.compile(
            r"CREATE\s+(?:OR\s+REPLACE\s+)?(FUNCTION|PROCEDURE)\s+(" + self._IDENT + r")",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            tipo = m.group(1).upper()
            nombre = self._limpiar_identificador(m.group(2).split(".")[-1])
            funcs.append({
                "nombre": nombre,
                "tipo":   tipo.lower(),
                "linea":  texto[:m.start()].count("\n") + 1,
            })
            self.catalogo["funciones"][nombre] = {"ruta": ruta_rel, "tipo": tipo.lower()}
        return funcs

    # ─── Extraer CREATE TRIGGER ─────────────────────────────────────────────
    def _extraer_triggers(self, texto: str) -> list:
        """Captura CREATE TRIGGER ... ON tabla."""
        triggers = []
        patron = re.compile(
            r"CREATE\s+TRIGGER\s+(" + self._IDENT + r").*?ON\s+(" + self._IDENT + r")",
            re.IGNORECASE | re.DOTALL
        )
        for m in patron.finditer(texto):
            triggers.append({
                "nombre": self._limpiar_identificador(m.group(1).split(".")[-1]),
                "tabla":  self._limpiar_identificador(m.group(2).split(".")[-1]),
                "linea":  texto[:m.start()].count("\n") + 1,
            })
        return triggers

    # ─── Extraer CREATE SEQUENCE ────────────────────────────────────────────
    def _extraer_secuencias(self, texto: str) -> list:
        """Captura CREATE SEQUENCE (PostgreSQL/Oracle)."""
        secs = []
        patron = re.compile(
            r"CREATE\s+SEQUENCE\s+(?:IF\s+NOT\s+EXISTS\s+)?(" + self._IDENT + r")",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            secs.append({
                "nombre": self._limpiar_identificador(m.group(1).split(".")[-1]),
                "linea":  texto[:m.start()].count("\n") + 1,
            })
        return secs

    # ─── Extraer políticas RLS de Supabase / PostgreSQL ─────────────────────
    def _extraer_policies(self, texto: str, ruta_rel: str) -> list:
        """Captura CREATE POLICY ... ON tabla."""
        policies = []
        patron = re.compile(
            r"CREATE\s+POLICY\s+(\w+|\"[^\"]+\")\s+ON\s+(" + self._IDENT + r")"
            r"(?:.*?FOR\s+(\w+))?",
            re.IGNORECASE | re.DOTALL
        )
        for m in patron.finditer(texto):
            tabla = self._limpiar_identificador(m.group(2).split(".")[-1])
            policy = {
                "nombre": self._limpiar_identificador(m.group(1)),
                "tabla":  tabla,
                "operacion": (m.group(3) or "ALL").upper(),
                "linea":  texto[:m.start()].count("\n") + 1,
            }
            policies.append(policy)
            self.catalogo["policies"].append({**policy, "ruta": ruta_rel})
        return policies

    # ─── Extraer FKs definidas vía ALTER TABLE ──────────────────────────────
    def _extraer_fks_alter(self, texto: str, ruta_rel: str) -> list:
        """
        Captura: ALTER TABLE x ADD [CONSTRAINT y] FOREIGN KEY (col)
                 REFERENCES otra(col)
        """
        fks = []
        patron = re.compile(
            r"ALTER\s+TABLE\s+(" + self._IDENT + r")\s+ADD\s+"
            r"(?:CONSTRAINT\s+\w+\s+)?FOREIGN\s+KEY\s*\(\s*(" + self._IDENT + r")\s*\)\s*"
            r"REFERENCES\s+(" + self._IDENT + r")\s*\(\s*(" + self._IDENT + r")\s*\)",
            re.IGNORECASE
        )
        for m in patron.finditer(texto):
            fks.append({
                "from_tabla": self._limpiar_identificador(m.group(1).split(".")[-1]),
                "from_col":   self._limpiar_identificador(m.group(2)),
                "to_tabla":   self._limpiar_identificador(m.group(3).split(".")[-1]),
                "to_col":     self._limpiar_identificador(m.group(4)),
                "ruta":       ruta_rel,
                "tipo":       "alter",
            })
        return fks

    # ─── Extraer operaciones DML por tabla ──────────────────────────────────
    def _extraer_dml(self, texto: str) -> dict:
        """
        Cuenta operaciones SELECT/INSERT/UPDATE/DELETE por tabla.
        Útil para entender qué tablas están "calientes" en el sistema.
        """
        dml = defaultdict(lambda: {"select": 0, "insert": 0, "update": 0, "delete": 0})
        patrones = {
            "select": re.compile(r"\bFROM\s+(" + self._IDENT + r")", re.IGNORECASE),
            "insert": re.compile(r"\bINSERT\s+INTO\s+(" + self._IDENT + r")", re.IGNORECASE),
            "update": re.compile(r"\bUPDATE\s+(" + self._IDENT + r")\s+SET", re.IGNORECASE),
            "delete": re.compile(r"\bDELETE\s+FROM\s+(" + self._IDENT + r")", re.IGNORECASE),
        }
        for op, patron in patrones.items():
            for m in patron.finditer(texto):
                tabla = self._limpiar_identificador(m.group(1).split(".")[-1]).lower()
                dml[tabla][op] += 1
        return {k: dict(v) for k, v in dml.items()}

    # ════════════════════════════════════════════════════════════════════════
    # UTILIDADES INTERNAS
    # ════════════════════════════════════════════════════════════════════════

    # ─── Extraer contenido entre paréntesis balanceados ─────────────────────
    def _extraer_parens_balanceados(self, texto: str, inicio: int) -> str:
        """
        Dado el índice de un ( de apertura, retorna el contenido hasta
        el ) correspondiente, manejando anidación.
        """
        if inicio >= len(texto) or texto[inicio] != "(":
            return ""
        nivel = 0
        for i in range(inicio, len(texto)):
            if texto[i] == "(":
                nivel += 1
            elif texto[i] == ")":
                nivel -= 1
                if nivel == 0:
                    return texto[inicio + 1 : i]
        return texto[inicio + 1 :]

    # ─── Split por separador respetando paréntesis ──────────────────────────
    def _split_top_level(self, texto: str, sep: str) -> list:
        """
        Hace split() pero solo cuando el separador está fuera de paréntesis.
        Necesario para separar columnas de CREATE TABLE sin romper tipos
        como DECIMAL(10, 2) o VARCHAR(255).
        """
        partes, actual, nivel = [], [], 0
        for c in texto:
            if c == "(":
                nivel += 1
            elif c == ")":
                nivel -= 1
            if c == sep and nivel == 0:
                partes.append("".join(actual))
                actual = []
            else:
                actual.append(c)
        if actual:
            partes.append("".join(actual))
        return partes

    # ─── Resultado vacío para errores ───────────────────────────────────────
    def _resultado_vacio(self, error: str = "") -> dict:
        return {
            "dialecto": "desconocido",
            "tablas": [], "vistas": [], "indices": [],
            "funciones": [], "triggers": [], "secuencias": [],
            "policies": [], "dml": {},
            "n_tablas": 0, "n_vistas": 0, "n_indices": 0,
            "n_funciones": 0, "n_triggers": 0, "n_policies": 0,
            "lineas": 0, "ok": False, "error": error,
        }

    # ─── Resumen agregado ───────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Métricas globales del análisis SQL."""
        return {
            "archivos_analizados": len(self.resultado),
            "total_tablas":        len(self.catalogo["tablas"]),
            "total_vistas":        len(self.catalogo["vistas"]),
            "total_funciones":     len(self.catalogo["funciones"]),
            "total_fks":           len(self.catalogo["fk"]),
            "total_policies_rls":  len(self.catalogo["policies"]),
        }



# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 8 — CODE-DB LINKER
# Cruza nombres de tablas (del SQL Parser) contra el código fuente
# Solo detecta patrones específicos de ORM/cliente DB para evitar colisiones
# Output: tabla → archivos que la usan, archivo → tablas que usa, huérfanas
# ════════════════════════════════════════════════════════════════════════════

class CodeDBLinker:
    """
    Motor de cruce código ↔ base de datos.
    Responsabilidades:
      - Recibir el catálogo de tablas del SQLParser
      - Buscar referencias en código fuente usando patrones específicos
        de ORMs y clientes DB (NO buscar el nombre pelado, eso da
        falsos positivos masivos con palabras como 'users' o 'data')
      - Construir mapa bidireccional tabla ↔ archivos
      - Identificar tablas huérfanas (definidas pero nunca usadas)
      - Identificar tablas referenciadas pero no definidas (pueden
        ser de un schema externo o un error tipográfico)
    """

    # ─── Inicialización con catálogo SQL + archivos del scanner ─────────────
    def __init__(self, catalogo_sql: dict, archivos: list, scanner=None):
        """
        Recibe:
          catalogo_sql: dict del SQLParser con la clave 'tablas'
          archivos:     lista de archivos del Scanner (para leer su contenido)
        """
        self.tablas_definidas = set(catalogo_sql.get("tablas", {}).keys())
        self.archivos         = archivos
        self.scanner          = scanner
        # Resultado bidireccional
        self.tabla_a_archivos = defaultdict(list)   # {tabla: [{ruta, framework, linea}]}
        self.archivo_a_tablas = defaultdict(set)    # {ruta: {tabla1, tabla2, ...}}
        # Tablas referenciadas pero NO en el catálogo (posible error)
        self.tablas_huerfanas_codigo = set()

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Recorre cada archivo de código y aplica los patrones de su lenguaje.
        """
        if not self.tablas_definidas and not self.archivos:
            return self._resultado_vacio()

        # Lenguajes que pueden contener queries o ORM
        lenguajes_relevantes = {
            "Python", "JavaScript", "TypeScript", "Java",
            "C#", "PHP", "Ruby", "R", "SQL",
        }

        for archivo in self.archivos:
            if archivo.get("lenguaje") not in lenguajes_relevantes:
                continue
            try:
                with open(archivo["ruta_abs"], "r",
                          encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception:
                continue

            referencias = self._buscar_referencias(
                contenido, archivo["lenguaje"], archivo["ruta_rel"]
            )
            for ref in referencias:
                tabla = ref["tabla"]
                self.tabla_a_archivos[tabla].append({
                    "ruta":      archivo["ruta_rel"],
                    "framework": ref["framework"],
                    "linea":     ref["linea"],
                })
                self.archivo_a_tablas[archivo["ruta_rel"]].add(tabla)
                # Marcar como huérfana si no está definida en SQL
                if self.tablas_definidas and tabla not in self.tablas_definidas:
                    self.tablas_huerfanas_codigo.add(tabla)

        return self._construir_resultado()

    # ════════════════════════════════════════════════════════════════════════
    # CATÁLOGO DE PATRONES POR FRAMEWORK
    # Cada patrón captura el NOMBRE DE LA TABLA en su grupo 1
    # ════════════════════════════════════════════════════════════════════════

    # ─── Patrones SQL crudo (válidos en cualquier lenguaje host) ────────────
    _PATRONES_SQL_CRUDO = [
        ("sql_from",   re.compile(r"\bFROM\s+[\"'`\[]?(\w+)[\"'`\]]?\s*(?:[,)\s]|$)", re.IGNORECASE)),
        ("sql_join",   re.compile(r"\bJOIN\s+[\"'`\[]?(\w+)[\"'`\]]?", re.IGNORECASE)),
        ("sql_into",   re.compile(r"\bINSERT\s+INTO\s+[\"'`\[]?(\w+)[\"'`\]]?", re.IGNORECASE)),
        ("sql_update", re.compile(r"\bUPDATE\s+[\"'`\[]?(\w+)[\"'`\]]?\s+SET", re.IGNORECASE)),
        ("sql_delete", re.compile(r"\bDELETE\s+FROM\s+[\"'`\[]?(\w+)[\"'`\]]?", re.IGNORECASE)),
    ]

    # ─── Patrones por framework / lenguaje ──────────────────────────────────
    _PATRONES_FRAMEWORK = {
        "Python": [
            # SQLAlchemy declarativo
            ("sqlalchemy_tablename", re.compile(r"__tablename__\s*=\s*[\"'](\w+)[\"']")),
            ("sqlalchemy_table",     re.compile(r"\bTable\(\s*[\"'](\w+)[\"']")),
            # Django ORM (db_table en Meta)
            ("django_db_table",      re.compile(r"\bdb_table\s*=\s*[\"'](\w+)[\"']")),
            # Supabase Python
            ("supabase_from",        re.compile(r"\.\s*from_?\s*\(\s*[\"'](\w+)[\"']\s*\)")),
            # psycopg2/asyncpg execute con strings
            ("psycopg_execute",      re.compile(r"execute\s*\(\s*[\"'](?:.*?\b(?:FROM|INTO|UPDATE|DELETE\s+FROM)\s+(\w+))", re.IGNORECASE | re.DOTALL)),
        ],
        "JavaScript": [
            # Supabase JS
            ("supabase_from",   re.compile(r"\.\s*from\s*\(\s*[\"'`](\w+)[\"'`]\s*\)")),
            # Knex
            ("knex_table",      re.compile(r"\bknex\s*\(\s*[\"'`](\w+)[\"'`]\s*\)")),
            # Sequelize
            ("sequelize_def",   re.compile(r"sequelize\.define\s*\(\s*[\"'`](\w+)[\"'`]")),
            ("sequelize_table", re.compile(r"tableName\s*:\s*[\"'`](\w+)[\"'`]")),
            # Prisma
            ("prisma_model",    re.compile(r"\bprisma\.(\w+)\.")),
        ],
        "TypeScript": [
            ("supabase_from",   re.compile(r"\.\s*from\s*\(\s*[\"'`](\w+)[\"'`]\s*\)")),
            ("knex_table",      re.compile(r"\bknex\s*\(\s*[\"'`](\w+)[\"'`]\s*\)")),
            ("typeorm_table",   re.compile(r"@Entity\s*\(\s*[\"'`](\w+)[\"'`]")),
            ("typeorm_name",    re.compile(r"@Entity\s*\(\s*\{\s*name\s*:\s*[\"'`](\w+)[\"'`]")),
            ("prisma_model",    re.compile(r"\bprisma\.(\w+)\.")),
        ],
        "Java": [
            # JPA / Hibernate
            ("jpa_table",       re.compile(r"@Table\s*\(\s*name\s*=\s*\"(\w+)\"")),
            # JdbcTemplate / strings con SQL
        ],
        "C#": [
            ("ef_table",        re.compile(r"\[Table\(\s*\"(\w+)\"\s*\)\]")),
            ("ef_dbset",        re.compile(r"DbSet<\w+>\s+(\w+)\s*\{\s*get")),
        ],
        "PHP": [
            # Laravel Eloquent
            ("eloquent_table",  re.compile(r"\$table\s*=\s*[\"'](\w+)[\"']")),
            ("db_table",        re.compile(r"DB::table\s*\(\s*[\"'](\w+)[\"']")),
        ],
        "Ruby": [
            # ActiveRecord (rare in scope but cubrimos)
            ("ar_table",        re.compile(r"self\.table_name\s*=\s*[\"'](\w+)[\"']")),
        ],
        "R": [
            # DBI / dplyr
            ("dbi_query",       re.compile(r"dbReadTable\s*\([^,]+,\s*[\"'](\w+)[\"']")),
            ("dplyr_tbl",       re.compile(r"\btbl\s*\([^,]+,\s*[\"'](\w+)[\"']")),
        ],
    }

    # ════════════════════════════════════════════════════════════════════════
    # MOTOR DE BÚSQUEDA
    # ════════════════════════════════════════════════════════════════════════

    # ─── Buscar todas las referencias a tablas en un archivo ────────────────
    def _buscar_referencias(
        self, contenido: str, lenguaje: str, ruta_rel: str
    ) -> list:
        """
        Aplica los patrones SQL crudos + los patrones específicos del lenguaje.
        Retorna lista de dicts {tabla, framework, linea}.
        """
        referencias = []

        # ─── Patrones SQL crudos (universales) ──────────────────────────────
        for nombre, patron in self._PATRONES_SQL_CRUDO:
            for m in patron.finditer(contenido):
                tabla = m.group(1).lower()
                # Filtros mínimos: que el match sea plausible
                if not self._es_nombre_tabla_valido(tabla):
                    continue
                referencias.append({
                    "tabla":     tabla,
                    "framework": nombre,
                    "linea":     contenido[:m.start()].count("\n") + 1,
                })

        # ─── Patrones específicos del lenguaje ──────────────────────────────
        for nombre, patron in self._PATRONES_FRAMEWORK.get(lenguaje, []):
            for m in patron.finditer(contenido):
                tabla = m.group(1).lower()
                if not self._es_nombre_tabla_valido(tabla):
                    continue
                referencias.append({
                    "tabla":     tabla,
                    "framework": nombre,
                    "linea":     contenido[:m.start()].count("\n") + 1,
                })

        return referencias

    # ─── Validar que un match parece ser un nombre de tabla real ────────────
    def _es_nombre_tabla_valido(self, nombre: str) -> bool:
        """
        Filtros mínimos para evitar capturar palabras reservadas o ruido.
        Conservador: solo descartamos lo que claramente no es una tabla.
        """
        if not nombre or len(nombre) < 2:
            return False
        # Palabras reservadas SQL frecuentes que aparecen tras FROM/JOIN/etc.
        reservadas = {
            "where", "group", "order", "having", "limit", "offset", "as",
            "select", "with", "case", "when", "then", "else", "end",
            "exists", "not", "null", "true", "false", "and", "or",
            "values", "set", "into", "lateral", "cross", "natural",
            "left", "right", "inner", "outer", "full", "join", "on", "using",
            "dual",  # Oracle
        }
        if nombre.lower() in reservadas:
            return False
        # No debe ser puro número
        if nombre.isdigit():
            return False
        return True

    # ════════════════════════════════════════════════════════════════════════
    # CONSTRUCCIÓN DEL RESULTADO FINAL
    # ════════════════════════════════════════════════════════════════════════

    # ─── Ensamblar el dict de salida ────────────────────────────────────────
    def _construir_resultado(self) -> dict:
        """
        Estructura final:
          - tabla_a_archivos: dict bidireccional
          - archivo_a_tablas: dict inverso
          - tablas_huerfanas_db: definidas en SQL pero nunca usadas
          - tablas_huerfanas_codigo: usadas en código pero no definidas
          - resumen: métricas globales
        """
        # Normalizar tablas a lowercase para comparación
        tablas_def_lower = {t.lower() for t in self.tablas_definidas}
        tablas_usadas_lower = set(self.tabla_a_archivos.keys())

        huerfanas_db = sorted(tablas_def_lower - tablas_usadas_lower)
        huerfanas_codigo = sorted(self.tablas_huerfanas_codigo)

        # Convertir defaultdicts a dicts normales para serialización
        tabla_archivos = {
            tabla: usos for tabla, usos in self.tabla_a_archivos.items()
        }
        archivo_tablas = {
            archivo: sorted(tablas)
            for archivo, tablas in self.archivo_a_tablas.items()
        }

        return {
            "tabla_a_archivos":         tabla_archivos,
            "archivo_a_tablas":         archivo_tablas,
            "tablas_huerfanas_db":      huerfanas_db,
            "tablas_huerfanas_codigo":  huerfanas_codigo,
            "resumen": {
                "tablas_definidas":      len(tablas_def_lower),
                "tablas_usadas":         len(tablas_usadas_lower),
                "tablas_definidas_y_usadas": len(tablas_def_lower & tablas_usadas_lower),
                "huerfanas_db":          len(huerfanas_db),
                "huerfanas_codigo":      len(huerfanas_codigo),
                "archivos_con_db":       len(archivo_tablas),
            },
        }

    # ─── Resultado vacío cuando no hay datos ────────────────────────────────
    def _resultado_vacio(self) -> dict:
        return {
            "tabla_a_archivos": {}, "archivo_a_tablas": {},
            "tablas_huerfanas_db": [], "tablas_huerfanas_codigo": [],
            "resumen": {
                "tablas_definidas": 0, "tablas_usadas": 0,
                "tablas_definidas_y_usadas": 0,
                "huerfanas_db": 0, "huerfanas_codigo": 0,
                "archivos_con_db": 0,
            },
        }





# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 9 — ETL/ELT DETECTOR
# Detecta pipelines de datos en el proyecto: Airflow, Prefect, Dagster, dbt
# Identifica transformaciones pandas/polars y modelos SQL con CTEs como pipeline
# Output: lista de pipelines detectados + resumen por framework
# ════════════════════════════════════════════════════════════════════════════

class ETLDetector:
    """
    Detector de pipelines ETL/ELT en el proyecto.
    Responsabilidades:
      - Detectar DAGs de Airflow, flows de Prefect, assets de Dagster
      - Identificar modelos dbt (SQL con ref()/source() y schema YAML)
      - Detectar transformaciones con pandas y polars
      - Detectar CTEs SQL complejas como señal de pipeline no trivial
      - Mapear framework, tipo y dependencias de cada pipeline
    """

    def __init__(self, archivos: list, scanner=None):
        """
        Recibe la lista completa de archivos del Scanner.
        Filtra internamente los que pueden contener patrones ETL.
        """
        self.archivos  = archivos
        self.scanner   = scanner
        self.resultado = {}      # {ruta_rel: {tipo_archivo, hallazgos, ok}}
        self.pipelines = []      # [{nombre, tipo, framework, archivo, n_tareas, deps}]

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Recorre todos los archivos y aplica detectores según el tipo.
        Retorna el resultado estructurado con resumen incluido.
        """
        for archivo in self.archivos:
            lang    = archivo.get("lenguaje", "")
            ext     = archivo.get("extension", "").lower()
            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]

            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception:
                continue

            if lang == "Python":
                self._analizar_python_etl(contenido, ruta_rel)
            elif lang == "SQL":
                self._analizar_sql_etl(contenido, ruta_rel)
            elif ext in (".yaml", ".yml"):
                self._analizar_yaml_etl(contenido, ruta_rel)

        return self._construir_resultado()

    # ════════════════════════════════════════════════════════════════════════
    # DETECTORES POR TIPO DE ARCHIVO
    # ════════════════════════════════════════════════════════════════════════

    # ─── Detector Python: Airflow, Prefect, Dagster, pandas, polars ─────────
    def _analizar_python_etl(self, contenido: str, ruta_rel: str) -> None:
        """
        Detecta patrones de orquestación y transformación en archivos Python.
        Registra hallazgos en self.resultado y self.pipelines.
        """
        hallazgos = []

        # ─── Airflow DAG ─────────────────────────────────────────────────────
        if re.search(r"\bDAG\s*\(|from\s+airflow", contenido):
            dag_names = re.findall(r"DAG\s*\(\s*[\"']([^\"']+)[\"']", contenido)
            task_ids  = re.findall(r"task_id\s*=\s*[\"']([^\"']+)[\"']", contenido)
            deps_raw  = re.findall(r"(\w+)\s*>>\s*(\w+)", contenido)
            hallazgos.append({
                "framework":    "airflow",
                "tipo":         "dag",
                "nombres":      dag_names,
                "tareas":       task_ids,
                "dependencias": deps_raw,
            })

        # ─── Prefect Flow ────────────────────────────────────────────────────
        if re.search(r"@flow|from\s+prefect\b", contenido):
            flows = re.findall(
                r"@flow\s*(?:\(.*?name\s*=\s*[\"']([^\"']+)[\"'])?",
                contenido, re.DOTALL
            )
            tasks = re.findall(
                r"@task\s*(?:\(.*?name\s*=\s*[\"']([^\"']+)[\"'])?",
                contenido, re.DOTALL
            )
            hallazgos.append({
                "framework":    "prefect",
                "tipo":         "flow",
                "nombres":      [n for n in flows if n],
                "tareas":       [n for n in tasks if n],
                "dependencias": [],
            })

        # ─── Dagster Asset/Op ────────────────────────────────────────────────
        if re.search(r"@asset|@op|from\s+dagster\b", contenido):
            assets = re.findall(r"@asset\s*\n\s*def\s+(\w+)", contenido)
            ops    = re.findall(r"@op\s*\n\s*def\s+(\w+)", contenido)
            hallazgos.append({
                "framework":    "dagster",
                "tipo":         "asset",
                "nombres":      assets,
                "tareas":       ops,
                "dependencias": [],
            })

        # ─── Transformaciones pandas ─────────────────────────────────────────
        pandas_ops = [op for op in [
            "read_csv", "read_parquet", "read_excel", "read_sql", "read_json",
            "to_csv", "to_parquet", "to_sql", "to_excel",
            "merge", "groupby", "pivot_table", "melt", "concat",
        ] if re.search(rf"\b{op}\s*\(", contenido)]
        if pandas_ops and re.search(r"import\s+pandas|from\s+pandas", contenido):
            hallazgos.append({
                "framework":    "pandas",
                "tipo":         "transformacion",
                "nombres":      [],
                "tareas":       pandas_ops,
                "dependencias": [],
            })

        # ─── Transformaciones polars ─────────────────────────────────────────
        polars_ops = [op for op in [
            "read_csv", "read_parquet", "read_json", "scan_csv", "scan_parquet",
            "write_csv", "write_parquet", "join", "group_by", "with_columns",
        ] if re.search(rf"(?:pl\.{op}|polars\.{op}|\b{op}\s*\()", contenido)]
        if polars_ops and re.search(r"import\s+polars|from\s+polars", contenido):
            hallazgos.append({
                "framework":    "polars",
                "tipo":         "transformacion",
                "nombres":      [],
                "tareas":       polars_ops,
                "dependencias": [],
            })

        # ─── Registrar si hay hallazgos ──────────────────────────────────────
        if hallazgos:
            self.resultado[ruta_rel] = {
                "tipo_archivo": "python_etl",
                "hallazgos":    hallazgos,
                "ok":           True,
            }
            for h in hallazgos:
                nombres = h["nombres"] or [ruta_rel.split("/")[-1].replace(".py", "")]
                for nombre in nombres:
                    self.pipelines.append({
                        "nombre":      nombre or ruta_rel,
                        "tipo":        h["tipo"],
                        "framework":   h["framework"],
                        "archivo":     ruta_rel,
                        "n_tareas":    len(h["tareas"]),
                        "dependencias": h["dependencias"],
                    })

    # ─── Detector SQL: modelos dbt, CTEs pipeline ────────────────────────────
    def _analizar_sql_etl(self, contenido: str, ruta_rel: str) -> None:
        """
        Detecta modelos dbt (ref/source Jinja) y CTEs complejas (3+ CTEs).
        """
        hallazgos = []

        # ─── Modelos dbt: ref() y source() ──────────────────────────────────
        refs    = re.findall(
            r"\{\{\s*ref\s*\(\s*['\"]([^'\"]+)['\"]\s*\)\s*\}\}", contenido
        )
        sources = re.findall(
            r"\{\{\s*source\s*\([^)]*['\"]([^'\"]+)['\"]\s*\)\s*\}\}", contenido
        )
        if refs or sources:
            nombre_modelo = ruta_rel.split("/")[-1].replace(".sql", "")
            hallazgos.append({
                "framework":    "dbt",
                "tipo":         "modelo",
                "nombres":      [nombre_modelo],
                "tareas":       refs,
                "dependencias": refs + sources,
            })

        # ─── CTEs complejas como señal de pipeline ───────────────────────────
        ctes = re.findall(r"WITH\s+(\w+)\s+AS\s*\(", contenido, re.IGNORECASE)
        if len(ctes) >= 3:
            hallazgos.append({
                "framework":    "sql_cte",
                "tipo":         "transformacion",
                "nombres":      ctes,
                "tareas":       ctes,
                "dependencias": [],
            })

        if hallazgos:
            self.resultado[ruta_rel] = {
                "tipo_archivo": "sql_etl",
                "hallazgos":    hallazgos,
                "ok":           True,
            }
            for h in hallazgos:
                nombre = (h["nombres"] or [ruta_rel])[0]
                self.pipelines.append({
                    "nombre":      nombre,
                    "tipo":        h["tipo"],
                    "framework":   h["framework"],
                    "archivo":     ruta_rel,
                    "n_tareas":    len(h["tareas"]),
                    "dependencias": h["dependencias"],
                })

    # ─── Detector YAML: dbt schema ───────────────────────────────────────────
    def _analizar_yaml_etl(self, contenido: str, ruta_rel: str) -> None:
        """
        Detecta configuraciones dbt en YAML (schema.yml, sources.yml).
        """
        hallazgos = []
        if (re.search(r"^\s*version\s*:\s*2", contenido, re.MULTILINE) and
                re.search(r"models:|sources:|seeds:", contenido)):
            modelos = re.findall(r"^\s+-\s+name:\s*(\w+)", contenido, re.MULTILINE)
            hallazgos.append({
                "framework":    "dbt_yaml",
                "tipo":         "schema",
                "nombres":      modelos,
                "tareas":       modelos,
                "dependencias": [],
            })
        if hallazgos:
            self.resultado[ruta_rel] = {
                "tipo_archivo": "yaml_etl",
                "hallazgos":    hallazgos,
                "ok":           True,
            }

    # ─── Construir resultado final ───────────────────────────────────────────
    def _construir_resultado(self) -> dict:
        """
        Ensambla el output final con resumen de frameworks y tipos detectados.
        """
        frameworks = Counter(p["framework"] for p in self.pipelines)
        tipos      = Counter(p["tipo"]      for p in self.pipelines)
        return {
            "archivos":  self.resultado,
            "pipelines": self.pipelines,
            "resumen": {
                "archivos_con_etl": len(self.resultado),
                "total_pipelines":  len(self.pipelines),
                "por_framework":    dict(frameworks),
                "por_tipo":         dict(tipos),
            },
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 10 — CLOUD MIGRATION AUDITOR
# Audita configuraciones cloud: Terraform, CloudFormation, Kubernetes, Docker
# Detecta proveedores: AWS, GCP, Azure, Cloudflare, Supabase, DigitalOcean
# Detecta credenciales hardcodeadas sin exponer el valor (solo línea y tipo)
# Output: inventario cloud + alertas de seguridad clasificadas CRÍTICO/ALTO
# ════════════════════════════════════════════════════════════════════════════

class CloudAuditor:
    """
    Auditor de configuración cloud del proyecto.
    Responsabilidades:
      - Detectar archivos .tf/.hcl (Terraform), YAML de CloudFormation
      - Detectar Dockerfiles y manifiestos Kubernetes
      - Inferir proveedores cloud por tipos de recursos Terraform
      - Escanear credenciales hardcodeadas en TODO el proyecto
      - Calcular inventario de recursos por proveedor
    """

    # ─── Patrones de credenciales hardcodeadas ───────────────────────────────
    _PATRONES_CREDENCIALES = [
        ("AWS Access Key ID",   re.compile(r"\bAKIA[0-9A-Z]{16}\b")),
        ("AWS Secret Key",      re.compile(
            r"(?i)aws.{0,20}(?:secret|key)\s*[=:]\s*[\"']?[A-Za-z0-9/+]{40}[\"']?"
        )),
        ("GitHub Token",        re.compile(
            r"\bghp_[0-9a-zA-Z]{36}\b|\bghs_[0-9a-zA-Z]{36}\b"
        )),
        ("Private Key Block",   re.compile(
            r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----"
        )),
        ("Generic API Key",     re.compile(
            r"(?i)(?:api[_\-]?key|apikey|access[_\-]?token|secret[_\-]?key)"
            r"\s*[=:]\s*[\"'][A-Za-z0-9_\-]{20,}[\"']"
        )),
        ("Password Hardcoded",  re.compile(
            r"(?i)(?:password|passwd|pwd)\s*[=:]\s*[\"'][^\"']{8,}[\"']"
        )),
        ("Connection String DB", re.compile(
            r"(?i)(?:mongodb|postgresql|postgres|mysql|redis)://[^\s\"'<>]+"
        )),
        ("Slack/Discord Token", re.compile(r"\bxox[bpoas]-[0-9A-Za-z\-]{10,}\b")),
    ]

    def __init__(self, archivos: list, scanner=None):
        """
        Recibe todos los archivos del scanner.
        Filtra internamente los relevantes para cloud y credenciales.
        """
        self.archivos       = archivos
        self.scanner        = scanner
        self.resultado      = {}   # {ruta_rel: análisis}
        self.alertas        = []   # [{tipo, archivo, linea, nivel}]
        self.recursos_cloud = []   # [{tipo, nombre, proveedor, archivo, herramienta}]

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Procesa archivos cloud y escanea credenciales en todo el proyecto.
        """
        ext_cloud   = {".tf", ".hcl"}
        nombres_docker = {
            "dockerfile", "dockerfile.dev", "dockerfile.prod",
            "dockerfile.staging", "dockerfile.local",
        }
        ext_scaneo  = {
            ".py", ".js", ".ts", ".java", ".cs", ".php",
            ".rb", ".go", ".yaml", ".yml", ".json", ".toml",
            ".ini", ".cfg", ".tf", ".hcl", ".sh", ".ps1",
        }

        for archivo in self.archivos:
            ext     = archivo.get("extension", "").lower()
            nombre  = archivo.get("nombre", "").lower()
            lang    = archivo.get("lenguaje", "")
            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]

            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception:
                continue

            # ─── Análisis estructural ────────────────────────────────────────
            if ext in ext_cloud:
                self._analizar_terraform(contenido, ruta_rel)
            elif nombre in nombres_docker:
                self._analizar_dockerfile(contenido, ruta_rel)
            elif lang in ("YAML", "JSON"):
                self._analizar_generic_cloud(contenido, ruta_rel)

            # ─── Escanear credenciales en archivos de texto ──────────────────
            if ext in ext_scaneo:
                self._escanear_credenciales(contenido, ruta_rel)

        return self._construir_resultado()

    # ─── Analizar archivos Terraform ─────────────────────────────────────────
    def _analizar_terraform(self, contenido: str, ruta_rel: str) -> None:
        """
        Extrae providers, resources, data sources, modules, variables y outputs.
        """
        providers    = re.findall(r'provider\s+"(\w+)"', contenido)
        resources    = re.findall(r'resource\s+"([\w_]+)"\s+"(\w+)"', contenido)
        data_sources = re.findall(r'data\s+"([\w_]+)"\s+"(\w+)"', contenido)
        modules      = re.findall(r'module\s+"(\w+)"', contenido)
        variables    = re.findall(r'variable\s+"(\w+)"', contenido)
        outputs      = re.findall(r'output\s+"(\w+)"', contenido)

        # Inferir proveedores cloud desde prefijo del tipo de recurso
        proveedores = set(providers)
        for tipo_recurso, _ in resources:
            p = self._inferir_proveedor(tipo_recurso)
            if p != "otro":
                proveedores.add(p)

        self.resultado[ruta_rel] = {
            "tipo":        "terraform",
            "providers":   sorted(proveedores),
            "resources":   [{"tipo": t, "nombre": n} for t, n in resources],
            "data_sources": [{"tipo": t, "nombre": n} for t, n in data_sources],
            "modules":     modules,
            "n_variables": len(variables),
            "n_outputs":   len(outputs),
            "n_recursos":  len(resources),
            "ok":          True,
        }
        for tipo_r, nombre_r in resources:
            self.recursos_cloud.append({
                "tipo":        tipo_r,
                "nombre":      nombre_r,
                "proveedor":   self._inferir_proveedor(tipo_r),
                "archivo":     ruta_rel,
                "herramienta": "terraform",
            })

    # ─── Analizar Dockerfile ─────────────────────────────────────────────────
    def _analizar_dockerfile(self, contenido: str, ruta_rel: str) -> None:
        """
        Detecta imagen base, puertos expuestos y variables de entorno.
        """
        imagenes = re.findall(r"^FROM\s+([^\s]+)",  contenido, re.MULTILINE | re.IGNORECASE)
        puertos  = re.findall(r"^EXPOSE\s+(\d+)",   contenido, re.MULTILINE | re.IGNORECASE)
        env_keys = re.findall(r"^ENV\s+(\w+)",      contenido, re.MULTILINE | re.IGNORECASE)

        self.resultado[ruta_rel] = {
            "tipo":         "dockerfile",
            "providers":    [],
            "imagenes_base": imagenes,
            "puertos":      puertos,
            "env_keys":     env_keys,
            "n_recursos":   1,
            "ok":           True,
        }

    # ─── Analizar YAML/JSON genérico (CloudFormation, Kubernetes) ────────────
    def _analizar_generic_cloud(self, contenido: str, ruta_rel: str) -> None:
        """
        Detecta CloudFormation templates y manifiestos Kubernetes.
        """
        # ─── AWS CloudFormation ──────────────────────────────────────────────
        if re.search(r'"AWSTemplateFormatVersion"|AWSTemplateFormatVersion:', contenido):
            recursos = re.findall(r"Type:\s*[\"']?(AWS::\w+::\w+)[\"']?", contenido)
            self.resultado[ruta_rel] = {
                "tipo":       "cloudformation",
                "providers":  ["aws"],
                "resources":  [{"tipo": r, "nombre": ""} for r in recursos],
                "n_recursos": len(recursos),
                "ok":         True,
            }
            for r in recursos:
                self.recursos_cloud.append({
                    "tipo": r, "nombre": "", "proveedor": "aws",
                    "archivo": ruta_rel, "herramienta": "cloudformation",
                })
            return

        # ─── Kubernetes ─────────────────────────────────────────────────────
        if re.search(
            r"apiVersion:|kind:\s*(Deployment|Service|Pod|Ingress"
            r"|StatefulSet|DaemonSet|ConfigMap|Secret|HorizontalPodAutoscaler)",
            contenido
        ):
            kind_m  = re.search(r"kind:\s*(\w+)", contenido)
            kind    = kind_m.group(1) if kind_m else "unknown"
            name_m  = re.search(r"name:\s*([\w\-]+)", contenido)
            nombre  = name_m.group(1) if name_m else ruta_rel
            self.resultado[ruta_rel] = {
                "tipo":       "kubernetes",
                "providers":  ["k8s"],
                "kind":       kind,
                "nombre":     nombre,
                "n_recursos": 1,
                "ok":         True,
            }

    # ─── Escanear credenciales hardcodeadas ──────────────────────────────────
    def _escanear_credenciales(self, contenido: str, ruta_rel: str) -> None:
        """
        Busca patrones de secretos sin almacenar el valor real.
        Clasifica en CRÍTICO o ALTO según el tipo de secreto.
        """
        criticos = {"AWS Access Key ID", "AWS Secret Key",
                    "GitHub Token", "Private Key Block", "Slack/Discord Token"}
        for tipo, patron in self._PATRONES_CREDENCIALES:
            for m in patron.finditer(contenido):
                linea = contenido[:m.start()].count("\n") + 1
                self.alertas.append({
                    "tipo":    tipo,
                    "archivo": ruta_rel,
                    "linea":   linea,
                    "nivel":   "CRÍTICO" if tipo in criticos else "ALTO",
                })

    # ─── Inferir proveedor cloud por prefijo del tipo de recurso ─────────────
    def _inferir_proveedor(self, tipo_recurso: str) -> str:
        """Mapea el prefijo del tipo de recurso Terraform al cloud provider."""
        tr = tipo_recurso.lower()
        if tr.startswith("aws_"):           return "aws"
        if tr.startswith("google_"):        return "gcp"
        if tr.startswith("azurerm_"):       return "azure"
        if tr.startswith("azuread_"):       return "azure"
        if tr.startswith("cloudflare_"):    return "cloudflare"
        if tr.startswith("supabase_"):      return "supabase"
        if tr.startswith("digitalocean_"):  return "digitalocean"
        if tr.startswith("hcloud_"):        return "hetzner"
        return "otro"

    # ─── Construir resultado final ───────────────────────────────────────────
    def _construir_resultado(self) -> dict:
        """
        Ensambla inventario cloud + alertas con resumen ejecutivo.
        """
        proveedores = set()
        tipos_arch  = Counter()
        for datos in self.resultado.values():
            for p in datos.get("providers", []):
                proveedores.add(p)
            tipos_arch[datos.get("tipo", "otro")] += 1

        n_criticas = sum(1 for a in self.alertas if a["nivel"] == "CRÍTICO")
        n_altas    = sum(1 for a in self.alertas if a["nivel"] == "ALTO")

        return {
            "archivos":       self.resultado,
            "alertas":        self.alertas,
            "recursos_cloud": self.recursos_cloud,
            "resumen": {
                "archivos_cloud":        len(self.resultado),
                "proveedores":           sorted(proveedores),
                "por_tipo_herramienta":  dict(tipos_arch),
                "total_recursos":        len(self.recursos_cloud),
                "alertas_criticas":      n_criticas,
                "alertas_altas":         n_altas,
                "total_alertas":         len(self.alertas),
            },
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 11 — DATA WAREHOUSE AUDITOR
# Analiza el modelo dimensional usando el catálogo de tablas del SQLParser
# Detecta: dim_*/fact_*/stg_*/mart_*/raw_*, star schema vs snowflake
# Detecta: SCD tipo 1/2/3 por estructura de columnas de historial
# Output: modelo dimensional inferido + score de calidad 0-100 + problemas
# ════════════════════════════════════════════════════════════════════════════

class DWHAuditor:
    """
    Auditor de modelo dimensional Data Warehouse.
    Responsabilidades:
      - Clasificar tablas por convención de nomenclatura DWH
      - Analizar tablas de dimensión (surrogate key, natural key, SCD)
      - Analizar tablas de hechos (métricas, FKs a dims, granularidad temporal)
      - Detectar tipo de schema (star, snowflake, dimensional simple)
      - Calcular score de calidad con penalizaciones por problemas comunes
      - Detectar tipo de SCD (1/2/3) por columnas de historial de vigencia
    """

    def __init__(self, catalogo_sql: dict, archivos_sql: list, scanner=None):
        """
        Recibe el catálogo de tablas construido por SQLParser.
        """
        self.catalogo     = catalogo_sql
        self.archivos_sql = archivos_sql
        self.scanner      = scanner

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Clasifica cada tabla del catálogo y construye el modelo dimensional.
        Retorna el modelo con score de calidad y lista de problemas detectados.
        """
        tablas = self.catalogo.get("tablas", {})
        if not tablas:
            return self._resultado_vacio()

        dims    = {}
        facts   = {}
        staging = {}
        marts   = {}
        raw     = {}
        otras   = {}

        # ─── Clasificar tablas por prefijo ───────────────────────────────────
        for nombre, datos in tablas.items():
            n        = nombre.lower()
            columnas = datos.get("columnas", [])

            if n.startswith(("dim_", "dimension_", "d_")):
                dims[nombre] = self._analizar_dimension(nombre, columnas, datos)
            elif n.startswith(("fact_", "fct_", "f_")):
                facts[nombre] = self._analizar_fact(nombre, columnas, datos)
            elif n.startswith(("stg_", "staging_", "stage_")):
                staging[nombre] = {
                    "nombre": nombre, "n_columnas": len(columnas),
                    "ruta": datos.get("ruta", ""),
                }
            elif n.startswith(("mart_", "dm_", "obt_", "wide_")):
                marts[nombre] = {
                    "nombre": nombre, "n_columnas": len(columnas),
                    "ruta": datos.get("ruta", ""),
                }
            elif n.startswith(("raw_", "landing_", "bronze_", "src_")):
                raw[nombre] = {
                    "nombre": nombre, "n_columnas": len(columnas),
                    "ruta": datos.get("ruta", ""),
                }
            else:
                otras[nombre] = {"nombre": nombre, "n_columnas": len(columnas)}

        # ─── Detectar schema y calcular score ───────────────────────────────
        tipo_schema      = self._detectar_tipo_schema(dims, facts)
        score, problemas = self._calcular_score(dims, facts)

        # ─── Detectar SCD para cada dimensión ───────────────────────────────
        for d in dims.values():
            d["scd_tipo"] = self._detectar_scd(d.get("columnas_raw", []))

        return {
            "dimensiones":   dims,
            "hechos":        facts,
            "staging":       staging,
            "marts":         marts,
            "raw":           raw,
            "otras":         otras,
            "tipo_schema":   tipo_schema,
            "score_calidad": score,
            "problemas":     problemas,
            "resumen": {
                "total_dims":    len(dims),
                "total_facts":   len(facts),
                "total_staging": len(staging),
                "total_marts":   len(marts),
                "total_raw":     len(raw),
                "tipo_schema":   tipo_schema,
                "score_calidad": score,
                "n_problemas":   len(problemas),
            },
        }

    # ─── Analizar una tabla de dimensión ─────────────────────────────────────
    def _analizar_dimension(self, nombre: str, columnas: list, datos: dict) -> dict:
        """
        Inspecciona si la dimensión tiene surrogate key, natural key
        y columnas de historial de vigencia (señal de SCD 2 o 3).
        """
        nombres_cols = [
            c if isinstance(c, str) else c.get("nombre", "")
            for c in columnas
        ]
        tiene_sk = any(
            c.endswith(("_key", "_sk")) or c in ("id", "pk")
            for c in nombres_cols
        )
        tiene_nk = any(
            c.endswith(("_id", "_code", "_nk", "_bk", "_natural"))
            for c in nombres_cols
        )
        tiene_fechas_scd = any(c in {
            "valid_from", "valid_to", "start_date", "end_date",
            "effective_date", "expiry_date", "is_current", "is_active",
            "current_flag", "row_is_current", "fecha_inicio", "fecha_fin",
        } for c in nombres_cols)
        return {
            "nombre":           nombre,
            "n_columnas":       len(columnas),
            "tiene_sk":         tiene_sk,
            "tiene_nk":         tiene_nk,
            "tiene_fechas_scd": tiene_fechas_scd,
            "columnas_raw":     nombres_cols,
            "scd_tipo":         None,
            "ruta":             datos.get("ruta", ""),
        }

    # ─── Analizar una tabla de hechos ─────────────────────────────────────────
    def _analizar_fact(self, nombre: str, columnas: list, datos: dict) -> dict:
        """
        Inspecciona si la tabla de hechos tiene FKs a dimensiones,
        métricas y una columna de fecha/periodo de granularidad.
        """
        nombres_cols = [
            c if isinstance(c, str) else c.get("nombre", "")
            for c in columnas
        ]
        sufijos_fk = ("_key", "_sk", "_fk", "_id")
        fks_cols   = [c for c in nombres_cols if c.endswith(sufijos_fk)]
        excluidas  = {"id", "created_at", "updated_at", "fecha", "date",
                      "periodo", "period", "load_date", "inserted_at"}
        metricas   = [
            c for c in nombres_cols
            if not c.endswith(sufijos_fk)
            and c not in excluidas
            and "date" not in c and "fecha" not in c
        ]
        tiene_fecha = any(
            c in {"date", "fecha", "periodo", "period", "event_date",
                  "transaction_date", "created_at", "snapshot_date"}
            or "date" in c or "fecha" in c
            for c in nombres_cols
        )
        return {
            "nombre":      nombre,
            "n_columnas":  len(columnas),
            "n_metricas":  len(metricas),
            "n_fks":       len(fks_cols),
            "tiene_fecha": tiene_fecha,
            "metricas":    metricas[:10],
            "fks_cols":    fks_cols,
            "ruta":        datos.get("ruta", ""),
        }

    # ─── Detectar tipo de SCD ─────────────────────────────────────────────────
    def _detectar_scd(self, columnas: list) -> int:
        """
        SCD 1: sin columnas de historial.
        SCD 2: columnas de fecha de vigencia o flag de fila activa.
        SCD 3: columnas prev_*/previous_* (historial limitado).
        """
        nombres = [c.lower() for c in columnas]
        scd2_indicators = {
            "valid_from", "valid_to", "start_date", "end_date",
            "effective_date", "expiry_date", "is_current", "is_active",
            "current_flag", "row_is_current", "fecha_inicio", "fecha_fin",
        }
        scd3_cols = [
            c for c in nombres
            if c.startswith("prev_") or c.startswith("previous_")
            or c.endswith("_previous") or c.endswith("_anterior")
        ]
        if any(c in scd2_indicators for c in nombres):
            return 2
        if scd3_cols:
            return 3
        return 1

    # ─── Detectar tipo de schema ──────────────────────────────────────────────
    def _detectar_tipo_schema(self, dims: dict, facts: dict) -> str:
        """Heurística para clasificar el schema dimensional."""
        if not dims and not facts:  return "sin_modelo_dimensional"
        if not facts:               return "solo_dimensiones"
        if not dims:                return "solo_hechos"
        if len(dims) >= 3 and len(facts) >= 1:  return "star_schema"
        return "dimensional_simple"

    # ─── Calcular score de calidad ────────────────────────────────────────────
    def _calcular_score(self, dims: dict, facts: dict) -> tuple:
        """
        Penaliza problemas comunes. Retorna (score_0_100, lista_problemas).
        """
        if not dims and not facts:
            return 0, [{"tipo": "sin_modelo",
                         "detalle": "No se detectaron tablas dim_* ni fact_*"}]
        score     = 100
        problemas = []

        for nombre, d in facts.items():
            if d["n_fks"] == 0:
                problemas.append({
                    "tipo": "fact_sin_fks", "tabla": nombre,
                    "detalle": f"Fact '{nombre}' sin columnas FK (_key/_sk/_id)",
                })
                score -= 10
            if not d["tiene_fecha"]:
                problemas.append({
                    "tipo": "fact_sin_fecha", "tabla": nombre,
                    "detalle": f"Fact '{nombre}' sin columna de fecha/período",
                })
                score -= 5

        for nombre, d in dims.items():
            if not d["tiene_sk"]:
                problemas.append({
                    "tipo": "dim_sin_sk", "tabla": nombre,
                    "detalle": f"Dim '{nombre}' sin surrogate key (_key/_sk)",
                })
                score -= 5
            if not d["tiene_nk"]:
                problemas.append({
                    "tipo": "dim_sin_nk", "tabla": nombre,
                    "detalle": f"Dim '{nombre}' sin natural key (_id/_code/_nk)",
                })
                score -= 3

        return max(0, score), problemas

    # ─── Resultado vacío ─────────────────────────────────────────────────────
    def _resultado_vacio(self) -> dict:
        return {
            "dimensiones": {}, "hechos": {}, "staging": {},
            "marts": {}, "raw": {}, "otras": {},
            "tipo_schema": "sin_modelo_dimensional",
            "score_calidad": 0, "problemas": [],
            "resumen": {
                "total_dims": 0, "total_facts": 0, "total_staging": 0,
                "total_marts": 0, "total_raw": 0,
                "tipo_schema": "sin_modelo_dimensional",
                "score_calidad": 0, "n_problemas": 0,
            },
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 12 — DATA PROFILER
# Perfila archivos .csv y .json sin dependencias externas (stdlib csv + json)
# Infiere tipos por columna, detecta PII, cardinalidad y outliers IQR manual
# Sugiere tipos SQL óptimos para cada columna
# Output: perfil completo por dataset + alertas de PII clasificadas
# ════════════════════════════════════════════════════════════════════════════

class DataProfiler:
    """
    Perfilador de archivos de datos CSV y JSON.
    Responsabilidades:
      - Leer CSV con el módulo csv stdlib (sin pandas ni polars)
      - Leer JSON arrays de objetos con el módulo json stdlib
      - Inferir tipo predominante por columna (bool/int/float/date/email/text)
      - Detectar nulos, cardinalidad y outliers (IQR manual)
      - Detectar PII por nombre de columna y por análisis de valores
      - Sugerir el tipo SQL más preciso para cada columna
      - Limitar análisis a MAX_FILAS_ANALISIS filas para performance
    """

    MAX_FILAS_ANALISIS = 10_000

    # ─── Expresiones regulares de tipos de datos ─────────────────────────────
    _RE_EMAIL  = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
    _RE_PHONE  = re.compile(r"^[+]?[\d\s\-().]{7,15}$")
    _RE_DATE   = re.compile(
        r"^\d{4}[-/]\d{2}[-/]\d{2}$|^\d{2}[-/]\d{2}[-/]\d{4}$"
        r"|^\d{4}[-/]\d{2}[-/]\d{2}[T ]\d{2}:\d{2}"
    )
    _RE_INT    = re.compile(r"^-?\d+$")
    _RE_FLOAT  = re.compile(r"^-?\d+[.,]\d+$")
    _RE_BOOL   = re.compile(
        r"^(?:true|false|yes|no|1|0|t|f|y|n|si|s[íi])$", re.IGNORECASE
    )
    _RE_CEDULA = re.compile(r"^\d{7,11}$")   # Cédula colombiana / doc de identidad LATAM

    # ─── Mapeo de nombres de columna a tipo PII ───────────────────────────────
    _PII_NOMBRES = {
        "email":     ["email", "correo", "mail", "e_mail", "e-mail"],
        "telefono":  ["phone", "telefono", "tel", "celular", "movil",
                      "mobile", "fono", "cellular"],
        "cedula":    ["cedula", "dni", "nit", "documento", "national_id",
                      "id_number", "cedula_ciudadania", "cc"],
        "nombre":    ["nombre", "name", "apellido", "lastname", "firstname",
                      "full_name", "nombre_completo"],
        "direccion": ["address", "direccion", "domicilio", "calle", "street",
                      "ciudad", "city"],
        "ip":        ["ip_address", "ip", "remote_addr", "client_ip"],
        "fecha_nac": ["birth_date", "fecha_nacimiento", "dob", "birthday"],
        "salario":   ["salary", "salario", "wage", "sueldo", "income"],
    }

    def __init__(self, archivos: list, scanner=None):
        """Filtra internamente solo CSV y JSON."""
        self.archivos  = [
            a for a in archivos
            if a.get("extension", "").lower() in {".csv", ".json"}
        ]
        self.scanner   = scanner
        self.resultado = {}

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Perfila cada archivo CSV o JSON.
        Salta archivos mayores a 50 MB para no bloquear el análisis.
        """
        LIMITE_BYTES = 50 * 1024 * 1024

        for archivo in self.archivos:
            ext      = archivo.get("extension", "").lower()
            ruta_abs = archivo["ruta_abs"]
            ruta_rel = archivo["ruta_rel"]
            tamanio  = archivo.get("tamanio_b", 0)

            if tamanio > LIMITE_BYTES:
                self.resultado[ruta_rel] = {
                    "ok": False,
                    "error": f"Archivo {tamanio // 1024 // 1024} MB > límite 50 MB",
                    "filas": 0, "columnas": [],
                }
                continue

            try:
                if ext == ".csv":
                    self.resultado[ruta_rel] = self._perfilar_csv(ruta_abs)
                elif ext == ".json":
                    self.resultado[ruta_rel] = self._perfilar_json(ruta_abs)
            except Exception as e:
                self.resultado[ruta_rel] = {
                    "ok": False, "error": str(e), "filas": 0, "columnas": [],
                }

        return {
            "archivos": self.resultado,
            "resumen":  self._construir_resumen(),
        }

    # ════════════════════════════════════════════════════════════════════════
    # LECTORES DE FORMATO
    # ════════════════════════════════════════════════════════════════════════

    # ─── Perfilar CSV con stdlib ──────────────────────────────────────────────
    def _perfilar_csv(self, ruta_abs: str) -> dict:
        """
        Lee el CSV detectando el dialecto automáticamente.
        Analiza hasta MAX_FILAS_ANALISIS filas.
        """
        import csv as csv_mod

        headers   = []
        filas_raw = []

        with open(ruta_abs, "r", encoding="utf-8", errors="replace", newline="") as f:
            muestra = f.read(8192)
            f.seek(0)
            try:
                dialecto = csv_mod.Sniffer().sniff(muestra)
                reader   = csv_mod.reader(f, dialecto)
            except Exception:
                reader = csv_mod.reader(f)

            for i, fila in enumerate(reader):
                if i == 0:
                    headers = fila
                else:
                    filas_raw.append(fila)
                if i > self.MAX_FILAS_ANALISIS:
                    break

        if not headers:
            return {"ok": False, "error": "CSV sin encabezados",
                    "filas": 0, "columnas": [], "formato": "csv"}

        total    = len(filas_raw)
        perfiles = []
        for idx, header in enumerate(headers):
            vals = [fila[idx] for fila in filas_raw if idx < len(fila)]
            perfiles.append(self._perfilar_columna(header, vals, total))

        alertas_pii = [c for c in perfiles if c.get("pii_detectado")]
        return {
            "ok":              True,
            "formato":         "csv",
            "filas":           total,
            "filas_truncadas": total >= self.MAX_FILAS_ANALISIS,
            "n_columnas":      len(headers),
            "columnas":        perfiles,
            "alertas_pii":     alertas_pii,
            "n_alertas_pii":   len(alertas_pii),
        }

    # ─── Perfilar JSON array de objetos ──────────────────────────────────────
    def _perfilar_json(self, ruta_abs: str) -> dict:
        """
        Lee el JSON (solo arrays de objetos homogéneos).
        Toma hasta MAX_FILAS_ANALISIS elementos.
        """
        with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
            raw = f.read(10 * 1024 * 1024)

        try:
            datos = json.loads(raw)
        except json.JSONDecodeError as e:
            return {"ok": False, "error": f"JSON inválido: {e}",
                    "filas": 0, "columnas": [], "formato": "json"}

        if not isinstance(datos, list) or not datos:
            return {"ok": False, "error": "JSON no es un array de objetos",
                    "filas": 0, "columnas": [], "formato": "json"}

        muestra = datos[:self.MAX_FILAS_ANALISIS]

        todas_keys: set = set()
        for obj in muestra:
            if isinstance(obj, dict):
                todas_keys.update(obj.keys())

        total    = len(muestra)
        perfiles = []
        for key in sorted(todas_keys):
            vals = [str(obj.get(key, "")) for obj in muestra if isinstance(obj, dict)]
            perfiles.append(self._perfilar_columna(key, vals, total))

        alertas_pii = [c for c in perfiles if c.get("pii_detectado")]
        return {
            "ok":              True,
            "formato":         "json",
            "filas":           total,
            "filas_truncadas": len(datos) > self.MAX_FILAS_ANALISIS,
            "n_columnas":      len(todas_keys),
            "columnas":        perfiles,
            "alertas_pii":     alertas_pii,
            "n_alertas_pii":   len(alertas_pii),
        }

    # ════════════════════════════════════════════════════════════════════════
    # MOTOR DE PERFILADO POR COLUMNA
    # ════════════════════════════════════════════════════════════════════════

    # ─── Perfil completo de una columna ──────────────────────────────────────
    def _perfilar_columna(self, nombre: str, valores: list, total_filas: int) -> dict:
        """
        Calcula tipo inferido, nulos, cardinalidad, PII y estadísticas.
        """
        nulos_set = {"", "null", "NULL", "None", "NA", "N/A", "nan", "NaN", "none"}
        n_total   = max(total_filas, len(valores), 1)
        n_nulos   = sum(1 for v in valores if v in nulos_set)
        no_nulos  = [v for v in valores if v not in nulos_set]
        n_unicos  = len(set(no_nulos))

        tipo_inf  = self._inferir_tipo(no_nulos[:300])
        pii_tipo  = self._detectar_pii(nombre, no_nulos[:100], tipo_inf)
        stats_num = self._estadisticas_numericas(tipo_inf, no_nulos[:5000])
        tipo_sql  = self._sugerir_tipo_sql(tipo_inf, no_nulos, nombre)

        return {
            "nombre":             nombre,
            "tipo_inferido":      tipo_inf,
            "tipo_sql_sugerido":  tipo_sql,
            "n_total":            n_total,
            "n_nulos":            n_nulos,
            "pct_nulos":          round(n_nulos / n_total * 100, 1),
            "n_unicos":           n_unicos,
            "cardinalidad":       round(n_unicos / max(len(no_nulos), 1), 4),
            "pii_detectado":      pii_tipo is not None,
            "pii_tipo":           pii_tipo,
            "estadisticas":       stats_num,
        }

    # ─── Inferir tipo predominante ────────────────────────────────────────────
    def _inferir_tipo(self, valores: list) -> str:
        """
        El tipo que supera el 60% de la muestra gana.
        Orden de prioridad: bool → int → float → date → email → text.
        """
        if not valores:
            return "vacia"
        conteos: dict = {
            "bool": 0, "integer": 0, "float": 0,
            "date": 0, "email": 0, "text": 0,
        }
        for v in valores:
            s = v.strip()
            if   self._RE_BOOL.match(s):  conteos["bool"]    += 1
            elif self._RE_INT.match(s):   conteos["integer"] += 1
            elif self._RE_FLOAT.match(s): conteos["float"]   += 1
            elif self._RE_DATE.match(s):  conteos["date"]    += 1
            elif self._RE_EMAIL.match(s): conteos["email"]   += 1
            else:                         conteos["text"]    += 1

        n = len(valores)
        for tipo, cnt in sorted(conteos.items(), key=lambda x: -x[1]):
            if cnt / n >= 0.60:
                return tipo
        return "text"

    # ─── Detectar PII con pipeline de 3 pasos (baja tasa de falsos positivos) ──
    def _detectar_pii(self, nombre_col: str, valores: list,
                      tipo_inf: str = "text") -> Optional[str]:
        """
        Pipeline de 3 pasos para detectar PII.
        Paso 1: Descartar columnas residuales y vacías.
        Paso 2: Heurística nombre+dtype → construye CANDIDATOS (no retorna aún).
        Paso 3: Validar con regex sobre muestra real con umbral 70%.
        Esto evita el falso positivo de columnas NOMBRE en listas de materiales,
        PRECIO_UNITARIO clasificado como cédula, y columnas Unnamed vacías.
        """
        # ─── Paso 1: Columnas residuales y vacías ────────────────────────────
        n_lower = nombre_col.lower().strip()
        if n_lower.startswith("unnamed:") or n_lower.startswith("unnamed "):
            return None
        if not valores:
            return None

        # ─── Paso 2: Heurística nombre + dtype (construye candidatos) ────────
        candidatos: list = []

        if tipo_inf in ("text", "email"):
            if any(p in n_lower for p in self._PII_NOMBRES["email"]):
                candidatos.append("email")

        if tipo_inf in ("text", "integer"):
            if any(p in n_lower for p in self._PII_NOMBRES["telefono"]):
                candidatos.append("telefono")

        # Cédula: solo integer o text, NUNCA float (precios tienen decimales)
        if tipo_inf in ("integer", "text"):
            if any(p in n_lower for p in self._PII_NOMBRES["cedula"]):
                candidatos.append("cedula")

        # Nombre: solo text y requiere validación de contenido (no productos)
        if tipo_inf == "text":
            if any(p in n_lower for p in self._PII_NOMBRES["nombre"]):
                candidatos.append("nombre")

        if tipo_inf == "text":
            if any(p in n_lower for p in self._PII_NOMBRES["direccion"]):
                candidatos.append("direccion")

        if tipo_inf in ("text", "integer"):
            if any(p in n_lower for p in self._PII_NOMBRES["ip"]):
                candidatos.append("ip")

        if tipo_inf in ("date", "text"):
            if any(p in n_lower for p in self._PII_NOMBRES["fecha_nac"]):
                candidatos.append("fecha_nac")

        # Salario: solo si el nombre es explícito (no flaggear columnas precio)
        if tipo_inf in ("integer", "float"):
            if any(p in n_lower for p in ["salary", "salario", "sueldo", "wage"]):
                candidatos.append("salario")

        if not candidatos:
            return None

        # ─── Paso 3: Validación con regex sobre muestra real (umbral 70%) ────
        muestra = [str(v).strip() for v in valores[:100] if v]
        if not muestra:
            return None
        n_m    = len(muestra)
        UMBRAL = 0.70

        re_nombre_persona = re.compile(
            r"^[A-ZÁÉÍÓÚÑÜA-Z][a-záéíóúñüa-z]+"
            r"(?:\s+[A-ZÁÉÍÓÚÑÜA-Z][a-záéíóúñüa-z]+)+$"
        )
        re_ip = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
        kw_dir = ["calle", "carrera", "avenida", "transversal",
                  "diagonal", "cra", "kr", "av", "street", "avenue", "road"]

        for tipo in candidatos:
            if tipo == "email":
                if sum(1 for v in muestra if self._RE_EMAIL.match(v)) / n_m >= UMBRAL:
                    return "email"
            elif tipo == "telefono":
                if sum(1 for v in muestra if self._RE_PHONE.match(v)) / n_m >= UMBRAL:
                    return "telefono"
            elif tipo == "cedula":
                if sum(1 for v in muestra if self._RE_CEDULA.match(v)) / n_m >= UMBRAL:
                    return "cedula"
            elif tipo == "nombre":
                # Persona = mínimo 2 palabras capitalizadas; producto = 1 palabra o técnico
                if sum(1 for v in muestra if re_nombre_persona.match(v)) / n_m >= UMBRAL:
                    return "nombre"
            elif tipo == "direccion":
                n_match = sum(1 for v in muestra if any(k in v.lower() for k in kw_dir))
                if n_match / n_m >= 0.50:   # umbral más bajo para direcciones
                    return "direccion"
            elif tipo == "ip":
                if sum(1 for v in muestra if re_ip.match(v)) / n_m >= UMBRAL:
                    return "ip"
            elif tipo == "fecha_nac":
                if sum(1 for v in muestra if self._RE_DATE.match(v)) / n_m >= UMBRAL:
                    return "fecha_nac"
            elif tipo == "salario":
                try:
                    nums = [float(v.replace(",", ".").replace("$", "")) for v in muestra]
                    if nums and (sum(nums) / len(nums)) > 100:
                        return "salario"
                except Exception:
                    pass

        return None

    # ─── Estadísticas numéricas sin numpy ────────────────────────────────────
    def _estadisticas_numericas(self, tipo_inf: str, valores: list) -> dict:
        """
        Min/max/media/mediana/outliers usando IQR manual (sort + percentil por índice).
        """
        if tipo_inf not in ("integer", "float") or not valores:
            return {}
        numeros = []
        for v in valores:
            try:
                numeros.append(float(v.replace(",", ".")))
            except Exception:
                pass
        if not numeros:
            return {}
        numeros.sort()
        n = len(numeros)
        q1 = numeros[n // 4]
        q3 = numeros[3 * n // 4]
        iqr = q3 - q1
        n_out = 0
        if iqr > 0:
            lb = q1 - 1.5 * iqr
            ub = q3 + 1.5 * iqr
            n_out = sum(1 for x in numeros if x < lb or x > ub)
        return {
            "min":        numeros[0],
            "max":        numeros[-1],
            "media":      round(sum(numeros) / n, 4),
            "mediana":    numeros[n // 2],
            "n_outliers": n_out,
        }

    # ─── Sugerir tipo SQL óptimo ──────────────────────────────────────────────
    def _sugerir_tipo_sql(self, tipo_inf: str, valores: list, nombre: str) -> str:
        """
        Retorna el tipo SQL más preciso para la columna.
        Para texto calcula max_len de la muestra y usa VARCHAR apropiado.
        """
        if tipo_inf == "bool":  return "BOOLEAN"
        if tipo_inf == "date":
            return "TIMESTAMP" if any("T" in v or " " in v for v in valores[:10]) else "DATE"
        if tipo_inf == "email": return "VARCHAR(320)"
        if tipo_inf == "vacia": return "TEXT"

        if tipo_inf == "integer":
            if not valores:
                return "INTEGER"
            try:
                maxval = max(abs(int(v)) for v in valores[:200] if self._RE_INT.match(v.strip()))
                if maxval <= 32_767:        return "SMALLINT"
                if maxval <= 2_147_483_647: return "INTEGER"
                return "BIGINT"
            except Exception:
                return "INTEGER"

        if tipo_inf == "float":
            decimales = []
            for v in valores[:100]:
                partes = re.split(r"[,.]", v)
                if len(partes) == 2:
                    decimales.append(len(partes[1]))
            max_dec = max(decimales) if decimales else 6
            return f"DECIMAL(18, {min(max_dec, 10)})"

        # text — inferir largo máximo para VARCHAR
        if not valores:
            return "TEXT"
        max_len = max((len(v) for v in valores[:200]), default=0)
        if max_len <= 50:    return f"VARCHAR({min(max_len * 2, 100)})"
        if max_len <= 255:   return "VARCHAR(255)"
        if max_len <= 1000:  return "VARCHAR(1000)"
        return "TEXT"

    # ─── Resumen agregado del profiling ──────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Resumen global del análisis de datos."""
        ok_list     = [d for d in self.resultado.values() if d.get("ok")]
        total_pii   = sum(d.get("n_alertas_pii", 0) for d in ok_list)
        total_filas = sum(d.get("filas", 0)          for d in ok_list)
        return {
            "archivos_perfilados": len(ok_list),
            "archivos_con_error":  len(self.resultado) - len(ok_list),
            "total_archivos":      len(self.resultado),
            "total_filas":         total_filas,
            "alertas_pii":         total_pii,
        }



# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# BLOQUE TRANSVERSAL — Framework Primary Detector
# Detecta el framework dominante del proyecto ANTES de evaluar métricas.
# Sin este contexto, "0 clases" en Streamlit se reporta como problema
# cuando en realidad es la arquitectura correcta para ese framework.
# ════════════════════════════════════════════════════════════════════════════

class FrameworkDetector:
    """
    Detecta el framework/paradigma dominante del proyecto.
    Responsabilidades:
      - Leer imports de todos los archivos Python del proyecto
      - Detectar frameworks por imports y archivos de configuración
      - Devolver framework primario + metadatos de versión
      - Ajustar umbrales de análisis según el framework
    """

    # ─── Pesos de cada framework por import ──────────────────────────────────
    _FRAMEWORKS_PYTHON = {
        "streamlit":  ["streamlit"],
        "fastapi":    ["fastapi"],
        "flask":      ["flask", "Flask"],
        "django":     ["django", "Django"],
        "dash":       ["dash", "plotly.dash"],
        "gradio":     ["gradio"],
        "airflow":    ["airflow"],
        "prefect":    ["prefect"],
        "dagster":    ["dagster"],
        "pytest":     ["pytest"],
        "notebook":   ["IPython", "ipywidgets"],
    }

    _FRAMEWORKS_JS = {
        "react":      ["react", "React"],
        "nextjs":     ["next", "next/router"],
        "vue":        ["vue", "Vue"],
        "angular":    ["@angular"],
        "express":    ["express"],
        "nestjs":     ["@nestjs"],
        "nuxt":       ["nuxt"],
    }

    _FRAMEWORKS_JAVA = {
        "spring":     ["org.springframework", "springframework"],
        "quarkus":    ["io.quarkus"],
        "micronaut":  ["io.micronaut"],
        "jakarta":    ["jakarta.enterprise"],
    }

    _FRAMEWORKS_CSHARP = {
        "aspnet":     ["Microsoft.AspNetCore", "System.Web.Mvc"],
        "blazor":     ["Microsoft.AspNetCore.Components"],
        "maui":       ["Microsoft.Maui"],
    }

    # ─── Umbrales ajustados por framework ────────────────────────────────────
    _UMBRALES = {
        "streamlit": {
            "lineas_funcion_critico": 400,   # Calibrado con QS Ingeniería: >400L = God-Function real
            "lineas_funcion_alto":    200,   # >200L = revisar extracción de submódulos
            "clases_cero_es_ok":      True,
            "oop_ratio_advertencia":  0.3,
            "cc_umbral_ui":           25,    # vista_*/modulo_*/render* — calibrado con datos reales
            "cc_umbral_logica":       10,    # funciones de utilidad/lógica — McCabe estándar
            "patrones_ui": r"^(vista_|modulo_|render|_render_)",
            "nota":                   "Calibrado con QS Ingeniería (~13K LOC). "
                                      "Dos poblaciones: UI overhead (umbral CC=25) y "
                                      "lógica/utilidades (umbral CC=10). "
                                      "Funciones >400 líneas = CRÍTICO (God-Function).",
        },
        "fastapi": {
            "lineas_funcion_critico": 200,
            "lineas_funcion_alto":    100,
            "clases_cero_es_ok":      True,
            "nota":                   "FastAPI usa handlers funcionales. Modelos son clases Pydantic.",
        },
        "flask": {
            "lineas_funcion_critico": 150,
            "lineas_funcion_alto":    80,
            "clases_cero_es_ok":      True,
            "nota":                   "Flask usa blueprints y funciones. Pocas clases es normal.",
        },
        "django": {
            "lineas_funcion_critico": 150,
            "lineas_funcion_alto":    80,
            "clases_cero_es_ok":      False,
            "nota":                   "Django es OOP. Views, Models y Forms son clases.",
        },
        "generic": {
            "lineas_funcion_critico": 300,
            "lineas_funcion_alto":    150,
            "clases_cero_es_ok":      None,
            "nota":                   "Sin framework dominante detectado.",
        },
    }

    def __init__(self, archivos: list, scanner=None):
        self.archivos = archivos
        self.scanner  = scanner

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def detectar(self) -> dict:
        """
        Analiza los archivos del proyecto y devuelve el framework dominante.
        Retorna: {framework, lenguaje_dominante, version_detectada, umbrales, nota}
        Lee el contenido de cada archivo directamente desde disco.
        """
        conteos: dict = {}

        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            ruta_abs = arch.get("ruta_abs", "")
            if not ruta_abs:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                continue

            if ext == ".py":
                conteos = self._contar_python(contenido, conteos)
            elif ext in (".js", ".ts", ".jsx", ".tsx"):
                conteos = self._contar_js(contenido, conteos)
            elif ext == ".java":
                conteos = self._contar_java(contenido, conteos)
            elif ext == ".cs":
                conteos = self._contar_csharp(contenido, conteos)

        # Detectar versión desde archivos de configuración
        version_info = self._detectar_versiones()

        # Determinar framework dominante
        framework, lenguaje = self._resolver_dominante(conteos)
        umbrales = self._UMBRALES.get(framework, self._UMBRALES["generic"])

        return {
            "framework":          framework,
            "lenguaje_dominante": lenguaje,
            "conteos":            conteos,
            "version_info":       version_info,
            "umbrales":           umbrales,
            "nota":               umbrales["nota"],
        }

    # ─── Contadores por lenguaje ──────────────────────────────────────────────
    def _contar_python(self, contenido: str, conteos: dict) -> dict:
        """Cuenta ocurrencias de imports de frameworks Python."""
        for fw, imports in self._FRAMEWORKS_PYTHON.items():
            for imp in imports:
                patron = re.compile(
                    r"^(?:import|from)\s+" + re.escape(imp), re.MULTILINE
                )
                hits = len(patron.findall(contenido))
                if hits:
                    conteos[fw] = conteos.get(fw, 0) + hits
        return conteos

    def _contar_js(self, contenido: str, conteos: dict) -> dict:
        """Cuenta imports/requires de frameworks JS/TS."""
        for fw, nombres in self._FRAMEWORKS_JS.items():
            for nombre in nombres:
                patron = re.compile(
                    r"""(?:import|require)\s*(?:\(?\s*['"])""" + re.escape(nombre),
                    re.MULTILINE,
                )
                hits = len(patron.findall(contenido))
                if hits:
                    conteos[fw] = conteos.get(fw, 0) + hits
        return conteos

    def _contar_java(self, contenido: str, conteos: dict) -> dict:
        """Cuenta imports de frameworks Java."""
        for fw, paquetes in self._FRAMEWORKS_JAVA.items():
            for pkg in paquetes:
                if pkg in contenido:
                    conteos[fw] = conteos.get(fw, 0) + contenido.count(pkg)
        return conteos

    def _contar_csharp(self, contenido: str, conteos: dict) -> dict:
        """Cuenta using de frameworks C#."""
        for fw, espacios in self._FRAMEWORKS_CSHARP.items():
            for esp in espacios:
                if esp in contenido:
                    conteos[fw] = conteos.get(fw, 0) + contenido.count(esp)
        return conteos

    # ─── Detección de versiones desde archivos de configuración ──────────────
    def _detectar_versiones(self) -> dict:
        """
        Lee manifiestos del proyecto para determinar versiones de lenguajes.
        Soporta: pom.xml (Java), package.json (Node), tsconfig.json,
                 pyproject.toml, .python-version, .csproj, composer.json
        """
        versiones: dict = {}
        archivos_conf = {
            a["ruta_rel"]: a
            for a in self.archivos
            if any(
                a.get("ruta_rel", "").endswith(n)
                for n in ("pom.xml", "package.json", "tsconfig.json",
                          "pyproject.toml", ".python-version",
                          "build.gradle", "composer.json")
            )
            or a.get("extension", "") == ".csproj"
        }

        for ruta_rel, arch in archivos_conf.items():
            ruta_abs_conf = arch.get("ruta_abs", "")
            nombre        = ruta_rel.lower()
            try:
                with open(ruta_abs_conf, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                contenido = ""

            if "pom.xml" in nombre:
                m = re.search(r"<java\.version>([\d.]+)</java\.version>", contenido)
                if m:
                    versiones["java"] = m.group(1)
                m2 = re.search(r"<source>([\d.]+)</source>", contenido)
                if m2 and "java" not in versiones:
                    versiones["java"] = m2.group(1)

            elif "package.json" in nombre:
                m = re.search(r'"node"\s*:\s*"([^"]+)"', contenido)
                if m:
                    versiones["node"] = m.group(1)

            elif "tsconfig.json" in nombre:
                m = re.search(r'"target"\s*:\s*"([^"]+)"', contenido)
                if m:
                    versiones["typescript_target"] = m.group(1)

            elif ".python-version" in nombre:
                versiones["python"] = contenido.strip()

            elif "pyproject.toml" in nombre:
                m = re.search(r'python\s*=\s*"([^"]+)"', contenido)
                if m:
                    versiones["python"] = m.group(1)

            elif nombre.endswith(".csproj"):
                m = re.search(r"<TargetFramework>(net[\w.]+)</TargetFramework>",
                              contenido)
                if m:
                    versiones["dotnet"] = m.group(1)

        return versiones

    # ─── Resolver framework dominante ────────────────────────────────────────
    def _resolver_dominante(self, conteos: dict) -> tuple:
        """
        Retorna (framework_dominante, lenguaje_dominante).
        El framework con más menciones en imports gana.
        """
        if not conteos:
            return ("generic", "desconocido")

        fw_dominante = max(conteos, key=lambda k: conteos[k])
        conteo_max   = conteos[fw_dominante]

        # Un solo import ya es suficiente para identificar el framework
        if conteo_max < 1:
            return ("generic", "python")

        lenguaje_map = {
            "streamlit": "python",  "fastapi": "python",   "flask": "python",
            "django": "python",     "dash": "python",       "gradio": "python",
            "airflow": "python",    "prefect": "python",    "dagster": "python",
            "react": "javascript",  "nextjs": "javascript", "vue": "javascript",
            "angular": "javascript","express": "javascript","nestjs": "javascript",
            "spring": "java",       "quarkus": "java",      "micronaut": "java",
            "aspnet": "csharp",     "blazor": "csharp",
        }
        lenguaje = lenguaje_map.get(fw_dominante, "python")
        return (fw_dominante, lenguaje)


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 13 — KPI ENGINE
# Detecta métricas de negocio (KPIs) en el código fuente.
# Busca patrones de agregación, variables de métricas y visualizaciones
# de indicadores. Permite a la IA entender qué mide el negocio.
# ════════════════════════════════════════════════════════════════════════════

class KPIEngine:
    """
    Motor de detección de KPIs de negocio.
    Responsabilidades:
      - Detectar variables y columnas con nombres de métricas de negocio
      - Identificar agregaciones (sum, count, mean, groupby, SQL AVG/SUM)
      - Detectar visualizaciones de métricas (st.metric, plotly indicators)
      - Clasificar KPIs por categoría (financiero, operativo, temporal)
      - Calcular densidad de KPIs por módulo
    """

    # ─── Patrones de nombres de métricas por categoría ───────────────────────
    _CATEGORIAS_KPI = {
        "financiero": [
            "revenue", "ingreso", "ingresos", "ventas", "venta", "facturacion",
            "factura", "cobro", "cobros", "pago", "pagos", "costo", "costos",
            "gasto", "gastos", "margen", "utilidad", "ganancia", "perdida",
            "presupuesto", "budget", "precio", "tarifa", "monto", "valor",
            "importe", "total", "subtotal", "descuento", "iva", "impuesto",
            "profit", "loss", "earning", "expense", "invoice", "payment",
            "salary", "salario", "sueldo", "anticipo", "deuda",
        ],
        "operativo": [
            "pedido", "pedidos", "orden", "ordenes", "solicitud", "solicitudes",
            "requisicion", "despacho", "entrega", "inventario", "stock",
            "cantidad", "unidades", "capacidad", "eficiencia", "productividad",
            "rendimiento", "throughput", "lead_time", "cycle_time",
            "order", "orders", "request", "delivery", "inventory", "quantity",
            "units", "capacity", "efficiency", "performance",
        ],
        "temporal": [
            "diario", "semanal", "mensual", "trimestral", "anual", "ytd",
            "mtd", "qtd", "acumulado", "historico", "tendencia", "proyeccion",
            "daily", "weekly", "monthly", "quarterly", "annual", "trend",
            "forecast", "projection", "cumulative", "rolling",
        ],
        "calidad": [
            "tasa_error", "error_rate", "precision", "recall", "accuracy",
            "satisfaccion", "nps", "churn", "retencion", "retention",
            "conversion", "bounce", "uptime", "disponibilidad",
        ],
    }

    # ─── Patrones de agregación ───────────────────────────────────────────────
    _RE_AGRUPACION_PY = re.compile(
        r"\.(?:groupby|agg|aggregate|pivot_table|resample)\s*\(",
        re.IGNORECASE,
    )
    _RE_AGG_FUNC_PY = re.compile(
        r"\.(?:sum|count|mean|median|std|var|min|max|nunique|cumsum|rolling)\s*\(",
        re.IGNORECASE,
    )
    _RE_AGG_SQL = re.compile(
        r"\b(?:SUM|COUNT|AVG|MIN|MAX|STDDEV|MEDIAN|GROUP\s+BY)\s*[\(\s]",
        re.IGNORECASE,
    )
    _RE_ST_METRIC = re.compile(
        r"st\.metric\s*\(",
        re.IGNORECASE,
    )
    _RE_PLOTLY_IND = re.compile(
        r"(?:go\.Indicator|px\.bar|px\.line|px\.scatter|"
        r"st\.bar_chart|st\.line_chart|st\.area_chart)\s*\(",
        re.IGNORECASE,
    )

    def __init__(self, archivos: list, scanner=None):
        self.archivos  = archivos
        self.scanner   = scanner
        self.resultado: dict = {}

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Analiza todos los archivos detectando KPIs.
        Retorna resumen global + detalle por archivo.
        """
        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")
            if not ruta_abs:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                continue
            if not contenido.strip():
                continue
            if ext in (".py", ".ipynb"):
                self.resultado[ruta_rel] = self._analizar_python(contenido, ruta_rel)
            elif ext in (".sql",):
                self.resultado[ruta_rel] = self._analizar_sql(contenido, ruta_rel)
            elif ext in (".js", ".ts", ".jsx", ".tsx"):
                self.resultado[ruta_rel] = self._analizar_js(contenido, ruta_rel)

        return {
            "archivos": self.resultado,
            "resumen":  self._construir_resumen(),
        }

    # ─── Análisis Python ─────────────────────────────────────────────────────
    def _analizar_python(self, contenido: str, ruta_rel: str) -> dict:
        """Detecta KPIs en archivos Python/Jupyter."""
        kpis_detectados = self._buscar_nombres_kpi(contenido)
        n_agrupaciones  = len(self._RE_AGRUPACION_PY.findall(contenido))
        n_agg_funcs     = len(self._RE_AGG_FUNC_PY.findall(contenido))
        n_st_metric     = len(self._RE_ST_METRIC.findall(contenido))
        n_viz_metrica   = len(self._RE_PLOTLY_IND.findall(contenido))

        return {
            "tipo":           "python",
            "kpis":           kpis_detectados,
            "n_kpis":         sum(len(v) for v in kpis_detectados.values()),
            "n_agrupaciones": n_agrupaciones,
            "n_agg_funcs":    n_agg_funcs,
            "n_st_metric":    n_st_metric,
            "n_viz_metrica":  n_viz_metrica,
            "score_densidad": n_agrupaciones + n_agg_funcs + n_st_metric,
        }

    # ─── Análisis SQL ─────────────────────────────────────────────────────────
    def _analizar_sql(self, contenido: str, ruta_rel: str) -> dict:
        """Detecta KPIs en archivos SQL."""
        kpis_detectados = self._buscar_nombres_kpi(contenido)
        n_agg_sql       = len(self._RE_AGG_SQL.findall(contenido))

        return {
            "tipo":           "sql",
            "kpis":           kpis_detectados,
            "n_kpis":         sum(len(v) for v in kpis_detectados.values()),
            "n_agrupaciones": n_agg_sql,
            "n_agg_funcs":    0,
            "n_st_metric":    0,
            "n_viz_metrica":  0,
            "score_densidad": n_agg_sql,
        }

    # ─── Análisis JS/TS ───────────────────────────────────────────────────────
    def _analizar_js(self, contenido: str, ruta_rel: str) -> dict:
        """Detecta KPIs en archivos JS/TS."""
        kpis_detectados = self._buscar_nombres_kpi(contenido)
        re_reduce = re.compile(r"\.(reduce|filter|map)\s*\(", re.IGNORECASE)
        n_reduce  = len(re_reduce.findall(contenido))
        return {
            "tipo":           "javascript",
            "kpis":           kpis_detectados,
            "n_kpis":         sum(len(v) for v in kpis_detectados.values()),
            "n_agrupaciones": n_reduce,
            "n_agg_funcs":    0,
            "n_st_metric":    0,
            "n_viz_metrica":  0,
            "score_densidad": n_reduce,
        }

    # ─── Búsqueda de nombres de KPI ──────────────────────────────────────────
    def _buscar_nombres_kpi(self, contenido: str) -> dict:
        """
        Busca ocurrencias de palabras clave de KPI en el contenido.
        Agrupa por categoría. Retorna {categoria: [keyword, ...]}
        """
        contenido_lower = contenido.lower()
        encontrados: dict = {}
        for categoria, palabras in self._CATEGORIAS_KPI.items():
            hits = [p for p in palabras if re.search(r"\b" + re.escape(p) + r"\b",
                                                      contenido_lower)]
            if hits:
                encontrados[categoria] = list(set(hits))
        return encontrados

    # ─── Resumen global ───────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Agrega métricas de todos los archivos."""
        total_kpis  = sum(d.get("n_kpis", 0)          for d in self.resultado.values())
        total_agg   = sum(d.get("n_agrupaciones", 0)   for d in self.resultado.values())
        total_viz   = sum(d.get("n_viz_metrica", 0)    for d in self.resultado.values())
        total_metric = sum(d.get("n_st_metric", 0)     for d in self.resultado.values())

        categorias_global: dict = {}
        for d in self.resultado.values():
            for cat, kws in d.get("kpis", {}).items():
                existing = set(categorias_global.get(cat, []))
                existing.update(kws)
                categorias_global[cat] = list(existing)

        top_archivos = sorted(
            self.resultado.items(),
            key=lambda x: x[1].get("score_densidad", 0),
            reverse=True,
        )[:5]

        return {
            "total_kpis_detectados": total_kpis,
            "total_agregaciones":    total_agg,
            "total_viz_metricas":    total_viz,
            "total_st_metric":       total_metric,
            "categorias":            categorias_global,
            "top_archivos_kpi":      [(a, d.get("score_densidad", 0))
                                      for a, d in top_archivos],
            "n_archivos_con_kpi":    sum(1 for d in self.resultado.values()
                                         if d.get("n_kpis", 0) > 0),
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 14 — VIZ MAPPER
# Inventario de visualizaciones del proyecto.
# Detecta qué librerías de visualización se usan, qué tipo de gráficos,
# y en qué funciones viven. Ayuda a auditar la capa de presentación.
# ════════════════════════════════════════════════════════════════════════════

class VizMapper:
    """
    Mapeador de visualizaciones en el código.
    Responsabilidades:
      - Detectar llamadas a matplotlib, plotly, seaborn, altair, bokeh
      - Detectar componentes Streamlit de visualización
      - Detectar D3.js y Chart.js en JS/HTML
      - Categorizar por tipo de gráfico (barras, línea, scatter, tabla, mapa)
      - Construir inventario de visualizaciones con contexto de función
    """

    # ─── Catálogo de visualizaciones por librería ─────────────────────────────
    _VIZ_PATRONES = {
        "matplotlib": {
            "re": re.compile(
                r"(?:plt|ax|axes)\."
                r"(?:plot|bar|barh|scatter|hist|pie|boxplot|violinplot|"
                r"imshow|contour|heatmap|stem|step|fill_between|"
                r"errorbar|quiver|streamplot)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_estatico",
        },
        "plotly": {
            "re": re.compile(
                r"(?:px|go)\."
                r"(?:bar|line|scatter|pie|histogram|box|violin|"
                r"heatmap|choropleth|scatter_mapbox|funnel|waterfall|"
                r"sunburst|treemap|indicator|candlestick|ohlc|"
                r"scatter_3d|surface|Figure)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_interactivo",
        },
        "seaborn": {
            "re": re.compile(
                r"sns\."
                r"(?:barplot|lineplot|scatterplot|histplot|boxplot|"
                r"violinplot|heatmap|clustermap|pairplot|jointplot|"
                r"regplot|lmplot|catplot|kdeplot|ecdfplot)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_estadistico",
        },
        "altair": {
            "re": re.compile(
                r"alt\.(?:Chart|layer|hconcat|vconcat|concat)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_declarativo",
        },
        "bokeh": {
            "re": re.compile(
                r"(?:figure|p)\."
                r"(?:line|circle|square|triangle|bar|vbar|hbar|"
                r"patch|image|text|annular_wedge)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_interactivo",
        },
        "streamlit": {
            "re": re.compile(
                r"st\."
                r"(?:line_chart|bar_chart|area_chart|scatter_chart|"
                r"map|plotly_chart|altair_chart|bokeh_chart|"
                r"vega_lite_chart|graphviz_chart|pyplot|image|"
                r"dataframe|table|metric)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "componente_streamlit",
        },
        "chartjs": {
            "re": re.compile(
                r"new\s+Chart\s*\(\s*|Chart\.register\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_web",
        },
        "d3": {
            "re": re.compile(
                r"d3\."
                r"(?:select|selectAll|append|attr|style|"
                r"scaleLinear|scaleBand|axisBottom|axisLeft)\s*\(",
                re.IGNORECASE,
            ),
            "tipo": "grafico_web_svg",
        },
    }

    def __init__(self, archivos: list, scanner=None):
        self.archivos  = archivos
        self.scanner   = scanner
        self.resultado: dict = {}

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def analizar(self) -> dict:
        """Mapea todas las visualizaciones del proyecto."""
        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")
            if not ruta_abs:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                continue
            if not contenido.strip():
                continue
            hallazgos = self._detectar_viz(contenido)
            if hallazgos:
                self.resultado[ruta_rel] = {
                    "libreria_dominante": self._libreria_dominante(hallazgos),
                    "n_total":            sum(v["count"] for v in hallazgos.values()),
                    "detalle":            hallazgos,
                }

        return {
            "archivos": self.resultado,
            "resumen":  self._construir_resumen(),
        }

    # ─── Detección de visualizaciones ────────────────────────────────────────
    def _detectar_viz(self, contenido: str) -> dict:
        """Retorna {libreria: {count, tipo}} para el contenido dado."""
        hallazgos: dict = {}
        for lib, cfg in self._VIZ_PATRONES.items():
            matches = cfg["re"].findall(contenido)
            if matches:
                hallazgos[lib] = {
                    "count": len(matches),
                    "tipo":  cfg["tipo"],
                }
        return hallazgos

    # ─── Librería dominante ───────────────────────────────────────────────────
    def _libreria_dominante(self, hallazgos: dict) -> str:
        """Retorna la librería con más ocurrencias."""
        if not hallazgos:
            return "ninguna"
        return max(hallazgos, key=lambda k: hallazgos[k]["count"])

    # ─── Resumen global ───────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Agrega visualizaciones de todo el proyecto."""
        por_libreria: dict = {}
        total_viz = 0
        for datos in self.resultado.values():
            for lib, info in datos.get("detalle", {}).items():
                cnt = info["count"]
                total_viz += cnt
                if lib not in por_libreria:
                    por_libreria[lib] = {"count": 0, "tipo": info["tipo"]}
                por_libreria[lib]["count"] += cnt

        libs_ordenadas = sorted(
            por_libreria.items(), key=lambda x: x[1]["count"], reverse=True
        )
        return {
            "total_visualizaciones": total_viz,
            "n_archivos_con_viz":    len(self.resultado),
            "librerias_usadas":      [l for l, _ in libs_ordenadas],
            "por_libreria":          por_libreria,
            "top_archivos":          sorted(
                self.resultado.items(),
                key=lambda x: x[1].get("n_total", 0),
                reverse=True,
            )[:5],
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 15 — QUERY INTEL
# Auditoría de calidad de queries SQL y ORM.
# Detecta patrones de consulta que degradan el rendimiento:
# N+1, SELECT *, queries sin límite, subqueries correlacionadas.
# ════════════════════════════════════════════════════════════════════════════

class QueryIntel:
    """
    Auditor de inteligencia de queries.
    Responsabilidades:
      - Detectar SELECT * sin cláusula WHERE (full scan)
      - Detectar N+1: query dentro de bucle (for/while)
      - Detectar queries sin LIMIT en tablas potencialmente grandes
      - Detectar subqueries correlacionadas en SQL
      - Detectar cadenas ORM sin paginación (.all() sin .limit())
      - Clasificar hallazgos por severidad CRÍTICO/ALTO/MEDIO
    """

    # ─── Patrones de queries problemáticas ───────────────────────────────────
    _RE_SELECT_STAR = re.compile(
        r"\bSELECT\s+\*\s+FROM\s+(\w+)",
        re.IGNORECASE | re.MULTILINE,
    )
    _RE_SELECT_SIN_WHERE = re.compile(
        r"\bSELECT\b(?:(?!\bWHERE\b).)*?(?:\bFROM\b)(?:(?!\bWHERE\b|\bLIMIT\b|\bJOIN\b).)*?;",
        re.IGNORECASE | re.DOTALL,
    )
    _RE_ORM_ALL_SIN_LIMIT = re.compile(
        r"\.(?:all|fetch_all|execute)\s*\(\s*\)(?!\s*\[)",
        re.IGNORECASE,
    )
    _RE_SUPABASE_SIN_LIMIT = re.compile(
        r"\.select\s*\([^)]*\)(?:(?!\.limit\s*\()(?!\.single\s*\()(?!\.maybeSingle\s*\().)*?\.execute\s*\(\)",
        re.IGNORECASE | re.DOTALL,
    )
    _RE_FOR_LOOP = re.compile(
        r"^\s*(?:for|while)\s+.+:$",
        re.MULTILINE,
    )
    _RE_QUERY_EN_LINEA = re.compile(
        r"\.(?:execute|query|fetchone|fetchall|find|findOne|findAll|"
        r"select|table|from_)\s*\(",
        re.IGNORECASE,
    )
    _RE_SUBQUERY_CORR = re.compile(
        r"\bSELECT\b.+\bWHERE\b.+\b(?:IN|EXISTS)\b\s*\(\s*SELECT\b",
        re.IGNORECASE | re.DOTALL,
    )

    def __init__(self, archivos: list, scanner=None):
        self.archivos  = archivos
        self.scanner   = scanner
        self.hallazgos: list = []

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def analizar(self) -> dict:
        """Audita queries en todos los archivos del proyecto."""
        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")
            if not ruta_abs:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                continue
            if not contenido.strip():
                continue
            if ext == ".sql":
                self._auditar_sql_puro(contenido, ruta_rel)
            elif ext in (".py",):
                self._auditar_python_orm(contenido, ruta_rel)
            elif ext in (".js", ".ts", ".jsx", ".tsx"):
                self._auditar_js_orm(contenido, ruta_rel)

        return {
            "hallazgos": self.hallazgos,
            "resumen":   self._construir_resumen(),
        }

    # ─── Auditoría SQL puro ───────────────────────────────────────────────────
    def _auditar_sql_puro(self, contenido: str, ruta_rel: str) -> None:
        """Detecta problemas en archivos .sql."""
        # SELECT *
        for m in self._RE_SELECT_STAR.finditer(contenido):
            linea = contenido[:m.start()].count("\n") + 1
            self.hallazgos.append({
                "severidad": "MEDIO",
                "tipo":      "select_star",
                "archivo":   ruta_rel,
                "linea":     linea,
                "tabla":     m.group(1),
                "mensaje":   f"SELECT * en tabla '{m.group(1)}'. "
                             "Especificar columnas reduce transferencia de datos.",
                "accion":    "Reemplazar SELECT * por las columnas necesarias.",
            })
        # Subquery correlacionada
        for m in self._RE_SUBQUERY_CORR.finditer(contenido):
            linea = contenido[:m.start()].count("\n") + 1
            self.hallazgos.append({
                "severidad": "ALTO",
                "tipo":      "subquery_correlacionada",
                "archivo":   ruta_rel,
                "linea":     linea,
                "tabla":     "",
                "mensaje":   "Subquery correlacionada detectada (IN/EXISTS con SELECT anidado). "
                             "Se ejecuta una vez por fila del query externo.",
                "accion":    "Reemplazar con JOIN o CTE para mejorar rendimiento.",
            })

    # ─── Auditoría Python ORM ─────────────────────────────────────────────────
    def _auditar_python_orm(self, contenido: str, ruta_rel: str) -> None:
        """Detecta N+1 y queries sin límite en Python."""
        lineas = contenido.split("\n")
        n      = len(lineas)
        indent_for: list = []

        for i, linea in enumerate(lineas):
            # Detectar inicio de bucle
            if self._RE_FOR_LOOP.match(linea):
                indent_nivel = len(linea) - len(linea.lstrip())
                indent_for.append((i, indent_nivel))

            # Detectar query en línea
            if self._RE_QUERY_EN_LINEA.search(linea):
                indent_linea = len(linea) - len(linea.lstrip())
                # Verificar si está dentro de un bucle activo
                dentro_de_bucle = False
                for loop_linea, loop_indent in indent_for:
                    if i > loop_linea and indent_linea > loop_indent:
                        dentro_de_bucle = True
                        break
                if dentro_de_bucle:
                    self.hallazgos.append({
                        "severidad": "CRITICO",
                        "tipo":      "n_plus_1",
                        "archivo":   ruta_rel,
                        "linea":     i + 1,
                        "tabla":     "",
                        "mensaje":   f"Query dentro de bucle en línea {i + 1}. "
                                     "Patrón N+1: se ejecuta 1 query por iteración.",
                        "accion":    "Extraer la query fuera del bucle. "
                                     "Usar IN clause o JOIN para obtener todos los registros de una vez.",
                    })

        # Detectar .all() / .execute() sin .limit()
        for m in self._RE_ORM_ALL_SIN_LIMIT.finditer(contenido):
            linea = contenido[:m.start()].count("\n") + 1
            self.hallazgos.append({
                "severidad": "ALTO",
                "tipo":      "sin_limite",
                "archivo":   ruta_rel,
                "linea":     linea,
                "tabla":     "",
                "mensaje":   f"Query con .all()/.execute() sin .limit() en línea {linea}. "
                             "Puede traer toda la tabla a memoria.",
                "accion":    "Agregar .limit(n) o paginación para tablas grandes.",
            })

    # ─── Auditoría JS/TS ORM ─────────────────────────────────────────────────
    def _auditar_js_orm(self, contenido: str, ruta_rel: str) -> None:
        """Detecta patrones problemáticos en JS/TS (Prisma, Sequelize, Supabase)."""
        lineas = contenido.split("\n")
        re_for_js = re.compile(r"^\s*(?:for|while|forEach|map|filter)\s*[(\[]")

        indent_for: list = []
        for i, linea in enumerate(lineas):
            if re_for_js.match(linea):
                indent_for.append((i, len(linea) - len(linea.lstrip())))
            if self._RE_QUERY_EN_LINEA.search(linea):
                indent_linea = len(linea) - len(linea.lstrip())
                for loop_linea, loop_indent in indent_for:
                    if i > loop_linea and indent_linea > loop_indent:
                        self.hallazgos.append({
                            "severidad": "CRITICO",
                            "tipo":      "n_plus_1",
                            "archivo":   ruta_rel,
                            "linea":     i + 1,
                            "tabla":     "",
                            "mensaje":   f"Query dentro de bucle/map en línea {i + 1} (N+1).",
                            "accion":    "Usar findMany con where IN o include para cargar relaciones en una query.",
                        })
                        break

    # ─── Resumen global ───────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Agrega hallazgos por severidad."""
        por_severidad: dict = {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0}
        por_tipo: dict      = {}
        for h in self.hallazgos:
            sev = h.get("severidad", "BAJO")
            por_severidad[sev] = por_severidad.get(sev, 0) + 1
            t = h.get("tipo", "otro")
            por_tipo[t]        = por_tipo.get(t, 0) + 1
        return {
            "total_hallazgos":  len(self.hallazgos),
            "por_severidad":    por_severidad,
            "por_tipo":         por_tipo,
            "n_archivos_con_problemas": len(set(h["archivo"] for h in self.hallazgos)),
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 16 — FLOW ANALYZER (Business Process Reconstructor)
# La joya de la corona. Reconstruye automáticamente flujos de negocio
# desde el código fuente sin que nadie describa qué es el flujo.
# Detecta máquinas de estado, gaps (estados escritos que nadie lee) y
# bypasses estructurales (saltos en la cadena lógica de aprobación).
# Genera sintaxis Mermaid.js para visualización en el dashboard.
# ════════════════════════════════════════════════════════════════════════════

class FlowAnalyzer:
    """
    Reconstructor de procesos de negocio desde código fuente.
    Responsabilidades:
      - Construir un Shared State Registry: (tabla, campo) → {writers, readers}
      - Detectar campos de estado (estado/status/step/etapa/phase)
      - Extraer valores literales de strings asignados a esos campos
      - Mapear qué módulo produce cada valor y qué módulo lo consume
      - Reconstruir cadenas de transición: ModA → [ValorX] → ModB → [ValorY] → ModC
      - Detectar GAPS: valores escritos que ningún módulo consume
      - Detectar BYPASSES: rutas que omiten un estado intermedio
      - Generar sintaxis Mermaid.js del grafo de flujo
    """

    # ─── Patrones de campos de estado ────────────────────────────────────────
    _RE_CAMPO_ESTADO = re.compile(
        r"\b(?:estado|status|step|etapa|phase|stage|estado_\w+|"
        r"workflow_state|current_state|current_step|current_status)\b",
        re.IGNORECASE,
    )

    # ─── Escrituras a campo de estado (por ORM/framework) ────────────────────
    _PATRONES_WRITE = [
        # Python dict/kwarg: {"estado": "Valor"} o estado="Valor"
        re.compile(
            r"""["'](?:estado|status|step|etapa|phase|stage)["']\s*:\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # Asignación directa: obj.estado = "Valor"
        re.compile(
            r"""\.(?:estado|status|step|etapa|phase|stage)\s*=\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # Supabase/PostgREST: .update({"estado": "Valor"})
        re.compile(
            r"""\.update\s*\(\s*\{[^}]*["'](?:estado|status|step|etapa|phase|stage)["']\s*:\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # SQL UPDATE SET estado = 'Valor'
        re.compile(
            r"""SET\s+(?:estado|status|step|etapa|phase|stage)\s*=\s*['"]([^'"]+)['"]""",
            re.IGNORECASE,
        ),
        # Django ORM: .update(estado="Valor")
        re.compile(
            r"""\.update\s*\([^)]*(?:estado|status|step|etapa|phase|stage)\s*=\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # INSERT INTO ... VALUES con estado
        re.compile(
            r"""INSERT\b.+?['"]([A-Z][A-Za-z\s]+(?:Pendiente|Aprobad|Rechazad|Activ|Inactiv|Complet|Cancel)[^'"]{0,30})['"]""",
            re.IGNORECASE,
        ),
    ]

    # ─── Lecturas de campo de estado (filtros/condiciones) ────────────────────
    _PATRONES_READ = [
        # .eq("estado", "Valor") — Supabase
        re.compile(
            r"""\.eq\s*\(\s*["'](?:estado|status|step|etapa|phase|stage)["']\s*,\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # .filter(estado="Valor") — Django ORM
        re.compile(
            r"""\.(?:filter|get|exclude)\s*\([^)]*(?:estado|status|step|etapa|phase|stage)(?:__\w+)?\s*=\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # WHERE estado = 'Valor' — SQL
        re.compile(
            r"""WHERE\b.+?(?:estado|status|step|etapa|phase|stage)\s*=\s*['"]([^'"]+)['"]""",
            re.IGNORECASE,
        ),
        # .in_("estado", [...]) — Supabase
        re.compile(
            r"""\.in_\s*\(\s*["'](?:estado|status|step|etapa|phase|stage)["']\s*,\s*\[([^\]]+)\]""",
            re.IGNORECASE,
        ),
        # if obj.estado == "Valor" — Python comparison
        re.compile(
            r"""\.(?:estado|status|step|etapa|phase|stage)\s*==\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
        # Prisma/Sequelize: where: { estado: "Valor" }
        re.compile(
            r"""(?:estado|status|step|etapa|phase|stage)\s*:\s*["']([^"']+)["']""",
            re.IGNORECASE,
        ),
    ]

    # ─── Palabras de estado de negocio (para filtrar ruido) ──────────────────
    _PALABRAS_ESTADO = {
        "pendiente", "aprobado", "aprobada", "rechazado", "rechazada",
        "activo", "activa", "inactivo", "inactiva", "completado", "completada",
        "cancelado", "cancelada", "procesando", "procesado", "revision",
        "validacion", "borrador", "publicado", "archivado", "cerrado",
        "abierto", "en_proceso", "girado", "pagado", "facturado",
        "pending", "approved", "rejected", "active", "inactive",
        "completed", "cancelled", "processing", "draft", "published",
        "closed", "open", "review", "validated",
    }

    def __init__(self, archivos: list, catalogo_sql: dict = None, scanner=None):
        self.archivos     = archivos
        self.catalogo_sql = catalogo_sql or {}
        self.scanner      = scanner
        # Shared State Registry: {valor_estado: {writers: [], readers: []}}
        self._registry: dict = {}
        # Mapa de tabla detectada por módulo
        self._modulo_tablas: dict = {}

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Construye el Shared State Registry y luego detecta flujos,
        gaps y bypasses. Genera sintaxis Mermaid para visualización.
        """
        # Paso 1: Construir el registro de estados
        for arch in self.archivos:
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")
            modulo   = self._nombre_modulo(ruta_rel)
            if not ruta_abs:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as _f:
                    contenido = _f.read()
            except Exception:
                continue
            if not contenido.strip():
                continue
            self._registrar_writes(contenido, modulo, ruta_rel)
            self._registrar_reads(contenido, modulo, ruta_rel)

        # Paso 2: Detectar flujos lineales (cadenas de transición)
        flujos    = self._detectar_flujos()

        # Paso 3: Detectar gaps y bypasses
        gaps      = self._detectar_gaps()
        bypasses  = self._detectar_bypasses(flujos)

        # Paso 4: Generar Mermaid
        mermaid   = self._generar_mermaid(flujos, gaps)

        return {
            "registry":     self._registry,
            "flujos":        flujos,
            "gaps":          gaps,
            "bypasses":      bypasses,
            "mermaid":       mermaid,
            "resumen":       self._construir_resumen(flujos, gaps, bypasses),
        }

    # ─── Registro de escrituras (productores) ────────────────────────────────
    def _registrar_writes(self, contenido: str, modulo: str, ruta_rel: str) -> None:
        """Extrae todos los valores de estado que este módulo escribe."""
        for patron in self._PATRONES_WRITE:
            for m in patron.finditer(contenido):
                raw = m.group(1).strip()
                # Para patrón .in_() que retorna lista separada por comas
                valores = self._extraer_lista_valores(raw)
                for valor in valores:
                    valor_norm = valor.strip().strip("'\"")
                    if not self._es_valor_estado(valor_norm):
                        continue
                    linea = contenido[:m.start()].count("\n") + 1
                    if valor_norm not in self._registry:
                        self._registry[valor_norm] = {"writers": [], "readers": []}
                    entry = {"modulo": modulo, "archivo": ruta_rel, "linea": linea}
                    if entry not in self._registry[valor_norm]["writers"]:
                        self._registry[valor_norm]["writers"].append(entry)

    # ─── Registro de lecturas (consumidores) ─────────────────────────────────
    def _registrar_reads(self, contenido: str, modulo: str, ruta_rel: str) -> None:
        """Extrae todos los valores de estado que este módulo lee/filtra."""
        for patron in self._PATRONES_READ:
            for m in patron.finditer(contenido):
                raw = m.group(1).strip()
                valores = self._extraer_lista_valores(raw)
                for valor in valores:
                    valor_norm = valor.strip().strip("'\"")
                    if not self._es_valor_estado(valor_norm):
                        continue
                    linea = contenido[:m.start()].count("\n") + 1
                    if valor_norm not in self._registry:
                        self._registry[valor_norm] = {"writers": [], "readers": []}
                    entry = {"modulo": modulo, "archivo": ruta_rel, "linea": linea}
                    if entry not in self._registry[valor_norm]["readers"]:
                        self._registry[valor_norm]["readers"].append(entry)

    # ─── Detección de flujos lineales ────────────────────────────────────────
    def _detectar_flujos(self) -> list:
        """
        Construye cadenas de transición:
        ModA produce estado S → ModB consume S y produce S2 → ModC consume S2...
        Retorna lista de flujos, cada uno con nombre, pasos y módulos.
        """
        flujos: list = []
        visitados: set = set()

        for valor_inicio, info in self._registry.items():
            if valor_inicio in visitados:
                continue
            if not info["writers"]:
                continue

            # Intentar construir cadena desde este valor
            cadena = self._construir_cadena(valor_inicio, visitados)
            if len(cadena) >= 2:
                flujos.append({
                    "nombre":   self._nombre_flujo(cadena),
                    "pasos":    cadena,
                    "n_pasos":  len(cadena),
                    "modulos":  list(dict.fromkeys(
                        p["modulo_writer"] for p in cadena if p.get("modulo_writer")
                    )),
                    "tipo":     "lineal" if self._es_lineal(cadena) else "ramificado",
                })

        return sorted(flujos, key=lambda f: f["n_pasos"], reverse=True)

    def _construir_cadena(self, valor_inicio: str, visitados: set) -> list:
        """Construye la cadena de transiciones desde un valor de estado inicial."""
        cadena: list = []
        valor_actual = valor_inicio
        max_pasos    = 20   # Evitar ciclos infinitos

        for _ in range(max_pasos):
            if valor_actual not in self._registry:
                break
            info = self._registry[valor_actual]
            if not info["writers"]:
                break

            writer  = info["writers"][0]
            readers = info["readers"]

            paso = {
                "valor":          valor_actual,
                "modulo_writer":  writer["modulo"],
                "archivo_writer": writer["archivo"],
                "linea_writer":   writer["linea"],
                "readers":        readers,
                "tiene_reader":   len(readers) > 0,
            }
            cadena.append(paso)
            visitados.add(valor_actual)

            # Buscar siguiente valor: el que el primer reader produce
            siguiente = None
            for reader in readers:
                mod_reader = reader["modulo"]
                # ¿Este módulo produce algún otro valor?
                for v2, info2 in self._registry.items():
                    if v2 == valor_actual or v2 in visitados:
                        continue
                    if any(w["modulo"] == mod_reader for w in info2["writers"]):
                        siguiente = v2
                        break
                if siguiente:
                    break

            if not siguiente:
                break
            valor_actual = siguiente

        return cadena

    # ─── Detección de gaps ────────────────────────────────────────────────────
    def _detectar_gaps(self) -> list:
        """
        Un GAP es un valor de estado que se escribe pero ningún módulo
        diferente al escritor lo consume. El flujo se rompe ahí.
        """
        gaps: list = []
        for valor, info in self._registry.items():
            if not info["writers"]:
                continue
            modulos_writers = {w["modulo"] for w in info["writers"]}
            modulos_readers = {r["modulo"] for r in info["readers"]}
            # Readers que son diferentes a los writers
            readers_externos = modulos_readers - modulos_writers
            if not readers_externos:
                gaps.append({
                    "valor_estado":  valor,
                    "escrito_por":   info["writers"],
                    "n_writers":     len(info["writers"]),
                    "mensaje":       (
                        f"Estado '{valor}' escrito por "
                        f"{', '.join(modulos_writers)} "
                        "pero ningún módulo diferente lo consume. "
                        "Posible flujo roto."
                    ),
                    "severidad":     "CRITICO",
                })
        return gaps

    # ─── Detección de bypasses estructurales ─────────────────────────────────
    def _detectar_bypasses(self, flujos: list) -> list:
        """
        Un BYPASS es cuando un módulo puede escribir un estado avanzado
        sin haber pasado por un estado intermedio. Ejemplo: A escribe
        "Aprobado" directamente sin pasar por "Pendiente Revisión".
        """
        bypasses: list = []
        for flujo in flujos:
            pasos = flujo["pasos"]
            if len(pasos) < 3:
                continue
            for i, paso in enumerate(pasos[2:], start=2):
                writer_avanzado  = paso["modulo_writer"]
                # ¿Este mismo módulo también puede escribir el estado i-1?
                paso_intermedio  = pasos[i - 1]
                writer_intermedio = paso_intermedio["modulo_writer"]
                if writer_avanzado == writer_intermedio:
                    continue
                # ¿El writer_avanzado también escribe directamente desde paso i-2?
                paso_anterior    = pasos[i - 2]
                readers_anterior = {r["modulo"] for r in paso_anterior["readers"]}
                if writer_avanzado in readers_anterior:
                    bypasses.append({
                        "flujo":           flujo["nombre"],
                        "modulo":          writer_avanzado,
                        "estado_saltado":  paso_intermedio["valor"],
                        "estado_origen":   paso_anterior["valor"],
                        "estado_destino":  paso["valor"],
                        "mensaje":         (
                            f"Posible bypass en '{flujo['nombre']}': "
                            f"'{writer_avanzado}' puede ir de "
                            f"'{paso_anterior['valor']}' directo a "
                            f"'{paso['valor']}' sin pasar por "
                            f"'{paso_intermedio['valor']}'."
                        ),
                        "severidad":       "ALTO",
                    })
        return bypasses

    # ─── Generación de Mermaid.js ─────────────────────────────────────────────
    def _generar_mermaid(self, flujos: list, gaps: list) -> str:
        """
        Genera sintaxis Mermaid flowchart para el primer flujo detectado.
        Los gaps se marcan en rojo. Compatible con Mermaid.js v10+.
        """
        if not flujos and not gaps:
            return ""

        lineas = ["flowchart LR"]
        nodos_vistos: set = set()
        valores_gap  = {g["valor_estado"] for g in gaps}

        if flujos:
            flujo = flujos[0]
            pasos = flujo["pasos"]
            for i, paso in enumerate(pasos):
                mod_id = paso["modulo_writer"].replace(".", "_").replace("/", "_")
                val_id = "VAL_" + paso["valor"].replace(" ", "_")[:20]
                val_display = paso["valor"][:25]
                mod_display = paso["modulo_writer"][:20]

                if mod_id not in nodos_vistos:
                    lineas.append(f'    {mod_id}["{mod_display}"]')
                    nodos_vistos.add(mod_id)
                if val_id not in nodos_vistos:
                    if paso["valor"] in valores_gap:
                        lineas.append(
                            f'    {val_id}(["{val_display} GAP"])'
                        )
                        lineas.append(f'    style {val_id} fill:#fc8181')
                    else:
                        lineas.append(f'    {val_id}(("{val_display}"))')
                    nodos_vistos.add(val_id)

                lineas.append(f"    {mod_id} --> {val_id}")

                if paso["readers"]:
                    reader_mod = paso["readers"][0]["modulo"]
                    reader_id  = reader_mod.replace(".", "_").replace("/", "_")
                    if reader_id not in nodos_vistos:
                        lineas.append(f'    {reader_id}["{reader_mod[:20]}"]')
                        nodos_vistos.add(reader_id)
                    lineas.append(f"    {val_id} --> {reader_id}")

        # Agregar gaps sin flujo detectado
        for gap in gaps[:5]:
            val_id  = "GAP_" + gap["valor_estado"].replace(" ", "_")[:20]
            escritores = gap["escrito_por"]
            if escritores:
                mod_id = escritores[0]["modulo"].replace(".", "_").replace("/", "_")
                if val_id not in nodos_vistos:
                    lineas.append(f'    {val_id}(["{gap["valor_estado"][:20]} HUERFANO"])')
                    lineas.append(f"    style {val_id} fill:#fc8181")
                    nodos_vistos.add(val_id)
                if mod_id not in nodos_vistos:
                    lineas.append(
                        f'    {mod_id}["{escritores[0]["modulo"][:20]}"]'
                    )
                    nodos_vistos.add(mod_id)
                lineas.append(f"    {mod_id} --> {val_id}")

        return "\n".join(lineas)

    # ─── Utilidades ──────────────────────────────────────────────────────────
    def _es_valor_estado(self, valor: str) -> bool:
        """
        Determina si un string literal parece un valor de estado de negocio.
        Filtra strings técnicos (URLs, SQL, paths, variables).
        """
        if not valor or len(valor) < 3 or len(valor) > 80:
            return False
        # Descartar si contiene caracteres técnicos típicos (no es estado de negocio)
        for char in ("/", "\\", "<", ">", "{", "}", "[", "]", "=", "?", "&", "#"):
            if char in valor:
                return False
        # Descartar si parece SQL o URL
        for prefix in ("http", "www.", "SELECT", "INSERT", "UPDATE", "DELETE",
                        "FROM ", "WHERE", "table", "column", "field"):
            if valor.lower().startswith(prefix.lower()):
                return False
        # Verificar si contiene alguna palabra clave de estado
        valor_lower = valor.lower().replace("_", " ").replace("-", " ")
        for kw in self._PALABRAS_ESTADO:
            if kw in valor_lower:
                return True
        # Aceptar también si el nombre del campo fue explícito (Paso 1: Pendiente X)
        if any(w[0].isupper() and len(w) > 3 for w in valor.split()):
            return True
        return False

    def _extraer_lista_valores(self, raw: str) -> list:
        """
        Para el patrón .in_() que retorna: '"Valor1", "Valor2", ...'
        Extrae cada valor como elemento separado.
        """
        re_items = re.compile(r"""["']([^"']+)["']""")
        items    = re_items.findall(raw)
        return items if items else [raw]

    def _nombre_modulo(self, ruta_rel: str) -> str:
        """Extrae nombre limpio del módulo desde ruta relativa."""
        import os
        return os.path.splitext(os.path.basename(ruta_rel))[0]

    def _nombre_flujo(self, cadena: list) -> str:
        """Genera nombre automático para el flujo desde los valores detectados."""
        if not cadena:
            return "Flujo desconocido"
        primer_valor = cadena[0]["valor"]
        n_pasos      = len(cadena)
        return f"Flujo ({primer_valor[:30]}...) — {n_pasos} estados"

    def _es_lineal(self, cadena: list) -> bool:
        """Un flujo es lineal si cada paso tiene exactamente un reader."""
        return all(len(p["readers"]) <= 1 for p in cadena)

    # ─── Resumen global ───────────────────────────────────────────────────────
    def _construir_resumen(self, flujos: list, gaps: list, bypasses: list) -> dict:
        """Agrega métricas del análisis de flujo."""
        total_estados = len(self._registry)
        estados_ok    = sum(
            1 for v in self._registry.values()
            if v["writers"] and v["readers"]
        )
        return {
            "total_valores_estado":  total_estados,
            "estados_con_writer":    sum(1 for v in self._registry.values() if v["writers"]),
            "estados_con_reader":    sum(1 for v in self._registry.values() if v["readers"]),
            "estados_completos":     estados_ok,
            "n_flujos_detectados":   len(flujos),
            "n_gaps":                len(gaps),
            "n_bypasses":            len(bypasses),
            "flujos":                [
                {"nombre": f["nombre"], "n_pasos": f["n_pasos"],
                 "tipo": f["tipo"], "modulos": f["modulos"]}
                for f in flujos[:10]
            ],
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 17 — Excel Intelligence
# Detecta y audita archivos Excel/CSV del proyecto.
# Con openpyxl (opcional): fórmulas, hojas, rangos nombrados, macros VBA.
# Sin openpyxl: análisis estructural por nombre de archivo y extensión.
# ════════════════════════════════════════════════════════════════════════════

class ExcelIntelligence:
    """
    Auditor de archivos Excel y CSV del proyecto.
    Responsabilidades:
      - Inventariar todos los .xlsx, .xls, .csv, .xlsm del proyecto
      - Con openpyxl: auditar fórmulas, hojas, rangos, macros (xlsm)
      - Detectar patrones de riesgo: fórmulas volátiles, refs externas,
        hojas ocultas, contraseñas, macros VBA
      - Sin openpyxl: análisis estructural por metadatos del archivo
    """

    _EXTENSIONES = {".xlsx", ".xls", ".xlsm", ".xlsb", ".csv", ".tsv", ".ods"}

    _FORMULAS_VOLATILES = {
        "NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT",
        "INFO", "CELL", "AHORA", "HOY", "ALEATORIO", "DESREF", "INDIRECTO",
    }

    _PATRONES_RIESGO = {
        "referencia_externa": re.compile(r"\[.*?\]|'[A-Za-z]:\\", re.IGNORECASE),
        "formula_array":      re.compile(r"\{="),
        "hyperlink":          re.compile(r"HYPERLINK|HIPERVINCULO", re.IGNORECASE),
    }

    def __init__(self, archivos: list, scanner=None):
        self.archivos = [a for a in archivos
                         if a.get("extension", "").lower() in self._EXTENSIONES]
        self.scanner  = scanner
        self.resultado: dict = {}

    def analizar(self) -> dict:
        """
        Analiza todos los archivos Excel/CSV del proyecto.
        Intenta importar openpyxl — si no está disponible, usa análisis básico.
        """
        if not self.archivos:
            return {"archivos": {}, "resumen": self._resumen_vacio()}

        try:
            import openpyxl as _openpyxl_mod
            _openpyxl = _openpyxl_mod
        except ImportError:
            _openpyxl = None

        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")
            tam_kb   = arch.get("tamanio_kb", 0)

            if ext in (".csv", ".tsv"):
                self.resultado[ruta_rel] = self._analizar_csv(ruta_abs, tam_kb)
            elif _openpyxl and ext in (".xlsx", ".xlsm", ".ods"):
                self.resultado[ruta_rel] = self._analizar_xlsx(
                    ruta_abs, tam_kb, _openpyxl, es_macro=(ext == ".xlsm")
                )
            else:
                self.resultado[ruta_rel] = {
                    "tipo":       ext.lstrip(".").upper(),
                    "tamanio_kb": tam_kb,
                    "analisis":   "basico",
                    "openpyxl":   False,
                    "problemas":  (
                        [{"tipo": "formato_legacy", "severidad": "BAJO",
                          "detalle": "Formato .xls legacy — migrar a .xlsx"}]
                        if ext == ".xls" else []
                    ),
                    "nota": ("Instala openpyxl para análisis profundo: "
                             "pip install openpyxl")
                             if ext not in (".xls", ".xlsb") else
                             "Formato binario — openpyxl no soportado",
                    "ok": True,
                }

        return {"archivos": self.resultado, "resumen": self._construir_resumen()}

    # ─── Análisis CSV / TSV ───────────────────────────────────────────────────
    def _analizar_csv(self, ruta_abs: str, tam_kb: float) -> dict:
        """Lee las primeras líneas del CSV para extraer metadatos básicos."""
        problemas = []
        n_cols    = 0
        separador = ","
        try:
            with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                lineas = [f.readline() for _ in range(10)]
            primera = lineas[0] if lineas else ""
            if primera.count(";") > primera.count(","):
                separador = ";"
            elif "\t" in primera:
                separador = "\t"
            n_cols = len(primera.split(separador))
            if any(c in primera for c in ("Ã", "â€", "Â")):
                problemas.append({
                    "tipo": "encoding_incorrecto", "severidad": "MEDIO",
                    "detalle": "Posible mojibake en encabezado — revisar encoding",
                })
            if tam_kb > 50_000:
                problemas.append({
                    "tipo": "archivo_muy_grande", "severidad": "ALTO",
                    "detalle": f"CSV de {tam_kb:.0f} KB — considerar Parquet o BD",
                })
        except Exception as e:
            problemas.append({"tipo": "error_lectura", "severidad": "BAJO",
                               "detalle": str(e)})
        return {
            "tipo": "CSV", "tamanio_kb": tam_kb, "separador": separador,
            "n_columnas": n_cols, "analisis": "csv_basico",
            "openpyxl": False, "problemas": problemas, "ok": True,
        }

    # ─── Análisis XLSX / XLSM con openpyxl ───────────────────────────────────
    def _analizar_xlsx(self, ruta_abs: str, tam_kb: float,
                       openpyxl, es_macro: bool) -> dict:
        """Abre con openpyxl y audita estructura y fórmulas."""
        problemas = []
        hojas_info = []
        n_formulas = n_formulas_volatiles = n_refs_ext = n_ocultas = 0

        try:
            wb = openpyxl.load_workbook(
                ruta_abs, read_only=True, data_only=False, keep_vba=es_macro
            )
        except Exception as e:
            return {
                "tipo": "XLSM" if es_macro else "XLSX", "tamanio_kb": tam_kb,
                "analisis": "error_apertura", "openpyxl": True,
                "problemas": [{"tipo": "error_apertura", "severidad": "ALTO",
                                "detalle": str(e)}],
                "ok": False,
            }

        for nombre_hoja in wb.sheetnames:
            ws      = wb[nombre_hoja]
            oculta  = ws.sheet_state != "visible"
            n_form_hoja = 0
            if oculta:
                n_ocultas += 1
                problemas.append({
                    "tipo": "hoja_oculta", "severidad": "MEDIO",
                    "detalle": f"Hoja '{nombre_hoja}' está oculta",
                })
            revisadas = 0
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if revisadas > 1000:
                        break
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        n_form_hoja += 1
                        n_formulas  += 1
                        fu = cell.value.upper()
                        if any(vf in fu for vf in self._FORMULAS_VOLATILES):
                            n_formulas_volatiles += 1
                        if self._PATRONES_RIESGO["referencia_externa"].search(cell.value):
                            n_refs_ext += 1
                    revisadas += 1
                if revisadas > 1000:
                    break
            hojas_info.append({
                "nombre": nombre_hoja, "oculta": oculta,
                "filas": ws.max_row or 0, "cols": ws.max_column or 0,
                "formulas": n_form_hoja,
            })

        n_rangos = 0
        try:
            n_rangos = len(list(wb.defined_names))
        except Exception:
            pass
        wb.close()

        if n_formulas_volatiles > 10:
            problemas.append({
                "tipo": "formulas_volatiles", "severidad": "ALTO",
                "detalle": (f"{n_formulas_volatiles} fórmulas volátiles "
                            "(NOW/TODAY/RAND/OFFSET) — degradan rendimiento"),
            })
        elif n_formulas_volatiles > 0:
            problemas.append({
                "tipo": "formulas_volatiles", "severidad": "MEDIO",
                "detalle": f"{n_formulas_volatiles} fórmulas volátiles detectadas",
            })
        if n_refs_ext > 0:
            problemas.append({
                "tipo": "referencias_externas", "severidad": "ALTO",
                "detalle": f"{n_refs_ext} referencias a archivos externos — riesgo de vínculos rotos",
            })
        if es_macro:
            problemas.append({
                "tipo": "macros_vba", "severidad": "MEDIO",
                "detalle": "Archivo .xlsm con macros VBA — revisar seguridad antes de distribuir",
            })
        if tam_kb > 10_000:
            problemas.append({
                "tipo": "archivo_pesado", "severidad": "BAJO",
                "detalle": f"Archivo de {tam_kb:.0f} KB — considerar optimización",
            })

        return {
            "tipo":                 "XLSM" if es_macro else "XLSX",
            "tamanio_kb":           tam_kb,
            "n_hojas":              len(hojas_info),
            "n_hojas_ocultas":      n_ocultas,
            "n_formulas":           n_formulas,
            "n_formulas_volatiles": n_formulas_volatiles,
            "n_refs_externas":      n_refs_ext,
            "n_rangos_nombrados":   n_rangos,
            "tiene_macros":         es_macro,
            "hojas":                hojas_info,
            "analisis":             "openpyxl_completo",
            "openpyxl":             True,
            "problemas":            problemas,
            "ok":                   True,
        }

    # ─── Resumen global ───────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        total        = len(self.resultado)
        n_xlsx       = sum(1 for d in self.resultado.values() if d.get("tipo") in ("XLSX","XLSM"))
        n_csv        = sum(1 for d in self.resultado.values() if d.get("tipo") == "CSV")
        n_macros     = sum(1 for d in self.resultado.values() if d.get("tiene_macros"))
        n_form       = sum(d.get("n_formulas", 0) for d in self.resultado.values())
        n_vol        = sum(d.get("n_formulas_volatiles", 0) for d in self.resultado.values())
        total_probs  = sum(len(d.get("problemas", [])) for d in self.resultado.values())
        con_probs    = sum(1 for d in self.resultado.values() if d.get("problemas"))
        usa_openpyxl = any(d.get("openpyxl") for d in self.resultado.values())
        return {
            "total_archivos":         total,
            "n_xlsx":                 n_xlsx,
            "n_csv":                  n_csv,
            "n_macros_vba":           n_macros,
            "total_formulas":         n_form,
            "formulas_volatiles":     n_vol,
            "archivos_con_problemas": con_probs,
            "total_problemas":        total_probs,
            "usa_openpyxl":           usa_openpyxl,
        }

    def _resumen_vacio(self) -> dict:
        return {
            "total_archivos": 0, "n_xlsx": 0, "n_csv": 0, "n_macros_vba": 0,
            "total_formulas": 0, "formulas_volatiles": 0,
            "archivos_con_problemas": 0, "total_problemas": 0, "usa_openpyxl": False,
        }

# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 18 — STRATUM Score (sistema de severidad ponderada)
# Reemplaza el conteo plano de "N problemas" con un score 0–100 ponderado.
# 4 niveles: CRÍTICO / ALTO / MEDIO / BAJO con impacto diferenciado.
# ════════════════════════════════════════════════════════════════════════════

class StratumScore:
    """
    Motor de puntuación ponderada del proyecto.
    Consolida todos los findings de todos los bloques y produce:
      - Score global 0-100 (100 = sin problemas)
      - Distribución por severidad
      - Top 10 problemas críticos priorizados
      - Badge de salud: EXCELENTE / BUENO / ATENCIÓN / CRÍTICO
    """

    # Pesos de penalización por severidad
    _PESOS = {
        "CRITICO": 8,
        "ALTO":    4,
        "MEDIO":   2,
        "BAJO":    1,
    }

    # Límites para normalizar (score = 100 - min(penalizacion/max_pen * 100, 100))
    _PEN_MAX = 200

    # Badges por rango de score
    _BADGES = [
        (90, "EXCELENTE",  "#68d391", "✅"),
        (70, "BUENO",      "#9ae6b4", "🟢"),
        (50, "ATENCIÓN",   "#f6ad55", "🟡"),
        (30, "DEFICIENTE", "#fc8181", "🔴"),
        ( 0, "CRÍTICO",    "#e53e3e", "🚨"),
    ]

    def __init__(self, resultados: dict):
        """
        resultados: dict con todos los resultado_* de StratumApp.
        Claves esperadas: 'ast', 'heur', 'sql', 'link', 'etl', 'cloud',
                         'dwh', 'profiler', 'kpi', 'viz', 'query', 'flow',
                         'excel', 'framework'
        """
        self.resultados = resultados

    def calcular(self) -> dict:
        """
        Consolida todos los findings y calcula:
          - score calibrado (umbrales por framework — datos reales)
          - score estándar  (McCabe=10 universal — para comparación)
        El delta entre ambos = ruido eliminado por la calibración de framework.

        PEN_MAX dinámico: escala con funciones+métodos del proyecto.
        Fórmula: max(200, total_units * peso_CRITICO)
        Semántica: score=0 cuando cada función tiene >= 1 hallazgo CRÍTICO.
        """
        findings = self._recolectar_findings()
        penalizacion = sum(
            self._PESOS.get(f.get("severidad", "BAJO"), 1)
            for f in findings
        )

        # ── PEN_MAX dinámico: evita colapso a 0 en proyectos grandes ─────
        r_ast    = self.resultados.get("ast")
        pen_max  = self._PEN_MAX          # fallback mínimo (sin archivos Python)
        if r_ast:
            res_ast     = r_ast.get("resumen", {})
            total_units = (
                res_ast.get("total_funciones", 0) +
                res_ast.get("total_metodos",   0)
            )
            if total_units > 0:
                pen_max = max(self._PEN_MAX, total_units * self._PESOS["CRITICO"])

        score = max(0, round(100 - (penalizacion / pen_max) * 100))

        por_severidad = {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0}
        for f in findings:
            sev = f.get("severidad", "BAJO")
            por_severidad[sev] = por_severidad.get(sev, 0) + 1

        # Top 10: primero CRITICO, luego ALTO, luego MEDIO
        orden_sev = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}
        top10 = sorted(
            findings,
            key=lambda f: orden_sev.get(f.get("severidad", "BAJO"), 3)
        )[:10]

        badge_txt, badge_color, badge_icon = "ATENCIÓN", "#f6ad55", "🟡"
        for umbral, txt, color, icon in self._BADGES:
            if score >= umbral:
                badge_txt, badge_color, badge_icon = txt, color, icon
                break

        # ── Score estándar (McCabe=10 universal) — solo para comparación ────
        # Usa el histograma CC del AST para estimar cuántas funciones adicionales
        # se flaggearían si no hubiera calibración de framework.
        score_estandar   = score
        fn_calibradas    = 0   # funciones salvadas por calibración
        if r_ast:
            histo = r_ast.get("resumen", {}).get("histograma_cc", {})
            if histo:
                # zona_gris: CC 11-20, no flaggeadas por calibración UI (umbral=25)
                # Estas se convertirían en ALTO con McCabe=10 universal
                fn_calibradas  = histo.get("zona_gris_n", 0)
                pen_extra      = fn_calibradas * self._PESOS["ALTO"]
                pen_estandar   = penalizacion + pen_extra
                score_estandar = max(0, round(
                    100 - (pen_estandar / pen_max) * 100
                ))

        return {
            "score":             score,
            "score_estandar":    score_estandar,
            "fn_calibradas":     fn_calibradas,
            "delta_score":       score - score_estandar,
            "badge":             badge_txt,
            "badge_color":       badge_color,
            "badge_icon":        badge_icon,
            "penalizacion":      penalizacion,
            "pen_max":           pen_max,
            "total_findings":    len(findings),
            "por_severidad":     por_severidad,
            "top10":             top10,
            "findings":          findings,
        }

    # ─── Recolector universal de findings ────────────────────────────────────
    def _recolectar_findings(self) -> list:
        """
        Extrae problemas de todos los módulos activos.
        Normaliza el formato a {severidad, tipo, origen, detalle, archivo?, linea?}
        """
        findings = []

        # AST Engine Python
        r_ast = self.resultados.get("ast")
        if r_ast:
            for ruta, datos in r_ast.get("archivos", {}).items():
                for p in datos.get("problemas", []):
                    findings.append({
                        "severidad": p.get("severidad", "MEDIO"),
                        "tipo":      p.get("tipo", ""),
                        "origen":    "AST Python",
                        "archivo":   ruta,
                        "linea":     p.get("linea", ""),
                        "detalle":   p.get("detalle", ""),
                    })

        # SQL Parser
        r_sql = self.resultados.get("sql")
        if r_sql:
            for ruta, datos in r_sql.get("archivos", {}).items():
                for p in datos.get("problemas", []):
                    findings.append({
                        "severidad": p.get("severidad", "MEDIO"),
                        "tipo":      p.get("tipo", ""),
                        "origen":    "SQL Parser",
                        "archivo":   ruta,
                        "linea":     p.get("linea", ""),
                        "detalle":   p.get("detalle", ""),
                    })

        # ETL Detector
        r_etl = self.resultados.get("etl")
        if r_etl:
            for h in r_etl.get("hallazgos", []):
                sev = "ALTO" if h.get("tipo_problema") else "BAJO"
                findings.append({
                    "severidad": sev,
                    "tipo":      h.get("tipo_problema", "etl_hallazgo"),
                    "origen":    "ETL Detector",
                    "archivo":   h.get("archivo", ""),
                    "linea":     h.get("linea", ""),
                    "detalle":   h.get("detalle", ""),
                })

        # Cloud Auditor
        r_cloud = self.resultados.get("cloud")
        if r_cloud:
            for h in r_cloud.get("hallazgos", []):
                findings.append({
                    "severidad": h.get("severidad", "MEDIO"),
                    "tipo":      h.get("tipo", "cloud_hallazgo"),
                    "origen":    "Cloud Auditor",
                    "archivo":   h.get("archivo", ""),
                    "linea":     h.get("linea", ""),
                    "detalle":   h.get("descripcion", h.get("detalle", "")),
                })

        # DWH Auditor
        r_dwh = self.resultados.get("dwh")
        if r_dwh:
            for p in r_dwh.get("problemas", []):
                findings.append({
                    "severidad": p.get("severidad", "MEDIO"),
                    "tipo":      p.get("tipo", "dwh_problema"),
                    "origen":    "DWH Auditor",
                    "archivo":   p.get("tabla", ""),
                    "linea":     "",
                    "detalle":   p.get("detalle", ""),
                })

        # Query Intel (ya tiene severidades propias)
        r_query = self.resultados.get("query")
        if r_query:
            for h in r_query.get("hallazgos", []):
                findings.append({
                    "severidad": h.get("severidad", "MEDIO"),
                    "tipo":      h.get("tipo", "query_hallazgo"),
                    "origen":    "Query Intel",
                    "archivo":   h.get("archivo", ""),
                    "linea":     h.get("linea", ""),
                    "detalle":   h.get("mensaje", h.get("detalle", "")),
                })

        # Excel Intelligence
        r_excel = self.resultados.get("excel")
        if r_excel:
            for ruta, datos in r_excel.get("archivos", {}).items():
                for p in datos.get("problemas", []):
                    findings.append({
                        "severidad": p.get("severidad", "BAJO"),
                        "tipo":      p.get("tipo", "excel_problema"),
                        "origen":    "Excel Intelligence",
                        "archivo":   ruta,
                        "linea":     "",
                        "detalle":   p.get("detalle", ""),
                    })

        # Flow Analyzer — gaps = CRITICO, bypasses = ALTO
        r_flow = self.resultados.get("flow")
        if r_flow:
            for g in r_flow.get("gaps", []):
                findings.append({
                    "severidad": "CRITICO",
                    "tipo":      "flujo_roto",
                    "origen":    "Flow Analyzer",
                    "archivo":   ", ".join(
                        w.get("modulo", "") for w in g.get("escrito_por", [])[:2]
                    ),
                    "linea":     "",
                    "detalle":   g.get("mensaje", "Estado huérfano en flujo de negocio"),
                })
            for b in r_flow.get("bypasses", []):
                findings.append({
                    "severidad": "ALTO",
                    "tipo":      "bypass_estructural",
                    "origen":    "Flow Analyzer",
                    "archivo":   b.get("modulo", ""),
                    "linea":     "",
                    "detalle":   b.get("mensaje", "Bypass estructural detectado"),
                })

        # Streamlit Cache Detector
        r_stc = self.resultados.get("st_cache")
        if r_stc and r_stc.get("aplica"):
            for h in r_stc.get("hallazgos", []):
                findings.append({
                    "severidad": h.get("severidad", "MEDIO"),
                    "tipo":      h.get("tipo",      "streamlit_cache"),
                    "origen":    "Streamlit Cache",
                    "archivo":   h.get("archivo", ""),
                    "linea":     h.get("linea",   ""),
                    "detalle":   h.get("detalle", ""),
                })

        # Security Auditor — todos los hallazgos con severidad propia
        r_sec = self.resultados.get("security")
        if r_sec:
            for h in r_sec.get("hallazgos", []):
                findings.append({
                    "severidad": h.get("severidad", "ALTO"),
                    "tipo":      h.get("tipo",      "seguridad"),
                    "origen":    f"Security [{h.get('categoria','General')}]",
                    "archivo":   h.get("archivo", ""),
                    "linea":     h.get("linea",   ""),
                    "detalle":   h.get("detalle", ""),
                })

        return findings


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 19 — Quick Fix Engine
# Para cada finding del STRATUM Score, sugiere una acción concreta,
# el esfuerzo estimado y un snippet de código cuando aplica.
# ════════════════════════════════════════════════════════════════════════════

class QuickFixEngine:
    """
    Motor de sugerencias de corrección para findings de STRATUM Score.
    Responsabilidades:
      - Mapear cada tipo de finding a una receta de fix
      - Estimar esfuerzo: MINUTOS / HORAS / DIAS
      - Generar snippet de código sugerido cuando aplica
      - Priorizar por ratio impacto/esfuerzo
    """

    # Recetas: tipo_problema → {accion, esfuerzo, snippet_template}
    _RECETAS = {
        "funcion_muy_larga": {
            "accion":   "Dividir en subfunciones de responsabilidad única",
            "esfuerzo": "HORAS",
            "snippet":  (
                "# Extraer sección lógica a subfunción\n"
                "def _procesar_seccion_A(datos):\n"
                "    \"\"\"Responsabilidad específica.\"\"\"\n"
                "    ...\n\n"
                "def funcion_original(datos):\n"
                "    resultado_a = _procesar_seccion_A(datos)\n"
                "    ..."
            ),
        },
        "funcion_larga": {
            "accion":   "Revisar si la función tiene más de una responsabilidad",
            "esfuerzo": "HORAS",
            "snippet":  None,
        },
        "complejidad_alta": {
            "accion":   "Simplificar lógica: extraer guard clauses, usar diccionarios de dispatch",
            "esfuerzo": "HORAS",
            "snippet":  (
                "# Guard clause — salida temprana reduce anidamiento\n"
                "def funcion(x):\n"
                "    if not x:\n"
                "        return None  # guard\n"
                "    # lógica principal sin else\n"
                "    ..."
            ),
        },
        "import_no_usado": {
            "accion":   "Eliminar import — reduce tiempo de carga",
            "esfuerzo": "MINUTOS",
            "snippet":  "# Eliminar la línea del import",
        },
        "funcion_no_usada": {
            "accion":   "Verificar si es API pública o código muerto; si es muerto, eliminar",
            "esfuerzo": "MINUTOS",
            "snippet":  None,
        },
        "flujo_roto": {
            "accion":   "Asignar módulo consumidor al estado huérfano o eliminar el estado",
            "esfuerzo": "HORAS",
            "snippet":  (
                "# Agregar handler para el estado huérfano\n"
                "if estado == 'ESTADO_HUERFANO':\n"
                "    # procesar transición\n"
                "    estado = 'SIGUIENTE_ESTADO'"
            ),
        },
        "bypass_estructural": {
            "accion":   "Agregar validación del estado previo antes de la transición",
            "esfuerzo": "HORAS",
            "snippet":  (
                "# Validar estado previo obligatorio\n"
                "ESTADOS_VALIDOS_PREVIOS = {'estado_A', 'estado_B'}\n"
                "if registro.estado not in ESTADOS_VALIDOS_PREVIOS:\n"
                "    raise ValueError(f'Transición inválida desde {registro.estado}')"
            ),
        },
        "formulas_volatiles": {
            "accion":   "Reemplazar NOW()/TODAY() por valores fijos o calcular en Python",
            "esfuerzo": "MINUTOS",
            "snippet":  (
                "# En lugar de fórmula NOW() en celda, calcular en Python:\n"
                "from datetime import datetime\n"
                "ws['A1'] = datetime.now().strftime('%Y-%m-%d %H:%M')"
            ),
        },
        "referencias_externas": {
            "accion":   "Consolidar datos en un solo archivo o usar fuente única (BD/API)",
            "esfuerzo": "DIAS",
            "snippet":  None,
        },
        "macros_vba": {
            "accion":   "Migrar lógica VBA a Python; distribuir solo .xlsx sin macros",
            "esfuerzo": "DIAS",
            "snippet":  None,
        },
        "n_plus_1": {
            "accion":   "Mover query fuera del loop; usar consulta con IN o JOIN",
            "esfuerzo": "HORAS",
            "snippet":  (
                "# ❌ N+1 — query dentro del loop\n"
                "for item in items:\n"
                "    obj = db.query(Model).filter_by(id=item.id).first()\n\n"
                "# ✅ Una sola query\n"
                "ids = [item.id for item in items]\n"
                "objs = db.query(Model).filter(Model.id.in_(ids)).all()\n"
                "mapa = {o.id: o for o in objs}"
            ),
        },
        "select_star": {
            "accion":   "Reemplazar SELECT * por columnas específicas",
            "esfuerzo": "MINUTOS",
            "snippet":  (
                "-- ❌  SELECT * FROM tabla\n"
                "-- ✅  SELECT id, nombre, fecha FROM tabla"
            ),
        },
        "archivo_muy_grande": {
            "accion":   "Particionar el archivo o migrar a base de datos / Parquet",
            "esfuerzo": "DIAS",
            "snippet":  (
                "# Leer CSV en chunks con pandas\n"
                "import pandas as pd\n"
                "for chunk in pd.read_csv('archivo.csv', chunksize=10_000):\n"
                "    procesar(chunk)"
            ),
        },
        "hoja_oculta": {
            "accion":   "Revisar si la hoja oculta es necesaria; documentar propósito",
            "esfuerzo": "MINUTOS",
            "snippet":  None,
        },
    }

    # Orden de esfuerzo para priorización (menor esfuerzo primero)
    _ORDEN_ESFUERZO = {"MINUTOS": 0, "HORAS": 1, "DIAS": 2}

    def __init__(self, findings: list):
        self.findings = findings

    def generar(self) -> dict:
        """
        Genera sugerencias de fix para cada finding.
        Retorna lista ordenada por ratio impacto/esfuerzo.
        """
        sugerencias = []
        orden_sev = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}

        for f in self.findings:
            tipo = f.get("tipo", "")
            receta = self._RECETAS.get(tipo)
            if not receta:
                continue
            sugerencias.append({
                "severidad":  f.get("severidad", "BAJO"),
                "tipo":       tipo,
                "origen":     f.get("origen", ""),
                "archivo":    f.get("archivo", ""),
                "linea":      f.get("linea", ""),
                "problema":   f.get("detalle", ""),
                "accion":     receta["accion"],
                "esfuerzo":   receta["esfuerzo"],
                "snippet":    receta.get("snippet"),
            })

        # Priorizar: CRITICO+MINUTOS primero, BAJO+DIAS último
        sugerencias.sort(key=lambda s: (
            orden_sev.get(s["severidad"], 3),
            self._ORDEN_ESFUERZO.get(s["esfuerzo"], 2),
        ))

        resumen_esfuerzo = {"MINUTOS": 0, "HORAS": 0, "DIAS": 0}
        for s in sugerencias:
            resumen_esfuerzo[s["esfuerzo"]] = resumen_esfuerzo.get(s["esfuerzo"], 0) + 1

        return {
            "sugerencias":       sugerencias,
            "total_sugerencias": len(sugerencias),
            "resumen_esfuerzo":  resumen_esfuerzo,
            "quick_wins":        [s for s in sugerencias
                                  if s["esfuerzo"] == "MINUTOS"],
        }

# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 20 — Pattern Intelligence Capa 1 (Coverage Score)
# El HeuristicEngine calcula ratio de estructuras detectadas vs señales brutas.
# Si ratio < 60% → advertencia de cobertura insuficiente.
# ════════════════════════════════════════════════════════════════════════════

class CoverageScoreEngine:
    """
    Capa 1 del Pattern Intelligence System.
    Evalúa qué tan bien STRATUM está cubriendo el código analizado.
    Si detecta pocas estructuras respecto a las señales brutas presentes,
    emite advertencia de cobertura — el usuario sabe que puede haber
    más código del que STRATUM está viendo.
    """

    # Señales brutas por lenguaje (keywords que indican estructura)
    _SENALES_BRUTAS = {
        "Python":     [r"^def ", r"^class ", r"^async def "],
        "JavaScript": [r"\bfunction\b", r"=>\s*\{", r"\bclass\b"],
        "TypeScript": [r"\bfunction\b", r"=>\s*\{", r"\bclass\b", r"\binterface\b"],
        "Java":       [r"\bclass\b", r"\binterface\b", r"\bvoid\b", r"\bpublic\b.*\("],
        "C#":         [r"\bclass\b", r"\binterface\b", r"\bvoid\b", r"\bpublic\b.*\("],
        "PHP":        [r"\bfunction\b", r"\bclass\b"],
        "R":          [r"<-\s*function\s*\(", r"=\s*function\s*\("],
    }

    _UMBRAL_COBERTURA = 0.60  # < 60% → advertencia

    def __init__(self, archivos: list, resultado_ast: dict = None,
                 resultado_heur: dict = None, scanner=None):
        self.archivos      = archivos
        self.resultado_ast = resultado_ast or {}
        self.resultado_heur = resultado_heur or {}
        self.scanner       = scanner

    def analizar(self) -> dict:
        """
        Para cada archivo analizado, calcula el ratio cobertura.
        Retorna advertencias cuando la cobertura es insuficiente.
        """
        advertencias = []
        coberturas   = {}

        for arch in self.archivos:
            ext      = arch.get("extension", "").lower()
            lenguaje = arch.get("lenguaje", "")
            ruta_rel = arch["ruta_rel"]
            ruta_abs = arch.get("ruta_abs", "")

            patrones = self._SENALES_BRUTAS.get(lenguaje, [])
            if not patrones or not ruta_abs:
                continue

            # Contar señales brutas en el archivo
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    lineas = f.readlines()
            except Exception:
                continue

            n_senales = 0
            for linea in lineas:
                for patron in patrones:
                    if re.search(patron, linea):
                        n_senales += 1
                        break

            if n_senales == 0:
                continue

            # Contar estructuras detectadas por AST o Heuristic
            n_detectadas = 0
            if lenguaje == "Python" and self.resultado_ast:
                datos = self.resultado_ast.get("archivos", {}).get(ruta_rel, {})
                n_detectadas = (
                    datos.get("n_funciones", 0) +
                    datos.get("n_clases", 0) +
                    datos.get("n_metodos", 0)
                )
            elif self.resultado_heur:
                datos = self.resultado_heur.get("archivos", {}).get(ruta_rel, {})
                n_detectadas = (
                    len(datos.get("funciones", [])) +
                    len(datos.get("clases", []))
                )

            ratio = n_detectadas / n_senales if n_senales > 0 else 1.0
            coberturas[ruta_rel] = {
                "lenguaje":    lenguaje,
                "n_senales":   n_senales,
                "n_detectadas": n_detectadas,
                "ratio":       round(ratio, 2),
            }

            if ratio < self._UMBRAL_COBERTURA and n_senales >= 5:
                advertencias.append({
                    "archivo":    ruta_rel,
                    "lenguaje":   lenguaje,
                    "ratio":      round(ratio * 100, 1),
                    "senales":    n_senales,
                    "detectadas": n_detectadas,
                    "mensaje":    (
                        f"Cobertura de análisis: {round(ratio*100,1)}% "
                        f"({n_detectadas}/{n_senales} estructuras detectadas) — "
                        f"posible sintaxis no reconocida o versión no soportada"
                    ),
                })

        # Ratio global
        total_senales    = sum(c["n_senales"]    for c in coberturas.values())
        total_detectadas = sum(c["n_detectadas"] for c in coberturas.values())
        ratio_global = (
            round(total_detectadas / total_senales * 100, 1)
            if total_senales > 0 else 100.0
        )

        return {
            "advertencias":   advertencias,
            "coberturas":     coberturas,
            "ratio_global":   ratio_global,
            "total_archivos_evaluados": len(coberturas),
            "n_advertencias": len(advertencias),
            "cobertura_ok":   ratio_global >= (self._UMBRAL_COBERTURA * 100),
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 21 — Pattern Intelligence Capa 2 (Language Version Detection)
# Lee manifiestos del proyecto para determinar la versión exacta de
# cada lenguaje y activar patrones versionados cuando aplique.
# ════════════════════════════════════════════════════════════════════════════

class LanguageVersionDetector:
    """
    Capa 2 del Pattern Intelligence System.
    Lee archivos de configuración del proyecto para detectar versiones
    de lenguajes y frameworks. Emite advertencias si STRATUM no tiene
    patrones específicos para esa versión.
    """

    # Versiones con soporte de patrones propio en STRATUM
    _VERSIONES_SOPORTADAS = {
        "python":     ["3.8", "3.9", "3.10", "3.11", "3.12", "3.13"],
        "node":       ["16", "18", "20", "22"],
        "java":       ["8", "11", "17", "21"],
        "dotnet":     ["6.0", "7.0", "8.0", "9.0"],
        "typescript": ["4", "5"],
        "php":        ["7.4", "8.0", "8.1", "8.2", "8.3"],
    }

    # Manifiestos que revelan versiones — nombre_patron → lenguaje
    _MANIFIESTOS = {
        "pyproject.toml":    "python",
        ".python-version":   "python",
        "Pipfile":           "python",
        "setup.cfg":         "python",
        "package.json":      "node",
        ".nvmrc":            "node",
        ".node-version":     "node",
        "tsconfig.json":     "typescript",
        "pom.xml":           "java",
        "build.gradle":      "java",
        "build.gradle.kts":  "java",
        ".java-version":     "java",
        "composer.json":     "php",
        "global.json":       "dotnet",
    }

    # Extensiones adicionales que implican lenguaje
    _EXT_MANIFIESTOS = {
        ".csproj":  "dotnet",
        ".fsproj":  "dotnet",
        ".vbproj":  "dotnet",
    }

    def __init__(self, archivos: list, scanner=None):
        self.archivos = archivos
        self.scanner  = scanner

    def analizar(self) -> dict:
        """
        Lee manifiestos del proyecto y extrae versiones de lenguajes.
        Cruza con _VERSIONES_SOPORTADAS y emite advertencias si hay gaps.
        """
        versiones_detectadas: dict = {}
        advertencias: list         = []
        manifiestos_encontrados: list = []

        for arch in self.archivos:
            nombre  = arch.get("nombre", "").lower()
            ext     = arch.get("extension", "").lower()
            ruta_abs = arch.get("ruta_abs", "")
            ruta_rel = arch.get("ruta_rel", "")

            lenguaje_man = (
                self._MANIFIESTOS.get(nombre) or
                self._EXT_MANIFIESTOS.get(ext)
            )
            if not lenguaje_man or not ruta_abs:
                continue

            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception:
                continue

            version = self._extraer_version(nombre, ext, contenido, lenguaje_man)
            if version:
                manifiestos_encontrados.append({
                    "archivo":  ruta_rel,
                    "lenguaje": lenguaje_man,
                    "version":  version,
                })
                # Guardar la más específica (puede haber múltiples manifiestos)
                if lenguaje_man not in versiones_detectadas:
                    versiones_detectadas[lenguaje_man] = version

        # Cruzar con versiones soportadas y generar advertencias
        for lang, version in versiones_detectadas.items():
            soportadas = self._VERSIONES_SOPORTADAS.get(lang, [])
            soportada = any(version.startswith(s) for s in soportadas)
            if not soportada and soportadas:
                advertencias.append({
                    "lenguaje": lang,
                    "version":  version,
                    "mensaje":  (
                        f"{lang.capitalize()} {version} detectado — "
                        f"STRATUM tiene patrones para {', '.join(soportadas[-3:])}. "
                        f"Análisis puede ser incompleto para sintaxis nueva."
                    ),
                })

        return {
            "versiones":              versiones_detectadas,
            "manifiestos_encontrados": manifiestos_encontrados,
            "advertencias":           advertencias,
            "n_lenguajes_detectados": len(versiones_detectadas),
            "n_advertencias":         len(advertencias),
        }

    # ─── Extracción de versión por tipo de manifiesto ─────────────────────────
    def _extraer_version(self, nombre: str, ext: str,
                         contenido: str, lenguaje: str) -> str:
        """Extrae la versión del lenguaje del contenido del manifiesto."""
        version = ""
        try:
            if nombre == "pyproject.toml":
                m = re.search(r'python\s*=\s*"([^"]+)"', contenido)
                if not m:
                    m = re.search(r'python_requires\s*=\s*["\']([^"\']+)["\']', contenido)
                if m:
                    version = m.group(1).lstrip(">=~^")

            elif nombre in (".python-version", ".node-version", ".nvmrc",
                            ".java-version"):
                version = contenido.strip().lstrip("v")

            elif nombre == "pipfile":
                m = re.search(r'python_version\s*=\s*["\']([^"\']+)["\']', contenido)
                if m:
                    version = m.group(1)

            elif nombre == "package.json":
                m = re.search(r'"node"\s*:\s*"([^"]+)"', contenido)
                if m:
                    version = m.group(1).lstrip(">=~^v")

            elif nombre == "tsconfig.json":
                m = re.search(r'"target"\s*:\s*"es(\d+)"', contenido, re.IGNORECASE)
                if m:
                    version = m.group(1)

            elif nombre == "pom.xml":
                m = re.search(r"<java\.version>([\d.]+)</java\.version>", contenido)
                if not m:
                    m = re.search(r"<source>([\d.]+)</source>", contenido)
                if m:
                    version = m.group(1)

            elif nombre in ("build.gradle", "build.gradle.kts"):
                m = re.search(r"sourceCompatibility\s*=\s*['\"]?([\d.]+)", contenido)
                if not m:
                    m = re.search(r"JavaVersion\.VERSION_([\d_]+)", contenido)
                    if m:
                        version = m.group(1).replace("_", ".")
                if m and not version:
                    version = m.group(1)

            elif nombre == "composer.json":
                m = re.search(r'"php"\s*:\s*"([^"]+)"', contenido)
                if m:
                    version = m.group(1).lstrip(">=~^")

            elif nombre == "global.json":
                m = re.search(r'"version"\s*:\s*"([^"]+)"', contenido)
                if m:
                    version = m.group(1)

            elif ext in (".csproj", ".fsproj", ".vbproj"):
                m = re.search(
                    r"<TargetFramework(?:Version)?>(net[\w.]+)</", contenido
                )
                if m:
                    version = m.group(1).replace("net", "")

            elif nombre == "setup.cfg":
                m = re.search(r"python_requires\s*=\s*>=\s*([\d.]+)", contenido)
                if m:
                    version = m.group(1)

        except Exception:
            pass

        return version.strip() if version else ""



# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 23 — SECURITY AUDITOR
# Detecta secretos hardcodeados, inyección SQL, patrones inseguros,
# deserialización unsafe y configuración expuesta en todo el codebase.
# Sin dependencias externas — análisis por patrones regex calibrados.
# ════════════════════════════════════════════════════════════════════════════

class SecurityAuditor:
    """
    Auditor de seguridad estático — Bloque 23.
    Categorías: secretos hardcodeados, inyección SQL,
    código inseguro, deserialización, config expuesta.
    Sin dependencias externas.
    """

    # ── Secretos hardcodeados ─────────────────────────────────────────────
    # \x22 = "   \x27 = '   evita conflicto con delimitadores de string Python
    _PAT_SECRETOS = [
        (re.compile(
            r'(?i)\b(password|passwd|pwd|contrasena)\s*[=:]\s*[\x22\x27]'
            r'(?!<|>|\$\{|\{\{|xxx|test|pass|change|example|placeholder|your_)'
            r'[^\x22\x27]{6,}[\x22\x27]'
        ), "CRITICO", "password_hardcodeado"),
        (re.compile(
            r'(?i)\b(api_key|apikey|api_secret|auth_token|access_token|'
            r'bearer_token|private_key)\s*[=:]\s*[\x22\x27][^\x22\x27]{12,}[\x22\x27]'
        ), "CRITICO", "token_api_hardcodeado"),
        (re.compile(r'(?<![A-Z0-9])(AKIA|ASIA|AROA)[A-Z0-9]{16}(?![A-Z0-9])'),
         "CRITICO", "aws_access_key_expuesto"),
        (re.compile(
            r'[\x22\x27]ey[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}'
            r'\.[A-Za-z0-9_-]{10,}[\x22\x27]'
        ), "CRITICO", "jwt_token_hardcodeado"),
        (re.compile(
            r'(?i)(postgres|mysql|mongodb|redis|sqlalchemy)[^"\'\n]*://'
            r'[^:\s]{1,}:[^@\s]{6,}@'
        ), "CRITICO", "database_url_con_credenciales"),
        (re.compile(
            r'(?i)SECRET_KEY\s*=\s*[\x22\x27]'
            r'(?!<|>|\$\{|\{\{|django-insecure|change|your|example)'
            r'[^\x22\x27]{16,}[\x22\x27]'
        ), "ALTO", "secret_key_expuesto"),
    ]

    # ── Inyección SQL ─────────────────────────────────────────────────────
    _PAT_SQL = [
        (re.compile(
            r'(?i)f[\x22\x27].*?(SELECT|INSERT|UPDATE|DELETE|DROP|CREATE)\b.*?\{'
        ), "CRITICO", "sql_injection_fstring"),
        (re.compile(
            r'(?i)[\x22\x27].*?(SELECT|INSERT|UPDATE|DELETE)\b.*?[\x22\x27]'
            r'\s*\.format\s*\('
        ), "CRITICO", "sql_injection_format"),
        (re.compile(
            r'(?i)[\x22\x27].*?(SELECT|INSERT|UPDATE|DELETE)\b.*?[\x22\x27]'
            r'\s*%\s*[(\w]'
        ), "ALTO", "sql_injection_percent"),
        (re.compile(
            r'(?i)cursor\.execute\s*\(\s*[\x22\x27].*?\+\s*\w'
        ), "ALTO", "sql_inject_cursor_concat"),
    ]

    # ── Código inseguro ───────────────────────────────────────────────────
    _PAT_INSEGURO = [
        (re.compile(r'\beval\s*\('),  "CRITICO", "uso_eval"),
        (re.compile(r'\bexec\s*\('),  "ALTO",    "uso_exec"),
        (re.compile(
            r'subprocess\.(call|run|Popen)\s*\([^)]*shell\s*=\s*True'
        ), "CRITICO", "subprocess_shell_true"),
        (re.compile(r'\bos\.system\s*\('), "ALTO", "os_system"),
        (re.compile(r'\bos\.popen\s*\('),  "ALTO", "os_popen"),
        (re.compile(r'\b__import__\s*\('), "ALTO", "import_dinamico"),
    ]

    # ── Deserialización insegura ──────────────────────────────────────────
    _PAT_DESERIAL = [
        (re.compile(r'\bpickle\.(loads|load)\s*\('),  "CRITICO", "pickle_inseguro"),
        (re.compile(r'\bmarshal\.(loads|load)\s*\('), "CRITICO", "marshal_inseguro"),
        (re.compile(r'\byaml\.load\s*\([^,)]+\)'),   "ALTO",    "yaml_load_sin_loader"),
        (re.compile(r'\bjsonpickle\.decode\s*\('),    "ALTO",    "jsonpickle_decode"),
    ]

    # ── Config expuesta (solo en archivos de settings) ────────────────────
    _PAT_CONFIG = [
        (re.compile(r'(?i)\bDEBUG\s*=\s*True\b'),  "ALTO",  "debug_en_produccion"),
        (re.compile(r'(?i)ALLOWED_HOSTS\s*=\s*\[\s*[\x22\x27]?\*[\x22\x27]?\s*\]'),
         "ALTO",  "allowed_hosts_wildcard"),
        (re.compile(r'(?i)CORS_ORIGIN_ALLOW_ALL\s*=\s*True'), "ALTO",  "cors_allow_all"),
        (re.compile(r'(?i)SECURE_SSL_REDIRECT\s*=\s*False'),  "MEDIO", "ssl_redirect_off"),
        (re.compile(r'(?i)SESSION_COOKIE_SECURE\s*=\s*False'),"MEDIO", "cookie_insegura"),
    ]

    _EXTS_ANALIZAR   = {".py", ".cfg", ".ini", ".env", ".yaml", ".yml", ".toml", ".json"}
    _SETTINGS_FILES  = {
        "settings.py", "local_settings.py", "config.py", "configuration.py",
        "secrets.py", "credentials.py", ".env", "env.py", "app_config.py",
    }
    _CATEGORIA = {
        "password_hardcodeado":        "Secretos",
        "token_api_hardcodeado":       "Secretos",
        "aws_access_key_expuesto":     "Secretos",
        "jwt_token_hardcodeado":       "Secretos",
        "database_url_con_credenciales": "Secretos",
        "secret_key_expuesto":         "Secretos",
        "sql_injection_fstring":       "Inyección SQL",
        "sql_injection_format":        "Inyección SQL",
        "sql_injection_percent":       "Inyección SQL",
        "sql_inject_cursor_concat":    "Inyección SQL",
        "uso_eval":                    "Código inseguro",
        "uso_exec":                    "Código inseguro",
        "subprocess_shell_true":       "Código inseguro",
        "os_system":                   "Código inseguro",
        "os_popen":                    "Código inseguro",
        "import_dinamico":             "Código inseguro",
        "pickle_inseguro":             "Deserialización",
        "marshal_inseguro":            "Deserialización",
        "yaml_load_sin_loader":        "Deserialización",
        "jsonpickle_decode":           "Deserialización",
        "debug_en_produccion":         "Config expuesta",
        "allowed_hosts_wildcard":      "Config expuesta",
        "cors_allow_all":              "Config expuesta",
        "ssl_redirect_off":            "Config expuesta",
        "cookie_insegura":             "Config expuesta",
    }

    def __init__(self, archivos: list, scanner=None):
        self.archivos  = archivos
        self.scanner   = scanner
        self.resultado = {}
        self.hallazgos = []

    def analizar(self) -> dict:
        if not self.archivos:
            return {"archivos": {}, "hallazgos": [], "resumen": self._resumen_vacio()}

        for arch in self.archivos:
            ruta_abs = arch.get("ruta_abs", "")
            ruta_rel = arch.get("ruta_rel", ruta_abs)
            ext      = arch.get("extension", "").lower()

            if ext not in self._EXTS_ANALIZAR:
                continue

            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as fh:
                    contenido = fh.read()
            except Exception:
                continue

            if not contenido.strip():
                continue

            nombre   = os.path.basename(ruta_abs).lower()
            lineas   = contenido.splitlines()
            hallazgos_arch = []

            # Secretos: todos los archivos de configuración y Python
            self._scan(lineas, ruta_rel, self._PAT_SECRETOS,  hallazgos_arch)

            if ext == ".py":
                self._scan(lineas, ruta_rel, self._PAT_SQL,       hallazgos_arch)
                self._scan(lineas, ruta_rel, self._PAT_INSEGURO,  hallazgos_arch)
                self._scan(lineas, ruta_rel, self._PAT_DESERIAL,  hallazgos_arch)
                if nombre in self._SETTINGS_FILES:
                    self._scan(lineas, ruta_rel, self._PAT_CONFIG, hallazgos_arch)

            if hallazgos_arch:
                self.resultado[ruta_rel] = hallazgos_arch
                self.hallazgos.extend(hallazgos_arch)

        return {
            "archivos":  self.resultado,
            "hallazgos": self.hallazgos,
            "resumen":   self._construir_resumen(),
        }

    def _scan(self, lineas, ruta_rel, patrones, destino):
        """Aplica patrones línea a línea con filtros anti-falso-positivo."""
        _FP_KEYWORDS = (
            "test_", "_test", "example", "placeholder", "your_",
            "<your", "todo:", "fixme:", "noqa", "# noqa",
            "mock_", "_mock", "fake_", "dummy_",
        )
        for num, linea in enumerate(lineas, 1):
            stripped = linea.strip()
            if not stripped or stripped.startswith("#") or stripped.startswith("//"):
                continue
            sl = stripped.lower()
            if any(kw in sl for kw in _FP_KEYWORDS):
                continue

            for patron, severidad, tipo in patrones:
                if patron.search(linea):
                    if any(h["linea"] == num and h["tipo"] == tipo for h in destino):
                        continue
                    destino.append({
                        "tipo":      tipo,
                        "severidad": severidad,
                        "archivo":   ruta_rel,
                        "linea":     num,
                        "detalle":   stripped[:120],
                        "categoria": self._CATEGORIA.get(tipo, "General"),
                    })
                    break   # un hallazgo por línea

    def _construir_resumen(self) -> dict:
        por_cat = {}
        por_sev = {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0}
        for h in self.hallazgos:
            por_cat[h["categoria"]] = por_cat.get(h["categoria"], 0) + 1
            por_sev[h["severidad"]] = por_sev.get(h["severidad"],  0) + 1

        criticos = por_sev["CRITICO"]
        nivel = (
            "CRITICO" if criticos >= 5 else
            "ALTO"    if criticos >= 1 or por_sev["ALTO"] >= 3 else
            "MEDIO"   if por_sev["ALTO"] >= 1 or por_sev["MEDIO"] >= 5 else
            "BAJO"
        )
        return {
            "total_hallazgos":     len(self.hallazgos),
            "archivos_afectados":  len(self.resultado),
            "por_severidad":       por_sev,
            "por_categoria":       por_cat,
            "nivel_riesgo":        nivel,
            "tiene_secretos":      por_cat.get("Secretos",      0) > 0,
            "tiene_sql_injection": por_cat.get("Inyección SQL", 0) > 0,
        }

    def _resumen_vacio(self) -> dict:
        return {
            "total_hallazgos": 0, "archivos_afectados": 0,
            "por_severidad":   {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0},
            "por_categoria": {}, "nivel_riesgo": "BAJO",
            "tiene_secretos": False, "tiene_sql_injection": False,
        }

# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 28 — REPORT GENERATOR
# Genera un resumen ejecutivo HTML exportable/imprimible con los hallazgos
# consolidados de todos los módulos STRATUM. Sin dependencias externas.
# ════════════════════════════════════════════════════════════════════════════

class ReportGenerator:
    """
    Generador de reporte ejecutivo exportable.

    Produce un archivo HTML standalone con:
    - Portada: nombre del proyecto, fecha, score global, badge
    - Resumen ejecutivo: top hallazgos, distribución de severidad
    - Tabla de módulos activos con estado
    - Sección de recomendaciones priorizadas
    - Estilo print-friendly (sin dependencias CDN en el resumen)
    """

    _COLORES = {
        "CRITICO": "#e53e3e",
        "ALTO":    "#f6ad55",
        "MEDIO":   "#fefcbf",
        "BAJO":    "#68d391",
        "OK":      "#68d391",
    }

    def __init__(self, stratum_app):
        """
        stratum_app: instancia de StratumApp con todos los resultado_* ya
        populados por ejecutar_analisis().
        """
        self.app = stratum_app

    # ── Punto de entrada ────────────────────────────────────────────────────
    def generar(self) -> dict:
        """
        Construye el reporte y lo guarda en el directorio del proyecto.
        Retorna: {
            "ruta_reporte": str,
            "resumen":      dict con métricas clave,
        }
        """
        metricas = self._recopilar_metricas()
        html     = self._construir_html(metricas)

        ruta_out = os.path.join(
            self.app.ruta_proyecto,
            f"stratum_reporte_{datetime.date.today().isoformat()}.html",
        )
        with open(ruta_out, "w", encoding="utf-8") as fh:
            fh.write(html)

        return {
            "ruta_reporte": ruta_out,
            "resumen":      metricas,
        }

    # ── Recopilación de métricas ────────────────────────────────────────────
    def _recopilar_metricas(self) -> dict:
        app = self.app
        nombre = os.path.basename(os.path.abspath(app.ruta_proyecto))

        # Score
        score_data  = app.resultado_score  or {}
        score       = score_data.get("score",      0)
        badge       = score_data.get("badge",      "N/A")
        badge_color = score_data.get("badge_color","#a0aec0")
        por_sev     = score_data.get("por_severidad",
                                     {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0})
        top10       = score_data.get("top10",      [])

        # AST
        r_ast   = app.resultado_ast or {}
        res_ast = r_ast.get("resumen", {})

        # Framework
        fw_data  = app.resultado_framework or {}
        fw       = fw_data.get("framework",  "N/A")
        fw_ver   = fw_data.get("version",    "")

        # Security
        r_sec   = app.resultado_security or {}
        sec_res = r_sec.get("resumen", {})

        # Coverage
        r_cov   = app.resultado_coverage or {}

        # Heuristic
        r_heur  = app.resultado_heur or {}
        res_h   = r_heur.get("resumen", {})

        # Quick Fix
        r_qf    = app.resultado_quickfix or {}

        modulos = []
        def _modulo(nombre_m, resultado, n_problemas_key="total_problemas"):
            activo = resultado is not None
            n_prob = 0
            if activo and isinstance(resultado, dict):
                res = resultado.get("resumen", {})
                n_prob = res.get(n_problemas_key, 0) or res.get("total_hallazgos", 0)
            modulos.append({"nombre": nombre_m, "activo": activo,
                            "problemas": n_prob})

        _modulo("AST Python",           app.resultado_ast,       "total_problemas")
        _modulo("Heuristic Engine",     app.resultado_heur,      "total_problemas")
        _modulo("SQL Analyzer",         app.resultado_sql,       "total_problemas")
        _modulo("Link Checker",         app.resultado_link,      "total_archivos")
        _modulo("ETL Detector",         app.resultado_etl,       "total_pipelines")
        _modulo("Cloud Auditor",        app.resultado_cloud,     "total_recursos")
        _modulo("DWH Auditor",          app.resultado_dwh,       "total_tablas")
        _modulo("Data Profiler",        app.resultado_profiler,  "total_archivos")
        _modulo("KPI Engine",           app.resultado_kpi,       "total_kpis")
        _modulo("Viz Mapper",           app.resultado_viz,       "total_visualizaciones")
        _modulo("Query Intel",          app.resultado_query,     "total_funciones")
        _modulo("Flow Analyzer",        app.resultado_flow,      "n_gaps")
        _modulo("Excel Intelligence",   app.resultado_excel,     "total_problemas")
        _modulo("Security Auditor",     app.resultado_security,  "total_hallazgos")
        _modulo("Coverage Score",       app.resultado_coverage,  "n_advertencias")
        _modulo("Language Versions",    app.resultado_version,   "n_advertencias")
        _modulo("Framework Detector",   app.resultado_framework, "n_frameworks")

        # Recomendaciones top (desde Quick Fix, máx 5)
        recomendaciones = []
        if r_qf:
            for s in r_qf.get("sugerencias", [])[:5]:
                recomendaciones.append({
                    "esfuerzo":  s.get("esfuerzo",  "HORAS"),
                    "severidad": s.get("severidad", "MEDIO"),
                    "tipo":      s.get("tipo",      ""),
                    "accion":    s.get("accion",    ""),
                    "archivo":   s.get("archivo",   ""),
                })

        return {
            "nombre":           nombre,
            "fecha":            datetime.date.today().isoformat(),
            "score":            score,
            "badge":            badge,
            "badge_color":      badge_color,
            "por_severidad":    por_sev,
            "top10":            top10,
            "framework":        fw,
            "framework_ver":    fw_ver,
            "total_clases":     res_ast.get("total_clases",    0),
            "total_funciones":  res_ast.get("total_funciones", 0),
            "total_metodos":    res_ast.get("total_metodos",   0),
            "total_archivos_py":res_h.get("total_archivos",    0),
            "sec_nivel":        sec_res.get("nivel_riesgo",    "N/A"),
            "sec_hallazgos":    sec_res.get("total_hallazgos", 0),
            "cobertura":        r_cov.get("ratio_global",      0),
            "modulos":          modulos,
            "recomendaciones":  recomendaciones,
        }

    # ── Construcción del HTML ──────────────────────────────────────────────
    def _construir_html(self, m: dict) -> str:
        color_score = m["badge_color"]
        por_sev     = m["por_severidad"]
        sec_color   = self._COLORES.get(m["sec_nivel"], "#a0aec0")

        # Distribución de severidad — barras simples
        total_findings = sum(por_sev.values()) or 1
        def _barra(sev, color):
            n   = por_sev.get(sev, 0)
            pct = round(n / total_findings * 100)
            return (
                f'<div style="display:flex;align-items:center;gap:0.5rem;margin:0.2rem 0">'
                f'<span style="width:4rem;font-size:0.8rem;color:{color}">{sev}</span>'
                f'<div style="flex:1;background:#e2e8f0;border-radius:2px;height:12px">'
                f'<div style="width:{pct}%;background:{color};height:12px;border-radius:2px"></div></div>'
                f'<span style="width:2.5rem;text-align:right;font-size:0.8rem">{n}</span>'
                f'</div>'
            )

        barras_sev = (
            _barra("CRITICO", "#e53e3e") +
            _barra("ALTO",    "#f6ad55") +
            _barra("MEDIO",   "#9ca3af") +
            _barra("BAJO",    "#6b7280")
        )

        # Top hallazgos
        rows_top = "".join(
            f'<tr>'
            f'<td style="color:{self._COLORES.get(h.get("severidad","BAJO"),"#9ca3af")};'
            f'font-weight:600">{h.get("severidad","")}</td>'
            f'<td>{h.get("tipo","")}</td>'
            f'<td style="color:#374151">{h.get("origen","")}</td>'
            f'<td style="font-size:0.78rem;color:#6b7280;word-break:break-all">'
            f'{h.get("archivo","")}</td>'
            f'<td style="font-size:0.78rem;color:#374151;word-break:break-word">'
            f'{h.get("detalle","")[:120]}</td>'
            f'</tr>'
            for h in m["top10"]
        ) or '<tr><td colspan=5 style="color:#6b7280">Sin hallazgos críticos</td></tr>'

        # Módulos
        rows_mod = "".join(
            f'<tr>'
            f'<td>{mod["nombre"]}</td>'
            f'<td style="color:{"#16a34a" if mod["activo"] else "#9ca3af"}">'
            f'{"✅ Activo" if mod["activo"] else "— No aplica"}</td>'
            f'<td style="text-align:right;color:{"#dc2626" if mod["problemas"]>0 else "#16a34a"}">'
            f'{mod["problemas"] if mod["activo"] else "—"}</td>'
            f'</tr>'
            for mod in m["modulos"]
        )

        # Recomendaciones
        rows_rec = "".join(
            f'<tr>'
            f'<td style="color:{"#16a34a" if r["esfuerzo"]=="MINUTOS" else "#d97706" if r["esfuerzo"]=="HORAS" else "#dc2626"};font-weight:600">'
            f'{r["esfuerzo"]}</td>'
            f'<td style="color:{self._COLORES.get(r["severidad"],"#9ca3af")};font-weight:600">'
            f'{r["severidad"]}</td>'
            f'<td>{r["tipo"]}</td>'
            f'<td style="font-size:0.78rem;color:#374151;word-break:break-word">'
            f'{r["accion"]}</td>'
            f'</tr>'
            for r in m["recomendaciones"]
        ) or '<tr><td colspan=4 style="color:#6b7280">Sin sugerencias generadas</td></tr>'

        return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>STRATUM — Reporte {m['nombre']} — {m['fecha']}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f8fafc;
          color: #1f2937; font-size: 14px; line-height: 1.6; }}
  .page {{ max-width: 960px; margin: 0 auto; padding: 2rem; }}
  .header {{ background: #1a1f2e; color: #e2e8f0; padding: 2rem;
             border-radius: 12px; margin-bottom: 1.5rem; }}
  .header h1 {{ font-size: 1.6rem; font-weight: 700; margin-bottom: 0.25rem; }}
  .header p {{ color: #718096; font-size: 0.85rem; }}
  .score-badge {{ display:inline-block; padding: 0.4rem 1.2rem;
                  border-radius: 999px; font-weight: 700; font-size: 1.1rem;
                  margin-top: 0.75rem; color: #1a1f2e;
                  background: {color_score}; }}
  .section {{ background: #fff; border: 1px solid #e5e7eb; border-radius: 10px;
              padding: 1.5rem; margin-bottom: 1.25rem;
              box-shadow: 0 1px 3px rgba(0,0,0,.06); }}
  .section h2 {{ font-size: 1rem; font-weight: 600; color: #374151;
                 border-bottom: 1px solid #f3f4f6; padding-bottom: 0.5rem;
                 margin-bottom: 1rem; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px,1fr));
           gap: 1rem; margin-bottom: 1rem; }}
  .card {{ background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 8px;
           padding: 0.75rem 1rem; }}
  .card-label {{ font-size: 0.7rem; color: #6b7280; text-transform: uppercase;
                 letter-spacing: .06em; margin-bottom: .25rem; }}
  .card-value {{ font-size: 1.5rem; font-weight: 700; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; }}
  td, th {{ padding: 0.4rem 0.6rem; text-align: left;
            border-bottom: 1px solid #f3f4f6; }}
  th {{ color: #6b7280; font-weight: 500; font-size: 0.72rem;
        text-transform: uppercase; background: #f9fafb; }}
  .footer {{ text-align: center; padding: 1.5rem; color: #9ca3af;
             font-size: 0.75rem; }}
  @media print {{
    body {{ background: #fff; }}
    .page {{ padding: 0; }}
    .header {{ border-radius: 0; }}
    .section {{ break-inside: avoid; }}
  }}
</style>
</head>
<body>
<div class="page">

  <div class="header">
    <h1>⬡ STRATUM System Intelligence Engine</h1>
    <p>Reporte de análisis estático · {m['nombre']} · {m['fecha']}</p>
    <div class="score-badge">{m['score']}/100 &nbsp; {m['badge']}</div>
  </div>

  <div class="section">
    <h2>Resumen ejecutivo</h2>
    <div class="grid">
      <div class="card">
        <div class="card-label">Score STRATUM</div>
        <div class="card-value" style="color:{color_score}">{m['score']}</div>
      </div>
      <div class="card">
        <div class="card-label">Framework</div>
        <div class="card-value" style="font-size:1.1rem">{m['framework']} {m['framework_ver']}</div>
      </div>
      <div class="card">
        <div class="card-label">Archivos Python</div>
        <div class="card-value">{m['total_archivos_py']}</div>
      </div>
      <div class="card">
        <div class="card-label">Funciones</div>
        <div class="card-value">{m['total_funciones']}</div>
      </div>
      <div class="card">
        <div class="card-label">Seguridad</div>
        <div class="card-value" style="color:{sec_color}">{m['sec_nivel']}</div>
      </div>
      <div class="card">
        <div class="card-label">Cobertura</div>
        <div class="card-value">{m['cobertura']}%</div>
      </div>
    </div>
    <h2 style="margin-top:1rem">Distribución de hallazgos por severidad</h2>
    {barras_sev}
  </div>

  <div class="section">
    <h2>Top 10 hallazgos prioritarios</h2>
    <table>
      <tr><th>Severidad</th><th>Tipo</th><th>Origen</th>
          <th>Archivo</th><th>Detalle</th></tr>
      {rows_top}
    </table>
  </div>

  <div class="section">
    <h2>Recomendaciones (Quick Wins primero)</h2>
    <table>
      <tr><th>Esfuerzo</th><th>Severidad</th><th>Tipo</th><th>Acción</th></tr>
      {rows_rec}
    </table>
  </div>

  <div class="section">
    <h2>Estado de módulos STRATUM</h2>
    <table>
      <tr><th>Módulo</th><th>Estado</th><th style="text-align:right">Hallazgos</th></tr>
      {rows_mod}
    </table>
  </div>

  <div class="footer">
    Generado por STRATUM System Intelligence Engine v{STRATUM_VERSION}<br>
    {m['fecha']} &mdash; {m['nombre']}
  </div>

</div>
</body>
</html>"""




# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 22 — STREAMLIT CACHE DETECTOR
# Analiza proyectos Streamlit buscando antipatrones de caché y session_state.
# Solo activo cuando FrameworkDetector detecta Streamlit como framework.
# ════════════════════════════════════════════════════════════════════════════

class StreamlitCacheDetector:
    """
    Detector de antipatrones de caché y estado en proyectos Streamlit.

    Detecta:
    - Funciones costosas sin decorador de caché (heurística por nombre/contenido)
    - @st.cache_data con objetos mutables en argumentos (bug clásico)
    - @st.cache_resource usado donde debería ser @st.cache_data (y viceversa)
    - st.session_state antipatterns: acceso sin .get() / sin inicialización
    - st.experimental_cache (deprecated, debe migrarse)
    - Loops sobre st.session_state sin copia (mutación durante iteración)
    - Variables de conexión DB fuera de @st.cache_resource
    """

    # ── Patrones por categoría ────────────────────────────────────────────
    # Funciones costosas típicas que deberían estar cacheadas
    _RE_COSTOSA = re.compile(
        r'(?i)^def\s+(cargar?|load|fetch|get_data|read_|query_|compute_|calcul|'
        r'entrenar?|train|predict|infer|transform|process_data|prepare)',
        re.MULTILINE,
    )
    # Decoradores de caché presentes
    _RE_CACHE_DECO = re.compile(
        r'@st\.(cache_data|cache_resource|cache)\s*[\(\n]'
    )
    _RE_CACHE_DEPRECATED = re.compile(
        r'@st\.(experimental_memo|experimental_singleton|cache)\s*[\(\n]'
    )
    # session_state sin inicialización segura
    _RE_SS_UNSAFE = re.compile(
        r'st\.session_state\[[\x22\x27][^\x22\x27]+[\x22\x27]\]\s*(?!\s*=\s*st\.session_state)'
    )
    _RE_SS_GET = re.compile(r'st\.session_state\.get\s*\(')
    _RE_SS_IN  = re.compile(r'[\x22\x27][^\x22\x27]+[\x22\x27]\s+in\s+st\.session_state')
    # Conexiones DB fuera de cache_resource
    _RE_DB_CONN = re.compile(
        r'(?i)(psycopg2\.connect|mysql\.connector\.connect|'
        r'sqlite3\.connect|create_engine|MongoClient|redis\.Redis|'
        r'snowflake\.connector\.connect)\s*\('
    )
    _RE_CACHE_RES = re.compile(r'@st\.cache_resource')
    # Mutable default en @st.cache_data
    _RE_MUTABLE_ARG = re.compile(
        r'@st\.cache_data[^\n]*\n\s*def\s+\w+\s*\([^)]*(?:\[\]|\{\}|list\(\)|dict\(\))[^)]*\)'
    )

    def __init__(self, archivos: list, resultado_framework: dict = None, scanner=None):
        self.archivos    = archivos
        self.fw_resultado = resultado_framework or {}
        self.scanner     = scanner
        self.resultado   = {}
        self.hallazgos   = []

    def analizar(self) -> dict:
        # Solo analizar si el proyecto usa Streamlit
        fw = self.fw_resultado.get("framework", "").lower()
        if "streamlit" not in fw:
            return {"archivos": {}, "hallazgos": [], "resumen": self._resumen_vacio(),
                    "aplica": False}

        archivos_py = [a for a in self.archivos
                       if a.get("extension", "").lower() == ".py"]
        if not archivos_py:
            return {"archivos": {}, "hallazgos": [], "resumen": self._resumen_vacio(),
                    "aplica": True}

        for arch in archivos_py:
            ruta_abs = arch.get("ruta_abs", "")
            ruta_rel = arch.get("ruta_rel", ruta_abs)
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as fh:
                    contenido = fh.read()
            except Exception:
                continue

            if "streamlit" not in contenido and "st." not in contenido:
                continue   # no es un archivo Streamlit

            h_arch = []
            self._check_funciones_sin_cache(contenido, ruta_rel, h_arch)
            self._check_deprecated(contenido, ruta_rel, h_arch)
            self._check_session_state(contenido, ruta_rel, h_arch)
            self._check_db_fuera_cache(contenido, ruta_rel, h_arch)
            self._check_mutable_args(contenido, ruta_rel, h_arch)

            if h_arch:
                self.resultado[ruta_rel] = h_arch
                self.hallazgos.extend(h_arch)

        return {
            "archivos":  self.resultado,
            "hallazgos": self.hallazgos,
            "resumen":   self._construir_resumen(),
            "aplica":    True,
        }

    # ── Detectores individuales ───────────────────────────────────────────

    def _check_funciones_sin_cache(self, contenido, ruta, dest):
        """Funciones con nombre 'costoso' que no tienen decorador de caché."""
        lineas = contenido.splitlines()
        for num, linea in enumerate(lineas, 1):
            if self._RE_COSTOSA.match(linea.strip()):
                # Revisar las 3 líneas anteriores buscando @st.cache_*
                prev = "\n".join(lineas[max(0, num-4):num-1])
                if not self._RE_CACHE_DECO.search(prev):
                    nombre = re.search(r'def\s+(\w+)', linea)
                    fn_nombre = nombre.group(1) if nombre else "función"
                    dest.append({
                        "tipo":      "funcion_costosa_sin_cache",
                        "severidad": "ALTO",
                        "archivo":   ruta,
                        "linea":     num,
                        "detalle":   (f"{fn_nombre}() parece costosa pero no tiene "
                                      f"@st.cache_data ni @st.cache_resource"),
                        "accion":    f"Agregar @st.cache_data a {fn_nombre}()",
                    })

    def _check_deprecated(self, contenido, ruta, dest):
        """Uso de decoradores de caché deprecados."""
        for num, linea in enumerate(contenido.splitlines(), 1):
            if self._RE_CACHE_DEPRECATED.search(linea):
                dep = re.search(r'@st\.(\w+)', linea)
                dep_nombre = dep.group(1) if dep else "cache"
                dest.append({
                    "tipo":      "cache_deprecado",
                    "severidad": "MEDIO",
                    "archivo":   ruta,
                    "linea":     num,
                    "detalle":   f"@st.{dep_nombre} está deprecado desde Streamlit 1.18",
                    "accion":    "Migrar a @st.cache_data o @st.cache_resource",
                })

    def _check_session_state(self, contenido, ruta, dest):
        """Acceso a session_state sin verificación previa de existencia."""
        lineas = contenido.splitlines()
        for num, linea in enumerate(lineas, 1):
            stripped = linea.strip()
            if not ("session_state" in stripped):
                continue
            # Detectar: st.session_state["key"] en lectura SIN .get() ni 'key' in ss
            if self._RE_SS_UNSAFE.search(linea):
                # Verificar que la línea no sea inicialización ni .get()
                if "= " not in linea.split("session_state")[1][:10]:
                    if not self._RE_SS_GET.search(linea) and "if" not in stripped:
                        dest.append({
                            "tipo":      "session_state_sin_init",
                            "severidad": "MEDIO",
                            "archivo":   ruta,
                            "linea":     num,
                            "detalle":   stripped[:120],
                            "accion":    "Usar st.session_state.get('key', default) "
                                         "o verificar con 'key' in st.session_state",
                        })
            # Mutación durante iteración
            if ("for " in stripped and "session_state" in stripped and
                    ".items()" in stripped):
                dest.append({
                    "tipo":      "session_state_iter_mutacion",
                    "severidad": "BAJO",
                    "archivo":   ruta,
                    "linea":     num,
                    "detalle":   stripped[:120],
                    "accion":    "Iterar sobre dict(st.session_state).items() para evitar "
                                 "RuntimeError por cambio de tamaño durante iteración",
                })

    def _check_db_fuera_cache(self, contenido, ruta, dest):
        """Conexiones DB inicializadas fuera de @st.cache_resource."""
        lineas = contenido.splitlines()
        for num, linea in enumerate(lineas, 1):
            if self._RE_DB_CONN.search(linea):
                # Buscar @st.cache_resource en las 5 líneas anteriores
                prev = "\n".join(lineas[max(0, num-6):num-1])
                if not self._RE_CACHE_RES.search(prev):
                    conn = re.search(
                        r'(psycopg2|mysql|sqlite3|create_engine|MongoClient|'
                        r'redis|snowflake)', linea, re.IGNORECASE
                    )
                    nombre_conn = conn.group(1) if conn else "conexión"
                    dest.append({
                        "tipo":      "db_conexion_sin_cache_resource",
                        "severidad": "ALTO",
                        "archivo":   ruta,
                        "linea":     num,
                        "detalle":   (f"Conexión {nombre_conn} fuera de "
                                      f"@st.cache_resource — reconecta en cada rerun"),
                        "accion":    "Envolver la conexión en una función "
                                     "@st.cache_resource",
                    })

    def _check_mutable_args(self, contenido, ruta, dest):
        """@st.cache_data con argumentos mutables (lista/dict)."""
        for m in self._RE_MUTABLE_ARG.finditer(contenido):
            linea = contenido[:m.start()].count("\n") + 1
            dest.append({
                "tipo":      "cache_data_arg_mutable",
                "severidad": "ALTO",
                "archivo":   ruta,
                "linea":     linea,
                "detalle":   "Función @st.cache_data con argumento mutable ([] o {}) "
                             "— el caché no funcionará correctamente",
                "accion":    "Usar None como default y crear el mutable dentro de "
                             "la función, o usar tuple en lugar de list",
            })

    # ── Resumen ────────────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        por_tipo = {}
        por_sev  = {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0}
        for h in self.hallazgos:
            por_tipo[h["tipo"]] = por_tipo.get(h["tipo"], 0) + 1
            por_sev[h["severidad"]] = por_sev.get(h["severidad"], 0) + 1
        return {
            "total_hallazgos":    len(self.hallazgos),
            "archivos_afectados": len(self.resultado),
            "por_tipo":           por_tipo,
            "por_severidad":      por_sev,
            "sin_cache_n":        por_tipo.get("funcion_costosa_sin_cache", 0),
            "deprecated_n":       por_tipo.get("cache_deprecado",           0),
            "session_state_n":    por_tipo.get("session_state_sin_init",    0),
        }

    def _resumen_vacio(self) -> dict:
        return {
            "total_hallazgos": 0, "archivos_afectados": 0,
            "por_tipo": {}, "por_severidad": {"CRITICO":0,"ALTO":0,"MEDIO":0,"BAJO":0},
            "sin_cache_n": 0, "deprecated_n": 0, "session_state_n": 0,
        }




# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 24 — DEPENDENCY SCANNER
# Audita los manifiestos de dependencias del proyecto sin red
# Soporta: requirements*.txt, pyproject.toml, setup.py, Pipfile,
#          package.json, pom.xml, build.gradle, composer.json,
#          .csproj, Gemfile, go.mod
# Detecta: deps sin pinear, wildcards, duplicadas prod/dev, riesgo heurístico
# ════════════════════════════════════════════════════════════════════════════

class DependencyScanner:
    """
    Auditor de dependencias del proyecto.
    Responsabilidades:
      - Parsear manifiestos de dependencias de múltiples ecosistemas
      - Detectar dependencias sin versión fija (riesgo de builds no reproducibles)
      - Detectar versiones con rangos amplios (^, ~, *, >=sin techo)
      - Detectar paquetes conocidos como problemáticos por versión (heurística local)
      - Calcular inventario total y clasificación prod/dev/opcional
    """

    # ─── Heurística de paquetes con versiones históricamente problemáticas ───
    # Lista mínima mantenida localmente — sin red, sin CVE DB externa
    # Formato: {paquete: [(operador, version_limite, severidad, descripcion)]}
    _PAQUETES_RIESGO = {
        # Python
        "pillow":       [("<", "9.0.0",  "ALTO",   "Múltiples CVEs RCE en <9.0")],
        "cryptography": [("<", "41.0.0", "ALTO",   "CVE-2023-38325 en <41.0")],
        "requests":     [("<", "2.28.0", "MEDIO",  "Vulnerabilidad SSRF en versiones antiguas")],
        "urllib3":      [("<", "1.26.5", "ALTO",   "CVE-2021-33503 ReDoS en <1.26.5")],
        "paramiko":     [("<", "2.10.1", "ALTO",   "CVE-2022-24302 symlink en <2.10.1")],
        "werkzeug":     [("<", "2.2.3",  "ALTO",   "CVE-2023-25577 DoS en <2.2.3")],
        "flask":        [("<", "2.2.5",  "MEDIO",  "Actualizar a >=2.3 recomendado")],
        "django":       [("<", "3.2.0",  "ALTO",   "Versiones <3.2 sin soporte LTS")],
        "sqlalchemy":   [("<", "1.4.0",  "MEDIO",  "API legacy, migrar a 2.x")],
        "pyyaml":       [("<", "5.4.0",  "CRÍTICO","CVE-2020-14343 RCE yaml.load sin Loader")],
        "numpy":        [("<", "1.22.0", "MEDIO",  "Múltiples CVEs de buffer overflow")],
        # JS/Node
        "lodash":       [("<", "4.17.21","ALTO",   "CVE-2021-23337 prototype pollution")],
        "axios":        [("<", "0.21.2", "ALTO",   "CVE-2021-3749 ReDoS en <0.21.2")],
        "express":      [("<", "4.18.0", "MEDIO",  "Actualizar a >=4.18 recomendado")],
        "minimist":     [("<", "1.2.6",  "ALTO",   "CVE-2021-44906 prototype pollution")],
        # Java
        "log4j-core":   [("<", "2.17.1", "CRÍTICO","Log4Shell CVE-2021-44228")],
        "spring-core":  [("<", "5.3.18", "CRÍTICO","Spring4Shell CVE-2022-22965")],
    }

    # ─── Manifiestos soportados por ecosistema ───────────────────────────────
    _MANIFIESTOS = {
        "python":     ["requirements.txt", "requirements-dev.txt", "requirements-test.txt",
                       "requirements-prod.txt", "pyproject.toml", "setup.py",
                       "setup.cfg", "Pipfile"],
        "javascript": ["package.json"],
        "java":       ["pom.xml", "build.gradle", "build.gradle.kts"],
        "php":        ["composer.json"],
        "csharp":     [".csproj", "packages.config"],
        "ruby":       ["Gemfile"],
        "go":         ["go.mod"],
    }

    def __init__(self, ruta_proyecto: str, archivos: list):
        self.ruta_proyecto = Path(ruta_proyecto)
        self.archivos      = archivos
        self.resultado     = {}   # {manifiesto: análisis}
        self.inventario    = []   # lista plana de todas las deps
        self.alertas       = []   # deps con riesgo detectado

    # ─── Punto de entrada principal ─────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Busca manifiestos en el proyecto y audita cada uno.
        Retorna inventario consolidado + alertas de riesgo.
        """
        manifiestos_encontrados = self._encontrar_manifiestos()

        for ruta_abs, ecosistema in manifiestos_encontrados:
            nombre = Path(ruta_abs).name
            try:
                contenido = Path(ruta_abs).read_text(encoding="utf-8", errors="replace")
            except Exception as e:
                self.resultado[nombre] = {"ok": False, "error": str(e)}
                continue

            parser = getattr(self, f"_parsear_{ecosistema}", None)
            if parser:
                deps = parser(contenido, nombre)
                self.resultado[nombre] = {
                    "ecosistema": ecosistema,
                    "deps":       deps,
                    "n_total":    len(deps),
                    "n_prod":     sum(1 for d in deps if not d.get("dev")),
                    "n_dev":      sum(1 for d in deps if d.get("dev")),
                    "n_sin_version": sum(1 for d in deps if not d.get("version")),
                    "n_wildcard": sum(1 for d in deps if d.get("es_wildcard")),
                    "ok":         True,
                }
                self.inventario.extend(deps)

        self._evaluar_riesgos()
        return {
            "manifiestos":  self.resultado,
            "inventario":   self.inventario,
            "alertas":      self.alertas,
            "resumen":      self._construir_resumen(),
        }

    # ─── Buscar manifiestos en el árbol del proyecto ────────────────────────
    def _encontrar_manifiestos(self) -> list:
        """
        Recorre el proyecto buscando archivos de manifiesto conocidos.
        Ignora node_modules, venv y carpetas similares.
        """
        encontrados = []
        ignorar = {"node_modules", "venv", ".venv", "env", "__pycache__",
                   "dist", "build", "target", "vendor"}

        for raiz, carpetas, ficheros in os.walk(self.ruta_proyecto):
            carpetas[:] = [c for c in carpetas if c not in ignorar]
            for fichero in ficheros:
                for ecosistema, nombres in self._MANIFIESTOS.items():
                    for patron in nombres:
                        if fichero == patron or (
                            patron.startswith("requirements") and
                            fichero.startswith("requirements") and
                            fichero.endswith(".txt")
                        ) or (
                            patron == ".csproj" and fichero.endswith(".csproj")
                        ):
                            ruta_abs = os.path.join(raiz, fichero)
                            encontrados.append((ruta_abs, ecosistema))
        return encontrados

    # ════════════════════════════════════════════════════════════════════════
    # PARSERS POR ECOSISTEMA
    # ════════════════════════════════════════════════════════════════════════

    # ─── requirements*.txt ──────────────────────────────────────────────────
    def _parsear_python(self, contenido: str, nombre: str) -> list:
        """
        Parsea requirements.txt estándar y variantes.
        Soporta: pkg==1.0, pkg>=1.0, pkg~=1.0, pkg, pkg[extra]>=1.0
        """
        deps = []
        es_dev = any(x in nombre.lower() for x in ("dev", "test", "lint", "ci"))

        for linea in contenido.splitlines():
            linea = linea.strip()
            if not linea or linea.startswith(("#", "-r", "-c", "--")):
                continue
            # Extraer nombre y versión
            m = re.match(
                r"^([A-Za-z0-9_.\-]+)(?:\[[^\]]+\])?"
                r"\s*([=!<>~^]+\s*[\d.*]+(?:\s*,\s*[=!<>~^]+\s*[\d.*]+)*)?\s*",
                linea
            )
            if not m:
                continue
            paquete = m.group(1).lower().replace("-", "_")
            version_spec = (m.group(2) or "").strip()
            deps.append(self._normalizar_dep(
                paquete, version_spec, dev=es_dev, ecosistema="python"
            ))
        return deps

    # ─── pyproject.toml ─────────────────────────────────────────────────────
    def _parsear_python_toml(self, contenido: str, nombre: str) -> list:
        """
        Parsea [project] dependencies y [project.optional-dependencies].
        También soporta [tool.poetry.dependencies].
        """
        deps = []
        seccion_actual = None

        for linea in contenido.splitlines():
            s = linea.strip()
            if s.startswith("["):
                seccion_actual = s
                continue
            if not s or s.startswith("#"):
                continue
            # PEP 621: "paquete>=1.0"
            if seccion_actual and "dependencies" in seccion_actual.lower():
                # dentro de lista TOML: "paquete>=ver",
                m = re.search(
                    r'"([A-Za-z0-9_.\-]+)(?:\[[^\]]+\])?\s*([=!<>~^,\s\d.*]*)"',
                    s
                )
                if not m:
                    # clave = "valor" estilo poetry
                    m2 = re.match(
                        r'([A-Za-z0-9_.\-]+)\s*=\s*["\'^~>=]*([^"\'#\s]+)',
                        s
                    )
                    if m2 and m2.group(1) not in ("python", "requires"):
                        es_dev = "dev" in seccion_actual.lower()
                        deps.append(self._normalizar_dep(
                            m2.group(1).lower(), m2.group(2), dev=es_dev, ecosistema="python"
                        ))
                    continue
                es_dev = "dev" in seccion_actual.lower() or "optional" in seccion_actual.lower()
                deps.append(self._normalizar_dep(
                    m.group(1).lower(), m.group(2).strip(), dev=es_dev, ecosistema="python"
                ))
        return deps

    # ─── package.json ───────────────────────────────────────────────────────
    def _parsear_javascript(self, contenido: str, nombre: str) -> list:
        """
        Parsea dependencies y devDependencies de package.json.
        """
        deps = []
        try:
            pkg = json.loads(contenido)
        except json.JSONDecodeError:
            return deps

        for clave, es_dev in [("dependencies", False), ("devDependencies", True),
                               ("peerDependencies", False), ("optionalDependencies", False)]:
            for paquete, version_spec in pkg.get(clave, {}).items():
                deps.append(self._normalizar_dep(
                    paquete.lower(), str(version_spec), dev=es_dev, ecosistema="javascript"
                ))
        return deps

    # ─── pom.xml ────────────────────────────────────────────────────────────
    def _parsear_java(self, contenido: str, nombre: str) -> list:
        """
        Parsea dependencias de Maven pom.xml.
        Extrae groupId, artifactId, version, scope.
        """
        deps = []
        try:
            # Eliminar namespace para simplificar XPath
            contenido_clean = re.sub(r'\s+xmlns[^"]*"[^"]*"', "", contenido)
            contenido_clean = re.sub(r'<\?xml[^>]+\?>', "", contenido_clean)
            root = ET.fromstring(contenido_clean)
        except ET.ParseError:
            return deps

        for dep in root.iter("dependency"):
            artifact = dep.findtext("artifactId") or ""
            version  = dep.findtext("version") or ""
            scope    = dep.findtext("scope") or "compile"
            es_dev   = scope in ("test", "provided")
            # Limpiar variables Maven ${...}
            version = "" if version.startswith("${") else version
            deps.append(self._normalizar_dep(
                artifact.lower(), version, dev=es_dev, ecosistema="java"
            ))
        return deps

    # ─── composer.json ──────────────────────────────────────────────────────
    def _parsear_php(self, contenido: str, nombre: str) -> list:
        """Parsea require y require-dev de composer.json."""
        deps = []
        try:
            pkg = json.loads(contenido)
        except json.JSONDecodeError:
            return deps
        for clave, es_dev in [("require", False), ("require-dev", True)]:
            for paquete, version_spec in pkg.get(clave, {}).items():
                if paquete == "php":
                    continue
                deps.append(self._normalizar_dep(
                    paquete.lower(), str(version_spec), dev=es_dev, ecosistema="php"
                ))
        return deps

    # ─── .csproj ────────────────────────────────────────────────────────────
    def _parsear_csharp(self, contenido: str, nombre: str) -> list:
        """Parsea PackageReference de archivos .csproj."""
        deps = []
        for m in re.finditer(
            r'<PackageReference\s+Include="([^"]+)"\s+Version="([^"]*)"',
            contenido, re.IGNORECASE
        ):
            deps.append(self._normalizar_dep(
                m.group(1).lower(), m.group(2), dev=False, ecosistema="csharp"
            ))
        return deps

    # ─── Gemfile ────────────────────────────────────────────────────────────
    def _parsear_ruby(self, contenido: str, nombre: str) -> list:
        """Parsea gem 'nombre', 'version' de Gemfile."""
        deps = []
        grupo_actual = None
        for linea in contenido.splitlines():
            s = linea.strip()
            m_grupo = re.match(r"group\s*:(\w+)", s)
            if m_grupo:
                grupo_actual = m_grupo.group(1)
                continue
            if s == "end":
                grupo_actual = None
                continue
            m = re.match(r"gem\s+['\"]([^'\"]+)['\"](?:\s*,\s*['\"]([^'\"]+)['\"])?", s)
            if m:
                es_dev = grupo_actual in ("development", "test", "ci")
                deps.append(self._normalizar_dep(
                    m.group(1).lower(), m.group(2) or "", dev=es_dev, ecosistema="ruby"
                ))
        return deps

    # ─── go.mod ─────────────────────────────────────────────────────────────
    def _parsear_go(self, contenido: str, nombre: str) -> list:
        """Parsea require de go.mod."""
        deps = []
        en_require = False
        for linea in contenido.splitlines():
            s = linea.strip()
            if s.startswith("require ("):
                en_require = True
                continue
            if en_require and s == ")":
                en_require = False
                continue
            if en_require or s.startswith("require "):
                m = re.match(r"(?:require\s+)?([^\s]+)\s+v([^\s]+)", s)
                if m:
                    deps.append(self._normalizar_dep(
                        m.group(1).lower(), "v" + m.group(2),
                        dev="// indirect" in s, ecosistema="go"
                    ))
        return deps

    # ─── Alias para pyproject.toml y Pipfile ────────────────────────────────
    def _parsear_python_pyproject(self, c, n): return self._parsear_python_toml(c, n)
    def _parsear_python_pipfile(self, c, n):   return self._parsear_python(c, n)

    # ════════════════════════════════════════════════════════════════════════
    # NORMALIZACIÓN Y EVALUACIÓN
    # ════════════════════════════════════════════════════════════════════════

    # ─── Normalizar una dependencia al formato común ─────────────────────────
    def _normalizar_dep(
        self, nombre: str, version_spec: str,
        dev: bool = False, ecosistema: str = ""
    ) -> dict:
        """
        Crea un dict normalizado para cualquier ecosistema.
        Evalúa si la versión es wildcard o no está fijada.
        """
        version_spec = (version_spec or "").strip()
        es_wildcard  = bool(re.search(r"[*x]|\.\*", version_spec)) if version_spec else False
        sin_version  = not version_spec
        # Rango amplio: solo >= sin techo (>=1.0 sin ,<2.0)
        rango_amplio = bool(
            re.match(r"^>=[\d.]+$", version_spec) or
            re.match(r"^\^", version_spec) or       # npm ^
            re.match(r"^~(?!=)", version_spec)       # npm ~ (no ~= Python)
        ) if version_spec else False

        return {
            "nombre":       nombre,
            "version":      version_spec,
            "dev":          dev,
            "ecosistema":   ecosistema,
            "es_wildcard":  es_wildcard,
            "sin_version":  sin_version,
            "rango_amplio": rango_amplio,
            "alertas":      [],
        }

    # ─── Evaluar riesgos sobre el inventario completo ────────────────────────
    def _evaluar_riesgos(self) -> None:
        """
        Cruza el inventario con la lista de paquetes de riesgo conocidos.
        También marca deps sin versión y wildcards como riesgo MEDIO/BAJO.
        """
        for dep in self.inventario:
            nombre = dep["nombre"].replace("-", "_")

            # ─── Sin versión especificada ───────────────────────────────────
            if dep["sin_version"]:
                alerta = {
                    "severidad": "MEDIO",
                    "paquete":   dep["nombre"],
                    "tipo":      "sin_version",
                    "detalle":   f"'{dep['nombre']}' sin versión fija — build no reproducible",
                }
                dep["alertas"].append(alerta)
                self.alertas.append(alerta)

            # ─── Versión wildcard ───────────────────────────────────────────
            elif dep["es_wildcard"]:
                alerta = {
                    "severidad": "MEDIO",
                    "paquete":   dep["nombre"],
                    "tipo":      "wildcard",
                    "detalle":   f"'{dep['nombre']}' usa wildcard '{dep['version']}' — puede instalar versión insegura",
                }
                dep["alertas"].append(alerta)
                self.alertas.append(alerta)

            # ─── Rango amplio (solo prod) ───────────────────────────────────
            elif dep["rango_amplio"] and not dep["dev"]:
                alerta = {
                    "severidad": "BAJO",
                    "paquete":   dep["nombre"],
                    "tipo":      "rango_amplio",
                    "detalle":   f"'{dep['nombre']}' con rango amplio '{dep['version']}' en prod",
                }
                dep["alertas"].append(alerta)
                self.alertas.append(alerta)

            # ─── Paquetes con versiones históricamente problemáticas ────────
            if nombre in self._PAQUETES_RIESGO and dep.get("version"):
                version_actual = re.search(r"[\d.]+", dep["version"])
                if not version_actual:
                    continue
                va = version_actual.group(0)
                for op, version_limite, severidad, desc in self._PAQUETES_RIESGO[nombre]:
                    try:
                        v_actual = tuple(int(x) for x in va.split(".")[:3])
                        v_limite = tuple(int(x) for x in version_limite.split(".")[:3])
                        riesgo = False
                        if op == "<"  and v_actual < v_limite: riesgo = True
                        if op == "<=" and v_actual <= v_limite: riesgo = True
                        if riesgo:
                            alerta = {
                                "severidad": severidad,
                                "paquete":   dep["nombre"],
                                "tipo":      "version_vulnerable",
                                "detalle":   f"'{dep['nombre']}=={va}' — {desc}",
                                "version_minima_segura": version_limite,
                            }
                            dep["alertas"].append(alerta)
                            self.alertas.append(alerta)
                    except (ValueError, AttributeError):
                        pass

    # ─── Resumen agregado ───────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        """Métricas globales de dependencias."""
        por_ecosistema = defaultdict(int)
        for dep in self.inventario:
            por_ecosistema[dep["ecosistema"]] += 1

        por_severidad = defaultdict(int)
        for a in self.alertas:
            por_severidad[a["severidad"]] += 1

        return {
            "n_manifiestos":     len(self.resultado),
            "n_deps_total":      len(self.inventario),
            "n_deps_prod":       sum(1 for d in self.inventario if not d.get("dev")),
            "n_deps_dev":        sum(1 for d in self.inventario if d.get("dev")),
            "n_sin_version":     sum(1 for d in self.inventario if d["sin_version"]),
            "n_wildcard":        sum(1 for d in self.inventario if d["es_wildcard"]),
            "n_alertas":         len(self.alertas),
            "alertas_criticas":  por_severidad.get("CRÍTICO", 0),
            "alertas_altas":     por_severidad.get("ALTO", 0),
            "alertas_medias":    por_severidad.get("MEDIO", 0),
            "por_ecosistema":    dict(por_ecosistema),
        }


# ════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 25 — PERFORMANCE PROFILER
# Anti-patrones de rendimiento NO cubiertos por QueryIntel
# QueryIntel: N+1 ORM/SQL, SELECT *, unbounded queries
# Este bloque: anti-patrones Python/JS a nivel de aplicación
# ════════════════════════════════════════════════════════════════════════════

class PerformanceProfiler:
    """
    Detector de anti-patrones de rendimiento a nivel de aplicación.
    Responsabilidades:
      - Detectar concatenación de strings en bucles (usar join)
      - Detectar I/O síncrono dentro de funciones async
      - Detectar ausencia de paginación en endpoints que devuelven listas
      - Detectar índices faltantes (tablas muy usadas sin índice definido)
      - Detectar creación de objetos grandes en memoria sin streaming
      - Detectar conexiones DB instanciadas sin pool (fuera de cache)
    """

    # ─── Inicialización ──────────────────────────────────────────────────────
    def __init__(self, archivos: list, resultado_sql=None, resultado_link=None):
        """
        archivos:      lista del scanner (todos los archivos)
        resultado_sql: output del SQLParser para cruzar índices
        resultado_link: output del CodeDBLinker para tablas más usadas
        """
        self.archivos       = archivos
        self.resultado_sql  = resultado_sql or {}
        self.resultado_link = resultado_link or {}
        self.hallazgos      = []

    # ─── Punto de entrada principal ──────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Escanea archivos de código buscando anti-patrones de rendimiento.
        """
        lenguajes_py  = {"Python"}
        lenguajes_web = {"JavaScript", "TypeScript"}

        for arch in self.archivos:
            lang = arch.get("lenguaje", "")
            ruta_abs = arch.get("ruta_abs", "")
            ruta_rel = arch.get("ruta_rel", "")
            if lang not in lenguajes_py | lenguajes_web:
                continue
            try:
                with open(ruta_abs, "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
            except Exception:
                continue

            if lang in lenguajes_py:
                self._analizar_python(contenido, ruta_rel)
            if lang in lenguajes_web:
                self._analizar_js(contenido, ruta_rel)

        # ─── Cruce con SQL: tablas muy usadas sin índice definido ────────────
        self._detectar_indices_faltantes()

        return {
            "hallazgos": self.hallazgos,
            "resumen":   self._construir_resumen(),
        }

    # ─── Análisis Python ─────────────────────────────────────────────────────
    def _analizar_python(self, contenido: str, ruta: str) -> None:
        """
        Detecta anti-patrones de rendimiento en código Python.
        """
        lineas = contenido.splitlines()

        # ─── Concatenación de strings en bucle ──────────────────────────────
        # Patrón: variable += "texto" o variable = variable + "texto" dentro de for/while
        en_bucle = False
        inicio_bucle = 0
        nivel_indent_bucle = 0
        for i, linea in enumerate(lineas, 1):
            stripped = linea.lstrip()
            indent   = len(linea) - len(stripped)
            if re.match(r"^(for|while)\s+", stripped):
                en_bucle = True
                inicio_bucle = i
                nivel_indent_bucle = indent
            elif en_bucle and indent <= nivel_indent_bucle and stripped and not stripped.startswith("#"):
                en_bucle = False
            if en_bucle and re.search(r"\w+\s*\+=\s*['\"]|\w+\s*=\s*\w+\s*\+\s*['\"]", stripped):
                self._agregar(ruta, i, "MEDIO", "string_concat_en_bucle",
                    "Concatenación de string en bucle — usar lista + ''.join() para O(n) en vez de O(n²)")

        # ─── I/O síncrono en función async ──────────────────────────────────
        en_async = False
        for i, linea in enumerate(lineas, 1):
            s = linea.strip()
            if re.match(r"^async\s+def\s+", s):
                en_async = True
            elif re.match(r"^def\s+|^class\s+", s):
                en_async = False
            if en_async:
                # time.sleep es el caso más claro
                if re.search(r"\btime\.sleep\s*\(", s):
                    self._agregar(ruta, i, "ALTO", "io_sync_en_async",
                        "time.sleep() en función async — usar await asyncio.sleep()")
                # requests (bloqueante) en async
                if re.search(r"\brequests\.(get|post|put|delete|patch)\s*\(", s):
                    self._agregar(ruta, i, "ALTO", "io_sync_en_async",
                        "requests (síncrono) en función async — usar httpx/aiohttp con await")
                # open() sin aiofiles en async
                if re.search(r"\bopen\s*\(", s) and "aiofiles" not in contenido:
                    self._agregar(ruta, i, "BAJO", "io_sync_en_async",
                        "open() síncrono en función async — considerar aiofiles para I/O masivo")

        # ─── Lectura completa de archivo grande sin chunking ─────────────────
        for i, linea in enumerate(lineas, 1):
            s = linea.strip()
            if re.search(r"\.read\(\)\s*$|\.readlines\(\)", s):
                # Solo flaggear si NO hay un size limit conocido
                if not re.search(r"\.read\(\d+\)", s):
                    self._agregar(ruta, i, "BAJO", "lectura_sin_chunking",
                        "Lectura completa sin límite — puede cargar archivos grandes en memoria")

        # ─── Conexión DB instanciada fuera de cache/pool ─────────────────────
        for i, linea in enumerate(lineas, 1):
            s = linea.strip()
            # psycopg2.connect o create_engine fuera de @st.cache_resource o singleton
            if re.search(r"psycopg2\.connect\s*\(|create_engine\s*\(|MongoClient\s*\(", s):
                # Heurística: si no está precedida por @st.cache_resource en las 3 líneas previas
                contexto = " ".join(lineas[max(0, i-4):i-1])
                if "@st.cache_resource" not in contexto and "@lru_cache" not in contexto:
                    self._agregar(ruta, i, "MEDIO", "conexion_sin_pool",
                        "Conexión DB instanciada sin caché/pool — puede crear conexión en cada request")

        # ─── Ausencia de paginación en Streamlit con st.dataframe grande ─────
        for i, linea in enumerate(lineas, 1):
            s = linea.strip()
            if re.search(r"st\.dataframe\(|st\.table\(", s):
                # Si no hay .head() o .limit() antes
                if not re.search(r"\.head\s*\(|\.limit\s*\(|\[:[\d]+\]", s):
                    self._agregar(ruta, i, "BAJO", "dataframe_sin_limite",
                        "st.dataframe() sin .head(n) — puede renderizar miles de filas")

    # ─── Análisis JavaScript/TypeScript ──────────────────────────────────────
    def _analizar_js(self, contenido: str, ruta: str) -> None:
        """
        Detecta anti-patrones de rendimiento en JS/TS.
        """
        lineas = contenido.splitlines()

        for i, linea in enumerate(lineas, 1):
            s = linea.strip()

            # ─── await dentro de forEach (no funciona como se espera) ───────
            if re.search(r"\.forEach\s*\(.*async", s) or (
                re.search(r"\.forEach\s*\(", s) and "await" in s
            ):
                self._agregar(ruta, i, "ALTO", "await_en_foreach",
                    "await dentro de forEach no espera las promesas — usar for...of o Promise.all()")

            # ─── Promise.all sobre array en loop ────────────────────────────
            # (ya está bien, pero detectar el antipatrón inverso: await en loop sin Promise.all)
            if re.search(r"for\s*\(|for\s+\w+\s+of\s+|for\s+\w+\s+in\s+", s):
                # buscar await en las 3 líneas siguientes dentro del loop
                ctx = " ".join(lineas[i:min(i+3, len(lineas))])
                if "await" in ctx and "Promise.all" not in contenido[max(0, contenido.find(linea)-100):]:
                    self._agregar(ruta, i, "MEDIO", "await_secuencial_en_loop",
                        "await secuencial en bucle — considerar Promise.all() para paralelizar")

            # ─── console.log en producción ───────────────────────────────────
            if re.search(r"\bconsole\.(log|warn|error)\s*\(", s):
                self._agregar(ruta, i, "BAJO", "console_log_produccion",
                    "console.log() en código — eliminar o reemplazar con logger para producción")

    # ─── Detectar tablas muy usadas sin índice definido ─────────────────────
    def _detectar_indices_faltantes(self) -> None:
        """
        Cruza las tablas más referenciadas en código (del Linker) con los
        índices definidos en SQL (del SQLParser). Tablas frecuentes sin
        índice son candidatas a degradación de performance.
        """
        if not self.resultado_sql or not self.resultado_link:
            return

        # Obtener tablas con índices definidos
        tablas_con_indice = set()
        for datos_arch in self.resultado_sql.get("archivos", {}).values():
            for idx in datos_arch.get("indices", []):
                tablas_con_indice.add(idx.get("tabla", "").lower())
            # La PK cuenta como índice implícito
            for tabla in datos_arch.get("tablas", []):
                for col in tabla.get("columnas", []):
                    if col.get("primary"):
                        tablas_con_indice.add(tabla["nombre"].lower())

        # Tablas más usadas en código
        tabla_a_archivos = self.resultado_link.get("tabla_a_archivos", {})
        for tabla, usos in tabla_a_archivos.items():
            if len(usos) >= 3 and tabla not in tablas_con_indice:
                self._agregar(
                    "análisis_cross_módulo", 0, "MEDIO", "indice_faltante",
                    f"Tabla '{tabla}' referenciada en {len(usos)} archivos pero sin índice definido en SQL"
                )

    # ─── Agregar un hallazgo normalizado ────────────────────────────────────
    def _agregar(self, ruta: str, linea: int, severidad: str,
                 tipo: str, detalle: str) -> None:
        """Añade un hallazgo al listado con formato estándar."""
        self.hallazgos.append({
            "ruta": ruta, "linea": linea,
            "severidad": severidad, "tipo": tipo, "detalle": detalle,
        })

    # ─── Resumen ─────────────────────────────────────────────────────────────
    def _construir_resumen(self) -> dict:
        por_severidad = defaultdict(int)
        por_tipo      = defaultdict(int)
        for h in self.hallazgos:
            por_severidad[h["severidad"]] += 1
            por_tipo[h["tipo"]] += 1
        return {
            "total_hallazgos": len(self.hallazgos),
            "por_severidad":   dict(por_severidad),
            "por_tipo":        dict(por_tipo),
        }


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 26 — DOCUMENTATION SCORE
# Evalúa la calidad de documentación del proyecto
# Usa los datos ya calculados por ASTEnginePython (docstring flags)
# No re-lee los archivos — consume resultados existentes
# ════════════════════════════════════════════════════════════════════════════

class DocumentationScore:
    """
    Evaluador de calidad de documentación.
    Responsabilidades:
      - Calcular ratio de docstrings (funciones, clases, módulos)
      - Evaluar presencia y calidad del README
      - Calcular ratio comentarios/líneas de código
      - Detectar funciones complejas sin docstring (mayor riesgo)
      - Producir score 0-100 y lista de gaps
    """

    def __init__(self, resultado_ast: dict, archivos: list, ruta_proyecto: str):
        """
        resultado_ast: output del ASTEnginePython (ya calculado)
        archivos:      lista del scanner para calcular ratio comentarios
        ruta_proyecto: para buscar README
        """
        self.resultado_ast  = resultado_ast or {}
        self.archivos       = archivos
        self.ruta_proyecto  = Path(ruta_proyecto)
        self.hallazgos      = []

    # ─── Punto de entrada ────────────────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Calcula el score de documentación y lista los gaps.
        """
        stats_docstrings = self._analizar_docstrings()
        stats_readme     = self._analizar_readme()
        stats_comentarios= self._analizar_comentarios()

        score = self._calcular_score(stats_docstrings, stats_readme, stats_comentarios)

        return {
            "score":            score,
            "docstrings":       stats_docstrings,
            "readme":           stats_readme,
            "comentarios":      stats_comentarios,
            "hallazgos":        self.hallazgos,
            "resumen": {
                "score":              score,
                "badge":              self._badge(score),
                "total_hallazgos":    len(self.hallazgos),
            },
        }

    # ─── Analizar cobertura de docstrings ────────────────────────────────────
    def _analizar_docstrings(self) -> dict:
        """
        Recorre el output del ASTEngine y cuenta funciones/clases con/sin docstring.
        Penaliza más las funciones con complejidad alta sin documentar.
        """
        total_funciones = 0
        con_docstring   = 0
        total_clases    = 0
        clases_con_doc  = 0
        sin_doc_complejas = []   # funciones complejidad>5 sin docstring

        for ruta, datos in self.resultado_ast.get("archivos", {}).items():
            if not datos.get("ok"):
                continue

            for func in datos.get("funciones", []):
                total_funciones += 1
                if func.get("docstring"):
                    con_docstring += 1
                elif func.get("complejidad", 1) > 5:
                    sin_doc_complejas.append({
                        "ruta": ruta, "nombre": func["nombre"],
                        "linea": func["linea"], "complejidad": func["complejidad"],
                    })
                    self.hallazgos.append({
                        "severidad": "MEDIO",
                        "tipo": "funcion_compleja_sin_docstring",
                        "detalle": f"'{func['nombre']}' (complejidad {func['complejidad']}) sin docstring",
                        "ruta": ruta, "linea": func["linea"],
                    })

            for cls in datos.get("clases", []):
                total_clases += 1
                if cls.get("docstring"):
                    clases_con_doc += 1
                else:
                    self.hallazgos.append({
                        "severidad": "BAJO",
                        "tipo": "clase_sin_docstring",
                        "detalle": f"Clase '{cls['nombre']}' sin docstring",
                        "ruta": ruta, "linea": cls.get("linea", 0),
                    })

            for cls in datos.get("clases", []):
                for met in cls.get("metodos", []):
                    total_funciones += 1
                    if met.get("docstring"):
                        con_docstring += 1
                    elif met.get("complejidad", 1) > 5:
                        sin_doc_complejas.append({
                            "ruta": ruta,
                            "nombre": f"{cls['nombre']}.{met['nombre']}",
                            "linea": met["linea"],
                            "complejidad": met["complejidad"],
                        })

        ratio = round(con_docstring / total_funciones * 100, 1) if total_funciones else 0

        return {
            "total_funciones":      total_funciones,
            "con_docstring":        con_docstring,
            "sin_docstring":        total_funciones - con_docstring,
            "ratio_pct":            ratio,
            "total_clases":         total_clases,
            "clases_con_doc":       clases_con_doc,
            "sin_doc_complejas":    sin_doc_complejas[:10],
        }

    # ─── Analizar README ─────────────────────────────────────────────────────
    def _analizar_readme(self) -> dict:
        """
        Busca README.md / README.rst / README.txt y evalúa su calidad.
        """
        readme_path = None
        for nombre in ("README.md", "readme.md", "README.rst", "README.txt", "README"):
            candidato = self.ruta_proyecto / nombre
            if candidato.exists():
                readme_path = candidato
                break

        if not readme_path:
            self.hallazgos.append({
                "severidad": "ALTO",
                "tipo": "readme_ausente",
                "detalle": "No se encontró README en la raíz del proyecto",
                "ruta": ".", "linea": 0,
            })
            return {"presente": False, "score": 0}

        try:
            contenido = readme_path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return {"presente": True, "score": 10}

        lineas       = [l for l in contenido.splitlines() if l.strip()]
        n_lineas     = len(lineas)
        n_secciones  = len(re.findall(r"^#+\s+", contenido, re.MULTILINE))
        tiene_codigo = bool(re.search(r"```|~~~", contenido))
        tiene_install= bool(re.search(r"instal|install|pip\s|npm\s|setup", contenido, re.IGNORECASE))
        tiene_uso    = bool(re.search(r"uso|usage|ejemplo|example|quickstart", contenido, re.IGNORECASE))

        score = 0
        if n_lineas >= 10:  score += 30
        if n_secciones >= 3: score += 20
        if tiene_codigo:    score += 20
        if tiene_install:   score += 15
        if tiene_uso:       score += 15

        if score < 50:
            self.hallazgos.append({
                "severidad": "MEDIO",
                "tipo": "readme_incompleto",
                "detalle": f"README con score {score}/100 — añadir secciones de instalación, uso y ejemplos de código",
                "ruta": readme_path.name, "linea": 0,
            })

        return {
            "presente":       True,
            "archivo":        readme_path.name,
            "n_lineas":       n_lineas,
            "n_secciones":    n_secciones,
            "tiene_codigo":   tiene_codigo,
            "tiene_install":  tiene_install,
            "tiene_uso":      tiene_uso,
            "score":          score,
        }

    # ─── Ratio comentarios/código ─────────────────────────────────────────────
    def _analizar_comentarios(self) -> dict:
        """
        Para archivos Python: ratio líneas de comentario (#) vs total.
        """
        total_lineas     = 0
        lineas_comentario= 0
        lineas_codigo    = 0

        py_archivos = [a for a in self.archivos if a.get("lenguaje") == "Python"]

        for arch in py_archivos:
            try:
                with open(arch["ruta_abs"], "r", encoding="utf-8", errors="replace") as f:
                    for linea in f:
                        s = linea.strip()
                        total_lineas += 1
                        if not s:
                            continue
                        if s.startswith("#"):
                            lineas_comentario += 1
                        else:
                            lineas_codigo += 1
            except Exception:
                pass

        ratio = round(lineas_comentario / max(1, lineas_codigo) * 100, 1)

        if ratio < 5 and lineas_codigo > 200:
            self.hallazgos.append({
                "severidad": "BAJO",
                "tipo": "bajo_ratio_comentarios",
                "detalle": f"Ratio comentarios/código: {ratio}% — por debajo del mínimo recomendado (5%)",
                "ruta": "proyecto", "linea": 0,
            })

        return {
            "total_lineas":      total_lineas,
            "lineas_comentario": lineas_comentario,
            "lineas_codigo":     lineas_codigo,
            "ratio_pct":         ratio,
        }

    # ─── Calcular score final de documentación ───────────────────────────────
    def _calcular_score(self, docstrings: dict, readme: dict, comentarios: dict) -> int:
        """
        Score 0-100 ponderado:
          40% → ratio docstrings funciones
          30% → calidad README
          20% → ratio comentarios
          10% → clases documentadas
        """
        score_ds  = min(100, docstrings.get("ratio_pct", 0))
        score_rm  = readme.get("score", 0)
        score_cm  = min(100, comentarios.get("ratio_pct", 0) * 5)   # 20% = 100pts
        ratio_cls = (
            docstrings.get("clases_con_doc", 0) /
            max(1, docstrings.get("total_clases", 1)) * 100
        )

        score = int(
            score_ds * 0.40 +
            score_rm * 0.30 +
            score_cm * 0.20 +
            ratio_cls * 0.10
        )
        return max(0, min(100, score))

    def _badge(self, score: int) -> str:
        if score >= 80: return "EXCELENTE"
        if score >= 60: return "BUENO"
        if score >= 40: return "MEJORABLE"
        if score >= 20: return "DEFICIENTE"
        return "CRÍTICO"


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 27 — TEST COVERAGE DETECTOR
# Detecta la presencia, calidad y cobertura estimada de tests
# No ejecuta los tests — analiza estáticamente
# ════════════════════════════════════════════════════════════════════════════

class TestCoverageDetector:
    """
    Detector de cobertura de tests (análisis estático).
    Responsabilidades:
      - Encontrar archivos de test por convención de nombres
      - Detectar framework de testing activo
      - Estimar ratio tests/funciones de producción
      - Identificar funciones críticas (alta complejidad) sin test asociado
      - Detectar configuración de coverage.py
    """

    # ─── Patrones de archivos de test por lenguaje ───────────────────────────
    _PATRONES_TEST = {
        "Python":     [r"^test_.*\.py$", r".*_test\.py$"],
        "JavaScript": [r".*\.test\.(js|jsx)$", r".*\.spec\.(js|jsx)$"],
        "TypeScript": [r".*\.test\.(ts|tsx)$", r".*\.spec\.(ts|tsx)$"],
        "Java":       [r".*Test\.java$", r".*Tests\.java$", r".*IT\.java$"],
        "C#":         [r".*Tests?\.cs$", r".*Specs?\.cs$"],
        "PHP":        [r".*Test\.php$"],
        "Ruby":       [r".*_spec\.rb$", r".*_test\.rb$"],
    }

    # ─── Patrones para contar funciones de test dentro de un archivo ─────────
    _PATRONES_FUNCION_TEST = {
        "Python":     re.compile(r"^\s*def\s+(test_\w+)\s*\(", re.MULTILINE),
        "JavaScript": re.compile(r"\b(it|test)\s*\(\s*['\"`]", re.MULTILINE),
        "TypeScript": re.compile(r"\b(it|test)\s*\(\s*['\"`]", re.MULTILINE),
        "Java":       re.compile(r"@Test\b", re.MULTILINE),
        "C#":         re.compile(r"\[Test(?:Method)?\]", re.MULTILINE),
        "PHP":        re.compile(r"public\s+function\s+test\w+\s*\(", re.MULTILINE),
        "Ruby":       re.compile(r"\bit\s+['\"]|def\s+test_", re.MULTILINE),
    }

    # ─── Frameworks de testing detectables ───────────────────────────────────
    _FRAMEWORKS_TEST = {
        "pytest":     re.compile(r"import\s+pytest|@pytest\.|conftest"),
        "unittest":   re.compile(r"import\s+unittest|TestCase"),
        "jest":       re.compile(r"from\s+['\"]@jest|require.*jest|describe\s*\("),
        "mocha":      re.compile(r"require.*mocha|describe\s*\(.*done\)"),
        "jasmine":    re.compile(r"jasmine\.|describe\s*\("),
        "junit":      re.compile(r"org\.junit\.|@RunWith"),
        "xunit":      re.compile(r"using\s+Xunit|using\s+NUnit"),
        "rspec":      re.compile(r"require\s+['\"]rspec"),
    }

    def __init__(self, archivos: list, resultado_ast: dict, ruta_proyecto: str):
        self.archivos       = archivos
        self.resultado_ast  = resultado_ast or {}
        self.ruta_proyecto  = Path(ruta_proyecto)
        self.archivos_test  = []
        self.hallazgos      = []

    # ─── Punto de entrada ────────────────────────────────────────────────────
    def analizar(self) -> dict:
        """
        Analiza la suite de tests del proyecto.
        """
        self._identificar_archivos_test()
        stats_tests        = self._contar_funciones_test()
        frameworks         = self._detectar_frameworks()
        config_coverage    = self._detectar_coverage_config()
        funciones_sin_test = self._detectar_funciones_criticas_sin_test()
        ratio              = self._calcular_ratio(stats_tests)

        return {
            "archivos_test":        self.archivos_test,
            "n_archivos_test":      len(self.archivos_test),
            "n_funciones_test":     stats_tests["total"],
            "frameworks":           frameworks,
            "tiene_coverage_cfg":   config_coverage,
            "ratio_estimado_pct":   ratio,
            "funciones_criticas_sin_test": funciones_sin_test,
            "hallazgos":            self.hallazgos,
            "resumen": {
                "n_archivos_test":    len(self.archivos_test),
                "n_funciones_test":   stats_tests["total"],
                "ratio_estimado_pct": ratio,
                "frameworks":         frameworks,
                "tiene_coverage_cfg": config_coverage,
                "total_hallazgos":    len(self.hallazgos),
            },
        }

    # ─── Identificar archivos de test ────────────────────────────────────────
    def _identificar_archivos_test(self) -> None:
        """
        Clasifica archivos como test según patrones por lenguaje.
        También detecta carpetas tests/, test/, __tests__/.
        """
        for arch in self.archivos:
            nombre = arch["nombre"]
            lang   = arch.get("lenguaje", "")
            ruta_rel = arch.get("ruta_rel", "")

            # Carpeta de tests
            partes = Path(ruta_rel).parts
            en_carpeta_test = any(
                p.lower() in ("tests", "test", "__tests__", "spec", "specs")
                for p in partes[:-1]
            )

            if en_carpeta_test:
                self.archivos_test.append(arch)
                continue

            for patron in self._PATRONES_TEST.get(lang, []):
                if re.match(patron, nombre, re.IGNORECASE):
                    self.archivos_test.append(arch)
                    break

    # ─── Contar funciones de test ────────────────────────────────────────────
    def _contar_funciones_test(self) -> dict:
        """
        Lee cada archivo de test y cuenta las funciones/casos de prueba.
        """
        total = 0
        por_archivo = {}

        for arch in self.archivos_test:
            lang = arch.get("lenguaje", "Python")
            patron = self._PATRONES_FUNCION_TEST.get(lang)
            if not patron:
                continue
            try:
                with open(arch["ruta_abs"], "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
                n = len(patron.findall(contenido))
                total += n
                por_archivo[arch["ruta_rel"]] = n
            except Exception:
                pass

        return {"total": total, "por_archivo": por_archivo}

    # ─── Detectar frameworks de testing ──────────────────────────────────────
    def _detectar_frameworks(self) -> list:
        """
        Busca señales de frameworks de testing en archivos de test.
        """
        detectados = set()
        for arch in self.archivos_test[:20]:   # muestra primeros 20
            try:
                with open(arch["ruta_abs"], "r", encoding="utf-8", errors="replace") as f:
                    contenido = f.read()
                for fw, patron in self._FRAMEWORKS_TEST.items():
                    if patron.search(contenido):
                        detectados.add(fw)
            except Exception:
                pass
        # También buscar en archivos de config
        for nombre in ("pytest.ini", "setup.cfg", "pyproject.toml", "jest.config.js",
                        "jest.config.ts", ".mocharc.js", "karma.conf.js"):
            if (self.ruta_proyecto / nombre).exists():
                if "pytest" in nombre:    detectados.add("pytest")
                if "jest" in nombre:      detectados.add("jest")
                if "mocha" in nombre:     detectados.add("mocha")
        return sorted(detectados)

    # ─── Detectar configuración de coverage ──────────────────────────────────
    def _detectar_coverage_config(self) -> bool:
        """
        Busca archivos de configuración de cobertura de código.
        """
        configs = [".coveragerc", "coverage.ini", "codecov.yml", ".codecov.yml",
                   "coveralls.yml", ".coveralls.yml"]
        for cfg in configs:
            if (self.ruta_proyecto / cfg).exists():
                return True
        # También buscar [tool.coverage] en pyproject.toml
        pyp = self.ruta_proyecto / "pyproject.toml"
        if pyp.exists():
            try:
                if "[tool.coverage" in pyp.read_text():
                    return True
            except Exception:
                pass
        return False

    # ─── Funciones críticas sin test ─────────────────────────────────────────
    def _detectar_funciones_criticas_sin_test(self) -> list:
        """
        Identifica funciones con complejidad > 8 en producción que no tienen
        un archivo de test hermano (mismo nombre con prefijo test_).
        """
        archivos_con_test = {
            re.sub(r"^test_|_test$", "", Path(a["ruta_rel"]).stem.lower())
            for a in self.archivos_test
        }

        sin_test = []
        for ruta_rel, datos in self.resultado_ast.get("archivos", {}).items():
            if not datos.get("ok"):
                continue
            modulo = Path(ruta_rel).stem.lower()
            tiene_test = modulo in archivos_con_test or f"test_{modulo}" in archivos_con_test

            for func in datos.get("funciones", []):
                if func.get("complejidad", 1) > 8 and not tiene_test:
                    sin_test.append({
                        "ruta":        ruta_rel,
                        "funcion":     func["nombre"],
                        "linea":       func["linea"],
                        "complejidad": func["complejidad"],
                    })
                    self.hallazgos.append({
                        "severidad": "MEDIO",
                        "tipo": "funcion_critica_sin_test",
                        "detalle": (
                            f"'{func['nombre']}' (complejidad {func['complejidad']}) "
                            f"en {ruta_rel} sin archivo de test asociado"
                        ),
                        "ruta": ruta_rel, "linea": func["linea"],
                    })

        # Warning si no hay tests en absoluto
        if not self.archivos_test:
            self.hallazgos.append({
                "severidad": "ALTO",
                "tipo": "sin_tests",
                "detalle": "El proyecto no tiene ningún archivo de test detectado",
                "ruta": ".", "linea": 0,
            })

        return sin_test[:20]

    # ─── Calcular ratio estimado ──────────────────────────────────────────────
    def _calcular_ratio(self, stats_tests: dict) -> float:
        """
        Ratio estimado: funciones_test / funciones_produccion.
        No es cobertura real (requeriría ejecutar coverage.py).
        """
        n_tests = stats_tests.get("total", 0)
        n_prod  = sum(
            d.get("n_funciones", 0) + d.get("n_metodos", 0)
            for d in self.resultado_ast.get("archivos", {}).values()
            if d.get("ok")
        )
        if n_prod == 0:
            return 0.0
        return round(min(100.0, n_tests / n_prod * 100), 1)


# ════════════════════════════════════════════════════════════════════════════
# BLOQUE TRANSVERSAL — Pattern Intelligence Capa 3
# Patrones externos en ~/.stratum/patterns/<lang>.json
# Tienen prioridad sobre los patrones internos del HeuristicEngine
# Única operación de red de STRATUM: --update-patterns (opt-in explícito)
# ════════════════════════════════════════════════════════════════════════════

class PatternIntelligenceCapa3:
    """
    Sistema de patrones externos actualizables.
    Responsabilidades:
      - Cargar patrones JSON de ~/.stratum/patterns/ si existen
      - Inyectarlos en HeuristicEngine con prioridad sobre los internos
      - Proveer el mecanismo de descarga --update-patterns (opt-in)
      - Reportar qué patrones están activos y su fuente
    """

    _URL_BASE = "https://raw.githubusercontent.com/stratumengine/patterns/main"
    _LANGS    = ["javascript", "typescript", "java", "csharp", "php", "r"]

    def __init__(self, ruta_trabajo: Path):
        self.ruta_patterns = ruta_trabajo / "patterns"
        self.patrones_cargados = {}   # {lang: [patrones]}

    # ─── Cargar patrones externos si existen ────────────────────────────────
    def cargar(self) -> dict:
        """
        Lee archivos JSON de ~/.stratum/patterns/ para cada lenguaje.
        Si no existen, retorna dict vacío (sin afectar el análisis).
        """
        if not self.ruta_patterns.exists():
            return {}

        for lang in self._LANGS:
            ruta_json = self.ruta_patterns / f"{lang}.json"
            if ruta_json.exists():
                try:
                    datos = json.loads(ruta_json.read_text(encoding="utf-8"))
                    if isinstance(datos, list):
                        self.patrones_cargados[lang] = datos
                except (json.JSONDecodeError, Exception):
                    pass   # JSON corrupto → ignorar, no romper el análisis

        return self.patrones_cargados

    # ─── Actualizar patrones desde el repositorio oficial ───────────────────
    def actualizar(self) -> dict:
        """
        ÚNICA operación de red en STRATUM — solo cuando el usuario
        ejecuta explícitamente: python3 stratum.py --update-patterns

        Descarga los JSONs de patrones del repositorio oficial a
        ~/.stratum/patterns/. Si falla la red, avisa pero no aborta.
        """
        import urllib.request
        self.ruta_patterns.mkdir(parents=True, exist_ok=True)

        resultados = {}
        for lang in self._LANGS:
            url  = f"{self._URL_BASE}/{lang}.json"
            dest = self.ruta_patterns / f"{lang}.json"
            try:
                req = urllib.request.urlopen(url, timeout=10)
                contenido = req.read().decode("utf-8")
                # Validar que es JSON válido antes de guardar
                json.loads(contenido)
                dest.write_text(contenido, encoding="utf-8")
                resultados[lang] = "actualizado"
            except Exception as e:
                resultados[lang] = f"error: {e}"

        return resultados

    # ─── Reportar estado de patrones ────────────────────────────────────────
    def reporte_estado(self) -> dict:
        """
        Retorna qué patrones externos están disponibles y su fecha.
        """
        estado = {}
        if not self.ruta_patterns.exists():
            return {"disponibles": False, "patrones": {}}

        for lang in self._LANGS:
            ruta_json = self.ruta_patterns / f"{lang}.json"
            if ruta_json.exists():
                try:
                    mtime = ruta_json.stat().st_mtime
                    fecha = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d")
                    datos = json.loads(ruta_json.read_text())
                    estado[lang] = {"fecha": fecha, "n_patrones": len(datos)}
                except Exception:
                    estado[lang] = {"fecha": "?", "n_patrones": 0}

        return {
            "disponibles": bool(estado),
            "patrones":    estado,
            "ruta":        str(self.ruta_patterns),
        }


# BLOQUE ORQ — ORQUESTADOR Y PUNTO DE ENTRADA
# Gestiona argumentos CLI, inicializa todos los módulos y arranca el servidor
# Este bloque NO tiene número porque es el coordinador, no una capacidad.
# Los bloques numerados (0-28) son funcionalidades del cronograma STRATUM.
# ════════════════════════════════════════════════════════════════════════════

class StratumApp:
    """
    Orquestador principal de STRATUM.
    Inicializa y coordina todos los módulos del sistema.
    En sesiones futuras, aquí se agregan los nuevos bloques.
    """

    def __init__(self, ruta_proyecto: str, puerto: int, modo_red: bool):
        self.ruta_proyecto = ruta_proyecto
        self.puerto        = puerto
        self.modo_red      = modo_red

        # Carpeta de trabajo de STRATUM (fuera del proyecto analizado)
        self.ruta_trabajo  = Path.home() / ".stratum"
        self.ruta_trabajo.mkdir(exist_ok=True)

        # Inicializar módulos core
        self.seguridad     = SeguridadStratum()
        self.logger        = SeguridadStratum.configurar_logging(self.ruta_trabajo)
        self.licencias     = SistemaLicencias(self.ruta_trabajo)
        self.scanner       = ScannerUniversal(ruta_proyecto)

        # Carpeta de caché para acelerar re-análisis
        self.ruta_cache    = self.ruta_trabajo / "cache"
        self.ruta_cache.mkdir(exist_ok=True)

        # Estado del análisis (se llena al ejecutar)
        self.resultado_scan    = None
        self.resultado_ast     = None    # Bloque 5: AST Engine Python
        self.resultado_heur    = None    # Bloque 6: Heuristic Engine multi-lang
        self.resultado_sql     = None    # Bloque 7: SQL Parser
        self.resultado_link    = None    # Bloque 8: Code-DB Linker
        self.resultado_etl     = None    # Bloque 9: ETL/ELT Detector
        self.resultado_cloud   = None    # Bloque 10: Cloud Auditor
        self.resultado_dwh     = None    # Bloque 11: DWH Auditor
        self.resultado_profiler = None   # Bloque 12: Data Profiler
        self.resultado_framework= None   # Transversal: Framework Detector
        self.resultado_kpi      = None   # Bloque 13: KPI Engine
        self.resultado_viz      = None   # Bloque 14: Viz Mapper
        self.resultado_query    = None   # Bloque 15: Query Intel
        self.resultado_flow     = None   # Bloque 16: Flow Analyzer
        self.resultado_excel    = None   # Bloque 17: Excel Intelligence
        self.resultado_score    = None   # Bloque 18: STRATUM Score
        self.resultado_quickfix = None   # Bloque 19: Quick Fix Engine
        self.resultado_coverage = None   # Bloque 20: Coverage Score
        self.resultado_version  = None   # Bloque 21: Language Version
        self.resultado_security = None   # Bloque 23: Security Auditor
        self.resultado_reporte  = None   # Bloque 28: Report Generator
        self.resultado_st_cache = None   # Bloque 22: Streamlit Cache
        self.resultado_deps     = None   # Bloque 24: Dependency Scanner
        self.resultado_perf     = None   # Bloque 25: Performance Profiler
        self.resultado_docs     = None   # Bloque 26: Documentation Score
        self.resultado_tests    = None   # Bloque 27: Test Coverage Detector
        self.patron_capa3       = None   # Pattern Intelligence Capa 3

    # ─── Mostrar banner de bienvenida ─────────────────────────────────────────
    def _mostrar_banner(self) -> None:
        """Imprime el banner de STRATUM con información del sistema."""
        estado = self.licencias.obtener_estado()
        hash_actual = self.seguridad.calcular_hash_propio()

        print("\n")
        print("  ╔══════════════════════════════════════════════════════════╗")
        print(f"  ║  STRATUM System Intelligence Engine  v{STRATUM_VERSION}          ║")
        print(f"  ║  {STRATUM_AUTOR:<56}  ║")
        print("  ╠══════════════════════════════════════════════════════════╣")
        print(f"  ║  Tier    : {estado['tier']:<47}  ║")
        print(f"  ║  Estado  : {estado['modo']:<47}  ║")
        print(f"  ║  Región  : {estado['region']:<47}  ║")
        print(f"  ║  SHA-256 : {hash_actual[:47]:<47}  ║")
        print("  ╠══════════════════════════════════════════════════════════╣")
        print(f"  ║  {estado['mensaje']:<56}  ║")
        print("  ╚══════════════════════════════════════════════════════════╝")

        # Mostrar advertencia de early adopter si aplica
        if estado["early_adopter"] and estado["modo"] == "pago":
            print("\n  🎉 Early Adopter — 50% de descuento permanente aplicado")

        print()

    # ─── Ejecutar el análisis completo del proyecto ───────────────────────────
    def ejecutar_analisis(self) -> None:
        """
        Coordina la ejecución de todos los módulos de análisis.
        Sesión 1: Scanner. Sesión 2: AST + Heuristic + SQL + Linker.
        Sesión 3: ETL + Cloud + DWH + Data Profiler.
        Próximas sesiones: KPIs, ML, Excel, Score, Reportes.
        """
        self.logger.info(f"Iniciando análisis: {self.ruta_proyecto}")

        # ─── BLOQUE 4: Scanner Universal ────────────────────────────────────
        self.resultado_scan = self.scanner.escanear()
        archivos = self.scanner.archivos

        # ─── BLOQUE 5: AST Engine Python ────────────────────────────────────
        archivos_py = self.scanner.filtrar_por_lenguaje("Python")
        if archivos_py:
            print(f"\n  🧠 Analizando {len(archivos_py)} archivo(s) Python con AST...")
            # Detectar framework PRIMERO para pasar umbrales al AST engine
            _fw_pre = FrameworkDetector(archivos, self.scanner).detectar()
            _umbrales_pre = _fw_pre.get("umbrales", {})

            self.resultado_ast = ASTEnginePython(
                archivos_py, self.scanner,
                umbrales_fw=_umbrales_pre
            ).analizar()
            r = self.resultado_ast["resumen"]
            print(f"     {r['total_clases']} clases · {r['total_funciones']} funciones · "
                  f"{r['complejidad_total']} complejidad · {r['total_problemas']} problemas")

        # ─── BLOQUE 6: Heuristic Engine multi-lenguaje ──────────────────────
        archivos_heur = [
            a for a in archivos
            if a["lenguaje"] in ("JavaScript", "TypeScript", "Java", "C#", "PHP", "R")
        ]
        if archivos_heur:
            print(f"\n  🔍 Analizando {len(archivos_heur)} archivo(s) JS/TS/Java/C#/PHP/R...")
            self.resultado_heur = HeuristicEngine(archivos, self.scanner).analizar()
            r = self.resultado_heur["resumen"]
            for lang, stats in r["por_lenguaje"].items():
                print(f"     {lang}: {stats['archivos']} archivos · "
                      f"{stats['clases']} clases · {stats['funciones']+stats['metodos']} funciones")

        # ─── BLOQUE 7: SQL Parser multi-dialecto ────────────────────────────
        archivos_sql = self.scanner.filtrar_por_lenguaje("SQL")
        if archivos_sql:
            print(f"\n  🗄️  Analizando {len(archivos_sql)} archivo(s) SQL...")
            self.resultado_sql = SQLParser(archivos, self.scanner).analizar()
            r = self.resultado_sql["resumen"]
            print(f"     {r['total_tablas']} tablas · {r['total_vistas']} vistas · "
                  f"{r['total_funciones']} funciones · {r['total_fks']} FKs · "
                  f"{r['total_policies_rls']} políticas RLS")

        # ─── BLOQUE 8: Code-DB Linker ───────────────────────────────────────
        if self.resultado_sql:
            print(f"\n  🔗 Cruzando código con base de datos...")
            self.resultado_link = CodeDBLinker(
                self.resultado_sql["catalogo"], archivos, self.scanner
            ).analizar()
            r = self.resultado_link["resumen"]
            print(f"     {r['archivos_con_db']} archivos referencian DB · "
                  f"{r['huerfanas_db']} tablas huérfanas · "
                  f"{r['huerfanas_codigo']} referencias sin definición")

        # ─── BLOQUE 9: ETL/ELT Detector ─────────────────────────────────────
        print(f"\n  🔄 Detectando pipelines ETL/ELT...")
        self.resultado_etl = ETLDetector(archivos, self.scanner).analizar()
        r = self.resultado_etl["resumen"]
        if r["total_pipelines"] > 0:
            fws = ", ".join(r["por_framework"].keys())
            print(f"     {r['archivos_con_etl']} archivos ETL · "
                  f"{r['total_pipelines']} pipelines · frameworks: {fws}")
        else:
            print(f"     Sin pipelines ETL detectados")

        # ─── BLOQUE 10: Cloud Auditor ────────────────────────────────────────
        print(f"\n  ☁️  Auditando configuración cloud...")
        self.resultado_cloud = CloudAuditor(archivos, self.scanner).analizar()
        r = self.resultado_cloud["resumen"]
        provs = ", ".join(r["proveedores"]) if r["proveedores"] else "ninguno"
        print(f"     {r['archivos_cloud']} archivos cloud · Proveedores: {provs} · "
              f"Alertas: {r['total_alertas']} ({r['alertas_criticas']} críticas)")

        # ─── BLOQUE 11: DWH Auditor ──────────────────────────────────────────
        if self.resultado_sql:
            archivos_sql = self.scanner.filtrar_por_lenguaje("SQL")
            print(f"\n  🏗️  Auditando modelo dimensional (DWH)...")
            self.resultado_dwh = DWHAuditor(
                self.resultado_sql["catalogo"], archivos_sql, self.scanner
            ).analizar()
            r = self.resultado_dwh["resumen"]
            print(f"     {r['total_dims']} dimensiones · {r['total_facts']} hechos · "
                  f"Schema: {r['tipo_schema']} · Score: {r['score_calidad']}/100")

        # ─── BLOQUE 12: Data Profiler ────────────────────────────────────────
        archivos_datos = [
            a for a in archivos
            if a.get("extension", "").lower() in {".csv", ".json"}
        ]
        if archivos_datos:
            print(f"\n  📊 Perfilando {len(archivos_datos)} archivo(s) de datos...")
            self.resultado_profiler = DataProfiler(archivos_datos, self.scanner).analizar()
            r = self.resultado_profiler["resumen"]
            print(f"     {r['archivos_perfilados']} perfilados · "
                  f"{r['total_filas']} filas · {r['alertas_pii']} alertas PII")

        # ─── TRANSVERSAL: Framework Primary Detector ────────────────────────
        print("\n  🔍 Detectando framework primario...")
        # Reusar resultado pre-calculado si existe, evita doble lectura de archivos
        self.resultado_framework = (
            _fw_pre if "_fw_pre" in dir() else
            FrameworkDetector(archivos, self.scanner).detectar()
        )
        fw = self.resultado_framework.get('framework', 'generic')
        print(f"     Framework detectado: {fw}")

        # ─── BLOQUE 13: KPI Engine ────────────────────────────────────────────
        print("\n  📈 Detectando KPIs de negocio...")
        self.resultado_kpi = KPIEngine(archivos, self.scanner).analizar()
        r_kpi = self.resultado_kpi["resumen"]
        print(f"     {r_kpi['n_archivos_con_kpi']} archivos con KPIs · "
              f"{r_kpi['total_kpis_detectados']} métricas · "
              f"{r_kpi['total_agregaciones']} agregaciones")

        # ─── BLOQUE 14: Viz Mapper ────────────────────────────────────────────
        print("\n  🗺️  Mapeando visualizaciones...")
        self.resultado_viz = VizMapper(archivos, self.scanner).analizar()
        r_viz = self.resultado_viz["resumen"]
        libs  = ', '.join(r_viz.get('librerias_usadas', [])[:4]) or 'ninguna'
        print(f"     {r_viz['total_visualizaciones']} visualizaciones · Libs: {libs}")

        # ─── BLOQUE 15: Query Intel ───────────────────────────────────────────
        print("\n  🔎 Auditando calidad de queries...")
        self.resultado_query = QueryIntel(archivos, self.scanner).analizar()
        r_query = self.resultado_query["resumen"]
        criticos = r_query['por_severidad'].get('CRITICO', 0)
        print(f"     {r_query['total_hallazgos']} hallazgos · "
              f"{criticos} críticos")

        # ─── BLOQUE 16: Flow Analyzer ─────────────────────────────────────────
        print("\n  🌊 Reconstruyendo flujos de negocio...")
        catalogo = self.resultado_sql.get('catalogo', {}) if self.resultado_sql else {}
        self.resultado_flow = FlowAnalyzer(archivos, catalogo, self.scanner).analizar()
        r_flow = self.resultado_flow["resumen"]
        print(f"     {r_flow['n_flujos_detectados']} flujos · "
              f"{r_flow['n_gaps']} gaps · {r_flow['n_bypasses']} bypasses")

        # ─── BLOQUE 17: Excel Intelligence ─────────────────────────────────
        archivos_excel = [a for a in archivos
                          if a.get("extension","").lower()
                          in {".xlsx",".xls",".xlsm",".xlsb",".csv",".tsv",".ods"}]
        if archivos_excel:
            print("\n  📊 Auditando archivos Excel/CSV...")
            self.resultado_excel = ExcelIntelligence(archivos, self.scanner).analizar()
            r_excel = self.resultado_excel["resumen"]
            print(f"     {r_excel['total_archivos']} archivos · "
                  f"{r_excel['total_problemas']} problemas")

        # ─── BLOQUE 23: Security Auditor ────────────────────────────────────
        print("\n  🔒 Ejecutando Security Auditor...")
        self.resultado_security = SecurityAuditor(archivos, self.scanner).analizar()
        r_sec = self.resultado_security["resumen"]
        print(f"     {r_sec['total_hallazgos']} hallazgos · "
              f"nivel: {r_sec['nivel_riesgo']} · "
              f"archivos afectados: {r_sec['archivos_afectados']}")

        # ─── BLOQUE 22: Streamlit Cache Detector ─────────────────────────────
        self.resultado_st_cache = StreamlitCacheDetector(
            archivos, self.resultado_framework, self.scanner
        ).analizar()
        if self.resultado_st_cache.get('aplica'):
            r_stc = self.resultado_st_cache['resumen']
            print(f"\n  ⚡ Streamlit Cache: {r_stc['total_hallazgos']} hallazgos · "
                  f"{r_stc['sin_cache_n']} sin caché · "
                  f"{r_stc['deprecated_n']} deprecados")

        # ─── BLOQUE 20+21: Pattern Intelligence ─────────────────────────────
        print("\n  🔬 Analizando cobertura de patrones y versiones...")
        self.resultado_coverage = CoverageScoreEngine(
            archivos,
            resultado_ast=self.resultado_ast,
            resultado_heur=self.resultado_heur,
            scanner=self.scanner,
        ).analizar()
        self.resultado_version = LanguageVersionDetector(
            archivos, self.scanner
        ).analizar()
        n_adv_cov = self.resultado_coverage["n_advertencias"]
        n_ver     = self.resultado_version["n_lenguajes_detectados"]
        print(f"     Cobertura global: {self.resultado_coverage['ratio_global']}% · "
              f"{n_ver} lenguajes detectados · {n_adv_cov} advertencias")

        # ─── BLOQUE 24: Dependency Scanner ──────────────────────────────────
        print("\n  📦 Escaneando dependencias...")
        self.resultado_deps = DependencyScanner(
            self.ruta_proyecto, archivos
        ).analizar()
        n_dep_alerta = self.resultado_deps["resumen"]["total_alertas"]
        n_dep_pkg    = self.resultado_deps["resumen"]["total_paquetes"]
        print(f"     {n_dep_pkg} paquetes · {n_dep_alerta} alertas de seguridad")

        # ─── BLOQUE 25: Performance Profiler ─────────────────────────────────
        print("\n  ⚡ Perfilando rendimiento...")
        self.resultado_perf = PerformanceProfiler(
            archivos,
            resultado_sql=self.resultado_sql,
            resultado_link=self.resultado_link,
        ).analizar()
        n_perf = self.resultado_perf["resumen"]["total_hallazgos"]
        print(f"     {n_perf} hallazgos de rendimiento")

        # ─── BLOQUE 26: Documentation Score ──────────────────────────────────
        print("\n  📝 Evaluando documentación...")
        self.resultado_docs = DocumentationScore(
            self.resultado_ast, archivos, self.ruta_proyecto
        ).analizar()
        doc_score = self.resultado_docs["score"]
        print(f"     Score documentación: {doc_score}/100")

        # ─── BLOQUE 27: Test Coverage Detector ───────────────────────────────
        print("\n  🧪 Detectando cobertura de tests...")
        self.resultado_tests = TestCoverageDetector(
            archivos, self.resultado_ast, self.ruta_proyecto
        ).analizar()
        n_test_fn = self.resultado_tests["resumen"]["total_funciones_test"]
        tc_ratio  = self.resultado_tests["resumen"]["ratio_archivos_test_pct"]
        print(f"     {n_test_fn} funciones de test · {tc_ratio}% archivos son tests")

        # ─── Pattern Intelligence Capa 3 ─────────────────────────────────────
        self.patron_capa3 = PatternIntelligenceCapa3(
            Path(self.ruta_proyecto)
        ).cargar()

        # ─── BLOQUE 18: STRATUM Score ────────────────────────────────────────
        print("\n  🏆 Calculando STRATUM Score...")
        self.resultado_score = StratumScore({
            "ast":       self.resultado_ast,
            "heur":      self.resultado_heur,
            "sql":       self.resultado_sql,
            "link":      self.resultado_link,
            "etl":       self.resultado_etl,
            "cloud":     self.resultado_cloud,
            "dwh":       self.resultado_dwh,
            "profiler":  self.resultado_profiler,
            "kpi":       self.resultado_kpi,
            "viz":       self.resultado_viz,
            "query":     self.resultado_query,
            "flow":      self.resultado_flow,
            "excel":    self.resultado_excel,
            "security":  self.resultado_security,
            "st_cache":  self.resultado_st_cache,
            "framework": self.resultado_framework,
            "deps":      self.resultado_deps,
            "perf":      self.resultado_perf,
            "docs":      self.resultado_docs,
            "tests":     self.resultado_tests,
        }).calcular()
        print(f"     Score: {self.resultado_score['score']}/100 · "
              f"{self.resultado_score['badge_icon']} {self.resultado_score['badge']} · "
              f"{self.resultado_score['total_findings']} findings")

        # ─── BLOQUE 19: Quick Fix Engine ─────────────────────────────────────
        self.resultado_quickfix = QuickFixEngine(
            self.resultado_score["findings"]
        ).generar()
        n_qw = len(self.resultado_quickfix["quick_wins"])
        print(f"     {self.resultado_quickfix['total_sugerencias']} sugerencias · "
              f"{n_qw} quick wins (minutos)")

        # ─── BLOQUE 28: Report Generator ────────────────────────────────────
        try:
            self.resultado_reporte = ReportGenerator(self).generar()
            print(f"\n  📄 Reporte ejecutivo: "
                  f"{self.resultado_reporte['ruta_reporte']}")
        except Exception as _e_rep:
            print(f"\n  ⚠ Reporte no generado: {_e_rep}")

        # ─── Registrar sesión en el sistema de licencias ────────────────────
        self.licencias.registrar_sesion(
            self.ruta_proyecto,
            self.resultado_scan["estadisticas"]["total_archivos"]
        )

        self.logger.info(
            f"Scan completo: {self.resultado_scan['estadisticas']['total_archivos']} "
            f"archivos, {self.resultado_scan['estadisticas']['total_mb']} MB"
        )

        # Mostrar resumen final en consola
        self._mostrar_resumen_scan()

    # ─── Mostrar resumen del escaneo en consola ───────────────────────────────
    def _mostrar_resumen_scan(self) -> None:
        """Imprime un resumen legible del escaneo en la terminal."""
        if not self.resultado_scan:
            return

        stats = self.resultado_scan["estadisticas"]
        print("\n  📊 RESUMEN DEL PROYECTO")
        print(f"  {'─' * 50}")
        print(f"  Nombre    : {self.resultado_scan['nombre']}")
        print(f"  Archivos  : {stats['total_archivos']}")
        print(f"  Carpetas  : {stats['total_carpetas']}")
        print(f"  Tamaño    : {stats['total_mb']} MB")
        print(f"\n  Por categoría:")

        for cat, cantidad in sorted(
            stats["por_categoria"].items(), key=lambda x: -x[1]
        ):
            barra = "█" * min(cantidad, 30)
            print(f"    {cat:<12} {barra} {cantidad}")

        print()

    # ─── Iniciar el servidor web local ───────────────────────────────────────
    def iniciar_servidor(self) -> None:
        """
        Arranca el servidor HTTP local para la interfaz web.
        El servidor SOLO escucha en localhost salvo --red explícito.
        La interfaz web completa se construye en la Sesión 6.
        """
        host = self.seguridad.validar_host(HOST_LOCAL, self.modo_red)

        # Manejador HTTP básico — se reemplaza en Sesión 6 con UI completa
        app_ref = self

        class ManejadorHTTP(http.server.BaseHTTPRequestHandler):

            def do_GET(self):
                """Sirve la interfaz web básica con los resultados del análisis."""
                self.send_response(200)
                self.send_header("Content-type", "text/html; charset=utf-8")
                # Cabeceras de seguridad
                self.send_header("X-Content-Type-Options", "nosniff")
                self.send_header("X-Frame-Options", "DENY")
                self.send_header("Content-Security-Policy", "default-src 'self' 'unsafe-inline'")
                self.end_headers()

                html = app_ref._generar_html_basico()
                self.wfile.write(html.encode("utf-8"))

            def log_message(self, format, *args):
                """Suprimir logs del servidor en consola (van al archivo de auditoría)."""
                app_ref.logger.info(f"HTTP: {format % args}")

        try:
            servidor = http.server.HTTPServer((host, self.puerto), ManejadorHTTP)
            url = f"http://{host}:{self.puerto}"

            print(f"  🌐 Servidor iniciado: {url}")
            print(f"  🔒 Solo accesible desde este equipo")
            print(f"  ⌨️  Presiona Ctrl+C para detener\n")

            # Abrir navegador automáticamente
            if ABRIR_NAVEGADOR:
                threading.Timer(
                    1.0, lambda: webbrowser.open(url)
                ).start()

            servidor.serve_forever()

        except KeyboardInterrupt:
            print("\n\n  👋 STRATUM detenido. Hasta pronto.\n")
        except OSError as e:
            print(f"\n  ❌ Error al iniciar servidor en puerto {self.puerto}: {e}")
            print(f"     Intenta con otro puerto: python3 stratum.py --puerto 8080\n")

    # ─── Generar HTML básico de resultados ───────────────────────────────────
    def _generar_html_basico(self) -> str:
        """
        Genera la interfaz web básica con los resultados del escaneo.
        Esta función se expande en la Sesión 6 con la UI completa y D3.js.
        """
        stats = self.resultado_scan["estadisticas"] if self.resultado_scan else {}
        nombre = self.resultado_scan["nombre"] if self.resultado_scan else "Sin proyecto"
        arbol  = self.scanner.generar_arbol_visual() if self.resultado_scan else ""
        estado = self.licencias.obtener_estado()

        # Construir tabla de archivos por categoría
        rows_cat = ""
        for cat, n in sorted(stats.get("por_categoria", {}).items(), key=lambda x: -x[1]):
            rows_cat += f"<tr><td>{cat}</td><td>{n}</td></tr>"

        # Construir tabla de extensiones top
        top_ext = sorted(
            stats.get("por_extension", {}).items(), key=lambda x: -x[1]
        )[:10]
        rows_ext = ""
        for ext, n in top_ext:
            rows_ext += f"<tr><td>{ext}</td><td>{n}</td></tr>"

        return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>STRATUM — {nombre}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #0f1117; color: #e2e8f0; min-height: 100vh; }}
  .header {{ background: linear-gradient(135deg, #1a1f2e, #0d1117);
             border-bottom: 1px solid #2d3748; padding: 1.5rem 2rem; }}
  .header h1 {{ font-size: 1.5rem; font-weight: 600; color: #63b3ed; letter-spacing: -0.5px; }}
  .header p {{ font-size: 0.8rem; color: #718096; margin-top: 0.25rem; }}
  .badge {{ display: inline-block; padding: 0.25rem 0.75rem; border-radius: 999px;
            font-size: 0.7rem; font-weight: 600; letter-spacing: 0.05em;
            background: #1a365d; color: #63b3ed; margin-top: 0.5rem; }}
  .content {{ padding: 2rem; max-width: 1200px; margin: 0 auto; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
           gap: 1rem; margin-bottom: 2rem; }}
  .card {{ background: #1a1f2e; border: 1px solid #2d3748; border-radius: 12px;
           padding: 1.25rem; }}
  .card-label {{ font-size: 0.75rem; color: #718096; text-transform: uppercase;
                 letter-spacing: 0.08em; margin-bottom: 0.5rem; }}
  .card-value {{ font-size: 2rem; font-weight: 600; color: #63b3ed; }}
  .card-sub {{ font-size: 0.75rem; color: #4a5568; margin-top: 0.25rem; }}
  .section {{ background: #1a1f2e; border: 1px solid #2d3748; border-radius: 12px;
              padding: 1.5rem; margin-bottom: 1.5rem; }}
  .section h2 {{ font-size: 1rem; font-weight: 600; color: #e2e8f0; margin-bottom: 1rem;
                 padding-bottom: 0.5rem; border-bottom: 1px solid #2d3748; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
  td, th {{ padding: 0.5rem 0.75rem; text-align: left; border-bottom: 1px solid #2d3748; }}
  th {{ color: #718096; font-weight: 500; font-size: 0.75rem; text-transform: uppercase; }}
  td:last-child {{ text-align: right; color: #63b3ed; font-weight: 600; }}
  .detail-cell {{ font-size: 0.78rem; max-width: 320px; word-break: break-word;
                  white-space: normal; text-align: left !important;
                  color: #a0aec0; line-height: 1.45; }}
  .file-cell   {{ font-size: 0.8rem; max-width: 200px; word-break: break-all;
                  white-space: normal; text-align: left !important; color: #63b3ed; }}
  .tree {{ font-family: 'Fira Code', 'Cascadia Code', monospace; font-size: 0.8rem;
           color: #a0aec0; white-space: pre; overflow-x: auto; line-height: 1.6; }}
  .status-bar {{ background: #1a2744; border: 1px solid #2b4c8c; border-radius: 8px;
                 padding: 0.75rem 1rem; margin-bottom: 1.5rem; color: #63b3ed;
                 font-size: 0.85rem; }}
  .coming {{ background: #1a1f2e; border: 1px dashed #2d3748; border-radius: 12px;
             padding: 2rem; text-align: center; color: #4a5568; margin-bottom: 1rem; }}
  .coming span {{ font-size: 1.5rem; display: block; margin-bottom: 0.5rem; }}
  .footer {{ text-align: center; padding: 2rem; color: #2d3748; font-size: 0.75rem;
             border-top: 1px solid #1a1f2e; margin-top: 2rem; }}
</style>
<script type="module">
  async function initMermaid() {{
    try {{
      const m = await import(
        "https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs"
      );
      m.default.initialize({{
        startOnLoad: false, theme: "dark",
        themeVariables: {{
          background: "#1a202c", primaryColor: "#2b4c8c",
          primaryTextColor: "#e2e8f0", lineColor: "#4a5568"
        }}
      }});
      await m.default.run({{ querySelector: ".mermaid" }});
    }} catch(e) {{
      document.querySelectorAll(".mermaid").forEach(function(el) {{
        if (!el.querySelector("svg")) {{
          var code = el.textContent || "";
          el.innerHTML =
            "<pre style='color:#a0aec0;font-size:0.75rem;overflow-x:auto;" +
            "background:#0d1117;padding:1rem;border-radius:6px;" +
            "border:1px solid #2d3748;white-space:pre-wrap'>" +
            code.replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;") +
            "</pre><p style='color:#4a5568;font-size:0.72rem;margin:0.3rem 0 0'>" +
            "&#9888; Diagrama requiere internet para renderizar.</p>";
        }}
      }});
    }}
  }}
  if (document.readyState === "loading") {{
    document.addEventListener("DOMContentLoaded", initMermaid);
  }} else {{ initMermaid(); }}
</script>
</head>
<body>

<div class="header">
  <h1>⬡ STRATUM System Intelligence Engine</h1>
  <p>{STRATUM_AUTOR} · {STRATUM_MARCA}</p>
  <span class="badge">v{STRATUM_VERSION} · Sesión 5 · Módulos 0–21 activos</span>
</div>

<div class="content">

  <div class="status-bar">
    📡 {estado['mensaje']} &nbsp;|&nbsp; Tier: {estado['tier'].upper()} &nbsp;|&nbsp;
    Región: {estado['region']}
    {'&nbsp;|&nbsp; 🎉 Early Adopter — 50% desc. permanente' if estado['early_adopter'] else ''}
  </div>

  <div class="grid">
    <div class="card">
      <div class="card-label">Proyecto</div>
      <div class="card-value" style="font-size:1.2rem;padding-top:0.4rem">{nombre}</div>
    </div>
    <div class="card">
      <div class="card-label">Archivos analizados</div>
      <div class="card-value">{stats.get('total_archivos', 0)}</div>
    </div>
    <div class="card">
      <div class="card-label">Tamaño total</div>
      <div class="card-value">{stats.get('total_mb', 0)}</div>
      <div class="card-sub">Megabytes</div>
    </div>
    <div class="card">
      <div class="card-label">Carpetas</div>
      <div class="card-value">{stats.get('total_carpetas', 0)}</div>
    </div>
  </div>

  <div class="grid">
    <div class="section">
      <h2>📂 Archivos por categoría</h2>
      <table>
        <tr><th>Categoría</th><th>Cantidad</th></tr>
        {rows_cat}
      </table>
    </div>
    <div class="section">
      <h2>🔤 Top extensiones</h2>
      <table>
        <tr><th>Extensión</th><th>Archivos</th></tr>
        {rows_ext}
      </table>
    </div>
  </div>

  <div class="section">
    <h2>🌳 Estructura del proyecto</h2>
    <div class="tree">{arbol}</div>
  </div>

  {self._html_seccion_ast()}

  {self._html_seccion_heuristic()}

  {self._html_seccion_sql()}

  {self._html_seccion_linker()}

  {self._html_seccion_etl()}

  {self._html_seccion_cloud()}

  {self._html_seccion_dwh()}

  {self._html_seccion_profiler()}

  {self._html_seccion_framework()}

  {self._html_seccion_kpi()}

  {self._html_seccion_viz()}

  {self._html_seccion_query()}

  {self._html_seccion_flow()}

  {self._html_seccion_score()}

  {self._html_seccion_quickfix()}

  {self._html_seccion_excel()}

  {self._html_seccion_coverage()}

  {self._html_seccion_version()}

  {self._html_seccion_security()}

  {self._html_seccion_st_cache()}

  {self._html_seccion_deps()}

  {self._html_seccion_perf()}

  {self._html_seccion_docs()}

  {self._html_seccion_tests()}

  {self._html_seccion_capa3()}

  <div class="section">
    <h2>&#128679; M&oacute;dulos en construcci&oacute;n</h2>
    <div class="grid">
      <div class="coming"><span>&#128202;</span>Advanced Analytics<br><small>Sesi&oacute;n 8</small></div>
      <div class="coming"><span>&#127760;</span>Dashboard D3.js<br><small>Sesi&oacute;n 8</small></div>
      <div class="coming"><span>&#129302;</span>ML Pattern Detector<br><small>Sesi&oacute;n 8</small></div>
    </div>
  </div>

</div>

<div class="footer">
  STRATUM Engine v{STRATUM_VERSION} · {STRATUM_AUTOR}<br>
  {STRATUM_MARCA}<br>
  Este programa no envía datos a internet · Todo procesamiento es local
</div>

</body>
</html>"""

    # ─── Sección HTML: AST Engine Python ─────────────────────────────────────
    def _html_seccion_ast(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 5 (AST Python)."""
        if not self.resultado_ast:
            return ""
        r = self.resultado_ast["resumen"]
        if r["archivos_analizados"] == 0:
            return ""
        # Top 5 archivos por complejidad
        top = sorted(
            self.resultado_ast["archivos"].items(),
            key=lambda x: -x[1].get("complejidad_total", 0)
        )[:5]
        rows = "".join(
            f"<tr><td>{ruta}</td><td>{d.get('complejidad_total',0)}</td>"
            f"<td>{d.get('n_funciones',0)}+{d.get('n_metodos',0)}</td></tr>"
            for ruta, d in top
        )
        return f"""<div class="section">
      <h2>🧠 BLOQUE 5 — AST Engine Python</h2>
      <div class="grid">
        <div class="card"><div class="card-label">Archivos .py</div>
          <div class="card-value">{r['archivos_analizados']}</div></div>
        <div class="card"><div class="card-label">Clases</div>
          <div class="card-value">{r['total_clases']}</div></div>
        <div class="card"><div class="card-label">Funciones</div>
          <div class="card-value">{r['total_funciones']+r['total_metodos']}</div></div>
        <div class="card"><div class="card-label">Complejidad ciclomática total</div>
          <div class="card-value">{r['complejidad_total']}<span style="font-size:0.7rem;color:#a0aec0;display:block">prom {r.get('complejidad_promedio_fn','?')}/función · umbral CC>{r.get('umbral_cc_usado',10)}</span></div></div>
        <div class="card"><div class="card-label">Problemas detectados</div>
          <div class="card-value">{r['total_problemas']}</div></div>
      </div>
      <h2 style="margin-top:1rem">Top 5 archivos más complejos</h2>
      <table><tr><th>Archivo</th><th>Complejidad</th><th>Funciones</th></tr>{rows}</table>
      {self._html_histograma_cc(r)}
    </div>"""

    # ─── Sección HTML: Heuristic Engine multi-lenguaje ───────────────────────
    def _html_seccion_heuristic(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 6 (Heuristic)."""
        if not self.resultado_heur:
            return ""
        r = self.resultado_heur["resumen"]
        if r["archivos_analizados"] == 0:
            return ""
        rows = "".join(
            f"<tr><td>{lang}</td><td>{s['archivos']}</td><td>{s['clases']}</td>"
            f"<td>{s['funciones']+s['metodos']}</td><td>{s['imports']}</td></tr>"
            for lang, s in sorted(r["por_lenguaje"].items())
        )
        return f"""<div class="section">
      <h2>🔍 BLOQUE 6 — Heuristic Engine (JS/TS/Java/C#/PHP/R)</h2>
      <table>
        <tr><th>Lenguaje</th><th>Archivos</th><th>Clases</th><th>Funciones</th><th>Imports</th></tr>
        {rows}
      </table>
    </div>"""

    # ─── Sección HTML: SQL Parser ────────────────────────────────────────────
    def _html_seccion_sql(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 7 (SQL Parser)."""
        if not self.resultado_sql:
            return ""
        r   = self.resultado_sql["resumen"]
        cat = self.resultado_sql["catalogo"]
        sin_sql = r["total_tablas"] == 0 and r["total_vistas"] == 0
        nota_sin_sql = (
            '<p style="margin-top:0.75rem;color:#a0aec0;font-size:0.85rem">'
            '&#8505; No se encontraron archivos .sql en este proyecto. '
            'Si usa Supabase, SQLAlchemy, Django ORM o similar, las queries '
            'est&#225;n embebidas en c&#243;digo Python y son capturadas '
            'por el Code-DB Linker (Bloque 8).</p>'
        ) if sin_sql else ""
        tablas_html = "".join(
            f"<tr><td>{nombre}</td><td>{len(d.get('columnas',[]))}</td>"
            f"<td>{d.get('dialecto','?')}</td></tr>"
            for nombre, d in sorted(cat.get("tablas", {}).items())[:20]
        )
        tabla_sec = (
            '<h2 style="margin-top:1rem">Tablas detectadas (primeras 20)</h2>'
            '<table><tr><th>Tabla</th><th>Columnas</th><th>Dialecto</th></tr>'
            + (tablas_html or "<tr><td colspan=3>Sin tablas detectadas</td></tr>")
            + "</table>"
        ) if not sin_sql else ""
        return (
            '<div class="section">'
            '<h2>&#128451; BLOQUE 7 &mdash; SQL Parser multi-dialecto</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Tablas</div>'
            f'<div class="card-value">{r["total_tablas"]}</div></div>'
            '<div class="card"><div class="card-label">Vistas</div>'
            f'<div class="card-value">{r["total_vistas"]}</div></div>'
            '<div class="card"><div class="card-label">Funciones SQL</div>'
            f'<div class="card-value">{r["total_funciones"]}</div></div>'
            '<div class="card"><div class="card-label">Foreign Keys</div>'
            f'<div class="card-value">{r["total_fks"]}</div></div>'
            '<div class="card"><div class="card-label">Pol&#237;ticas RLS</div>'
            f'<div class="card-value">{r["total_policies_rls"]}</div></div>'
            "</div>"
            + nota_sin_sql
            + tabla_sec
            + "</div>"
        )

    # ─── Sección HTML: Code-DB Linker ────────────────────────────────────────
    def _html_seccion_linker(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 8 (Linker)."""
        if not self.resultado_link:
            return ""
        r = self.resultado_link["resumen"]
        huerfanas = self.resultado_link["tablas_huerfanas_db"][:10]
        usadas = sorted(
            self.resultado_link["tabla_a_archivos"].items(),
            key=lambda x: -len(x[1])
        )[:10]
        huerfanas_html = "".join(f"<tr><td>{t}</td></tr>" for t in huerfanas) or '<tr><td>Ninguna</td></tr>'
        usadas_html = "".join(
            f"<tr><td>{t}</td><td>{len(usos)}</td></tr>" for t, usos in usadas
        ) or '<tr><td>Sin referencias</td><td>0</td></tr>'
        return f"""<div class="section">
      <h2>🔗 BLOQUE 8 — Code-DB Linker</h2>
      <div class="grid">
        <div class="card"><div class="card-label">Tablas definidas</div>
          <div class="card-value">{r['tablas_definidas']}</div></div>
        <div class="card"><div class="card-label">Tablas usadas en código</div>
          <div class="card-value">{r['tablas_usadas']}</div></div>
        <div class="card"><div class="card-label">Huérfanas (sin uso)</div>
          <div class="card-value">{r['huerfanas_db']}</div></div>
        <div class="card"><div class="card-label">Sin definir (en código)</div>
          <div class="card-value">{r['huerfanas_codigo']}</div></div>
      </div>
      <div class="grid">
        <div class="section">
          <h2>Top 10 tablas más referenciadas</h2>
          <table><tr><th>Tabla</th><th>Archivos</th></tr>{usadas_html}</table>
        </div>
        <div class="section">
          <h2>Tablas huérfanas (definidas, no usadas)</h2>
          <table><tr><th>Tabla</th></tr>{huerfanas_html}</table>
        </div>
      </div>
    </div>"""

    # ─── Sección HTML: ETL/ELT Detector ──────────────────────────────────────
    def _html_seccion_etl(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 9 (ETL Detector)."""
        if not self.resultado_etl:
            return ""
        r = self.resultado_etl["resumen"]
        pipelines = self.resultado_etl.get("pipelines", [])[:15]
        if r["total_pipelines"] == 0:
            return ""
        rows_fw = "".join(
            f"<tr><td>{fw}</td><td>{cnt}</td></tr>"
            for fw, cnt in sorted(r["por_framework"].items(), key=lambda x: -x[1])
        )
        rows_pip = "".join(
            f"<tr><td>{p['nombre']}</td><td>{p['framework']}</td>"
            f"<td>{p['tipo']}</td><td>{p['n_tareas']}</td></tr>"
            for p in pipelines
        )
        return f"""<div class="section">
      <h2>🔄 BLOQUE 9 — ETL/ELT Detector</h2>
      <div class="grid">
        <div class="card"><div class="card-label">Archivos ETL</div>
          <div class="card-value">{r['archivos_con_etl']}</div></div>
        <div class="card"><div class="card-label">Pipelines detectados</div>
          <div class="card-value">{r['total_pipelines']}</div></div>
        <div class="card"><div class="card-label">Frameworks</div>
          <div class="card-value">{len(r['por_framework'])}</div></div>
      </div>
      <div class="grid">
        <div class="section">
          <h2>Por framework</h2>
          <table><tr><th>Framework</th><th>Pipelines</th></tr>{rows_fw}</table>
        </div>
        <div class="section">
          <h2>Pipelines detectados</h2>
          <table><tr><th>Nombre</th><th>Framework</th><th>Tipo</th><th>Tareas</th></tr>
          {rows_pip or '<tr><td colspan=4>Sin pipelines</td></tr>'}</table>
        </div>
      </div>
    </div>"""

    # ─── Sección HTML: Cloud Auditor ─────────────────────────────────────────
    def _html_seccion_cloud(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 10 (Cloud Auditor)."""
        if not self.resultado_cloud:
            return ""
        r = self.resultado_cloud["resumen"]
        alertas = self.resultado_cloud.get("alertas", [])[:20]
        if r["archivos_cloud"] == 0 and r["total_alertas"] == 0:
            return ""
        provs_html = " · ".join(
            f'<span style="color:#63b3ed">{p.upper()}</span>'
            for p in r["proveedores"]
        ) or "Sin proveedores detectados"
        rows_alert = "".join(
            f'<tr><td style="color:{"#fc8181" if a["nivel"]=="CRÍTICO" else "#f6ad55"}">'
            f'{a["nivel"]}</td><td>{a["tipo"]}</td>'
            f'<td>{a["archivo"]}</td><td>{a["linea"]}</td></tr>'
            for a in alertas
        )
        rows_tipo = "".join(
            f"<tr><td>{tipo}</td><td>{cnt}</td></tr>"
            for tipo, cnt in r["por_tipo_herramienta"].items()
        )
        return f"""<div class="section">
      <h2>☁️ BLOQUE 10 — Cloud Migration Auditor</h2>
      <div class="grid">
        <div class="card"><div class="card-label">Archivos cloud</div>
          <div class="card-value">{r['archivos_cloud']}</div></div>
        <div class="card"><div class="card-label">Recursos totales</div>
          <div class="card-value">{r['total_recursos']}</div></div>
        <div class="card"><div class="card-label">Alertas críticas</div>
          <div class="card-value" style="color:#fc8181">{r['alertas_criticas']}</div></div>
        <div class="card"><div class="card-label">Alertas altas</div>
          <div class="card-value" style="color:#f6ad55">{r['alertas_altas']}</div></div>
      </div>
      <p style="margin:0.75rem 0 0.5rem;color:#a0aec0;font-size:0.85rem">
        Proveedores: {provs_html}
      </p>
      <div class="grid">
        <div class="section">
          <h2>Por herramienta</h2>
          <table><tr><th>Tipo</th><th>Archivos</th></tr>
          {rows_tipo or '<tr><td colspan=2>Sin archivos cloud</td></tr>'}</table>
        </div>
        <div class="section">
          <h2>Alertas de seguridad (primeras 20)</h2>
          <table><tr><th>Nivel</th><th>Tipo</th><th>Archivo</th><th>Línea</th></tr>
          {rows_alert or '<tr><td colspan=4>✅ Sin credenciales hardcodeadas</td></tr>'}</table>
        </div>
      </div>
    </div>"""

    # ─── Sección HTML: DWH Auditor ────────────────────────────────────────────
    def _html_seccion_dwh(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 11 (DWH Auditor)."""
        if not self.resultado_dwh:
            return ""
        r = self.resultado_dwh["resumen"]
        if r["total_dims"] == 0 and r["total_facts"] == 0:
            return ""
        dims  = self.resultado_dwh.get("dimensiones", {})
        facts = self.resultado_dwh.get("hechos", {})
        probs = self.resultado_dwh.get("problemas", [])[:10]
        rows_dims = "".join(
            f"<tr><td>{n}</td><td>{d.get('n_columnas',0)}</td>"
            f"<td>SCD {d.get('scd_tipo','?')}</td>"
            f"<td>{'✅' if d.get('tiene_sk') else '⚠️'}</td></tr>"
            for n, d in sorted(dims.items())[:15]
        )
        rows_facts = "".join(
            f"<tr><td>{n}</td><td>{d.get('n_metricas',0)}</td>"
            f"<td>{d.get('n_fks',0)}</td>"
            f"<td>{'✅' if d.get('tiene_fecha') else '⚠️'}</td></tr>"
            for n, d in sorted(facts.items())[:15]
        )
        rows_probs = "".join(
            f"<tr><td>{p['tipo']}</td><td>{p.get('tabla','')}</td>"
            f"<td style='font-size:0.8rem'>{p['detalle']}</td></tr>"
            for p in probs
        )
        score_color    = "#68d391" if r["score_calidad"] >= 70 else \
                         "#f6ad55" if r["score_calidad"] >= 40 else "#fc8181"
        schema_display = r["tipo_schema"].replace("_", " ").title()
        staging_total  = r["total_staging"] + r["total_marts"]
        sec_probs      = (
            '<div class="section"><h2>Problemas detectados</h2>'
            '<table><tr><th>Tipo</th><th>Tabla</th><th>Detalle</th></tr>'
            + rows_probs + "</table></div>"
        ) if probs else ""
        return f"""<div class="section">
      <h2>🏗️ BLOQUE 11 — Data Warehouse Auditor</h2>
      <div class="grid">
        <div class="card"><div class="card-label">Dimensiones</div>
          <div class="card-value">{r['total_dims']}</div></div>
        <div class="card"><div class="card-label">Hechos (facts)</div>
          <div class="card-value">{r['total_facts']}</div></div>
        <div class="card"><div class="card-label">Staging / Marts</div>
          <div class="card-value">{staging_total}</div></div>
        <div class="card"><div class="card-label">Schema detectado</div>
          <div class="card-value" style="font-size:0.9rem;padding-top:0.4rem">
            {schema_display}</div></div>
        <div class="card"><div class="card-label">Score calidad</div>
          <div class="card-value" style="color:{score_color}">{r['score_calidad']}/100</div></div>
      </div>
      <div class="grid">
        <div class="section">
          <h2>Dimensiones</h2>
          <table><tr><th>Nombre</th><th>Cols</th><th>SCD</th><th>SK</th></tr>
          {rows_dims or '<tr><td colspan=4>Sin dimensiones</td></tr>'}</table>
        </div>
        <div class="section">
          <h2>Hechos</h2>
          <table><tr><th>Nombre</th><th>Métricas</th><th>FKs</th><th>Fecha</th></tr>
          {rows_facts or '<tr><td colspan=4>Sin tablas de hechos</td></tr>'}</table>
        </div>
      </div>
      {sec_probs}
    </div>"""

    # ─── Sección HTML: Data Profiler ──────────────────────────────────────────
    def _html_seccion_profiler(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 12 (Data Profiler)."""
        if not self.resultado_profiler:
            return ""
        r = self.resultado_profiler["resumen"]
        if r["archivos_perfilados"] == 0:
            return ""
        archivos = self.resultado_profiler.get("archivos", {})
        # Construir tabla de archivos + columnas con PII
        rows_arch = ""
        all_pii   = []
        for ruta_rel, datos in sorted(archivos.items()):
            if not datos.get("ok"):
                continue
            pii_count = datos.get("n_alertas_pii", 0)
            rows_arch += (
                f"<tr><td>{ruta_rel}</td>"
                f"<td>{datos.get('formato','?').upper()}</td>"
                f"<td>{datos.get('filas',0):,}</td>"
                f"<td>{datos.get('n_columnas',0)}</td>"
                f"<td>{'⚠️ ' + str(pii_count) if pii_count else '✅'}</td></tr>"
            )
            for col in datos.get("alertas_pii", []):
                all_pii.append({**col, "archivo": ruta_rel})
        rows_pii = "".join(
            f"<tr><td>{p.get('archivo','')}</td><td>{p.get('nombre','')}</td>"
            f"<td>{p.get('pii_tipo','')}</td>"
            f"<td>{p.get('tipo_sql_sugerido','')}</td></tr>"
            for p in all_pii[:20]
        )
        pii_color = "#fc8181" if r["alertas_pii"] > 0 else "#68d391"
        sec_pii   = (
            '<h2 style="margin-top:1rem;color:#fc8181">&#9888; Columnas con PII detectado</h2>'
            "<table><tr><th>Archivo</th><th>Columna</th>"
            "<th>Tipo PII</th><th>SQL sugerido</th></tr>"
            + rows_pii + "</table>"
        ) if all_pii else ""
        total_filas_fmt = str(r["total_filas"])
        return (
            '<div class="section">'
            '<h2>&#128202; BLOQUE 12 &mdash; Data Profiler</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Archivos perfilados</div>'
            f'<div class="card-value">{r["archivos_perfilados"]}</div></div>'
            '<div class="card"><div class="card-label">Total filas</div>'
            f'<div class="card-value">{total_filas_fmt}</div></div>'
            '<div class="card"><div class="card-label">Alertas PII</div>'
            f'<div class="card-value" style="color:{pii_color}">{r["alertas_pii"]}</div></div>'
            "</div>"
            '<h2 style="margin-top:1rem">Archivos analizados</h2>'
            "<table><tr><th>Archivo</th><th>Formato</th><th>Filas</th>"
            "<th>Columnas</th><th>PII</th></tr>"
            + (rows_arch or '<tr><td colspan=5>Sin archivos de datos</td></tr>')
            + "</table>"
            + sec_pii
            + "</div>"
        )


    def _html_seccion_framework(self) -> str:
            """Tarjeta del dashboard con resultado del Framework Primary Detector."""
            if not self.resultado_framework:
                return ""
            r   = self.resultado_framework
            fw  = r.get("framework", "generic")
            lng = r.get("lenguaje_dominante", "desconocido")
            nota = r.get("nota", "")
            ver_items = r.get("version_info", {})
            vers = ", ".join(f"{k}: {v}" for k, v in ver_items.items()) or "No detectada"
            umbrales = r.get("umbrales", {})
            fn_crit  = umbrales.get("lineas_funcion_critico", "—")
            fn_alto  = umbrales.get("lineas_funcion_alto", "—")
            clases_ok = "✅ Correcto" if umbrales.get("clases_cero_es_ok") else (
                        "⚠️ Revisar"  if umbrales.get("clases_cero_es_ok") is False else "—")
            return (
                '<div class="section">'
                '<h2>&#128269; Framework Primario Detectado</h2>'
                '<div class="grid">'
                '<div class="card"><div class="card-label">Framework</div>'
                f'<div class="card-value">{fw.upper()}</div></div>'
                '<div class="card"><div class="card-label">Lenguaje dominante</div>'
                f'<div class="card-value">{lng}</div></div>'
                '<div class="card"><div class="card-label">Versión detectada</div>'
                f'<div class="card-value" style="font-size:0.85rem">{vers}</div></div>'
                '</div>'
                '<table style="margin-top:0.75rem">'
                '<tr><th>Umbral función crítica</th><th>Umbral función alta</th>'
                '<th>0 clases</th></tr>'
                f'<tr><td>&gt;{fn_crit} líneas = CRÍTICO</td>'
                f'<td>&gt;{fn_alto} líneas = ALTO</td>'
                f'<td>{clases_ok}</td></tr>'
                '</table>'
                f'<p style="margin-top:0.5rem;color:#a0aec0;font-size:0.85rem">{nota}</p>'
                '</div>'
            )

    def _html_seccion_kpi(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 13 (KPI Engine)."""
        if not self.resultado_kpi:
            return ""
        r = self.resultado_kpi["resumen"]
        if r["n_archivos_con_kpi"] == 0:
            return ""
        cats = r.get("categorias", {})
        rows_cats = "".join(
            f"<tr><td>{cat.capitalize()}</td>"
            f"<td>{', '.join(kws[:5])}</td></tr>"
            for cat, kws in sorted(cats.items())
        )
        top_archs = r.get("top_archivos_kpi", [])
        rows_top = "".join(
            f"<tr><td>{a}</td><td>{score}</td></tr>"
            for a, score in top_archs[:5]
        )
        return (
            '<div class="section">'
            '<h2>&#128200; BLOQUE 13 &mdash; KPI Engine</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Archivos con KPIs</div>'
            f'<div class="card-value">{r["n_archivos_con_kpi"]}</div></div>'
            '<div class="card"><div class="card-label">Métricas detectadas</div>'
            f'<div class="card-value">{r["total_kpis_detectados"]}</div></div>'
            '<div class="card"><div class="card-label">Agregaciones</div>'
            f'<div class="card-value">{r["total_agregaciones"]}</div></div>'
            '<div class="card"><div class="card-label">st.metric / Indicadores</div>'
            f'<div class="card-value">{r["total_st_metric"]}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Categorías de KPI detectadas</h2>'
            '<table><tr><th>Categoría</th><th>Palabras clave encontradas</th></tr>'
            + (rows_cats or '<tr><td colspan=2>Sin categorías</td></tr>') +
            '</table>'
            '<h2 style="margin-top:1rem">Top archivos por densidad de KPIs</h2>'
            '<table><tr><th>Archivo</th><th>Score densidad</th></tr>'
            + (rows_top or '<tr><td colspan=2>Sin datos</td></tr>') +
            '</table></div>'
        )

    def _html_seccion_viz(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 14 (Viz Mapper)."""
        if not self.resultado_viz:
            return ""
        r = self.resultado_viz["resumen"]
        if r["total_visualizaciones"] == 0:
            return ""
        por_lib = r.get("por_libreria", {})
        rows_lib = "".join(
            f"<tr><td>{lib}</td><td>{info['count']}</td>"
            f"<td>{info['tipo']}</td></tr>"
            for lib, info in sorted(
                por_lib.items(), key=lambda x: x[1]["count"], reverse=True
            )
        )
        top_archs = r.get("top_archivos", [])[:5]
        rows_arch = "".join(
            f"<tr><td>{a}</td><td>{d.get('libreria_dominante','?')}</td>"
            f"<td>{d.get('n_total',0)}</td></tr>"
            for a, d in top_archs
        )
        return (
            '<div class="section">'
            '<h2>&#128506; BLOQUE 14 &mdash; Viz Mapper</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Total visualizaciones</div>'
            f'<div class="card-value">{r["total_visualizaciones"]}</div></div>'
            '<div class="card"><div class="card-label">Archivos con viz</div>'
            f'<div class="card-value">{r["n_archivos_con_viz"]}</div></div>'
            '<div class="card"><div class="card-label">Librerías usadas</div>'
            f'<div class="card-value" style="font-size:0.85rem">'
            f'{", ".join(r["librerias_usadas"][:4]) or "ninguna"}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Por librería de visualización</h2>'
            '<table><tr><th>Librería</th><th>Usos</th><th>Tipo</th></tr>'
            + (rows_lib or '<tr><td colspan=3>Sin visualizaciones</td></tr>') +
            '</table>'
            '<h2 style="margin-top:1rem">Top archivos</h2>'
            '<table><tr><th>Archivo</th><th>Librería dominante</th><th>Total</th></tr>'
            + (rows_arch or '<tr><td colspan=3>Sin datos</td></tr>') +
            '</table></div>'
        )

    def _html_seccion_query(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 15 (Query Intel)."""
        if not self.resultado_query:
            return ""
        r = self.resultado_query["resumen"]
        if r["total_hallazgos"] == 0:
            return ""
        por_sev  = r.get("por_severidad", {})
        criticos = por_sev.get("CRITICO", 0)
        altos    = por_sev.get("ALTO",    0)
        medios   = por_sev.get("MEDIO",   0)
        color_total = "#fc8181" if criticos > 0 else (
                      "#f6ad55" if altos   > 0 else "#68d391")
        hallazgos = self.resultado_query.get("hallazgos", [])
        rows_h = "".join(
            f'<tr><td style="color:{"#fc8181" if h["severidad"]=="CRITICO" else "#f6ad55" if h["severidad"]=="ALTO" else "#fefcbf"}">'
            f'{h["severidad"]}</td>'
            f'<td>{h["tipo"]}</td>'
            f'<td>{h["archivo"]}</td>'
            f'<td>{h.get("linea","")}</td>'
            f'<td style="font-size:0.78rem">{h["mensaje"][:80]}</td></tr>'
            for h in hallazgos[:15]
        )
        return (
            '<div class="section">'
            '<h2>&#128270; BLOQUE 15 &mdash; Query Intel</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Total hallazgos</div>'
            f'<div class="card-value" style="color:{color_total}">'
            f'{r["total_hallazgos"]}</div></div>'
            '<div class="card"><div class="card-label">Críticos (N+1)</div>'
            f'<div class="card-value" style="color:#fc8181">{criticos}</div></div>'
            '<div class="card"><div class="card-label">Altos</div>'
            f'<div class="card-value" style="color:#f6ad55">{altos}</div></div>'
            '<div class="card"><div class="card-label">Medios</div>'
            f'<div class="card-value" style="color:#fefcbf">{medios}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Detalle de hallazgos</h2>'
            '<table><tr><th>Severidad</th><th>Tipo</th><th>Archivo</th>'
            '<th>Línea</th><th>Mensaje</th></tr>'
            + (rows_h or '<tr><td colspan=5>Sin hallazgos</td></tr>') +
            '</table></div>'
        )

    def _html_seccion_flow(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 16 (Flow Analyzer)."""
        if not self.resultado_flow:
            return ""
        r = self.resultado_flow["resumen"]
        if r["total_valores_estado"] == 0:
            return ""
        flujos   = r.get("flujos", [])
        gaps     = self.resultado_flow.get("gaps", [])
        bypasses = self.resultado_flow.get("bypasses", [])
        mermaid  = self.resultado_flow.get("mermaid", "")

        color_gaps = "#fc8181" if gaps     else "#68d391"
        color_byp  = "#f6ad55" if bypasses else "#68d391"

        rows_flujos = "".join(
            f"<tr><td>{f['nombre']}</td><td>{f['n_pasos']}</td>"
            f"<td>{f['tipo']}</td>"
            f"<td style='font-size:0.78rem'>{', '.join(f['modulos'][:3])}</td></tr>"
            for f in flujos[:8]
        )
        rows_gaps = "".join(
            f'<tr><td style="color:#fc8181">{g["valor_estado"]}</td>'
            f'<td>{", ".join(w["modulo"] for w in g["escrito_por"][:2])}</td>'
            f'<td style="font-size:0.78rem">{g["mensaje"][:80]}</td></tr>'
            for g in gaps[:10]
        )

        mermaid_block = ""
        if mermaid:
            mermaid_block = (
                '<h2 style="margin-top:1rem">Grafo de flujo</h2>'
                '<div class="mermaid" style="background:#1a202c;padding:1rem;'
                'border-radius:8px;overflow-x:auto">'
                + mermaid +
                '</div>'
            )

        return (
            '<div class="section">'
            '<h2>&#127754; BLOQUE 16 &mdash; Flow Analyzer</h2>'
            '<div class="grid">'
            '<div class="card"><div class="card-label">Valores de estado</div>'
            f'<div class="card-value">{r["total_valores_estado"]}</div></div>'
            '<div class="card"><div class="card-label">Flujos detectados</div>'
            f'<div class="card-value">{r["n_flujos_detectados"]}</div></div>'
            '<div class="card"><div class="card-label">Gaps (flujos rotos)</div>'
            f'<div class="card-value" style="color:{color_gaps}">'
            f'{r["n_gaps"]}</div></div>'
            '<div class="card"><div class="card-label">Bypasses estructurales</div>'
            f'<div class="card-value" style="color:{color_byp}">'
            f'{r["n_bypasses"]}</div></div>'
            '</div>'
            + (
                '<h2 style="margin-top:1rem">Flujos de negocio reconstruidos</h2>'
                '<table><tr><th>Flujo</th><th>Pasos</th><th>Tipo</th>'
                '<th>Módulos</th></tr>'
                + (rows_flujos or '<tr><td colspan=4>Sin flujos detectados</td></tr>')
                + '</table>'
                if flujos else ''
            )
            + (
                '<h2 style="margin-top:1rem;color:#fc8181">&#9888; Gaps detectados</h2>'
                '<table><tr><th>Estado huérfano</th><th>Escrito por</th>'
                '<th>Diagnóstico</th></tr>'
                + rows_gaps +
                '</table>'
                if gaps else ''
            )
            + mermaid_block
            + '</div>'
        )

# ============================================================================
# PUNTO DE ENTRADA PRINCIPAL
# ============================================================================


    # ─── Sección HTML: STRATUM Score ──────────────────────────────────────────
    def _html_seccion_score(self) -> str:
        """Tarjeta principal del dashboard con STRATUM Score ponderado."""
        if not self.resultado_score:
            return ""
        r = self.resultado_score
        score        = r["score"]
        badge        = r["badge"]
        badge_color  = r["badge_color"]
        badge_icon   = r["badge_icon"]
        por_sev      = r["por_severidad"]
        top10        = r["top10"]

        color_score = badge_color
        rows_top = "".join(
            f'<tr><td style="color:{"#fc8181" if f["severidad"]=="CRITICO" else "#f6ad55" if f["severidad"]=="ALTO" else "#fefcbf"}">'
            f'{f["severidad"]}</td>'
            f'<td>{f["tipo"]}</td>'
            f'<td>{f["origen"]}</td>'
            f'<td class="file-cell" title="{f.get("archivo","")}">{f.get("archivo","")}</td>'
            f'<td class="detail-cell" title="{f.get("detalle","")}">{f.get("detalle","")}</td></tr>'
            for f in top10
        )
        score_est   = r.get("score_estandar", score)
        fn_cal      = r.get("fn_calibradas", 0)
        delta       = r.get("delta_score", 0)
        hay_delta   = delta > 0 and fn_cal > 0

        comparativo_html = ""
        if hay_delta:
            comparativo_html = (
                f'<div style="margin-top:0.75rem;padding:0.75rem;'
                f'border-radius:6px;background:#1a202c;border:1px solid #4a5568">'
                f'<div style="display:flex;gap:2rem;align-items:center;flex-wrap:wrap">'
                f'<div><span style="color:#a0aec0;font-size:0.78rem">'
                f'Score estándar (McCabe=10 universal)</span><br>'
                f'<span style="color:#f6ad55;font-size:1.4rem;font-weight:bold">'
                f'{score_est}/100</span></div>'
                f'<div style="color:#68d391;font-size:1.8rem">→</div>'
                f'<div><span style="color:#a0aec0;font-size:0.78rem">'
                f'Score calibrado (framework-aware)</span><br>'
                f'<span style="color:{color_score};font-size:1.4rem;font-weight:bold">'
                f'{score}/100</span></div>'
                f'<div style="color:#68d391;font-size:1.4rem;font-weight:bold">'
                f'+{delta} pts</div>'
                f'</div>'
                f'<p style="color:#68d391;font-size:0.8rem;margin:0.4rem 0 0">'
                f'&#128202; {fn_cal} funciones UI (vista_*/modulo_*/render*) con CC 11–20 '
                f'no se cuentan como errores — calibrado con datos reales de QS Ingeniería. '
                f'Una herramienta genérica las reportaría como {fn_cal} falsos positivos.</p>'
                f'</div>'
            )

        return (
            '<div class="section">'
            '<h2>&#11088; BLOQUE 18 &mdash; STRATUM Score</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Score calibrado</div>'
            f'<div class="card-value" style="color:{color_score};font-size:2rem">'
            f'{score}/100</div></div>'
            f'<div class="card"><div class="card-label">Estado</div>'
            f'<div class="card-value" style="color:{badge_color}">'
            f'{badge_icon} {badge}</div></div>'
            f'<div class="card"><div class="card-label">Críticos</div>'
            f'<div class="card-value" style="color:#fc8181">'
            f'{por_sev.get("CRITICO",0)}</div></div>'
            f'<div class="card"><div class="card-label">Altos</div>'
            f'<div class="card-value" style="color:#f6ad55">'
            f'{por_sev.get("ALTO",0)}</div></div>'
            f'<div class="card"><div class="card-label">Medios</div>'
            f'<div class="card-value" style="color:#fefcbf">'
            f'{por_sev.get("MEDIO",0)}</div></div>'
            f'<div class="card"><div class="card-label">Bajos</div>'
            f'<div class="card-value" style="color:#a0aec0">'
            f'{por_sev.get("BAJO",0)}</div></div>'
            '</div>'
            + comparativo_html +
            '<h2 style="margin-top:1rem">Top 10 hallazgos prioritarios</h2>'
            '<table><tr><th>Severidad</th><th>Tipo</th><th>Origen</th>'
            '<th>Archivo</th><th>Detalle</th></tr>'
            + (rows_top or '<tr><td colspan=5>Sin hallazgos — proyecto en excelente estado</td></tr>')
            + '</table></div>'
        )

    # ─── Sección HTML: Quick Fix Engine ───────────────────────────────────────
    def _html_seccion_quickfix(self) -> str:
        """Tarjeta del dashboard con sugerencias de Quick Fix."""
        if not self.resultado_quickfix:
            return ""
        r = self.resultado_quickfix
        if r["total_sugerencias"] == 0:
            return ""
        esfuerzo = r["resumen_esfuerzo"]
        quick_wins = r["quick_wins"][:5]
        todas = r["sugerencias"][:15]

        rows_qw = "".join(
            f'<tr><td style="color:#68d391">⚡ MINUTOS</td>'
            f'<td style="color:{"#fc8181" if s["severidad"]=="CRITICO" else "#f6ad55"}">'
            f'{s["severidad"]}</td>'
            f'<td>{s["tipo"]}</td>'
            f'<td class="file-cell" title="{s.get("archivo","")}">{s.get("archivo","")}</td>'
            f'<td class="detail-cell" title="{s["accion"]}">{s["accion"]}</td></tr>'
            for s in quick_wins
        )
        rows_todas = "".join(
            f'<tr><td style="color:{"#68d391" if s["esfuerzo"]=="MINUTOS" else "#f6ad55" if s["esfuerzo"]=="HORAS" else "#fc8181"}">'
            f'{s["esfuerzo"]}</td>'
            f'<td style="color:{"#fc8181" if s["severidad"]=="CRITICO" else "#f6ad55" if s["severidad"]=="ALTO" else "#fefcbf"}">'
            f'{s["severidad"]}</td>'
            f'<td>{s["tipo"]}</td>'
            f'<td class="file-cell" title="{s.get("archivo","")}">{s.get("archivo","")}</td>'
            f'<td class="detail-cell" title="{s["accion"]}">{s["accion"]}</td></tr>'
            for s in todas
        )
        return (
            '<div class="section">'
            '<h2>&#128295; BLOQUE 19 &mdash; Quick Fix Engine</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Total sugerencias</div>'
            f'<div class="card-value">{r["total_sugerencias"]}</div></div>'
            f'<div class="card"><div class="card-label">⚡ Quick Wins (minutos)</div>'
            f'<div class="card-value" style="color:#68d391">{esfuerzo.get("MINUTOS",0)}</div></div>'
            f'<div class="card"><div class="card-label">🔧 Horas de trabajo</div>'
            f'<div class="card-value" style="color:#f6ad55">{esfuerzo.get("HORAS",0)}</div></div>'
            f'<div class="card"><div class="card-label">📅 Días de trabajo</div>'
            f'<div class="card-value" style="color:#fc8181">{esfuerzo.get("DIAS",0)}</div></div>'
            '</div>'
            + (
                '<h2 style="margin-top:1rem;color:#68d391">⚡ Quick Wins — Acción inmediata</h2>'
                '<table><tr><th>Esfuerzo</th><th>Severidad</th><th>Tipo</th>'
                '<th>Archivo</th><th>Acción</th></tr>'
                + rows_qw +
                '</table>'
                if quick_wins else ''
            )
            + '<h2 style="margin-top:1rem">Todas las sugerencias (top 15)</h2>'
            '<table><tr><th>Esfuerzo</th><th>Severidad</th><th>Tipo</th>'
            '<th>Archivo</th><th>Acción</th></tr>'
            + (rows_todas or '<tr><td colspan=5>Sin sugerencias</td></tr>')
            + '</table></div>'
        )

    # ─── Sección HTML: Excel Intelligence ────────────────────────────────────
    def _html_seccion_excel(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 17 (Excel Intelligence)."""
        if not self.resultado_excel:
            return ""
        r = self.resultado_excel["resumen"]
        if r["total_archivos"] == 0:
            return ""
        archivos = self.resultado_excel.get("archivos", {})
        rows = "".join(
            f"<tr><td>{ruta}</td>"
            f"<td>{d.get('tipo','?')}</td>"
            f"<td>{d.get('tamanio_kb',0):.0f} KB</td>"
            f"<td>{d.get('n_hojas','') or d.get('n_columnas','')}</td>"
            f"<td>{len(d.get('problemas',[]))}</td>"
            f"<td style='font-size:0.78rem;color:#f6ad55'>"
            f"{d['problemas'][0]['detalle'][:60] if d.get('problemas') else '✅'}</td></tr>"
            for ruta, d in list(archivos.items())[:10]
        )
        return (
            '<div class="section">'
            '<h2>&#128217; BLOQUE 17 &mdash; Excel Intelligence</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Total archivos</div>'
            f'<div class="card-value">{r["total_archivos"]}</div></div>'
            f'<div class="card"><div class="card-label">XLSX / XLSM</div>'
            f'<div class="card-value">{r["n_xlsx"]}</div></div>'
            f'<div class="card"><div class="card-label">CSV / TSV</div>'
            f'<div class="card-value">{r["n_csv"]}</div></div>'
            f'<div class="card"><div class="card-label">Fórmulas detectadas</div>'
            f'<div class="card-value">{r["total_formulas"]}</div></div>'
            f'<div class="card"><div class="card-label">Fórmulas volátiles</div>'
            f'<div class="card-value" style="color:{"#f6ad55" if r["formulas_volatiles"]>0 else "#68d391"}">'
            f'{r["formulas_volatiles"]}</div></div>'
            f'<div class="card"><div class="card-label">Macros VBA</div>'
            f'<div class="card-value" style="color:{"#f6ad55" if r["n_macros_vba"]>0 else "#68d391"}">'
            f'{r["n_macros_vba"]}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Detalle de archivos</h2>'
            '<table><tr><th>Archivo</th><th>Tipo</th><th>Tamaño</th>'
            '<th>Hojas/Cols</th><th>Problemas</th><th>Primer hallazgo</th></tr>'
            + (rows or '<tr><td colspan=6>Sin archivos Excel/CSV</td></tr>')
            + '</table>'
            + ('' if r["usa_openpyxl"] else
               '<p style="color:#a0aec0;font-size:0.85rem;margin-top:0.5rem">'
               '⚠️ openpyxl no instalado — análisis básico. '
               'Instala con: <code>pip install openpyxl</code></p>')
            + '</div>'
        )

    # ─── Sección HTML: Pattern Intelligence (Coverage + Version) ─────────────
    def _html_seccion_coverage(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 20 (Coverage Score)."""
        if not self.resultado_coverage:
            return ""
        r = self.resultado_coverage
        if r["total_archivos_evaluados"] == 0:
            return ""
        adv = r["advertencias"][:8]
        color_ratio = "#68d391" if r["cobertura_ok"] else "#f6ad55"
        rows_adv = "".join(
            f'<tr><td>{a["archivo"]}</td>'
            f'<td>{a["lenguaje"]}</td>'
            f'<td style="color:#f6ad55">{a["ratio"]}%</td>'
            f'<td style="font-size:0.78rem">{a["mensaje"][:80]}</td></tr>'
            for a in adv
        )
        return (
            '<div class="section">'
            '<h2>&#128270; BLOQUE 20 &mdash; Pattern Intelligence &mdash; Coverage Score</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Cobertura global</div>'
            f'<div class="card-value" style="color:{color_ratio}">'
            f'{r["ratio_global"]}%</div></div>'
            f'<div class="card"><div class="card-label">Archivos evaluados</div>'
            f'<div class="card-value">{r["total_archivos_evaluados"]}</div></div>'
            f'<div class="card"><div class="card-label">Advertencias</div>'
            f'<div class="card-value" style="color:{"#f6ad55" if adv else "#68d391"}">'
            f'{r["n_advertencias"]}</div></div>'
            '</div>'
            + (
                '<h2 style="margin-top:1rem;color:#f6ad55">⚠️ Archivos con cobertura baja</h2>'
                '<table><tr><th>Archivo</th><th>Lenguaje</th>'
                '<th>Cobertura</th><th>Diagnóstico</th></tr>'
                + rows_adv + '</table>'
                if adv else
                '<p style="color:#68d391;margin-top:0.5rem">✅ Cobertura de análisis suficiente en todos los archivos</p>'
            )
            + '</div>'
        )

    def _html_seccion_version(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 21 (Language Version)."""
        if not self.resultado_version:
            return ""
        r = self.resultado_version
        if r["n_lenguajes_detectados"] == 0:
            return ""
        versiones = r["versiones"]
        advertencias = r["advertencias"]
        rows_ver = "".join(
            f"<tr><td>{lang.capitalize()}</td>"
            f"<td>{ver}</td>"
            f"<td style='color:{'#fc8181' if any(a['lenguaje']==lang for a in advertencias) else '#68d391'}'>"
            f"{'⚠️ Sin patrones específicos' if any(a['lenguaje']==lang for a in advertencias) else '✅ Soportado'}</td></tr>"
            for lang, ver in versiones.items()
        )
        rows_adv = "".join(
            f'<tr><td>{a["lenguaje"].capitalize()}</td>'
            f'<td>{a["version"]}</td>'
            f'<td style="font-size:0.78rem;color:#f6ad55">{a["mensaje"][:100]}</td></tr>'
            for a in advertencias
        )
        return (
            '<div class="section">'
            '<h2>&#128195; BLOQUE 21 &mdash; Pattern Intelligence &mdash; Language Versions</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Lenguajes detectados</div>'
            f'<div class="card-value">{r["n_lenguajes_detectados"]}</div></div>'
            f'<div class="card"><div class="card-label">Manifiestos leídos</div>'
            f'<div class="card-value">{len(r["manifiestos_encontrados"])}</div></div>'
            f'<div class="card"><div class="card-label">Advertencias</div>'
            f'<div class="card-value" style="color:{"#f6ad55" if advertencias else "#68d391"}">'
            f'{r["n_advertencias"]}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Versiones detectadas</h2>'
            '<table><tr><th>Lenguaje</th><th>Versión</th><th>Soporte STRATUM</th></tr>'
            + (rows_ver or '<tr><td colspan=3>Sin manifiestos encontrados</td></tr>')
            + '</table>'
            + (
                '<h2 style="margin-top:1rem;color:#f6ad55">⚠️ Versiones sin patrones específicos</h2>'
                '<table><tr><th>Lenguaje</th><th>Versión</th><th>Mensaje</th></tr>'
                + rows_adv + '</table>'
                if advertencias else ''
            )
            + '</div>'
        )


    # ─── Sección HTML: Security Auditor ──────────────────────────────────────
    def _html_seccion_security(self) -> str:
        """Tarjeta del dashboard con resultados del Bloque 23 (Security Auditor)."""
        if not self.resultado_security:
            return ""
        r        = self.resultado_security["resumen"]
        hallazgos = self.resultado_security.get("hallazgos", [])
        if r["total_hallazgos"] == 0:
            return (
                '<div class="section">'
                '<h2>&#128274; BLOQUE 23 &mdash; Security Auditor</h2>'
                '<div style="color:#68d391;padding:1rem">&#10003; Sin hallazgos de seguridad '
                'detectados en el codebase analizado.</div>'
                '</div>'
            )

        nivel         = r["nivel_riesgo"]
        color_nivel   = {"CRITICO": "#e53e3e", "ALTO": "#f6ad55",
                         "MEDIO": "#fefcbf", "BAJO": "#68d391"}.get(nivel, "#a0aec0")
        por_sev       = r["por_severidad"]
        por_cat       = r["por_categoria"]

        # Top 15 hallazgos, CRITICO primero
        orden = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}
        top   = sorted(hallazgos, key=lambda h: orden.get(h["severidad"], 3))[:15]

        rows_hall = "".join(
            f'<tr>'
            f'<td style="color:{"#fc8181" if h["severidad"]=="CRITICO" else "#f6ad55" if h["severidad"]=="ALTO" else "#fefcbf"}">'
            f'{h["severidad"]}</td>'
            f'<td style="color:#a0aec0;font-size:0.78rem">{h["categoria"]}</td>'
            f'<td class="file-cell" title="{h["archivo"]}">{h["archivo"]}</td>'
            f'<td style="color:#a0aec0">{h["linea"]}</td>'
            f'<td class="detail-cell" title="{h["detalle"]}">{h["detalle"]}</td>'
            f'</tr>'
            for h in top
        )

        # Distribución por categoría
        rows_cat = "".join(
            f'<tr><td>{cat}</td>'
            f'<td style="color:#63b3ed;font-weight:600">{n}</td></tr>'
            for cat, n in sorted(por_cat.items(), key=lambda x: -x[1])
        )

        alertas = []
        if r["tiene_secretos"]:
            alertas.append("🔑 Secretos / credenciales hardcodeadas detectadas — ROTACIÓN INMEDIATA recomendada")
        if r["tiene_sql_injection"]:
            alertas.append("💉 Patrones de inyección SQL detectados — revisar y parametrizar queries")
        alertas_html = "".join(
            f'<div style="background:#2d1515;border-left:3px solid #e53e3e;'
            f'padding:0.5rem 0.75rem;margin:0.25rem 0;font-size:0.83rem;color:#fc8181">'
            f'{a}</div>' for a in alertas
        ) if alertas else ""

        return (
            '<div class="section">'
            '<h2>&#128274; BLOQUE 23 &mdash; Security Auditor</h2>'
            + alertas_html +
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Nivel de riesgo</div>'
            f'<div class="card-value" style="color:{color_nivel}">{nivel}</div></div>'
            f'<div class="card"><div class="card-label">Total hallazgos</div>'
            f'<div class="card-value">{r["total_hallazgos"]}</div></div>'
            f'<div class="card"><div class="card-label">Archivos afectados</div>'
            f'<div class="card-value">{r["archivos_afectados"]}</div></div>'
            f'<div class="card"><div class="card-label">Críticos</div>'
            f'<div class="card-value" style="color:#fc8181">{por_sev["CRITICO"]}</div></div>'
            f'<div class="card"><div class="card-label">Altos</div>'
            f'<div class="card-value" style="color:#f6ad55">{por_sev["ALTO"]}</div></div>'
            f'<div class="card"><div class="card-label">Medios / Bajos</div>'
            f'<div class="card-value" style="color:#fefcbf">'
            f'{por_sev["MEDIO"] + por_sev["BAJO"]}</div></div>'
            '</div>'
            '<div style="display:grid;grid-template-columns:1fr 2fr;gap:1rem;margin-top:1rem">'
            '<div>'
            '<h2>Por categoría</h2>'
            '<table><tr><th>Categoría</th><th>N</th></tr>'
            + (rows_cat or '<tr><td colspan=2>—</td></tr>') +
            '</table>'
            '</div>'
            '<div>'
            '<h2>Hallazgos (top 15)</h2>'
            '<table><tr><th>Sev.</th><th>Categoría</th><th>Archivo</th>'
            '<th>L#</th><th>Detalle</th></tr>'
            + (rows_hall or '<tr><td colspan=5>—</td></tr>') +
            '</table>'
            '</div>'
            '</div>'
            '</div>'
        )


    # ─── Sub-sección: histograma CC ───────────────────────────────────────────

    # ─── Sección HTML: Streamlit Cache Detector ─────────────────────
    def _html_seccion_st_cache(self) -> str:
        """Tarjeta del dashboard — Bloque 22 Streamlit Cache."""
        if not self.resultado_st_cache:
            return ""
        r_full = self.resultado_st_cache
        if not r_full.get("aplica"):
            return ""
        r = r_full["resumen"]
        hallazgos = r_full.get("hallazgos", [])
        if r["total_hallazgos"] == 0:
            return (
                '<div class="section">'
                '<h2>&#9889; BLOQUE 22 &mdash; Streamlit Cache Detector</h2>'
                '<div style="color:#68d391;padding:1rem">'
                '&#10003; Sin antipatrones de caché detectados.</div></div>'
            )
        c_sev = {"CRITICO": "#fc8181", "ALTO": "#f6ad55",
                 "MEDIO": "#fefcbf", "BAJO": "#a0aec0"}
        orden  = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}
        top    = sorted(hallazgos, key=lambda h: orden.get(h["severidad"], 3))[:12]
        rows = "".join(
            '<tr>'
            f'<td style="color:{c_sev.get(h["severidad"],"#a0aec0")}">{h["severidad"]}</td>'
            f'<td style="font-size:0.78rem;color:#a0aec0">{h["tipo"].replace("_"," ")}</td>'
            f'<td class="file-cell" title="{h["archivo"]}">{h["archivo"]}</td>'
            f'<td style="color:#a0aec0">{h["linea"]}</td>'
            f'<td class="detail-cell">{h["detalle"]}</td>'
            '</tr>'
            for h in top
        )
        return (
            '<div class="section">'
            '<h2>&#9889; BLOQUE 22 &mdash; Streamlit Cache Detector</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Total hallazgos</div>'
            f'<div class="card-value">{r["total_hallazgos"]}</div></div>'
            f'<div class="card"><div class="card-label">Sin @cache</div>'
            f'<div class="card-value" style="color:#f6ad55">{r["sin_cache_n"]}</div></div>'
            f'<div class="card"><div class="card-label">Deprecados</div>'
            f'<div class="card-value" style="color:#fc8181">{r["deprecated_n"]}</div></div>'
            f'<div class="card"><div class="card-label">session_state issues</div>'
            f'<div class="card-value" style="color:#fefcbf">{r["session_state_n"]}</div></div>'
            '</div>'
            '<h2 style="margin-top:1rem">Hallazgos</h2>'
            '<table><tr><th>Sev.</th><th>Tipo</th><th>Archivo</th>'
            '<th>L#</th><th>Detalle</th></tr>'
            + (rows or '<tr><td colspan=5>Sin hallazgos</td></tr>')
            + '</table></div>'
        )

    # ─── Sección HTML: Dependency Scanner ──────────────────────────────────
    def _html_seccion_deps(self) -> str:
        """Tarjeta del dashboard — Bloque 24 Dependency Scanner."""
        if not self.resultado_deps:
            return ""
        r = self.resultado_deps
        res = r.get("resumen", {})
        alertas = r.get("alertas", [])
        if res.get("total_paquetes", 0) == 0:
            return (
                '<div class="section">'
                '<h2>&#128230; BLOQUE 24 &mdash; Dependency Scanner</h2>'
                '<div style="color:#a0aec0;padding:1rem">No se encontraron manifiestos de dependencias.</div>'
                '</div>'
            )
        c_sev = {"CRITICO": "#fc8181", "ALTO": "#f6ad55", "MEDIO": "#fefcbf", "BAJO": "#a0aec0"}
        orden = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}
        top = sorted(alertas, key=lambda a: orden.get(a.get("severidad", "BAJO"), 3))[:15]
        rows = "".join(
            '<tr>'
            f'<td style="color:{c_sev.get(a.get("severidad","BAJO"),"#a0aec0")}">{a.get("severidad","")}</td>'
            f'<td style="color:#e2e8f0">{a.get("paquete","")}</td>'
            f'<td style="color:#a0aec0;font-size:0.8rem">{a.get("version_instalada","")}</td>'
            f'<td class="detail-cell">{a.get("descripcion","")}</td>'
            '</tr>'
            for a in top
        )
        eco_counts = {}
        for pkg in r.get("inventario", []):
            e = pkg.get("ecosistema", "?")
            eco_counts[e] = eco_counts.get(e, 0) + 1
        eco_html = " &bull; ".join(f'{e}: {n}' for e, n in sorted(eco_counts.items()))
        return (
            '<div class="section">'
            '<h2>&#128230; BLOQUE 24 &mdash; Dependency Scanner</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Paquetes</div>'
            f'<div class="card-value">{res.get("total_paquetes",0)}</div></div>'
            f'<div class="card"><div class="card-label">Alertas</div>'
            f'<div class="card-value" style="color:#fc8181">{res.get("total_alertas",0)}</div></div>'
            f'<div class="card"><div class="card-label">Manifiestos</div>'
            f'<div class="card-value">{res.get("total_manifiestos",0)}</div></div>'
            f'<div class="card"><div class="card-label">Dev deps</div>'
            f'<div class="card-value" style="color:#a0aec0">{res.get("total_dev",0)}</div></div>'
            '</div>'
            + (f'<div style="color:#a0aec0;font-size:0.82rem;margin:.5rem 0">{eco_html}</div>' if eco_html else "")
            + ('<h2 style="margin-top:1rem">Alertas de seguridad</h2>'
               '<table><tr><th>Sev.</th><th>Paquete</th><th>Versión</th><th>Descripción</th></tr>'
               + (rows or '<tr><td colspan=4 style="color:#68d391">&#10003; Sin alertas de seguridad</td></tr>')
               + '</table>' if alertas else
               '<div style="color:#68d391;padding:.5rem 0">&#10003; Sin alertas de seguridad detectadas.</div>')
            + '</div>'
        )

    # ─── Sección HTML: Performance Profiler ─────────────────────────────────
    def _html_seccion_perf(self) -> str:
        """Tarjeta del dashboard — Bloque 25 Performance Profiler."""
        if not self.resultado_perf:
            return ""
        r = self.resultado_perf
        res = r.get("resumen", {})
        hallazgos = r.get("hallazgos", [])
        if res.get("total_hallazgos", 0) == 0:
            return (
                '<div class="section">'
                '<h2>&#9889; BLOQUE 25 &mdash; Performance Profiler</h2>'
                '<div style="color:#68d391;padding:1rem">&#10003; Sin antipatrones de rendimiento detectados.</div>'
                '</div>'
            )
        c_sev = {"CRITICO": "#fc8181", "ALTO": "#f6ad55", "MEDIO": "#fefcbf", "BAJO": "#a0aec0"}
        orden = {"CRITICO": 0, "ALTO": 1, "MEDIO": 2, "BAJO": 3}
        top = sorted(hallazgos, key=lambda h: orden.get(h.get("severidad","BAJO"), 3))[:15]
        rows = "".join(
            '<tr>'
            f'<td style="color:{c_sev.get(h.get("severidad","BAJO"),"#a0aec0")}">{h.get("severidad","")}</td>'
            f'<td style="font-size:0.78rem;color:#a0aec0">{h.get("tipo","").replace("_"," ")}</td>'
            f'<td class="file-cell" title="{h.get("archivo","")}">{h.get("archivo","")}</td>'
            f'<td style="color:#a0aec0">{h.get("linea","")}</td>'
            f'<td class="detail-cell">{h.get("detalle","")}</td>'
            '</tr>'
            for h in top
        )
        por_tipo = res.get("por_tipo", {})
        tipo_html = " &bull; ".join(
            f'{t.replace("_"," ")}: {n}' for t, n in sorted(por_tipo.items(), key=lambda x: -x[1])[:5]
        ) if por_tipo else ""
        return (
            '<div class="section">'
            '<h2>&#9889; BLOQUE 25 &mdash; Performance Profiler</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Hallazgos</div>'
            f'<div class="card-value" style="color:#f6ad55">{res.get("total_hallazgos",0)}</div></div>'
            f'<div class="card"><div class="card-label">Críticos</div>'
            f'<div class="card-value" style="color:#fc8181">{res.get("criticos",0)}</div></div>'
            f'<div class="card"><div class="card-label">Altos</div>'
            f'<div class="card-value" style="color:#f6ad55">{res.get("altos",0)}</div></div>'
            f'<div class="card"><div class="card-label">Archivos afectados</div>'
            f'<div class="card-value">{res.get("archivos_afectados",0)}</div></div>'
            '</div>'
            + (f'<div style="color:#a0aec0;font-size:0.82rem;margin:.5rem 0">{tipo_html}</div>' if tipo_html else "")
            + '<h2 style="margin-top:1rem">Hallazgos</h2>'
            '<table><tr><th>Sev.</th><th>Tipo</th><th>Archivo</th><th>L#</th><th>Detalle</th></tr>'
            + rows + '</table></div>'
        )

    # ─── Sección HTML: Documentation Score ──────────────────────────────────
    def _html_seccion_docs(self) -> str:
        """Tarjeta del dashboard — Bloque 26 Documentation Score."""
        if not self.resultado_docs:
            return ""
        r = self.resultado_docs
        score = r.get("score", 0)
        color_score = "#68d391" if score >= 70 else ("#f6ad55" if score >= 40 else "#fc8181")
        docstrings = r.get("docstrings", {})
        readme     = r.get("readme", {})
        comentarios= r.get("comentarios", {})
        hallazgos  = r.get("hallazgos", [])
        rows = "".join(
            '<tr>'
            f'<td style="color:#f6ad55">MEDIO</td>'
            f'<td class="file-cell">{h.get("archivo","")}</td>'
            f'<td class="detail-cell">{h.get("mensaje","")}</td>'
            '</tr>'
            for h in hallazgos[:12]
        )
        return (
            '<div class="section">'
            '<h2>&#128221; BLOQUE 26 &mdash; Documentation Score</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Score global</div>'
            f'<div class="card-value" style="color:{color_score}">{score}/100</div></div>'
            f'<div class="card"><div class="card-label">Docstrings</div>'
            f'<div class="card-value">{docstrings.get("score",0)}/100</div></div>'
            f'<div class="card"><div class="card-label">README</div>'
            f'<div class="card-value">{readme.get("score",0)}/100</div></div>'
            f'<div class="card"><div class="card-label">Comentarios</div>'
            f'<div class="card-value">{comentarios.get("score",0)}/100</div></div>'
            '</div>'
            + (f'<div style="color:#a0aec0;font-size:0.82rem;margin:.4rem 0">'
               f'Cobertura docstrings: {docstrings.get("cobertura_pct",0)}% &bull; '
               f'Ratio comentarios/código: {comentarios.get("ratio_pct",0)}%'
               '</div>' )
            + ('<h2 style="margin-top:1rem">Oportunidades de mejora</h2>'
               '<table><tr><th>Sev.</th><th>Archivo</th><th>Detalle</th></tr>'
               + (rows or '<tr><td colspan=3 style="color:#68d391">&#10003; Documentación completa</td></tr>')
               + '</table>' if hallazgos else
               '<div style="color:#68d391;padding:.5rem 0">&#10003; Documentación completa.</div>')
            + '</div>'
        )

    # ─── Sección HTML: Test Coverage Detector ───────────────────────────────
    def _html_seccion_tests(self) -> str:
        """Tarjeta del dashboard — Bloque 27 Test Coverage Detector."""
        if not self.resultado_tests:
            return ""
        r = self.resultado_tests
        res = r.get("resumen", {})
        frameworks = r.get("frameworks_detectados", [])
        criticas_sin_test = r.get("funciones_criticas_sin_test", [])
        ratio = res.get("ratio_archivos_test_pct", 0)
        color_ratio = "#68d391" if ratio >= 20 else ("#f6ad55" if ratio >= 5 else "#fc8181")
        fw_html = ", ".join(frameworks) if frameworks else "Ninguno detectado"
        rows = "".join(
            '<tr>'
            f'<td style="color:#fc8181">ALTO</td>'
            f'<td class="file-cell">{fn.get("archivo","")}</td>'
            f'<td style="color:#e2e8f0">{fn.get("funcion","")}</td>'
            '</tr>'
            for fn in criticas_sin_test[:10]
        )
        return (
            '<div class="section">'
            '<h2>&#129514; BLOQUE 27 &mdash; Test Coverage Detector</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Archivos test / total</div>'
            f'<div class="card-value" style="color:{color_ratio}">'
            f'{res.get("archivos_test",0)} / {res.get("total_archivos",0)}</div></div>'
            f'<div class="card"><div class="card-label">% archivos son tests</div>'
            f'<div class="card-value" style="color:{color_ratio}">{ratio}%</div></div>'
            f'<div class="card"><div class="card-label">Funciones de test</div>'
            f'<div class="card-value">{res.get("total_funciones_test",0)}</div></div>'
            f'<div class="card"><div class="card-label">Coverage config</div>'
            f'<div class="card-value" style="color:{"#68d391" if res.get("tiene_coverage_config") else "#fc8181"}">'
            f'{"&#10003;" if res.get("tiene_coverage_config") else "&#10007;"}</div></div>'
            '</div>'
            f'<div style="color:#a0aec0;font-size:0.82rem;margin:.4rem 0">'
            f'Frameworks: {fw_html}</div>'
            + ('<h2 style="margin-top:1rem">Funciones críticas sin test</h2>'
               '<table><tr><th>Sev.</th><th>Archivo</th><th>Función</th></tr>'
               + (rows or '<tr><td colspan=3 style="color:#68d391">&#10003; Todas las funciones críticas tienen test</td></tr>')
               + '</table>' if criticas_sin_test else
               '<div style="color:#68d391;padding:.5rem 0">&#10003; No se detectaron funciones críticas sin cobertura.</div>')
            + '</div>'
        )

    # ─── Sección HTML: Pattern Intelligence Capa 3 ──────────────────────────
    def _html_seccion_capa3(self) -> str:
        """Tarjeta del dashboard — Pattern Intelligence Capa 3 (patrones externos)."""
        if not self.patron_capa3:
            return ""
        r = self.patron_capa3
        if not r.get("activo"):
            return ""
        langs   = r.get("lenguajes_cargados", [])
        total_p = r.get("total_patrones", 0)
        version = r.get("version", "—")
        fecha   = r.get("ultima_actualizacion", "—")
        if not langs or total_p == 0:
            return ""
        return (
            '<div class="section">'
            '<h2>&#127888; Pattern Intelligence Capa 3</h2>'
            '<div class="grid">'
            f'<div class="card"><div class="card-label">Patrones externos</div>'
            f'<div class="card-value" style="color:#9ae6b4">{total_p}</div></div>'
            f'<div class="card"><div class="card-label">Lenguajes</div>'
            f'<div class="card-value">{len(langs)}</div></div>'
            f'<div class="card"><div class="card-label">Versión</div>'
            f'<div class="card-value" style="color:#a0aec0;font-size:1rem">{version}</div></div>'
            f'<div class="card"><div class="card-label">Actualizado</div>'
            f'<div class="card-value" style="color:#a0aec0;font-size:0.9rem">{fecha[:10] if fecha != "—" else "—"}</div></div>'
            '</div>'
            f'<div style="color:#a0aec0;font-size:0.82rem;margin:.4rem 0">'
            f'Lenguajes: {", ".join(langs)} &bull; '
            f'Actualizar: <code>python3 stratum.py --update-patterns</code></div>'
            '</div>'
        )

    def _html_histograma_cc(self, r: dict) -> str:
        """
        Barra visual de distribución de complejidad ciclomática (CC).
        Rangos: 1-5 simple, 6-10 aceptable, 11-20 zona gris, 21-50 crítico, 50+.
        """
        histo = r.get("histograma_cc")
        if not histo:
            return ""
        rangos        = histo.get("rangos", {})
        top_complejas = histo.get("top_complejas", [])
        zona_gris_n   = histo.get("zona_gris_n", 0)
        criticas_n    = histo.get("criticas_n",  0)

        total_fn = sum(rangos.values()) or 1
        colores  = {
            "1-5":   ("#68d391", "Simple"),
            "6-10":  ("#9ae6b4", "Aceptable"),
            "11-20": ("#f6ad55", "Zona gris"),
            "21-50": ("#fc8181", "Crítico"),
            "50+":   ("#e53e3e", "Muy crítico"),
        }

        barras = ""
        for rango, n in rangos.items():
            color, etiqueta = colores.get(rango, ("#a0aec0", rango))
            pct = round(n / total_fn * 100)
            barras += (
                f'<div style="display:flex;align-items:center;gap:0.5rem;'
                f'margin-bottom:0.3rem">'
                f'<span style="width:4rem;font-size:0.75rem;color:#718096">{rango}</span>'
                f'<div style="flex:1;background:#2d3748;border-radius:3px;height:14px">'
                f'<div style="width:{pct}%;background:{color};height:14px;'
                f'border-radius:3px;min-width:{2 if n else 0}px"></div></div>'
                f'<span style="width:3rem;text-align:right;font-size:0.75rem;'
                f'color:{color}">{n}</span>'
                f'<span style="font-size:0.7rem;color:#4a5568">{etiqueta}</span>'
                f'</div>'
            )

        top_rows = "".join(
            f'<tr><td style="font-size:0.78rem">{fn["nombre"]}</td>'
            f'<td style="color:#fc8181;font-weight:600">{fn["cc"]}</td>'
            f'<td style="font-size:0.78rem">{fn["archivo"]}</td></tr>'
            for fn in top_complejas[:8]
        )

        alerta_zona = ""
        if zona_gris_n > 0:
            alerta_zona = (
                f'<p style="color:#f6ad55;font-size:0.78rem;margin:0.5rem 0 0">'
                f'&#9888; {zona_gris_n} funciones en zona gris (CC 11-20): '
                f'no son críticas con umbral Streamlit calibrado, '
                f'pero serían ALTO con McCabe=10 universal.</p>'
            )

        return (
            '<div style="margin-top:1rem">'
            '<h2>Distribución de Complejidad Ciclomática</h2>'
            + barras
            + alerta_zona
            + (
                '<h2 style="margin-top:0.75rem">Top funciones más complejas</h2>'
                '<table><tr><th>Función</th><th>CC</th><th>Archivo</th></tr>'
                + top_rows + '</table>'
                if top_complejas else ""
            )
            + '</div>'
        )


# ============================================================================
# PUNTO DE ENTRADA PRINCIPAL
# ============================================================================

def main() -> None:
    """STRATUM System Intelligence Engine — punto de entrada CLI."""
    parser = argparse.ArgumentParser(
        prog="stratum",
        description=f"STRATUM System Intelligence Engine v{STRATUM_VERSION}",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "ruta", nargs="?", default=".",
        help="Ruta al proyecto a analizar (default: directorio actual)",
    )
    parser.add_argument(
        "--puerto", "-p", type=int, default=8501,
        help="Puerto del servidor web (default: 8501)",
    )
    parser.add_argument(
        "--red", "-r", action="store_true",
        help="Exponer en red local (default: solo localhost)",
    )
    parser.add_argument(
        "--hash", action="store_true",
        help="Mostrar SHA-256 del archivo para verificación de integridad",
    )
    parser.add_argument(
        "--version", "-v", action="version",
        version=f"STRATUM v{STRATUM_VERSION}",
    )
    parser.add_argument(
        "--self-test", action="store_true",
        help="Ejecutar suite de tests unitarios internos de STRATUM",
    )
    parser.add_argument(
        "--update-patterns", action="store_true",
        help="Descargar/actualizar patrones externos (requiere internet)",
    )

    args = parser.parse_args()

    if args.hash:
        seg      = SeguridadStratum()
        resultado = seg.verificar_integridad()
        print(f"\n  SHA-256: {resultado['hash_actual']}")
        print("  Comparte este hash con tus clientes para verificar integridad.\n")
        return

    if getattr(args, "self_test", False):
        import sys as _sys
        _sys.exit(_ejecutar_self_test())

    if getattr(args, "update_patterns", False):
        from pathlib import Path as _Path
        _capa3 = PatternIntelligenceCapa3(_Path("."))
        _res   = _capa3.actualizar()
        print(f"\n  🌐 Patrones actualizados: {_res.get('actualizados', 0)} lenguajes")
        print(f"  Ruta: {_res.get('ruta', '~/.stratum/patterns')}\n")
        return

    app = StratumApp(
        ruta_proyecto=args.ruta,
        puerto=args.puerto,
        modo_red=args.red,
    )
    app._mostrar_banner()
    app.ejecutar_analisis()
    app.iniciar_servidor()



# ════════════════════════════════════════════════════════════════════════════
# BLOQUE 27.5 — STRATUM SELF-TEST SUITE
# Suite de tests unitarios para el propio motor STRATUM.
# Ejecutar con: python3 stratum.py --self-test
# Sin dependencias externas (stdlib unittest únicamente).
# ════════════════════════════════════════════════════════════════════════════

class _StratumTestBase(unittest.TestCase):
    """Clase base con helpers comunes a todos los test cases."""

    # ─── Crear estructura de fixtures en directorio temporal ────────────────
    @staticmethod
    def _make_fixture_dir(archivos: dict) -> str:
        """
        Crea un directorio temporal con los archivos dados.
        archivos: {nombre_relativo: contenido_str}
        Retorna la ruta absoluta del directorio creado.
        """
        tmpdir = tempfile.mkdtemp(prefix="stratum_test_")
        for nombre, contenido in archivos.items():
            ruta = os.path.join(tmpdir, nombre)
            os.makedirs(os.path.dirname(ruta), exist_ok=True)
            with open(ruta, "w", encoding="utf-8") as f:
                f.write(contenido)
        return tmpdir


# ════════════════════════════════════════════════════════════════════════════
# TEST: ScannerUniversal
# ════════════════════════════════════════════════════════════════════════════

class TestScannerUniversal(_StratumTestBase):
    """Tests para el módulo ScannerUniversal (Bloque 0)."""

    def test_detecta_lenguaje_python(self):
        tmpdir = self._make_fixture_dir({"app.py": "x = 1\n"})
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        lenguajes = {a["lenguaje"] for a in resultado["archivos"]}
        self.assertIn("Python", lenguajes)

    def test_detecta_lenguaje_javascript(self):
        tmpdir = self._make_fixture_dir({"index.js": "const x = 1;\n"})
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        lenguajes = {a["lenguaje"] for a in resultado["archivos"]}
        self.assertIn("JavaScript", lenguajes)

    def test_excluye_node_modules(self):
        tmpdir = self._make_fixture_dir({
            "src/app.py": "x = 1\n",
            "node_modules/lib.js": "const y = 2;\n",
        })
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        rutas = [a["ruta_rel"] for a in resultado["archivos"]]
        self.assertFalse(any("node_modules" in r for r in rutas))

    def test_excluye_venv(self):
        tmpdir = self._make_fixture_dir({
            "main.py": "pass\n",
            "venv/lib/python3.10/site-packages/pkg.py": "pass\n",
        })
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        rutas = [a["ruta_rel"] for a in resultado["archivos"]]
        self.assertFalse(any("venv" in r for r in rutas))

    def test_estadisticas_basicas(self):
        tmpdir = self._make_fixture_dir({
            "a.py": "x = 1\ny = 2\n",
            "b.py": "z = 3\n",
        })
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        self.assertEqual(resultado["estadisticas"]["total_archivos"], 2)
        self.assertGreater(resultado["estadisticas"]["total_lineas"], 0)

    def test_directorio_vacio(self):
        import tempfile
        tmpdir = tempfile.mkdtemp(prefix="stratum_empty_")
        scanner = ScannerUniversal(tmpdir)
        resultado = scanner.escanear()
        self.assertEqual(resultado["estadisticas"]["total_archivos"], 0)


# ════════════════════════════════════════════════════════════════════════════
# TEST: ASTEngine (Bloque 5)
# ════════════════════════════════════════════════════════════════════════════

class TestASTEngine(_StratumTestBase):
    """Tests para el motor AST de Python."""

    def _get_archivos(self, tmpdir):
        scanner = ScannerUniversal(tmpdir)
        return scanner.escanear()["archivos"]

    def test_detecta_funciones(self):
        tmpdir = self._make_fixture_dir({
            "calc.py": "def suma(a, b):\n    return a + b\n\ndef resta(a, b):\n    return a - b\n"
        })
        archivos = self._get_archivos(tmpdir)
        resultado = ASTEnginePython(archivos).analizar()
        self.assertEqual(resultado["resumen"]["total_funciones"], 2)

    def test_detecta_clases(self):
        tmpdir = self._make_fixture_dir({
            "models.py": "class Usuario:\n    pass\n\nclass Producto:\n    pass\n"
        })
        archivos = self._get_archivos(tmpdir)
        resultado = ASTEnginePython(archivos).analizar()
        self.assertEqual(resultado["resumen"]["total_clases"], 2)

    def test_complejidad_ciclomatica_simple(self):
        """Una función sin branches tiene CC = 1."""
        tmpdir = self._make_fixture_dir({
            "simple.py": "def saludo(nombre):\n    return f'Hola {nombre}'\n"
        })
        archivos = self._get_archivos(tmpdir)
        resultado = ASTEnginePython(archivos).analizar()
        fns = resultado.get("funciones", [])
        self.assertTrue(any(f["cc"] == 1 for f in fns))

    def test_complejidad_ciclomatica_con_branches(self):
        """Función con if/elif/else sube la CC."""
        codigo = (
            "def clasificar(x):\n"
            "    if x > 100:\n"
            "        return 'grande'\n"
            "    elif x > 50:\n"
            "        return 'mediano'\n"
            "    elif x > 10:\n"
            "        return 'pequeño'\n"
            "    else:\n"
            "        return 'micro'\n"
        )
        tmpdir = self._make_fixture_dir({"classify.py": codigo})
        archivos = self._get_archivos(tmpdir)
        resultado = ASTEnginePython(archivos).analizar()
        fns = resultado.get("funciones", [])
        clasificar_fn = next((f for f in fns if f["nombre"] == "clasificar"), None)
        self.assertIsNotNone(clasificar_fn)
        self.assertGreater(clasificar_fn["cc"], 1)

    def test_detecta_funciones_sin_docstring(self):
        tmpdir = self._make_fixture_dir({
            "nodoc.py": "def sin_doc():\n    return 42\n"
        })
        archivos = self._get_archivos(tmpdir)
        resultado = ASTEnginePython(archivos).analizar()
        fns = resultado.get("funciones", [])
        sin_doc = [f for f in fns if not f.get("tiene_docstring")]
        self.assertTrue(len(sin_doc) >= 1)

    def test_archivo_con_syntax_error_no_rompe(self):
        """El engine no debe lanzar excepción ante código inválido."""
        tmpdir = self._make_fixture_dir({
            "broken.py": "def foo(:\n    pass\n"
        })
        archivos = self._get_archivos(tmpdir)
        try:
            resultado = ASTEnginePython(archivos).analizar()
            # Debe retornar resultado vacío/parcial, no lanzar
            self.assertIsInstance(resultado, dict)
        except Exception as e:
            self.fail(f"ASTEnginePython lanzó excepción ante código roto: {e}")

    def test_no_lee_contenido_via_scanner(self):
        """Los archivos del scanner solo tienen metadatos, no contenido."""
        tmpdir = self._make_fixture_dir({"check.py": "x = 1\n"})
        archivos = self._get_archivos(tmpdir)
        for arch in archivos:
            self.assertNotIn("contenido", arch,
                "Scanner no debe cargar contenido de archivos")


# ════════════════════════════════════════════════════════════════════════════
# TEST: SecurityAuditor (Bloque 23)
# ════════════════════════════════════════════════════════════════════════════

class TestSecurityAuditor(_StratumTestBase):
    """Tests para el auditor de seguridad."""

    def _run(self, archivos_fixture: dict) -> dict:
        tmpdir = self._make_fixture_dir(archivos_fixture)
        scanner = ScannerUniversal(tmpdir)
        archivos = scanner.escanear()["archivos"]
        return SecurityAuditor(archivos, scanner).analizar()

    def test_detecta_sql_injection(self):
        codigo = 'query = "SELECT * FROM users WHERE id = " + user_id\n'
        resultado = self._run({"dao.py": codigo})
        tipos = [h["tipo"] for h in resultado.get("hallazgos", [])]
        self.assertTrue(any("sql" in t.lower() or "injection" in t.lower() for t in tipos),
                        f"No detectó SQL injection. Tipos hallados: {tipos}")

    def test_detecta_pii_email(self):
        codigo = 'email = "usuario@ejemplo.com"\nprint(email)\n'
        resultado = self._run({"user.py": codigo})
        tipos = [h["tipo"] for h in resultado.get("hallazgos", [])]
        self.assertTrue(
            any("pii" in t.lower() or "email" in t.lower() for t in tipos),
            f"No detectó PII email. Tipos: {tipos}"
        )

    def test_detecta_hardcoded_secret(self):
        codigo = 'SECRET_KEY = "s3cr3t_k3y_hardcoded_12345"\nAPI_KEY = "abcdef1234567890"\n'
        resultado = self._run({"config.py": codigo})
        tipos = [h["tipo"] for h in resultado.get("hallazgos", [])]
        self.assertTrue(
            any("secret" in t.lower() or "key" in t.lower() or "hardcoded" in t.lower()
                for t in tipos),
            f"No detectó secret hardcoded. Tipos: {tipos}"
        )

    def test_codigo_limpio_sin_hallazgos_criticos(self):
        codigo = (
            "import os\n\n"
            "def obtener_usuario(user_id: int) -> dict:\n"
            "    \"\"\"Obtiene usuario por ID.\"\"\"\n"
            "    return {\"id\": user_id, \"nombre\": \"Test\"}\n"
        )
        resultado = self._run({"clean.py": codigo})
        criticos = [h for h in resultado.get("hallazgos", [])
                    if h.get("severidad") == "CRITICO"]
        self.assertEqual(len(criticos), 0,
            f"Código limpio no debería tener hallazgos CRÍTICOS: {criticos}")

    def test_resultado_tiene_resumen(self):
        resultado = self._run({"empty.py": "pass\n"})
        self.assertIn("resumen", resultado)
        self.assertIn("total_hallazgos", resultado["resumen"])


# ════════════════════════════════════════════════════════════════════════════
# TEST: StratumScore (Bloque 18)
# ════════════════════════════════════════════════════════════════════════════

class TestStratumScore(unittest.TestCase):
    """Tests para el sistema de scoring ponderado."""

    def _score_vacio(self):
        """Score con todos los módulos vacíos → debe retornar 100."""
        return StratumScore({}).calcular()

    def _score_con_findings(self, n_criticos=0, n_altos=0, n_medios=0, n_bajos=0):
        """Construye findings sintéticos para probar el scoring."""
        findings = []
        for _ in range(n_criticos):
            findings.append({"severidad": "CRITICO", "tipo": "test_critico",
                             "archivo": "a.py", "linea": 1, "detalle": "x"})
        for _ in range(n_altos):
            findings.append({"severidad": "ALTO", "tipo": "test_alto",
                             "archivo": "a.py", "linea": 1, "detalle": "x"})
        for _ in range(n_medios):
            findings.append({"severidad": "MEDIO", "tipo": "test_medio",
                             "archivo": "a.py", "linea": 1, "detalle": "x"})
        for _ in range(n_bajos):
            findings.append({"severidad": "BAJO", "tipo": "test_bajo",
                             "archivo": "a.py", "linea": 1, "detalle": "x"})
        # Inyectamos findings via "security" que es procesado por _recolectar_findings
        datos = {"security": {"hallazgos": findings}}
        return StratumScore(datos).calcular()

    def test_score_sin_findings_es_100(self):
        resultado = self._score_vacio()
        self.assertEqual(resultado["score"], 100)

    def test_score_range_0_100(self):
        resultado = self._score_con_findings(n_criticos=50, n_altos=100)
        self.assertGreaterEqual(resultado["score"], 0)
        self.assertLessEqual(resultado["score"], 100)

    def test_criticos_reducen_mas_que_bajos(self):
        score_critico = self._score_con_findings(n_criticos=1)["score"]
        score_bajo    = self._score_con_findings(n_bajos=1)["score"]
        self.assertLess(score_critico, score_bajo,
            "Un hallazgo CRÍTICO debe reducir más el score que un BAJO")

    def test_badge_platinum_en_100(self):
        resultado = self._score_vacio()
        self.assertEqual(resultado["badge"], "EXCELENTE",
            f"Score 100 debería dar badge EXCELENTE, obtuvo: {resultado['badge']}")

    def test_badge_f_en_score_bajo(self):
        resultado = self._score_con_findings(n_criticos=200)
        self.assertIn(resultado["badge"], ["CRÍTICO", "DEFICIENTE", "ATENCIÓN"],
            f"200 hallazgos CRÍTICOS debería dar badge bajo, obtuvo: {resultado['badge']}")

    def test_resultado_tiene_keys_obligatorias(self):
        resultado = self._score_vacio()
        for key in ["score", "badge", "badge_icon", "total_findings", "findings"]:
            self.assertIn(key, resultado, f"Falta key '{key}' en resultado de StratumScore")

    def test_pen_max_dinamico(self):
        """PEN_MAX debe ser max(200, (total_funciones + total_metodos) × 8)."""
        # Con 100 funciones → PEN_MAX ≥ 800
        datos = {
            "ast": {
                "resumen": {"total_funciones": 100, "total_metodos": 0},
                "hallazgos": [],
            }
        }
        resultado = StratumScore(datos).calcular()
        # Solo verificamos que no explota y retorna score válido
        self.assertGreaterEqual(resultado["score"], 0)
        self.assertLessEqual(resultado["score"], 100)


# ════════════════════════════════════════════════════════════════════════════
# TEST: SQLParser (Bloque 7)
# ════════════════════════════════════════════════════════════════════════════

class TestSQLParser(_StratumTestBase):
    """Tests para el parser SQL multi-dialecto."""

    def _parse(self, sql_content: str) -> dict:
        tmpdir = self._make_fixture_dir({"schema.sql": sql_content})
        scanner = ScannerUniversal(tmpdir)
        archivos = scanner.escanear()["archivos"]
        return SQLParser(archivos).analizar()

    def test_detecta_create_table(self):
        sql = "CREATE TABLE usuarios (\n  id INTEGER PRIMARY KEY,\n  nombre TEXT\n);\n"
        resultado = self._parse(sql)
        self.assertIn("usuarios", resultado["tablas"])

    def test_detecta_multiples_tablas(self):
        sql = (
            "CREATE TABLE clientes (id INT);\n"
            "CREATE TABLE pedidos (id INT, cliente_id INT);\n"
            "CREATE TABLE productos (id INT, nombre TEXT);\n"
        )
        resultado = self._parse(sql)
        self.assertEqual(len(resultado["tablas"]), 3)

    def test_detecta_foreign_keys(self):
        sql = (
            "CREATE TABLE pedidos (\n"
            "  id INT PRIMARY KEY,\n"
            "  cliente_id INT,\n"
            "  FOREIGN KEY (cliente_id) REFERENCES clientes(id)\n"
            ");\n"
        )
        resultado = self._parse(sql)
        tablas = resultado["tablas"]
        if "pedidos" in tablas:
            fks = tablas["pedidos"].get("foreign_keys", [])
            self.assertTrue(len(fks) >= 1, "Debe detectar al menos 1 FK")

    def test_sin_tablas_no_falla(self):
        resultado = self._parse("-- Solo comentarios\nSELECT 1;\n")
        self.assertIsInstance(resultado, dict)
        self.assertIn("tablas", resultado)

    def test_resultado_tiene_resumen(self):
        resultado = self._parse("CREATE TABLE test (id INT);\n")
        self.assertIn("resumen", resultado)


# ════════════════════════════════════════════════════════════════════════════
# TEST: DependencyScanner (Bloque 24)
# ════════════════════════════════════════════════════════════════════════════

class TestDependencyScanner(_StratumTestBase):
    """Tests para el scanner de dependencias."""

    def _scan(self, archivos_fixture: dict) -> dict:
        tmpdir = self._make_fixture_dir(archivos_fixture)
        scanner = ScannerUniversal(tmpdir)
        archivos = scanner.escanear()["archivos"]
        return DependencyScanner(tmpdir, archivos).analizar()

    def test_parsea_requirements_txt(self):
        resultado = self._scan({
            "requirements.txt": "flask==2.3.0\nrequests>=2.28.0\nnumpy\n"
        })
        nombres = [p["nombre"] for p in resultado.get("inventario", [])]
        self.assertIn("flask", nombres)
        self.assertIn("requests", nombres)

    def test_parsea_package_json(self):
        import json
        pkg = json.dumps({
            "dependencies": {"express": "^4.18.0", "axios": "^1.0.0"},
            "devDependencies": {"jest": "^29.0.0"}
        })
        resultado = self._scan({"package.json": pkg})
        nombres = [p["nombre"] for p in resultado.get("inventario", [])]
        self.assertIn("express", nombres)

    def test_detecta_pyyaml_vulnerable(self):
        resultado = self._scan({
            "requirements.txt": "pyyaml==5.1.0\n"
        })
        alertas = resultado.get("alertas", [])
        paquetes_alerta = [a["paquete"].lower() for a in alertas]
        self.assertIn("pyyaml", paquetes_alerta,
            "pyyaml 5.1.0 debería generar alerta de seguridad (CVE-2020-14343)")

    def test_sin_manifiestos_retorna_vacio(self):
        resultado = self._scan({"app.py": "import os\n"})
        self.assertEqual(resultado["resumen"]["total_manifiestos"], 0)
        self.assertEqual(resultado["resumen"]["total_paquetes"], 0)

    def test_resumen_tiene_keys(self):
        resultado = self._scan({"requirements.txt": "flask==2.3.0\n"})
        for key in ["total_manifiestos", "total_paquetes", "total_alertas"]:
            self.assertIn(key, resultado["resumen"])


# ════════════════════════════════════════════════════════════════════════════
# TEST: DocumentationScore (Bloque 26)
# ════════════════════════════════════════════════════════════════════════════

class TestDocumentationScore(_StratumTestBase):
    """Tests para el evaluador de documentación."""

    def _score(self, archivos_fixture: dict) -> dict:
        tmpdir = self._make_fixture_dir(archivos_fixture)
        scanner = ScannerUniversal(tmpdir)
        archivos = scanner.escanear()["archivos"]
        resultado_ast = ASTEnginePython(archivos).analizar()
        return DocumentationScore(resultado_ast, archivos, tmpdir).analizar()

    def test_score_range_0_100(self):
        resultado = self._score({"undoc.py": "def foo():\n    pass\n"})
        self.assertGreaterEqual(resultado["score"], 0)
        self.assertLessEqual(resultado["score"], 100)

    def test_con_docstrings_score_mayor(self):
        sin_doc = self._score({"nodoc.py": "def foo():\n    pass\n"})
        con_doc = self._score({
            "doc.py": '"""Módulo documentado.\"\"\"\n\ndef foo():\n    """Retorna nada.\"\"\"\n    pass\n'
        })
        self.assertGreaterEqual(con_doc["score"], sin_doc["score"],
            "Módulo con docstrings debe tener score ≥ módulo sin docstrings")

    def test_readme_detectado(self):
        resultado = self._score({
            "README.md": "# Mi Proyecto\n\nEste es el README.\n",
            "app.py": "pass\n",
        })
        self.assertGreater(resultado["readme"]["score"], 0,
            "Presencia de README.md debe aumentar el score")

    def test_resultado_tiene_keys(self):
        resultado = self._score({"x.py": "pass\n"})
        for key in ["score", "docstrings", "readme", "comentarios"]:
            self.assertIn(key, resultado)


# ════════════════════════════════════════════════════════════════════════════
# TEST: TestCoverageDetector (Bloque 27)
# ════════════════════════════════════════════════════════════════════════════

class TestTestCoverageDetector(_StratumTestBase):
    """Tests para el detector de cobertura de tests."""

    def _detect(self, archivos_fixture: dict) -> dict:
        tmpdir = self._make_fixture_dir(archivos_fixture)
        scanner = ScannerUniversal(tmpdir)
        archivos = scanner.escanear()["archivos"]
        resultado_ast = ASTEnginePython(archivos).analizar()
        return TestCoverageDetector(archivos, resultado_ast, tmpdir).analizar()

    def test_detecta_archivos_test_pytest(self):
        resultado = self._detect({
            "app.py": "def suma(a, b):\n    return a + b\n",
            "test_app.py": "def test_suma():\n    assert suma(1, 2) == 3\n",
        })
        self.assertGreater(resultado["resumen"]["archivos_test"], 0)

    def test_cuenta_funciones_test(self):
        resultado = self._detect({
            "tests/test_calc.py": (
                "def test_suma():\n    pass\n"
                "def test_resta():\n    pass\n"
                "def test_mult():\n    pass\n"
            )
        })
        self.assertEqual(resultado["resumen"]["total_funciones_test"], 3)

    def test_detecta_framework_pytest(self):
        resultado = self._detect({
            "test_sample.py": "import pytest\n\ndef test_foo():\n    assert True\n"
        })
        self.assertIn("pytest", resultado.get("frameworks_detectados", []))

    def test_sin_tests_ratio_cero(self):
        resultado = self._detect({
            "app.py": "def foo():\n    pass\n",
            "utils.py": "def bar():\n    pass\n",
        })
        self.assertEqual(resultado["resumen"]["archivos_test"], 0)

    def test_detecta_coverage_config(self):
        resultado = self._detect({
            ".coveragerc": "[run]\nsource = src\n",
            "app.py": "pass\n",
        })
        self.assertTrue(resultado["resumen"]["tiene_coverage_config"],
            ".coveragerc debe activar tiene_coverage_config")

    def test_resumen_tiene_keys(self):
        resultado = self._detect({"x.py": "pass\n"})
        for key in ["archivos_test", "total_archivos", "total_funciones_test",
                    "ratio_archivos_test_pct", "tiene_coverage_config"]:
            self.assertIn(key, resultado["resumen"])


# ════════════════════════════════════════════════════════════════════════════
# RUNNER INTERNO — ejecutado por --self-test
# ════════════════════════════════════════════════════════════════════════════

def _ejecutar_self_test() -> int:
    """
    Corre la suite completa de unittest y retorna el exit code.
    0 = todos los tests pasaron; 1 = hubo fallos.
    """
    suite = unittest.TestSuite()
    loader = unittest.TestLoader()

    clases_test = [
        TestScannerUniversal,
        TestASTEngine,
        TestSecurityAuditor,
        TestStratumScore,
        TestSQLParser,
        TestDependencyScanner,
        TestDocumentationScore,
        TestTestCoverageDetector,
    ]

    for cls in clases_test:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    print("\n  🧪 STRATUM Self-Test Suite")
    print("  " + "─" * 54)

    runner = unittest.TextTestRunner(
        verbosity=2,
        stream=__import__("sys").stdout,
        descriptions=True,
        failfast=False,
    )
    resultado = runner.run(suite)

    n_ok   = resultado.testsRun - len(resultado.failures) - len(resultado.errors)
    n_fail = len(resultado.failures) + len(resultado.errors)

    print("\n  " + "─" * 54)
    print(f"  Tests ejecutados : {resultado.testsRun}")
    print(f"  ✅ Pasados        : {n_ok}")
    if n_fail:
        print(f"  ❌ Fallados       : {n_fail}")

    return 0 if resultado.wasSuccessful() else 1



if __name__ == "__main__":
    main()
